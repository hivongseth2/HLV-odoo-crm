from odoo import models, api, fields
from odoo.exceptions import UserError
import io
import base64


class HlvStockQuick(models.TransientModel):
    _name = "hlv.stock.quick"
    _description = "Xem ton kho theo nhom"

    @api.model
    def get_data(self, group_id, warehouse_ids, show_zero, include_outgoing=True, extra_cols=None):
        if not group_id:
            return {"lines": [], "total": 0.0, "outgoing_total": 0.0, "columns": []}
        extra_cols = extra_cols or []
        group = self.env["hlv.product.report.group"].browse(group_id)
        if warehouse_ids:
            warehouses = self.env["stock.warehouse"].browse(warehouse_ids)
            columns = [{"id": wh.id, "name": wh.name} for wh in warehouses]
        else:
            warehouses = []
            columns = []
        # Pre-compute outgoing location ids per warehouse (pack zone + output zone)
        wh_outgoing_locs = {}
        if include_outgoing and warehouses:
            for wh in warehouses:
                ids = []
                for loc in [wh.wh_pack_stock_loc_id, wh.wh_output_stock_loc_id]:
                    if loc:
                        children = self.env["stock.location"].search([("id", "child_of", loc.id)])
                        ids.extend(children.ids)
                wh_outgoing_locs[wh.id] = ids
        # Pre-compute extra column data
        product_ids_list = [p.id for p in group.product_ids]
        extra_data = {}
        _direct_price_fields = {
            "sale_price": "lst_price",
            "price_web": "x_studio_ga_web",
            "price_listed": "x_studio_ga_hng_nim_yt",
            "price_tmdt": "x_studio_gia_san_tmdt",
            "price_commercial": "x_studio_gi_bn_thng_mi",
        }
        _direct_keys = [k for k in extra_cols if k in _direct_price_fields]
        if _direct_keys:
            for product in group.product_ids:
                tmpl = product.product_tmpl_id
                d = extra_data.setdefault(product.id, {})
                for key in _direct_keys:
                    fname = _direct_price_fields[key]
                    d[key] = getattr(tmpl, fname, None) or getattr(product, fname, None) or 0.0
        if "purchase_price" in extra_cols:
            po_lines = self.env["purchase.order.line"].search([
                ("product_id", "in", product_ids_list),
                ("order_id.state", "in", ["purchase", "done"]),
            ], order="id desc")
            seen_pp = set()
            for pl in po_lines:
                pid = pl.product_id.id
                if pid not in seen_pp:
                    extra_data.setdefault(pid, {})["purchase_price"] = pl.price_unit
                    seen_pp.add(pid)
        if "sales_cycle" in extra_cols:
            from datetime import datetime, timedelta
            from collections import defaultdict
            date_from = (datetime.now() - timedelta(days=90)).strftime("%Y-%m-%d %H:%M:%S")
            so_lines = self.env["sale.order.line"].search([
                ("product_id", "in", product_ids_list),
                ("order_id.state", "in", ["sale", "done"]),
                ("order_id.date_order", ">=", date_from),
            ])
            sale_order_sets = defaultdict(set)
            for sl in so_lines:
                sale_order_sets[sl.product_id.id].add(sl.order_id.id)
            for pid in product_ids_list:
                count = len(sale_order_sets.get(pid, set()))
                extra_data.setdefault(pid, {})["sales_cycle"] = round(90.0 / count, 1) if count > 0 else None
        if "avg_cost" in extra_cols:
            for product in group.product_ids:
                # standard_price is AVCO / manual cost on product.product (company-dependent)
                cost = product.with_company(self.env.company).standard_price
                extra_data.setdefault(product.id, {})["avg_cost"] = cost or 0.0
        if "incoming_qty" in extra_cols:
            # confirmed/assigned purchase moves not yet done
            in_moves = self.env["stock.move"].read_group(
                [
                    ("product_id", "in", product_ids_list),
                    ("state", "in", ["waiting", "confirmed", "assigned"]),
                    ("location_dest_id.usage", "=", "internal"),
                    ("location_id.usage", "!=", "internal"),
                ],
                ["product_id", "product_qty:sum"],
                ["product_id"],
            )
            for row in in_moves:
                pid = row["product_id"][0]
                extra_data.setdefault(pid, {})["incoming_qty"] = row["product_qty"]
        if "reserved_qty" in extra_cols:
            # confirmed/assigned sale moves going out, not yet done
            out_moves = self.env["stock.move"].read_group(
                [
                    ("product_id", "in", product_ids_list),
                    ("state", "in", ["waiting", "confirmed", "assigned"]),
                    ("location_id.usage", "=", "internal"),
                    ("location_dest_id.usage", "!=", "internal"),
                ],
                ["product_id", "product_qty:sum"],
                ["product_id"],
            )
            for row in out_moves:
                pid = row["product_id"][0]
                extra_data.setdefault(pid, {})["reserved_qty"] = row["product_qty"]
        lines = []
        total = 0.0
        outgoing_total = 0.0
        for product in group.product_ids.sorted("default_code"):
            if warehouses:
                col_qtys = []
                col_outgoing_qtys = []
                for wh in warehouses:
                    sq = product.with_context(location=wh.lot_stock_id.id).qty_available
                    col_qtys.append(sq)
                    oq = 0.0
                    if include_outgoing and wh_outgoing_locs.get(wh.id):
                        quants = self.env["stock.quant"].search([
                            ("product_id", "=", product.id),
                            ("location_id", "in", wh_outgoing_locs[wh.id]),
                            ("quantity", ">", 0),
                        ])
                        oq = sum(q.quantity for q in quants)
                    col_outgoing_qtys.append(oq)
                prod_total = sum(col_qtys)
                prod_outgoing = sum(col_outgoing_qtys)
            else:
                col_qtys = []
                col_outgoing_qtys = []
                prod_total = product.qty_available
                prod_outgoing = 0.0
            if not show_zero and prod_total == 0 and prod_outgoing == 0:
                continue
            total += prod_total
            outgoing_total += prod_outgoing
            line_extra = {key: extra_data.get(product.id, {}).get(key) for key in extra_cols}
            lines.append({
                "id": product.id,
                "code": product.default_code or "",
                "name": product.name,
                "uom": product.uom_id.name or "",
                "image_url": "/web/image/product.product/%d/image_128" % product.id,
                "col_qtys": col_qtys,
                "col_outgoing_qtys": col_outgoing_qtys,
                "total": prod_total,
                "outgoing_total": prod_outgoing,
                "extra": line_extra,
            })
        return {"lines": lines, "total": total, "outgoing_total": outgoing_total, "columns": columns}

    @api.model
    def export_excel(self, group_id, warehouse_ids, show_zero, include_outgoing=True, extra_cols=None):
        try:
            import xlsxwriter
        except ImportError:
            raise UserError("xlsxwriter chua duoc cai.")
        extra_cols = extra_cols or []
        data = self.get_data(group_id, warehouse_ids, show_zero, include_outgoing, extra_cols)
        columns = data["columns"]
        lines = data["lines"]
        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output, {"in_memory": True})
        ws = wb.add_worksheet("Ton kho")
        fh = wb.add_format({"bold": True, "bg_color": "#1a2639", "font_color": "#ffffff", "border": 1, "align": "center", "valign": "vcenter"})
        fc = wb.add_format({"border": 1, "font_name": "Courier New", "font_size": 9, "font_color": "#546e7a"})
        ft = wb.add_format({"border": 1})
        fn = wb.add_format({"border": 1, "num_format": "#,##0.##", "align": "right", "font_color": "#198754", "bold": True})
        f0 = wb.add_format({"border": 1, "num_format": "#,##0.##", "align": "right", "font_color": "#adb5bd"})
        fl = wb.add_format({"bold": True, "bg_color": "#e8f5e9", "font_color": "#155724", "border": 1, "align": "right"})
        fg = wb.add_format({"bold": True, "bg_color": "#e8f5e9", "font_color": "#198754", "border": 1, "num_format": "#,##0.##", "align": "right", "font_size": 12})
        fs = wb.add_format({"border": 1, "align": "center", "font_color": "#adb5bd"})
        fo = wb.add_format({"border": 1, "num_format": "#,##0.##", "align": "right", "font_color": "#e65100", "bold": True})
        fog = wb.add_format({"bold": True, "bg_color": "#fff8e1", "font_color": "#e65100", "border": 1, "num_format": "#,##0.##", "align": "right", "font_size": 12})
        n = len(columns)
        n_extra = len(extra_cols)
        has_outgoing = include_outgoing and n > 0
        extra_col_start = (5 + n + (1 if has_outgoing else 0)) if n else 5
        if n_extra:
            last_col = extra_col_start + n_extra - 1
        elif n:
            last_col = 4 + n + (1 if has_outgoing else 0)
        else:
            last_col = 4
        ws.merge_range(0, 0, 1, last_col, "B\u00c1O C\u00c1O T\u1ed2N KHO", wb.add_format({"bold": True, "font_size": 14, "align": "center", "valign": "vcenter"}))
        ws.set_row(0, 28)
        ws.set_row(1, 8)
        ws.set_column(0, 0, 5)
        ws.set_column(1, 1, 16)
        ws.set_column(2, 2, 45)
        ws.set_column(3, 3, 10)
        for i in range(n + 1):
            ws.set_column(4 + i, 4 + i, 16)
        if has_outgoing:
            ws.set_column(4 + n + 1, 4 + n + 1, 16)
        for j in range(n_extra):
            ws.set_column(extra_col_start + j, extra_col_start + j, 18)
        ws.set_row(2, 24)
        ws.write(2, 0, "#", fh)
        ws.write(2, 1, "M\u00e3 SP", fh)
        ws.write(2, 2, "T\u00ean s\u1ea3n ph\u1ea9m", fh)
        ws.write(2, 3, "\u0110VT", fh)
        if columns:
            for i, col in enumerate(columns):
                ws.write(2, 4 + i, col["name"], fh)
            ws.write(2, 4 + n, "T\u1ed4NG", fh)
            if has_outgoing:
                ws.write(2, 4 + n + 1, "\u0110\u00f3ng g\u00f3i/Out", fh)
        else:
            ws.write(2, 4, "T\u1ed3n kho", fh)
        _extra_labels = {
            "sale_price": "Gi\u00e1 b\u00e1n ",
            "price_web": "Gi\u00e1 Web",
            "price_listed": "Gi\u00e1 Ni\u00eam Y\u1ebft",
            "price_tmdt": "Gi\u00e1 S\u00e0n TM\u0110T",
            "price_commercial": "Gi\u00e1 Th\u01b0\u01a1ng M\u1ea1i",
            "purchase_price": "Gi\u00e1 mua",
            "sales_cycle": "Chu k\u1ef3 b\u00e1n (ng\u00e0y/\u0111\u01a1n)",
            "avg_cost": "Gi\u00e1 v\u1ed1n TB",
            "incoming_qty": "D\u1ef1 ki\u1ebfn nh\u1eadp",
            "reserved_qty": "D\u1ef1 ki\u1ebfn giao",
        }
        for j, ec in enumerate(extra_cols):
            ws.write(2, extra_col_start + j, _extra_labels.get(ec, ec), fh)
        f_money = wb.add_format({"border": 1, "num_format": "#,##0", "align": "right", "font_color": "#0d47a1"})
        f_cycle = wb.add_format({"border": 1, "num_format": "#,##0.0", "align": "right", "font_color": "#7b1fa2"})
        fl_extra = wb.add_format({"bg_color": "#e8f5e9", "border": 1})
        row = 3
        for idx, line in enumerate(lines):
            ws.write(row, 0, idx + 1, fs)
            ws.write(row, 1, line["code"], fc)
            ws.write(row, 2, line["name"], ft)
            ws.write(row, 3, line.get("uom", ""), ft)
            if columns:
                for i, qty in enumerate(line["col_qtys"]):
                    ws.write(row, 4 + i, qty, fn if qty > 0 else f0)
                ws.write(row, 4 + n, line["total"], fn if line["total"] > 0 else f0)
                if has_outgoing:
                    oqt = line.get("outgoing_total", 0)
                    ws.write(row, 4 + n + 1, oqt, fo if oqt > 0 else f0)
            else:
                ws.write(row, 4, line["total"], fn if line["total"] > 0 else f0)
            for j, ec in enumerate(extra_cols):
                val = line.get("extra", {}).get(ec)
                if val is None:
                    ws.write(row, extra_col_start + j, "-", ft)
                elif ec == "sales_cycle":
                    ws.write(row, extra_col_start + j, val, f_cycle)
                elif ec in ("incoming_qty", "reserved_qty"):
                    fq = wb.add_format({"border": 1, "num_format": "#,##0.##", "align": "right",
                                        "font_color": "#1565c0" if ec == "incoming_qty" else "#e65100", "bold": True})
                    ws.write(row, extra_col_start + j, val, fq if val > 0 else f0)
                else:
                    ws.write(row, extra_col_start + j, val, f_money)
            row += 1
        ws.merge_range(row, 0, row, 3, "T\u1ed4NG T\u1ed2N KHO", fl)
        if columns:
            for i in range(n):
                ct = sum(l["col_qtys"][i] for l in lines)
                ws.write(row, 4 + i, ct, fg)
            ws.write(row, 4 + n, data["total"], fg)
            if has_outgoing:
                ogt = data.get("outgoing_total", 0)
                ws.write(row, 4 + n + 1, ogt, fog if ogt > 0 else f0)
        else:
            ws.write(row, 4, data["total"], fg)
        for j in range(n_extra):
            ws.write(row, extra_col_start + j, "", fl_extra)
        wb.close()
        output.seek(0)
        att = self.env["ir.attachment"].create({
            "name": "ton_kho.xlsx",
            "type": "binary",
            "datas": base64.b64encode(output.read()).decode(),
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "res_model": self._name,
            "res_id": 0,
        })
        return att.id

    @api.model
    def get_product_cost_layers(self, product_id):
        """Return most recent PO-linked inbound layers that account for current on-hand qty."""
        product = self.env["product.product"].browse(product_id)
        on_hand_qty = product.qty_available

        # Newest first — accumulate until we reach on_hand_qty
        layers = self.env["stock.valuation.layer"].search(
            [
                ("product_id", "=", product_id),
                ("quantity", ">", 0),
                ("stock_move_id.purchase_line_id", "!=", False),
            ],
            order="create_date desc",
        )
        rows = []
        remaining = on_hand_qty
        company_currency = self.env.company.currency_id
        for lyr in layers:
            if remaining <= 0.001:
                break
            qty_take = min(lyr.quantity, remaining)
            remaining -= qty_take

            move = lyr.stock_move_id
            po_line = move.purchase_line_id if move else None
            picking = move.picking_id if move else None
            if po_line:
                currency = po_line.currency_id
                price_unit = po_line.price_unit
                if currency and currency != company_currency:
                    price_unit = currency._convert(
                        price_unit, company_currency,
                        self.env.company, lyr.create_date or fields.Date.today()
                    )
            else:
                price_unit = lyr.unit_cost
            rows.append({
                "date": lyr.create_date.strftime("%d/%m/%Y") if lyr.create_date else "",
                "reference": picking.name if picking else "",
                "po_name": po_line.order_id.name if po_line else "",
                "qty": qty_take,
                "unit_cost": price_unit,
                "value": round(qty_take * price_unit, 2),
                "uom": lyr.uom_id.name if lyr.uom_id else "",
            })

        # Reverse to show oldest → newest
        rows.reverse()
        total_qty = sum(r["qty"] for r in rows)
        total_value = sum(r["value"] for r in rows)
        computed_avg = total_value / total_qty if total_qty else 0.0
        return {
            "layers": rows,
            "total_qty": total_qty,
            "total_value": round(total_value, 2),
            "computed_avg": round(computed_avg, 2),
        }

    @api.model
    def get_product_pending_moves(self, product_id, key, warehouse_ids):
        """Return list of pending stock moves for incoming_qty or reserved_qty panel."""
        if key == "incoming_qty":
            domain = [
                ("product_id", "=", product_id),
                ("state", "in", ["waiting", "confirmed", "assigned"]),
                ("location_dest_id.usage", "=", "internal"),
                ("location_id.usage", "!=", "internal"),
            ]
        elif key == "reserved_qty":
            domain = [
                ("product_id", "=", product_id),
                ("state", "in", ["waiting", "confirmed", "assigned"]),
                ("location_id.usage", "=", "internal"),
                ("location_dest_id.usage", "!=", "internal"),
            ]
        else:
            return []
        state_labels = {"waiting": "Đang chờ", "confirmed": "Đã xác nhận", "assigned": "Sẵn sàng"}
        moves = self.env["stock.move"].search(domain, order="date asc", limit=100)
        result = []
        for m in moves:
            result.append({
                "picking_name": m.picking_id.name if m.picking_id else (m.name or ""),
                "origin": (m.picking_id.origin if m.picking_id else None) or m.origin or "",
                "state": state_labels.get(m.state, m.state),
                "qty": m.product_uom_qty,
                "uom": m.product_uom.name if m.product_uom else "",
                "date": m.date.strftime("%d/%m/%Y") if m.date else "",
                "partner": (m.picking_id.partner_id.name if m.picking_id and m.picking_id.partner_id else "") or "",
            })
        return result

    @api.model
    def get_group_products(self, group_id):
        group = self.env["hlv.product.report.group"].browse(group_id)
        result = []
        for p in group.product_ids.sorted("name"):
            result.append({"id": p.id, "name": p.name, "code": p.default_code or "", "image_url": "/web/image/product.product/%d/image_128" % p.id})
        return result

    @api.model
    def search_products(self, query, exclude_ids, offset=0):
        domain = [
            ("type", "in", ["consu", "product"]),
            "|",
            ("name", "ilike", query),
            ("default_code", "ilike", query),
            ("id", "not in", exclude_ids or []),
        ]
        total = self.env["product.product"].search_count(domain)
        products = self.env["product.product"].search(domain, limit=50, offset=offset, order="name")
        items = [{"id": p.id, "name": p.name, "code": p.default_code or "", "image_url": "/web/image/product.product/%d/image_128" % p.id} for p in products]
        return {"items": items, "total": total}

    @api.model
    def get_product_locations(self, product_id, warehouse_ids):
        outgoing_loc_ids = set()
        if warehouse_ids:
            warehouses = self.env["stock.warehouse"].browse(warehouse_ids)
            loc_ids = []
            for wh in warehouses:
                # Khu vuc stock chinh
                stock_locs = self.env["stock.location"].search([
                    ("id", "child_of", wh.lot_stock_id.id),
                    ("usage", "=", "internal"),
                ])
                loc_ids.extend(stock_locs.ids)
                # Khu vuc dong goi (pack zone) - chuan bi giao
                if wh.wh_pack_stock_loc_id:
                    pack_locs = self.env["stock.location"].search([
                        ("id", "child_of", wh.wh_pack_stock_loc_id.id),
                    ])
                    loc_ids.extend(pack_locs.ids)
                    outgoing_loc_ids.update(pack_locs.ids)
                # Khu vuc output - chuan bi giao
                if wh.wh_output_stock_loc_id:
                    out_locs = self.env["stock.location"].search([
                        ("id", "child_of", wh.wh_output_stock_loc_id.id),
                    ])
                    loc_ids.extend(out_locs.ids)
                    outgoing_loc_ids.update(out_locs.ids)
            loc_ids = list(set(loc_ids))
        else:
            locs = self.env["stock.location"].search([("usage", "=", "internal")])
            loc_ids = locs.ids
        quants = self.env["stock.quant"].search([
            ("product_id", "=", product_id),
            ("location_id", "in", loc_ids),
            ("quantity", ">", 0),
        ], order="quantity desc")
        result = []
        for q in quants:
            wh = q.location_id.warehouse_id
            result.append({
                "location": q.location_id.display_name,
                "warehouse": wh.name if wh else "",
                "qty": q.quantity,
                "outgoing": q.location_id.id in outgoing_loc_ids,
            })
        return result

    @api.model
    def import_products_from_excel(self, group_id, b64data):
        try:
            import openpyxl
        except ImportError:
            raise UserError("openpyxl ch\u01b0a \u0111\u01b0\u1ee3c c\u00e0i \u0111\u1eb7t.")
        import io as _io, base64 as _b64
        raw = _b64.b64decode(b64data)
        try:
            wb = openpyxl.load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as e:
            raise UserError("Kh\u00f4ng th\u1ec3 \u0111\u1ecdc file Excel: %s" % str(e))
        ws = wb.active
        codes = []
        skip_headers = {"m\u00e3 sp", "default_code", "ma sp", "code", "m\u00e3sp"}
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row and row[0] is not None:
                val = str(row[0]).strip()
                if val and val.lower() not in skip_headers:
                    codes.append(val)
        wb.close()
        if not codes:
            return {"added": [], "not_found": [], "already_in": [], "total": 0}
        group = self.env["hlv.product.report.group"].browse(group_id)
        existing_ids = set(group.product_ids.ids)
        added, not_found, already_in = [], [], []
        for code in codes:
            product = self.env["product.product"].search([("default_code", "=", code)], limit=1)
            if not product:
                not_found.append(code)
            elif product.id in existing_ids:
                already_in.append({"code": code, "name": product.name})
            else:
                group.write({"product_ids": [(4, product.id)]})
                existing_ids.add(product.id)
                added.append({"code": code, "name": product.name})
        return {"added": added, "not_found": not_found, "already_in": already_in, "total": len(codes)}

    def _detect_combo_for_move(self, move, sale_line):
        """Try to find the combo/kit parent for a move with price=0.
        Returns dict {name, code, price} or None.
        Tries sale_line.order_id first, then picking.sale_id as fallback.
        """
        try:
            BomLine = self.env.get("mrp.bom.line")
            if not BomLine:
                return None
            # Find kit BOMs that contain this product
            bom_lines = BomLine.search([
                ("product_id", "=", move.product_id.id),
                ("bom_id.type", "=", "phantom"),
            ], limit=10)
            if not bom_lines:
                return None
            kit_tmpl_ids = set(bom_lines.mapped("bom_id.product_tmpl_id").ids)

            # Resolve sale order: try sale_line first, then picking.sale_id
            order = False
            if sale_line:
                order = sale_line.order_id
            if not order:
                picking = move.picking_id
                if picking:
                    order = getattr(picking, "sale_id", False)

            if order:
                parent_line = order.order_line.filtered(
                    lambda l: l.product_id.product_tmpl_id.id in kit_tmpl_ids
                    and l.price_unit > 0
                )
                if parent_line:
                    pl = parent_line[0]
                    return {
                        "name": pl.product_id.name,
                        "code": pl.product_id.default_code or "",
                        "price": pl.price_unit,
                    }

            # BOM exists but can't find parent line — still mark as combo
            bom = bom_lines[0].bom_id
            return {
                "name": bom.product_tmpl_id.name,
                "code": bom.product_tmpl_id.default_code or "",
                "price": 0.0,
            }
        except Exception:
            pass
        return None

    @api.model
    def get_product_moves(self, product_id, warehouse_ids, date_from=None, date_to=None):
        from datetime import datetime, timedelta
        # UTC+7 (Asia/Ho_Chi_Minh)
        try:
            import pytz
            tz_vn = pytz.timezone("Asia/Ho_Chi_Minh")
            now_local = datetime.now(tz_vn)
        except ImportError:
            tz_vn = None
            now_local = datetime.utcnow() + timedelta(hours=7)

        if not date_from:
            date_from = now_local.strftime("%Y-%m-01")
        if not date_to:
            date_to = now_local.strftime("%Y-%m-%d")

        def local_to_utc_start(d_str):
            dt = datetime.strptime(d_str, "%Y-%m-%d")
            if tz_vn:
                import pytz as _pytz
                return tz_vn.localize(dt).astimezone(_pytz.utc).replace(tzinfo=None)
            return dt - timedelta(hours=7)

        def local_to_utc_end(d_str):
            dt = datetime.strptime(d_str + " 23:59:59", "%Y-%m-%d %H:%M:%S")
            if tz_vn:
                import pytz as _pytz
                return tz_vn.localize(dt).astimezone(_pytz.utc).replace(tzinfo=None)
            return dt - timedelta(hours=7)

        def utc_to_local_str(dt_utc):
            if not dt_utc:
                return ""
            if tz_vn:
                import pytz as _pytz
                dt = _pytz.utc.localize(dt_utc.replace(tzinfo=None)).astimezone(tz_vn)
            else:
                dt = dt_utc + timedelta(hours=7)
            return dt.strftime("%d/%m/%Y")

        date_from_utc = local_to_utc_start(date_from)
        date_to_utc = local_to_utc_end(date_to)
        date_from_str = date_from_utc.strftime("%Y-%m-%d %H:%M:%S")
        date_to_str = date_to_utc.strftime("%Y-%m-%d %H:%M:%S")

        # Warehouse stock locations — include pack + output zones so that
        # 3-step delivery moves (PICK stock→pack, PACK pack→output) are treated
        # as internal (both ends in set) and only OUT (output→customer) is counted.
        if warehouse_ids:
            warehouses = self.env["stock.warehouse"].browse(warehouse_ids)
            all_loc_ids = []
            for wh in warehouses:
                locs = self.env["stock.location"].search([
                    ("id", "child_of", wh.lot_stock_id.id),
                    ("usage", "=", "internal"),
                ])
                all_loc_ids.extend(locs.ids)
                if wh.wh_pack_stock_loc_id:
                    pack_locs = self.env["stock.location"].search([
                        ("id", "child_of", wh.wh_pack_stock_loc_id.id),
                    ])
                    all_loc_ids.extend(pack_locs.ids)
                if wh.wh_output_stock_loc_id:
                    out_locs = self.env["stock.location"].search([
                        ("id", "child_of", wh.wh_output_stock_loc_id.id),
                    ])
                    all_loc_ids.extend(out_locs.ids)
            all_loc_ids = list(set(all_loc_ids))
        else:
            all_loc_ids = self.env["stock.location"].search([("usage", "=", "internal")]).ids
        all_loc_set = set(all_loc_ids)

        product = self.env["product.product"].browse(product_id)

        # Opening balance using to_date context (tồn đầu kỳ)
        try:
            opening = product.with_context(
                location=all_loc_ids,
                to_date=date_from_str,
            ).qty_available
        except Exception:
            opening = 0.0

        # Moves in date range
        moves = self.env["stock.move"].search([
            ("product_id", "=", product_id),
            ("state", "=", "done"),
            ("date", ">=", date_from_str),
            ("date", "<=", date_to_str),
            "|",
            ("location_dest_id", "in", all_loc_ids),
            ("location_id", "in", all_loc_ids),
        ], order="date asc, id asc")

        result_moves = []
        running = opening

        for move in moves:
            is_dest = move.location_dest_id.id in all_loc_set
            is_src = move.location_id.id in all_loc_set
            if is_dest and is_src:
                continue  # internal transfer within warehouse

            qty = float(getattr(move, "quantity", 0) or getattr(move, "quantity_done", 0) or 0)
            if qty <= 0:
                continue

            if is_dest:
                move_type = "in"
                running += qty
                in_qty = qty
                out_qty = 0.0
            else:
                move_type = "out"
                running -= qty
                in_qty = 0.0
                out_qty = qty

            # Price + combo detection
            price = 0.0
            combo_info = None
            purchase_line = getattr(move, "purchase_line_id", False)
            sale_line = getattr(move, "sale_line_id", False)
            if purchase_line:
                price = purchase_line.price_unit or 0.0
            elif sale_line:
                # Phantom kit: sale_line points to the combo product, not the component
                if sale_line.product_id.id != move.product_id.id:
                    price = 0.0
                    combo_info = {
                        "name": sale_line.product_id.name,
                        "code": sale_line.product_id.default_code or "",
                        "price": sale_line.price_unit,
                    }
                else:
                    price = sale_line.price_unit or 0.0
            else:
                price = getattr(move, "price_unit", 0.0) or 0.0
            # Fallback combo detection for moves without sale_line (price still 0)
            if price == 0.0 and move_type == "out" and not combo_info:
                combo_info = self._detect_combo_for_move(move, sale_line)

            picking = move.picking_id
            partner = picking.partner_id if picking else False

            # Transit transfer detection (deltatech_picking_transit)
            is_transit = False
            transit_linked = ""  # tên phiếu liên kết bên kia
            if picking:
                source_tf = getattr(picking, "source_transfer_id", False)
                second_created = getattr(picking, "second_transfer_created", False)
                src_usage = move.location_id.usage
                dst_usage = move.location_dest_id.usage
                if source_tf:
                    # Bước 2: phiếu nhận từ transit
                    is_transit = True
                    transit_linked = source_tf.name or ""
                elif second_created:
                    # Bước 1: phiếu xuất sang transit, tìm phiếu bước 2
                    is_transit = True
                    step2 = self.env["stock.picking"].search(
                        [("source_transfer_id", "=", picking.id)], limit=1
                    )
                    transit_linked = step2.name if step2 else ""
                elif src_usage == "transit" or dst_usage == "transit":
                    is_transit = True

            # Build origin: show linked transit picking name if applicable
            if is_transit and transit_linked:
                origin = transit_linked
            else:
                origin = (picking.origin if picking else "") or ""

            # Description
            if is_transit:
                description = "Nhập chuyển kho" if move_type == "in" else "Xuất chuyển kho"
            else:
                description = "Nhập kho" if move_type == "in" else "Xuất kho"

            result_moves.append({
                "type": move_type,
                "is_transit": is_transit,
                "date": utc_to_local_str(move.date),
                "reference": (picking.name if picking else move.name) or "",
                "origin": origin,
                "description": description,
                "uom": move.product_uom.name or "",
                "price": price,
                "combo_info": combo_info,
                "in_qty": in_qty,
                "out_qty": out_qty,
                "balance": round(running, 4),
                "partner_code": (partner.ref or "") if partner else "",
                "partner_name": (partner.name or "") if partner else "",
            })

        return {
            "opening": opening,
            "moves": result_moves,
            "date_from": date_from,
            "date_to": date_to,
            "closing": round(running, 4),
        }

    @api.model
    def export_moves_excel(self, product_id, warehouse_ids, date_from=None, date_to=None):
        import io
        try:
            import xlsxwriter
        except ImportError:
            raise UserError("xlsxwriter ch\u01b0a \u0111\u01b0\u1ee3c c\u00e0i \u0111\u1eb7t.")

        data = self.get_product_moves(product_id, warehouse_ids, date_from, date_to)
        product = self.env["product.product"].browse(product_id)

        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output, {"in_memory": True})
        ws = wb.add_worksheet("S\u1ed5 chi ti\u1ebft")

        # Formats
        ftitle = wb.add_format({"bold": True, "font_size": 14, "align": "center", "valign": "vcenter", "font_color": "#1b5e20"})
        finfo  = wb.add_format({"font_size": 11, "italic": True, "font_color": "#546e7a", "border": 0})
        fh     = wb.add_format({"bold": True, "bg_color": "#33691e", "font_color": "#fff", "border": 1, "align": "center", "valign": "vcenter", "font_size": 11})
        ft     = wb.add_format({"border": 1})
        fmono  = wb.add_format({"border": 1, "font_name": "Courier New", "font_size": 9, "font_color": "#1a2639"})
        fdate  = wb.add_format({"border": 1, "align": "center", "font_color": "#546e7a"})
        fuom   = wb.add_format({"border": 1, "align": "center", "font_color": "#6c757d"})
        fprice = wb.add_format({"border": 1, "num_format": "#,##0", "align": "right", "font_color": "#5d4037"})
        fcombo = wb.add_format({"border": 1, "align": "left", "font_color": "#e65100", "italic": True, "font_size": 10})
        fqty_in  = wb.add_format({"border": 1, "num_format": "#,##0.##", "align": "right", "font_color": "#1565c0", "bold": True})
        fqty_out = wb.add_format({"border": 1, "num_format": "#,##0.##", "align": "right", "font_color": "#b71c1c", "bold": True})
        fqty_bal = wb.add_format({"border": 1, "num_format": "#,##0.##", "align": "right", "font_color": "#1b5e20", "bold": True})
        fqty_neg = wb.add_format({"border": 1, "num_format": "#,##0.##", "align": "right", "font_color": "#e53935", "bold": True})
        fempty   = wb.add_format({"border": 1})
        fopening = wb.add_format({"bold": True, "italic": True, "bg_color": "#e8f5e9", "font_color": "#1b5e20", "border": 1})
        fopening_qty = wb.add_format({"bold": True, "bg_color": "#c8e6c9", "font_color": "#1b5e20", "border": 1, "num_format": "#,##0.##", "align": "right"})
        fclosing = wb.add_format({"bold": True, "italic": True, "bg_color": "#a5d6a7", "font_color": "#1b5e20", "border": 1})
        fclosing_qty = wb.add_format({"bold": True, "bg_color": "#81c784", "font_color": "#1b5e20", "border": 1, "num_format": "#,##0.##", "align": "right", "font_size": 12})

        # Column widths
        ws.set_column(0, 0, 12)   # Ngay
        ws.set_column(1, 1, 20)   # So CT
        ws.set_column(2, 2, 20)   # So HD/Nguon
        ws.set_column(3, 3, 20)   # Dien giai
        ws.set_column(4, 4, 8)    # DVT
        ws.set_column(5, 5, 18)   # Don gia
        ws.set_column(6, 6, 12)   # Nhap SL
        ws.set_column(7, 7, 12)   # Xuat SL
        ws.set_column(8, 8, 12)   # Ton
        ws.set_column(9, 9, 14)   # Ma DT
        ws.set_column(10, 10, 28) # Ten DT

        # Title
        ws.set_row(0, 28)
        ws.set_row(1, 6)
        ws.merge_range(0, 0, 0, 10, "S\u1ed4 CHI TI\u1ebeT NH\u1eacP/XU\u1ea4T KHO", ftitle)
        ws.set_row(2, 18)
        ws.merge_range(2, 0, 2, 4,
            "S\u1ea3n ph\u1ea9m: %s [%s]" % (product.name, product.default_code or ""), finfo)
        ws.write(2, 5, "T\u1eeb: %s" % (data["date_from"] or ""), finfo)
        ws.write(2, 7, "\u0110\u1ebfn: %s" % (data["date_to"] or ""), finfo)

        # Header row
        ws.set_row(3, 22)
        for i, h in enumerate(["Ng\u00e0y", "S\u1ed1 ch\u1ee9ng t\u1eeb", "S\u1ed1 H\u0110/Ngu\u1ed3n",
                                "Di\u1ec5n gi\u1ea3i", "\u0110VT", "\u0110\u01a1n gi\u00e1",
                                "Nh\u1eadp SL", "Xu\u1ea5t SL", "T\u1ed3n",
                                "M\u00e3 \u0110T", "T\u00ean \u0111\u1ed1i t\u00e1c"]):
            ws.write(3, i, h, fh)

        row = 4

        # Opening
        ws.merge_range(row, 0, row, 7, "S\u1ed1 d\u01b0 \u0111\u1ea7u k\u1ef3", fopening)
        ws.write(row, 8, data["opening"], fopening_qty)
        ws.write(row, 9, "", fopening)
        ws.write(row, 10, "", fopening)
        row += 1

        # Move rows
        for mv in data["moves"]:
            label = ("\u2b07 NK " if mv["type"] == "in" else "\u2b06 XK ") + mv["reference"]
            ws.write(row, 0, mv["date"], fdate)
            ws.write(row, 1, label, fmono)
            ws.write(row, 2, mv["origin"], fmono)
            ws.write(row, 3, mv["description"], ft)
            ws.write(row, 4, mv["uom"], fuom)
            combo = mv.get("combo_info")
            if combo:
                combo_label = "Combo: %s" % (combo.get("code") or combo.get("name", ""))
                if combo.get("price"):
                    combo_label += " / %s\u20ab" % "{:,.0f}".format(combo["price"]).replace(",", ".")
                ws.write(row, 5, combo_label, fcombo)
            else:
                ws.write(row, 5, mv.get("price") or 0, fprice)
            if mv["in_qty"] > 0:
                ws.write(row, 6, mv["in_qty"], fqty_in)
                ws.write(row, 7, "", fempty)
            else:
                ws.write(row, 6, "", fempty)
                ws.write(row, 7, mv["out_qty"], fqty_out)
            bal = mv["balance"]
            ws.write(row, 8, bal, fqty_bal if bal >= 0 else fqty_neg)
            ws.write(row, 9, mv["partner_code"], fmono)
            ws.write(row, 10, mv["partner_name"], ft)
            row += 1

        # Closing
        ws.merge_range(row, 0, row, 7, "S\u1ed1 d\u01b0 cu\u1ed1i k\u1ef3", fclosing)
        ws.write(row, 8, data["closing"], fclosing_qty)
        ws.write(row, 9, "", fclosing)
        ws.write(row, 10, "", fclosing)

        wb.close()
        output.seek(0)

        pname = (product.default_code or product.name or "sp").replace("/", "-").replace(" ", "_")[:30]
        att = self.env["ir.attachment"].create({
            "name": "so_chi_tiet_%s.xlsx" % pname,
            "type": "binary",
            "datas": base64.b64encode(output.read()).decode(),
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "res_model": self._name,
            "res_id": 0,
        })
        return att.id

    @api.model
    def export_all_moves_excel(self, group_id, warehouse_ids, date_from=None, date_to=None):
        import io
        try:
            import xlsxwriter
        except ImportError:
            raise UserError("xlsxwriter ch\u01b0a \u0111\u01b0\u1ee3c c\u00e0i \u0111\u1eb7t.")

        group = self.env["hlv.product.report.group"].browse(group_id)
        products = group.product_ids.sorted("name")
        if not products:
            raise UserError("Nh\u00f3m kh\u00f4ng c\u00f3 s\u1ea3n ph\u1ea9m n\u00e0o.")

        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output, {"in_memory": True})

        # Shared formats
        ftitle    = wb.add_format({"bold": True, "font_size": 13, "align": "center", "valign": "vcenter", "font_color": "#1b5e20"})
        finfo     = wb.add_format({"font_size": 10, "italic": True, "font_color": "#546e7a"})
        fh        = wb.add_format({"bold": True, "bg_color": "#33691e", "font_color": "#fff", "border": 1, "align": "center", "valign": "vcenter", "font_size": 11})
        ft        = wb.add_format({"border": 1})
        fmono     = wb.add_format({"border": 1, "font_name": "Courier New", "font_size": 9, "font_color": "#1a2639"})
        fdate     = wb.add_format({"border": 1, "align": "center", "font_color": "#546e7a"})
        fuom      = wb.add_format({"border": 1, "align": "center", "font_color": "#6c757d"})
        fprice    = wb.add_format({"border": 1, "num_format": "#,##0", "align": "right", "font_color": "#5d4037"})
        fcombo    = wb.add_format({"border": 1, "align": "left", "font_color": "#e65100", "italic": True, "font_size": 10})
        fqty_in   = wb.add_format({"border": 1, "num_format": "#,##0.##", "align": "right", "font_color": "#1565c0", "bold": True})
        fqty_out  = wb.add_format({"border": 1, "num_format": "#,##0.##", "align": "right", "font_color": "#b71c1c", "bold": True})
        fqty_bal  = wb.add_format({"border": 1, "num_format": "#,##0.##", "align": "right", "font_color": "#1b5e20", "bold": True})
        fqty_neg  = wb.add_format({"border": 1, "num_format": "#,##0.##", "align": "right", "font_color": "#e53935", "bold": True})
        fempty    = wb.add_format({"border": 1})
        fopening  = wb.add_format({"bold": True, "italic": True, "bg_color": "#e8f5e9", "font_color": "#1b5e20", "border": 1})
        fopen_qty = wb.add_format({"bold": True, "bg_color": "#c8e6c9", "font_color": "#1b5e20", "border": 1, "num_format": "#,##0.##", "align": "right"})
        fclosing  = wb.add_format({"bold": True, "italic": True, "bg_color": "#a5d6a7", "font_color": "#1b5e20", "border": 1})
        fclose_qty= wb.add_format({"bold": True, "bg_color": "#81c784", "font_color": "#1b5e20", "border": 1, "num_format": "#,##0.##", "align": "right", "font_size": 12})
        HEADERS   = ["Ng\u00e0y", "S\u1ed1 ch\u1ee9ng t\u1eeb", "S\u1ed1 H\u0110/Ngu\u1ed3n",
                     "Di\u1ec5n gi\u1ea3i", "\u0110VT", "\u0110\u01a1n gi\u00e1",
                     "Nh\u1eadp SL", "Xu\u1ea5t SL", "T\u1ed3n",
                     "M\u00e3 \u0110T", "T\u00ean \u0111\u1ed1i t\u00e1c"]

        # Summary sheet (first sheet)
        ws_idx = wb.add_worksheet("T\u1ed5ng quan")
        ws_idx.set_column(0, 0, 8)
        ws_idx.set_column(1, 1, 18)
        ws_idx.set_column(2, 2, 40)
        ws_idx.set_column(3, 3, 10)
        ws_idx.set_column(4, 4, 12)
        ws_idx.set_column(5, 5, 12)
        ws_idx.set_column(6, 6, 12)
        ws_idx.set_row(0, 26)
        ws_idx.merge_range(0, 0, 0, 6, "S\u1ed4 CHI TI\u1ebeT NH\u1eacP/XU\u1ea4T KHO - %s" % group.name, ftitle)
        ws_idx.write(1, 0, "T\u1eeb: %s" % (date_from or ""), finfo)
        ws_idx.write(1, 2, "\u0110\u1ebfn: %s" % (date_to or ""), finfo)
        fhidx = wb.add_format({"bold": True, "bg_color": "#1a2639", "font_color": "#fff", "border": 1, "align": "center"})
        ftidx = wb.add_format({"border": 1, "num_format": "#,##0.##", "align": "right", "font_color": "#198754", "bold": True})
        ftidxb= wb.add_format({"border": 1, "align": "left"})
        ws_idx.set_row(2, 18)
        for i, h in enumerate(["#", "M\u00e3 SP", "T\u00ean s\u1ea3n ph\u1ea9m", "\u0110VT", "\u0110\u1ea7u k\u1ef3", "Ph\u00e1t sinh", "Cu\u1ed1i k\u1ef3"]):
            ws_idx.write(2, i, h, fhidx)

        idx_row = 3
        for seq, product in enumerate(products):
            def _write_product_sheet(prod, seq_no):
                data = self.get_product_moves(prod.id, warehouse_ids, date_from, date_to)
                sheet_name = ("%s" % (prod.default_code or prod.name or "SP%d" % seq_no))[:31]
                # Sanitize sheet name (Excel forbidden chars)
                for ch in [":", "\\", "/", "?", "*", "[", "]"]:
                    sheet_name = sheet_name.replace(ch, "_")
                ws = wb.add_worksheet(sheet_name)
                ws.set_column(0, 0, 12)
                ws.set_column(1, 1, 20)
                ws.set_column(2, 2, 20)
                ws.set_column(3, 3, 20)
                ws.set_column(4, 4, 8)
                ws.set_column(5, 5, 18)
                ws.set_column(6, 6, 12)
                ws.set_column(7, 7, 12)
                ws.set_column(8, 8, 12)
                ws.set_column(9, 9, 14)
                ws.set_column(10, 10, 28)
                ws.set_row(0, 26)
                ws.merge_range(0, 0, 0, 10,
                    "S\u1ed4 CHI TI\u1ebeT: %s [%s]" % (prod.name, prod.default_code or ""), ftitle)
                ws.write(1, 0, "T\u1eeb: %s" % (data["date_from"] or ""), finfo)
                ws.write(1, 4, "\u0110\u1ebfn: %s" % (data["date_to"] or ""), finfo)
                ws.set_row(2, 20)
                for ci, h in enumerate(HEADERS):
                    ws.write(2, ci, h, fh)
                r = 3
                ws.merge_range(r, 0, r, 7, "S\u1ed1 d\u01b0 \u0111\u1ea7u k\u1ef3", fopening)
                ws.write(r, 8, data["opening"], fopen_qty)
                ws.write(r, 9, "", fopening)
                ws.write(r, 10, "", fopening)
                r += 1
                net_in = 0.0
                net_out = 0.0
                for mv in data["moves"]:
                    label = ("\u2b07 NK " if mv["type"] == "in" else "\u2b06 XK ") + mv["reference"]
                    ws.write(r, 0, mv["date"], fdate)
                    ws.write(r, 1, label, fmono)
                    ws.write(r, 2, mv["origin"], fmono)
                    ws.write(r, 3, mv["description"], ft)
                    ws.write(r, 4, mv["uom"], fuom)
                    combo = mv.get("combo_info")
                    if combo:
                        combo_label = "Combo: %s" % (combo.get("code") or combo.get("name", ""))
                        if combo.get("price"):
                            combo_label += " / %s\u20ab" % "{:,.0f}".format(combo["price"]).replace(",", ".")
                        ws.write(r, 5, combo_label, fcombo)
                    else:
                        ws.write(r, 5, mv.get("price") or 0, fprice)
                    if mv["in_qty"] > 0:
                        ws.write(r, 6, mv["in_qty"], fqty_in)
                        ws.write(r, 7, "", fempty)
                        net_in += mv["in_qty"]
                    else:
                        ws.write(r, 6, "", fempty)
                        ws.write(r, 7, mv["out_qty"], fqty_out)
                        net_out += mv["out_qty"]
                    bal = mv["balance"]
                    ws.write(r, 8, bal, fqty_bal if bal >= 0 else fqty_neg)
                    ws.write(r, 9, mv["partner_code"], fmono)
                    ws.write(r, 10, mv["partner_name"], ft)
                    r += 1
                ws.merge_range(r, 0, r, 7, "S\u1ed1 d\u01b0 cu\u1ed1i k\u1ef3", fclosing)
                ws.write(r, 8, data["closing"], fclose_qty)
                ws.write(r, 9, "", fclosing)
                ws.write(r, 10, "", fclosing)
                return data["opening"], net_in + net_out, data["closing"]

            opening_val, total_moves, closing_val = _write_product_sheet(product, seq + 1)
            ws_idx.write(idx_row, 0, seq + 1, wb.add_format({"border": 1, "align": "center", "font_color": "#adb5bd"}))
            ws_idx.write(idx_row, 1, product.default_code or "", wb.add_format({"border": 1, "font_name": "Courier New", "font_size": 9}))
            ws_idx.write(idx_row, 2, product.name or "", ftidxb)
            ws_idx.write(idx_row, 3, product.uom_id.name if product.uom_id else "", wb.add_format({"border": 1, "align": "center"}))
            ws_idx.write(idx_row, 4, opening_val, ftidx)
            ws_idx.write(idx_row, 5, total_moves, ftidx)
            ws_idx.write(idx_row, 6, closing_val, ftidx)
            idx_row += 1

        wb.close()
        output.seek(0)
        fname = "so_chi_tiet_tat_ca_%s.xlsx" % (group.name or "nhom").replace(" ", "_")[:40]
        att = self.env["ir.attachment"].create({
            "name": fname,
            "type": "binary",
            "datas": base64.b64encode(output.read()).decode(),
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "res_model": self._name,
            "res_id": 0,
        })
        return att.id
