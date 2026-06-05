# -*- coding: utf-8 -*-

import base64
import json
import logging
import re
import unicodedata
import uuid
from datetime import date, datetime
from io import BytesIO

from odoo import api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class HlvExcelPurchaseFile(models.Model):
    _name = "hlv.excel.purchase.file"
    _description = "File Excel tra cứu sổ chi tiết mua hàng"
    _order = "write_date desc, id desc"

    name = fields.Char(string="Tên file/tra cứu", required=True, default="Sổ chi tiết mua hàng")
    active = fields.Boolean(default=True)
    file_data = fields.Binary(string="File Excel", attachment=True, required=True)
    file_name = fields.Char(string="Tên file")
    sheet_name = fields.Char(string="Sheet", help="Để trống sẽ lấy sheet đầu tiên.")
    header_row = fields.Integer(string="Dòng header", default=4, required=True)
    public_slug = fields.Char(string="Mã link công khai", copy=False, readonly=True, default=lambda self: uuid.uuid4().hex[:12])
    public_url = fields.Char(string="Link công khai", compute="_compute_public_url")
    access_password = fields.Char(string="Mật khẩu truy cập public")
    column_ids = fields.One2many("hlv.excel.purchase.column", "file_id", string="Cột Excel")
    line_ids = fields.One2many("hlv.excel.purchase.line", "file_id", string="Dữ liệu")
    line_count = fields.Integer(string="Số dòng", compute="_compute_line_count")
    last_import_date = fields.Datetime(string="Lần import cuối", readonly=True)

    @api.depends("public_slug")
    def _compute_public_url(self):
        base_url = self.env["ir.config_parameter"].sudo().get_param("web.base.url", default="") or ""
        for record in self:
            record.public_url = f"{base_url}/excel-purchase-search/{record.public_slug}" if record.public_slug else ""

    def _compute_line_count(self):
        counts = self.env["hlv.excel.purchase.line"].read_group(
            [("file_id", "in", self.ids)], ["file_id"], ["file_id"]
        )
        mapped = {item["file_id"][0]: item["file_id_count"] for item in counts}
        for record in self:
            record.line_count = mapped.get(record.id, 0)

    def _get_workbook_sheet(self):
        self.ensure_one()
        if not self.file_data:
            raise UserError("Chưa upload file Excel.")
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise UserError("Server chưa cài thư viện openpyxl để đọc Excel.") from exc

        try:
            workbook = load_workbook(BytesIO(base64.b64decode(self.file_data)), data_only=True, read_only=True)
        except Exception as exc:
            raise UserError(f"Không đọc được file Excel: {exc}") from exc

        if self.sheet_name:
            if self.sheet_name not in workbook.sheetnames:
                raise UserError(f"Không tìm thấy sheet '{self.sheet_name}'.")
            return workbook, workbook[self.sheet_name]
        return workbook, workbook[workbook.sheetnames[0]]

    def _format_cell_value(self, value):
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        if isinstance(value, date):
            return value.strftime("%d/%m/%Y")
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _normalize_keyword(self, value):
        value = (value or "").lower()
        value = re.sub(r"\s+", " ", value)
        value = value.strip()
        ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
        if ascii_value and ascii_value != value:
            return f"{value} {ascii_value}"
        return value

    def action_read_headers(self):
        for record in self:
            if record.header_row < 1:
                raise UserError("Dòng header phải lớn hơn 0.")
            workbook, sheet = record._get_workbook_sheet()
            try:
                header_cells = next(sheet.iter_rows(min_row=record.header_row, max_row=record.header_row, values_only=True))
            finally:
                workbook.close()

            record.column_ids.unlink()
            columns = []
            for index, value in enumerate(header_cells, start=1):
                header = record._format_cell_value(value)
                if not header:
                    continue
                columns.append({
                    "file_id": record.id,
                    "sequence": index,
                    "name": header,
                    "searchable": True,
                    "show_public": True,
                })
            if not columns:
                raise UserError("Không đọc được header tại dòng đã khai báo.")
            self.env["hlv.excel.purchase.column"].create(columns)
        return True

    def action_import_lines(self):
        Line = self.env["hlv.excel.purchase.line"]
        for record in self:
            if not record.column_ids:
                record.action_read_headers()

            searchable_sequences = set(record.column_ids.filtered("searchable").mapped("sequence"))
            if not searchable_sequences:
                raise UserError("Cần chọn ít nhất 1 cột được phép tìm kiếm.")

            column_by_sequence = {column.sequence: column.name for column in record.column_ids}
            max_sequence = max(column_by_sequence)
            workbook, sheet = record._get_workbook_sheet()
            rows_to_create = []
            try:
                for row_index, row in enumerate(
                    sheet.iter_rows(min_row=record.header_row + 1, max_col=max_sequence, values_only=True),
                    start=record.header_row + 1,
                ):
                    row_values = {}
                    search_parts = []
                    has_value = False
                    for sequence, header in column_by_sequence.items():
                        value = record._format_cell_value(row[sequence - 1] if sequence - 1 < len(row) else "")
                        if value:
                            has_value = True
                        row_values[str(sequence)] = value
                        if sequence in searchable_sequences and value:
                            search_parts.append(value)
                    if not has_value:
                        continue
                    rows_to_create.append({
                        "file_id": record.id,
                        "excel_row": row_index,
                        "row_json": json.dumps(row_values, ensure_ascii=False),
                        "search_text": record._normalize_keyword(" ".join(search_parts)),
                    })
            finally:
                workbook.close()

            Line.search([("file_id", "=", record.id)]).unlink()
            if rows_to_create:
                Line.create(rows_to_create)
            record.write({"last_import_date": fields.Datetime.now()})
            _logger.info("Imported %s Excel purchase rows for file %s", len(rows_to_create), record.id)
        return True

    def action_open_public_url(self):
        self.ensure_one()
        return {
            "type": "ir.actions.act_url",
            "url": f"/excel-purchase-search/{self.public_slug}",
            "target": "new",
        }


class HlvExcelPurchaseColumn(models.Model):
    _name = "hlv.excel.purchase.column"
    _description = "Cột tra cứu Excel mua hàng"
    _order = "file_id, sequence"

    file_id = fields.Many2one("hlv.excel.purchase.file", required=True, ondelete="cascade")
    sequence = fields.Integer(string="Số cột", required=True)
    name = fields.Char(string="Header", required=True)
    searchable = fields.Boolean(string="Cho phép tìm kiếm", default=True)
    show_public = fields.Boolean(string="Hiển thị public", default=True)


class HlvExcelPurchaseLine(models.Model):
    _name = "hlv.excel.purchase.line"
    _description = "Dòng tra cứu Excel mua hàng"
    _order = "excel_row asc, id asc"

    file_id = fields.Many2one("hlv.excel.purchase.file", required=True, ondelete="cascade", index=True)
    excel_row = fields.Integer(string="Dòng Excel", index=True)
    row_json = fields.Text(string="Dữ liệu JSON", required=True)
    search_text = fields.Text(string="Nội dung tìm kiếm", index=True)

    def get_row_values(self):
        self.ensure_one()
        try:
            return json.loads(self.row_json or "{}")
        except Exception:
            return {}
