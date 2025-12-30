
import base64
import io
import logging
from odoo import models, fields, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

class VendorImportWizard(models.TransientModel):
    _name = 'vendor.import.wizard'
    _description = 'Import/Update Vendors from Excel'

    excel_file = fields.Binary(string='Excel File', required=True)
    file_name = fields.Char(string='File Name')

    def action_import(self):
        self.ensure_one()
        if not self.excel_file:
            raise UserError(_("Please upload an Excel file."))

        try:
            import openpyxl
        except ImportError:
            raise UserError(_("openpyxl library is not installed."))

        try:
            file_content = base64.b64decode(self.excel_file)
            book = openpyxl.load_workbook(filename=io.BytesIO(file_content))
            sheet = book.active
        except Exception as e:
            raise UserError(_("Error reading Excel file: %s") % str(e))

        # Find header row
        header_row_idx = None
        headers = {}
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row and 'Mã nhà cung cấp' in [str(c).strip() for c in row if c]:
                header_row_idx = row_idx
                for col_idx, cell_value in enumerate(row):
                    if cell_value:
                        headers[str(cell_value).strip()] = col_idx
                break
        
        if header_row_idx is None:
            raise UserError(_("Could not find header row containing 'Mã nhà cung cấp'."))

        # Process rows
        Partner = self.env['res.partner']
        updated_count = 0
        
        # Define column mappings names based on user description
        col_name_ref = 'Mã nhà cung cấp'
        col_name_name = 'Tên nhà cung cấp'
        col_name_street = 'Địa chỉ'
        col_name_vat = 'Mã số thuế'
        col_name_phone = 'Điện thoại' # Updates both phone and mobile

        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            # Get Name
            if col_name_name not in headers:
                 continue # strict check? or just skip? 
            
            name_idx = headers.get(col_name_name)
            name = row[name_idx] if name_idx is not None and name_idx < len(row) else None
            
            if not name:
                continue
            
            name = str(name).strip()
            
            # Search partner
            partner = Partner.search([('name', '=', name)], limit=1)
            if not partner:
                _logger.info(f"Vendor not found: {name}")
                continue
            
            vals = {}
            
            # Update 'company_registry' (Mã nhà cung cấp)
            if col_name_ref in headers:
                ref_val = row[headers[col_name_ref]]
                if ref_val:
                     vals['company_registry'] = str(ref_val).strip()
            
            # Update 'street' (Địa chỉ)
            if col_name_street in headers:
                street_val = row[headers[col_name_street]]
                if street_val:
                    vals['street'] = str(street_val).strip()

            # Update 'vat' (Mã số thuế)
            if col_name_vat in headers:
                vat_val = row[headers[col_name_vat]]
                if vat_val:
                    vals['vat'] = str(vat_val).strip()
            
            # Update 'phone' and 'mobile'
            if col_name_phone in headers:
                phone_val = row[headers[col_name_phone]]
                if phone_val:
                    phone_str = str(phone_val).strip()
                    vals['phone'] = phone_str
                    vals['mobile'] = phone_str
            
            # Check "Công ty" in name to set is_company
            if "công ty" in name.lower():
                vals['is_company'] = True
                vals['company_type'] = 'company'

            if vals:
                partner.write(vals)
                updated_count += 1
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Success'),
                'message': _('Updated %s vendors successfully.') % updated_count,
                'type': 'success',
                'sticky': False,
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }
