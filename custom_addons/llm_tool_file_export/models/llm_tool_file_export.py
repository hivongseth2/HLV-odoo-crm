import base64
import csv
import io
import logging

from odoo import api, models

_logger = logging.getLogger(__name__)


class LLMToolFileExport(models.Model):
    _inherit = "llm.tool"

    @api.model
    def _get_available_implementations(self):
        implementations = super()._get_available_implementations()
        return implementations + [
            ("file_export", "File Export"),
        ]

    def file_export_execute(
        self,
        filename: str,
        headers: list[str],
        rows: list[list],
        file_type: str = "xlsx",
        sheet_name: str = "Sheet1",
    ) -> str:
        """Create and export a downloadable file (xlsx or csv) with structured data.
        Use this tool when you need to generate a file for the user to download,
        especially when presenting tabular data, price lists, or reports.
        If the user uploaded a file and asked to fill in data, read the original data,
        add/modify the information, then use this tool to export the updated file.

        Args:
            filename: Output filename (e.g., "bao_gia.xlsx", "danh_sach.csv")
            headers: List of column headers (e.g., ["SKU", "Tên sản phẩm", "Giá bán"])
            rows: List of data rows, each row is a list of cell values matching headers order
            file_type: File format - "xlsx" or "csv" (default: xlsx)
            sheet_name: Worksheet name for xlsx files (default: Sheet1)
        """
        if not headers:
            return "Error: headers cannot be empty"
        if not rows:
            return "Error: rows cannot be empty"

        if file_type == "xlsx":
            content, mimetype = self._export_create_xlsx(headers, rows, sheet_name)
            if not filename.endswith(".xlsx"):
                filename += ".xlsx"
        elif file_type == "csv":
            content, mimetype = self._export_create_csv(headers, rows)
            if not filename.endswith(".csv"):
                filename += ".csv"
        else:
            return f"Unsupported file type: {file_type}. Use 'xlsx' or 'csv'."

        # Create attachment
        attachment = self.env["ir.attachment"].create(
            {
                "name": filename,
                "datas": base64.b64encode(content).decode(),
                "mimetype": mimetype,
                "type": "binary",
            }
        )

        # Attach to the tool message if available
        message = self.env.context.get("message")
        if message:
            try:
                message.write({"attachment_ids": [(4, attachment.id)]})
                # Link attachment to the thread record
                if message.model and message.res_id:
                    attachment.write(
                        {
                            "res_model": message.model,
                            "res_id": message.res_id,
                        }
                    )
            except Exception as e:
                _logger.warning("Could not attach file to message: %s", e)

        return (
            f"File '{filename}' created successfully "
            f"({len(rows)} rows, {len(headers)} columns). "
            f"The file is attached to this message for download."
        )

    def _export_create_xlsx(self, headers, rows, sheet_name="Sheet1"):
        """Create an xlsx file from headers and rows."""
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                if col_idx > len(headers):
                    break
                cell_value = self._export_convert_value(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.border = thin_border
                # Right-align numbers
                if isinstance(cell_value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0" if isinstance(cell_value, int) else "#,##0.00"

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 3, 50)

        # Freeze header row
        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        return (
            output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _export_create_csv(self, headers, rows):
        """Create a CSV file from headers and rows."""
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows:
            # Ensure row doesn't exceed header count
            writer.writerow(row[: len(headers)])
        # UTF-8 BOM for Excel compatibility
        return output.getvalue().encode("utf-8-sig"), "text/csv"

    @staticmethod
    def _export_convert_value(value):
        """Try to convert string values to appropriate Python types for Excel."""
        if value is None:
            return None
        if isinstance(value, (int, float, bool)):
            return value
        if isinstance(value, str):
            stripped = value.strip().replace(",", "")
            # Try integer
            try:
                return int(stripped)
            except (ValueError, TypeError):
                pass
            # Try float
            try:
                return float(stripped)
            except (ValueError, TypeError):
                pass
        return value
