# -*- coding: utf-8 -*-
import base64
import re
from io import BytesIO

from odoo import _, api, fields, models
from odoo.exceptions import UserError


class HlvContactSplitWizard(models.TransientModel):
    _name = 'hlv.contact.split.wizard'
    _description = 'Tách liên hệ thành root company mới'

    source_partner_id = fields.Many2one(
        'res.partner',
        string='Liên hệ nguồn',
        required=True,
    )
    split_mode = fields.Selection([
        ('existing', 'Chuyển sang công ty có sẵn'),
        ('new', 'Tạo công ty mới'),
    ], string='Cách xử lý', default='existing', required=True)
    destination_partner_id = fields.Many2one(
        'res.partner',
        string='Công ty nhận dữ liệu',
        domain=[('parent_id', '=', False)],
    )
    new_name = fields.Char(string='Tên công ty mới')
    new_ref = fields.Char(string='Mã khách MISA')
    new_vat = fields.Char(string='Mã số thuế')
    excel_file = fields.Binary(string='File Excel đơn bán')
    excel_filename = fields.Char(string='Tên file')
    excel_row_count = fields.Integer(string='Số dòng đọc được', readonly=True)
    excel_match_count = fields.Integer(string='Số đơn tìm thấy', readonly=True)
    excel_message = fields.Text(string='Kết quả nạp file', readonly=True)
    child_partner_ids = fields.Many2many(
        'res.partner',
        string='Liên hệ con chuyển sang công ty mới',
    )
    sale_order_ids = fields.Many2many(
        'sale.order',
        string='Đơn bán chuyển sang công ty mới',
    )
    purchase_order_ids = fields.Many2many(
        'purchase.order',
        string='Đơn mua chuyển sang công ty mới',
    )
    new_partner_id = fields.Many2one(
        'res.partner',
        string='Công ty mới',
        readonly=True,
    )
    note = fields.Text(
        string='Ghi chú',
        default=(
            'Dùng khi cùng tên công ty nhưng MISA có nhiều mã khách. '
            'Root mới sẽ được tạo theo khóa MST + mã khách; ref và '
            'company_registry đều nhận mã khách MISA.'
        ),
        readonly=True,
    )

    @api.onchange('source_partner_id')
    def _onchange_source_partner_id(self):
        if not self.source_partner_id:
            return
        self.new_name = self.source_partner_id.name
        self.new_vat = self.source_partner_id.vat

    def _norm_code(self, value):
        return (value or '').strip().upper().replace(' ', '')

    def _partner_codes(self, partner):
        return set(filter(None, [
            self._norm_code(partner.ref),
            self._norm_code(partner.company_registry),
        ]))

    def _destination_partner(self):
        self.ensure_one()
        if self.split_mode == 'existing':
            if not self.destination_partner_id:
                raise UserError(_('Cần chọn công ty nhận dữ liệu.'))
            destination = self.destination_partner_id.sudo().commercial_partner_id or self.destination_partner_id.sudo()
            if destination.parent_id:
                destination = destination.parent_id
            return destination

        source = self.source_partner_id.sudo()
        code = (self.new_ref or '').strip()
        vat = (self.new_vat or '').strip()
        if not code:
            raise UserError(_('Cần nhập mã khách MISA để tách.'))

        Partner = self.env['res.partner'].sudo().with_context(active_test=False)
        domain = [
            ('parent_id', '=', False),
            '|',
            ('ref', '=', code),
            ('company_registry', '=', code),
        ]
        existing = Partner.search(domain).filtered(
            lambda p: self._norm_code(code) in self._partner_codes(p)
        )
        if vat:
            existing = existing.filtered(lambda p: not p.vat or (p.vat or '').strip() == vat)
        if existing:
            return existing[:1]

        vals = {
            'name': self.new_name,
            'is_company': True,
            'customer_rank': max(source.customer_rank, 1),
            'supplier_rank': source.supplier_rank,
            'ref': code,
            'company_registry': code,
            'vat': vat or False,
            'street': source.street,
            'street2': source.street2,
            'city': source.city,
            'state_id': source.state_id.id,
            'country_id': source.country_id.id,
            'phone': source.phone,
            'mobile': source.mobile,
            'email': source.email,
        }
        new_partner = Partner.create(vals)
        new_partner.message_post(body=_(
            'Tạo root company mới từ %s theo khóa %s-%s.'
        ) % (source.display_name, vat or '-', code))
        return new_partner

    def _split_order_values(self, value):
        value = str(value or '').strip()
        if not value:
            return []
        return [part.strip() for part in re.split(r'[\n,;]+', value) if part.strip()]

    def _excel_cell_text(self, value):
        if value is None:
            return ''
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value).strip()

    def _sale_order_search_domain(self, order_no):
        SaleOrder = self.env['sale.order']
        clauses = [('name', '=', order_no)]
        for field_name in ('client_order_ref', 'origin', 'misa_order_no', 'x_studio_misa_order_no'):
            if field_name in SaleOrder._fields:
                clauses.append((field_name, '=', order_no))
        if len(clauses) == 1:
            return clauses
        return ['|'] * (len(clauses) - 1) + clauses

    def action_load_sale_orders_from_excel(self):
        self.ensure_one()
        if not self.excel_file:
            raise UserError(_('Cần tải lên file Excel.'))

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_('Server chưa có thư viện openpyxl để đọc Excel.'))

        destination = self._destination_partner()
        destination_codes = self._partner_codes(destination)
        if not destination_codes:
            raise UserError(_('Công ty nhận dữ liệu chưa có ref/company_registry để đối chiếu cột I.'))

        try:
            payload = base64.b64decode(self.excel_file)
            workbook = load_workbook(BytesIO(payload), data_only=True, read_only=True)
        except Exception as exc:
            raise UserError(_('Không đọc được file Excel: %s') % exc)

        sheet = workbook.active
        SaleOrder = self.env['sale.order'].sudo()
        source = self.source_partner_id.sudo()
        selected_orders = SaleOrder.browse()
        row_count = 0
        matched_rows = 0
        missing = []
        wrong_source = []
        skipped_code = 0
        multiple = []

        for row_index, row in enumerate(sheet.iter_rows(), start=1):
            if row_index == 1:
                continue
            row_count += 1
            order_value = row[1].value if len(row) >= 2 else None
            code_value = row[8].value if len(row) >= 9 else None
            customer_code = self._norm_code(self._excel_cell_text(code_value))
            if customer_code not in destination_codes:
                skipped_code += 1
                continue

            for order_no in self._split_order_values(self._excel_cell_text(order_value)):
                orders = SaleOrder.search(self._sale_order_search_domain(order_no), order='id asc')
                if not orders:
                    missing.append(order_no)
                    continue
                if len(orders) > 1:
                    multiple.append('%s: %s' % (order_no, orders.ids))
                    continue
                order = orders[0]
                current_root = order.partner_id.commercial_partner_id or order.partner_id
                if source and current_root.id != source.id:
                    wrong_source.append('%s -> %s' % (order_no, current_root.display_name))
                    continue
                selected_orders |= order
                matched_rows += 1

        message = [
            _('Đã đọc %s dòng Excel.') % row_count,
            _('Bỏ qua %s dòng không khớp mã công ty nhận.') % skipped_code,
            _('Tìm thấy %s đơn bán hợp lệ.') % len(selected_orders),
        ]
        if missing:
            message.append(_('Không tìm thấy đơn: %s') % ', '.join(missing[:20]))
        if multiple:
            message.append(_('Đơn bị trùng kết quả: %s') % '; '.join(multiple[:10]))
        if wrong_source:
            message.append(_('Đơn không nằm trên công ty nguồn: %s') % '; '.join(wrong_source[:10]))

        self.write({
            'sale_order_ids': [(6, 0, selected_orders.ids)],
            'excel_row_count': row_count,
            'excel_match_count': matched_rows,
            'excel_message': '\n'.join(message),
        })
        view = self.env.ref('hlv_contact_refine.view_hlv_contact_split_wizard_form', raise_if_not_found=False)
        return {
            'type': 'ir.actions.act_window',
            'name': _('Tách / chuyển dữ liệu liên hệ'),
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'views': [(view.id if view else False, 'form')],
            'target': 'new',
        }

    def _move_sale_orders(self, orders, destination):
        if not orders:
            return
        vals = {'partner_id': destination.id}
        if 'partner_invoice_id' in orders._fields:
            vals['partner_invoice_id'] = destination.id
        if 'partner_shipping_id' in orders._fields:
            vals['partner_shipping_id'] = destination.id
        orders.sudo().write(vals)
        if 'picking_ids' in orders._fields:
            pickings = orders.mapped('picking_ids').filtered(lambda p: p.state != 'cancel')
            if pickings:
                pickings.sudo().write({'partner_id': destination.id})

    def action_split(self):
        self.ensure_one()
        new_partner = self._destination_partner()

        if self.child_partner_ids:
            self.child_partner_ids.sudo().write({'parent_id': new_partner.id})
        if self.sale_order_ids:
            self._move_sale_orders(self.sale_order_ids, new_partner)
        if self.purchase_order_ids:
            self.purchase_order_ids.sudo().write({'partner_id': new_partner.id})

        self.new_partner_id = new_partner.id
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'res.partner',
            'res_id': new_partner.id,
            'view_mode': 'form',
            'target': 'current',
        }
