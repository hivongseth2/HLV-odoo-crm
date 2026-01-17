# -*- coding: utf-8 -*-
import openpyxl
import logging

_logger = logging.getLogger(__name__)

def normalize(name):
    if not name: return ""
    name = str(name).lower().strip()
    # Remove prefixes
    prefixes = [
        'tỉnh ', 'thành phố ', 'quận ', 'huyện ', 'thị xã ', 
        'phường ', 'xã ', 'thị trấn ', 'tp. ', 'tp ', 'q. ', 'h. ', 'p. ', 'x. '
    ]
    for p in prefixes:
        if name.startswith(p):
            name = name[len(p):]
    return name.strip()

def run_import(env, excel_path):
    print("Opening Excel file: %s" % excel_path)
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print("Error opening Excel: %s" % e)
        return

    sheet = wb['MAPPING']
    
    # CHANGED: Use J&T Models instead of GHN
    Province = env['jnt.province']
    District = env['jnt.district']
    Ward = env['jnt.ward']
    
    updated_count = 0
    created_count = 0
    
    # Pre-loading might be too heavy if empty, but let's try
    provinces = Province.search([])
    prov_map = {normalize(p.name): p for p in provinces}
    
    print("Starting import from %s rows..." % sheet.max_row)
    
    # Cache districts
    dist_cache = {} # {prov_id: {norm_name: dist_record}}

    for row_idx, row in enumerate(sheet.iter_rows(min_row=3), 3):
        prov_name = row[0].value
        dist_name = row[1].value
        ward_raw = row[2].value # e.g. "Phường An Khánh-028TPT02"
        ward_clean = row[3].value # e.g. "Phường An Khánh"
        
        if not prov_name or not dist_name or not ward_raw:
            continue
            
        # Extract J&T code
        code = None
        if '-' in str(ward_raw):
            code = str(ward_raw).split('-')[-1].strip()
        
        if not code:
            continue
            
        n_prov = normalize(prov_name)
        province = prov_map.get(n_prov)
        
        # If province doesn't exist, create it (J&T dedicated data)
        if not province:
            province = Province.create({'name': prov_name})
            prov_map[n_prov] = province
            print("Created Province: %s" % prov_name)
            
        # Get/Cache Districts for this province
        if province.id not in dist_cache:
            districts = District.search([('province_id', '=', province.id)])
            dist_cache[province.id] = {normalize(d.name): d for d in districts}

        province_districts = dist_cache[province.id]
        n_dist = normalize(dist_name)
        district = province_districts.get(n_dist)
        
        if not district:
            # Create District if new
            district = District.create({'name': dist_name, 'province_id': province.id})
            province_districts[n_dist] = district
            # print("Created District: %s" % dist_name)
            
        # Find or Create Ward
        n_ward = normalize(ward_clean)
        # Check by code first if possible, or name
        ward = Ward.search([
            ('district_id', '=', district.id),
            '|', ('jnt_code', '=', code), 
            '|', ('name', '=', ward_clean), ('name', 'ilike', n_ward)
        ], limit=1)
        
        if ward:
            if ward.jnt_code != code:
                ward.write({'jnt_code': code})
                updated_count += 1
        else:
            Ward.create({
                'name': ward_clean,
                'jnt_code': code,
                'district_id': district.id
            })
            created_count += 1
            # print("Created Ward: %s" % ward_clean)

        if row_idx % 500 == 0:
            print("Processed %s rows... (Updated: %s, Created: %s)" % (row_idx, updated_count, created_count))

    print("Import finished.") 
    print("Updated codes: %s" % updated_count)
    print("Created new wards: %s" % created_count)
    
    env.cr.commit()
    print("Changes committed to database.")

# To run:
# from odoo.addons.hlv_delivery_jt.scripts.import_jnt_locations import run_import
# run_import(env, r'e:\project\odoo_HLV\HLV-odoo-crm\custom_addons\hlv_delivery_jt\J&T_Danh mục địa chỉ.xlsx')
