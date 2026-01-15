import openpyxl
import json
import os

excel_path = r'e:\project\odoo_HLV\HLV-odoo-crm\custom_addons\hlv_delivery_jt\J&T_Danh mục địa chỉ.xlsx'
output_path = r'e:\project\odoo_HLV\HLV-odoo-crm\custom_addons\hlv_delivery_jt\excel_inspect.json'

try:
    # Load without read_only to see if it helps with dimension detection
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print(f"Total sheets: {len(wb.sheetnames)}")
    
    all_sheet_info = {}
    for name in wb.sheetnames:
        sheet = wb[name]
        try:
            # Check dimensions
            dim = sheet.calculate_dimension()
            print(f"Sheet {name} dimension: {dim}")
            
            # Get first 5 rows to see sample data
            rows_data = []
            for i, row in enumerate(sheet.iter_rows(max_row=10, min_col=1, max_col=min(sheet.max_column or 10, 20))):
                rows_data.append([cell.value for cell in row])
            
            all_sheet_info[name] = {
                "dimension": dim,
                "max_col": sheet.max_column,
                "max_row": sheet.max_row,
                "sample": rows_data
            }
        except Exception as se:
            all_sheet_info[name] = {"error": str(se)}

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(all_sheet_info, f, ensure_ascii=False, indent=2)
    
    print(f"Success! Analyzed {len(wb.sheetnames)} sheets to {output_path}")
except Exception as e:
    print(f"Error: {str(e)}")
