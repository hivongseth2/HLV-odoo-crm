# -*- coding: utf-8 -*-
from odoo import fields, models, api, _
from odoo.exceptions import UserError
import base64
from datetime import datetime
from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    Workbook = None
    XLImage = None


class BBBGExcelExportWizard(models.TransientModel):
    _name = 'bbbg.excel.export.wizard'
    _description = 'Xuất Excel Biên Bản Bàn Giao'

    def _calculate_row_height(self, text, column_width, font_size=13):
        """
        Tính chiều cao dòng dựa trên độ dài text và chiều rộng cột
        Args:
            text: Nội dung text
            column_width: Tổng chiều rộng các cột đã merge (Excel units)
            font_size: Kích thước font
        Returns:
            Chiều cao dòng tính bằng points
        """
        if not text:
            return 15  # Default height
        
        # Ước tính số ký tự trên 1 dòng
        # Giảm xuống 0.75 để tính cho dấu tiếng Việt và padding
        chars_per_line = int(column_width * 0.75)
        
        # Tính số dòng cần thiết
        text_length = len(str(text))
        num_lines = max(1, (text_length + chars_per_line - 1) // chars_per_line)
        
        # Kiểm tra nếu có \n trong text
        if '\n' in str(text):
            num_lines = max(num_lines, str(text).count('\n') + 1)
        
        # Chiều cao = số dòng * font_size * 1.3 (tăng line spacing)
        row_height = num_lines * font_size * 1.3
        
        return max(18, row_height)  # Minimum 18 points

    def _get_active_picking(self):
        """Lấy picking từ context"""
        active_id = self._context.get('active_id') or self._context.get('active_ids', [False])[0]
        if active_id:
            return self.env['stock.picking'].browse(active_id)
        return False

    def action_export_excel(self):
        """Xuất file Excel Biên Bản Bàn Giao"""
        if not Workbook:
            raise UserError(_('Thư viện openpyxl chưa được cài đặt. Vui lòng chạy: pip install openpyxl'))

        picking = self._get_active_picking()
        if not picking:
            raise UserError(_('Không tìm thấy phiếu giao hàng'))

        # Lấy dữ liệu combo từ helper (giống BBGN)
        enriched_lines = self.env['hlv.report.helper'].get_enriched_lines_for_picking_combo(picking)

        # Tạo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'BBBG'

        # Định nghĩa styles
        header_font = Font(name='Times New Roman', size=14, bold=True)
        title_font = Font(name='Times New Roman', size=18, bold=True)
        normal_font = Font(name='Times New Roman', size=13)
        bold_font = Font(name='Times New Roman', size=13, bold=True)
        small_font = Font(name='Times New Roman', size=12)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=False)
        left_align_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Set column widths
        ws.column_dimensions['A'].width = 6    # STT
        ws.column_dimensions['B'].width = 40   # Tên hàng
        ws.column_dimensions['C'].width = 18   # Đơn vị tính
        ws.column_dimensions['D'].width = 13   # Số lượng
        ws.column_dimensions['E'].width = 22   # Ghi chú

        # ============= HEADER - LOGO VÀ THÔNG TIN CÔNG TY =============
        row = 1
        
        # Tính toán số dòng cần thiết cho thông tin công ty
        company_name_header = picking.company_id.name or 'CÔNG TY TNHH VI NA HOÀNG LONG VŨ'
        addr = picking.company_id.partner_id._display_address(without_company=True).replace('\n', ' ') or ''
        tax_text = f'Mã số thuế: {picking.company_id.vat or ""}'
        website_text = f'Website: {picking.company_id.website or ""}'
        
        # Tính chiều cao cho từng dòng thông tin công ty (C+D+E = 53 units)
        height_company_name = self._calculate_row_height(company_name_header, 53, 14)
        height_addr = self._calculate_row_height(addr, 53, 13)
        height_tax = self._calculate_row_height(tax_text, 53, 13)
        height_website = self._calculate_row_height(website_text, 53, 13)
        
        # Tổng chiều cao của 4 dòng thông tin
        total_info_height = height_company_name + height_addr + height_tax + height_website
        
        # ====== THÔNG TIN CÔNG TY (tự động điều chỉnh chiều cao) ======
        # Row 1: Tên công ty
        ws.merge_cells(f'C{row}:E{row}')
        ws[f'C{row}'] = company_name_header
        ws[f'C{row}'].font = header_font
        ws[f'C{row}'].alignment = left_align_wrap
        ws.row_dimensions[row].height = height_company_name

        # Row 2: Địa chỉ
        row += 1
        ws.merge_cells(f'C{row}:E{row}')
        ws[f'C{row}'] = addr
        ws[f'C{row}'].font = normal_font
        ws[f'C{row}'].alignment = left_align_wrap
        ws.row_dimensions[row].height = height_addr

        # Row 3: Mã số thuế
        row += 1
        ws.merge_cells(f'C{row}:E{row}')
        ws[f'C{row}'] = tax_text
        ws[f'C{row}'].font = normal_font
        ws[f'C{row}'].alignment = left_align_wrap
        ws.row_dimensions[row].height = height_tax

        # Row 4: Website
        row += 1
        ws.merge_cells(f'C{row}:E{row}')
        ws[f'C{row}'] = website_text
        ws[f'C{row}'].font = normal_font
        ws[f'C{row}'].alignment = left_align_wrap
        ws.row_dimensions[row].height = height_website

        # ====== LOGO (độc lập, overlay lên cột A:B) ======
        # Merge cells cho logo từ A1:B4 (không cố định chiều cao)
        ws.merge_cells('A1:B4')
        
        # Thêm logo với kích thước tính theo tổng chiều cao thực tế của 4 dòng
        if picking.company_id.logo and XLImage:
            try:
                logo_data = base64.b64decode(picking.company_id.logo)
                logo_stream = BytesIO(logo_data)
                img = XLImage(logo_stream)
                
                # Tính chiều cao THỰC TẾ của vùng logo (tổng 4 dòng đã tính)
                # 1 point = 1.333 pixels tại 96 DPI
                total_height_px = total_info_height * 1.333
                
                # Tính chiều rộng của vùng logo (A+B)
                col_a_width_px = ((256 * 6 + 18) / 256) * 7 + 5
                col_b_width_px = ((256 * 40 + 18) / 256) * 7 + 5
                total_width_px = col_a_width_px + col_b_width_px
                
                # Scale logo để vừa với ô, giữ tỷ lệ aspect ratio
                if getattr(img, "width", None) and getattr(img, "height", None) and img.width > 0 and img.height > 0:
                    # Tính tỷ lệ scale theo cả chiều rộng và cao, lấy giá trị nhỏ hơn
                    scale_height = (total_height_px * 0.85) / float(img.height)  # 85% để có padding
                    scale_width = (total_width_px * 0.85) / float(img.width)
                    scale = min(scale_height, scale_width)
                    
                    new_height = int(img.height * scale)
                    new_width = int(img.width * scale)
                    
                    img.height = new_height
                    img.width = new_width
                    
                    # Tính offset để căn giữa theo chiều dọc, bên trái cột B theo chiều ngang
                    offset_x = 0  # Bên trái cột B
                    offset_y = (total_height_px - new_height) / 2
                    
                    # Đặt anchor tại B1 (bên trái cột B)
                    img.anchor = 'B1'
                    ws.add_image(img)
                    
                    # Set offset (1 pixel ≈ 9525 EMU at 96 DPI)
                    if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                        img.anchor._from.colOff = int(offset_x * 9525)
                        img.anchor._from.rowOff = int(offset_y * 9525)
                else:
                    # Fallback nếu không có dimension
                    size = int(min(total_height_px, total_width_px) * 0.85)
                    img.height = size
                    img.width = size
                    
                    offset_x = 0
                    offset_y = (total_height_px - size) / 2
                    
                    img.anchor = 'B1'
                    ws.add_image(img)
                    
                    if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                        img.anchor._from.colOff = int(offset_x * 9525)
                        img.anchor._from.rowOff = int(offset_y * 9525)
                    
            except Exception as e:
                pass

        # Tiêu đề
        row += 2
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = 'BIÊN BẢN BÀN GIAO'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].alignment = center_align
        ws.row_dimensions[row].height = 35

        # Số phiếu và ngày
        row += 1
        ws.merge_cells(f'A{row}:E{row}')
        now = datetime.now()
        ws[f'A{row}'] = f'(Số phiếu xuất: {picking.name} – ngày {now.strftime("%d")} tháng {now.strftime("%m")} năm {now.strftime("%Y")})'
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = center_align

        # Thông tin hai bên - merge toàn bộ dòng để tránh mất nội dung
        row += 2
        
        # Đại diện bên nhận (A) + Tên công ty trong cùng 1 dòng
        ws.merge_cells(f'A{row}:E{row}')
        partner_name = picking.partner_id.commercial_partner_id.name if picking.partner_id else ''
        text_content = f'Đại diện bên nhận (A): {partner_name}'
        ws[f'A{row}'] = text_content
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].alignment = left_align_wrap
        # Tính chiều cao tự động: A+B+C+D+E = 6+40+18+13+22 = 99 units
        ws.row_dimensions[row].height = self._calculate_row_height(text_content, 99, 13)

        # Ông (Bà)
        row += 1
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = 'Ông (Bà):'
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = left_align
        ws.row_dimensions[row].height = 18

        # Địa chỉ bên A
        row += 1
        ws.merge_cells(f'A{row}:E{row}')
        p = picking.partner_id.commercial_partner_id or picking.partner_id
        p_addr = ''
        if p:
            p_addr = (p.street or '') + (p.street2 and (', ' + p.street2) or '')
        text_content = f'Địa chỉ: {p_addr}'
        ws[f'A{row}'] = text_content
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = left_align_wrap
        ws.row_dimensions[row].height = self._calculate_row_height(text_content, 99, 13)

        # Đại diện bên giao (B) + Tên công ty trong cùng 1 dòng
        row += 2
        ws.merge_cells(f'A{row}:E{row}')
        company_name = picking.company_id.name or ''
        text_content = f'Đại diện bên giao (B): {company_name}'
        ws[f'A{row}'] = text_content
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].alignment = left_align_wrap
        ws.row_dimensions[row].height = self._calculate_row_height(text_content, 99, 13)

        # Ông (Bà)
        row += 1
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = 'Ông (Bà):'
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = left_align
        ws.row_dimensions[row].height = 18

        # Địa chỉ bên B
        row += 1
        ws.merge_cells(f'A{row}:E{row}')
        addr_b = picking.company_id.partner_id._display_address(without_company=True).replace('\n', ', ') or ''
        text_content = f'Địa chỉ: {addr_b}'
        ws[f'A{row}'] = text_content
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = left_align_wrap
        ws.row_dimensions[row].height = self._calculate_row_height(text_content, 99, 13)

        # Bên B đã bàn giao
        row += 2
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = 'Bên B đã bàn giao cho bên A:'
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].alignment = left_align

        # Bảng hàng hóa - Header
        row += 1
        header_row = row
        headers = ['STT', 'Tên hàng', 'Đơn vị tính', 'Số lượng', 'Ghi chú']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')

        # Data rows - Xử lý combo: Parent hiển thị bình thường, Children thụt vào
        row += 1
        idx = 0
        
        for line_data in enriched_lines:
            line_type = line_data['type']
            is_child = line_data['is_combo_child']
            qty = line_data['qty'] or 0.0
            product_name = line_data.get('product_name', '')
            uom_name = line_data.get('uom', '')  # Helper trả về 'uom' chứ không phải 'uom_name'

            # Chỉ render khi có qty > 0
            if qty > 0:
                # STT - chỉ hiển thị cho parent và standalone
                cell = ws.cell(row=row, column=1)
                if not is_child:
                    idx += 1
                    cell.value = idx
                else:
                    cell.value = ''  # Combo child không có STT
                cell.alignment = center_align
                cell.border = thin_border
                cell.font = normal_font

                # Tên hàng - thụt vào nếu là combo child
                cell = ws.cell(row=row, column=2)
                if is_child:
                    # Thụt vào 2 spaces cho combo child
                    cell.value = f'  {product_name}'
                else:
                    cell.value = product_name
                cell.alignment = left_align_wrap
                cell.border = thin_border
                cell.font = normal_font

                # Đơn vị tính
                cell = ws.cell(row=row, column=3)
                cell.value = uom_name
                cell.alignment = center_align
                cell.border = thin_border
                cell.font = normal_font

                # Số lượng
                cell = ws.cell(row=row, column=4)
                cell.value = round(qty, 2)
                cell.number_format = '0.00'
                cell.alignment = center_align
                cell.border = thin_border
                cell.font = normal_font

                # Ghi chú (trống)
                cell = ws.cell(row=row, column=5)
                cell.border = thin_border
                cell.font = normal_font

                row += 1

        # Nếu không có dòng nào
        if idx == 0:
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws.cell(row=row, column=1)
            cell.value = 'Không có dòng hàng.'
            cell.font = Font(name='Times New Roman', size=13, italic=True)
            cell.alignment = center_align
            cell.border = thin_border
            row += 1

        # Chữ ký
        row += 3
        ws.merge_cells(f'A{row}:B{row}')
        signature_text = 'ĐẠI DIỆN BÊN NHẬN\n(Ký, họ tên)'
        ws[f'A{row}'] = signature_text
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].alignment = center_align
        # A+B = 6+40 = 46 units, có 2 dòng
        ws.row_dimensions[row].height = self._calculate_row_height(signature_text, 46, 13)

        ws.merge_cells(f'C{row}:E{row}')
        ws[f'C{row}'] = 'ĐẠI DIỆN BÊN GIAO\n(Ký, họ tên)'
        ws[f'C{row}'].font = bold_font
        ws[f'C{row}'].alignment = center_align

        # Spacing cho chữ ký
        row += 5

        # Cấu hình in PDF
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.print_area = f'A1:E{row}'
        
        ws.page_margins.left = 0.7
        ws.page_margins.right = 0.7
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3
        
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = False

        # Tạo file
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Tạo attachment
        filename = f'BBBG_{picking.name}_{now.strftime("%d%m%Y")}.xlsx'
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': 'stock.picking',
            'res_id': picking.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }