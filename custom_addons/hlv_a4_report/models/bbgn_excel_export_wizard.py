# -*- coding: utf-8 -*-
from odoo import fields, models, api, _
from odoo.exceptions import UserError
import base64
from datetime import datetime
from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    Workbook = None
    XLImage = None


class BBGNExcelExportWizard(models.TransientModel):
    _name = 'bbgn.excel.export.wizard'
    _description = 'Xuất Excel BBGN không giá'

    def _get_active_picking(self):
        """Lấy picking từ context"""
        active_id = self._context.get('active_id') or self._context.get('active_ids', [False])[0]
        if active_id:
            return self.env['stock.picking'].browse(active_id)
        return False

    def action_export_excel(self):
        """Xuất file Excel BBGN không giá"""
        if not Workbook:
            raise UserError(_('Thư viện openpyxl chưa được cài đặt. Vui lòng chạy: pip install openpyxl'))

        picking = self._get_active_picking()
        if not picking:
            raise UserError(_('Không tìm thấy phiếu giao hàng'))

        # Lấy dữ liệu combo từ helper
        enriched_lines = self.env['hlv.report.helper'].get_enriched_lines_for_picking_combo(picking)

        # Tạo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'BBGN'

        # Định nghĩa styles - FONT SIZE VỪA PHẢI (không dùng fitToPage nên giữ font đọc được)
        header_font = Font(name='Times New Roman', size=14, bold=True)
        title_font = Font(name='Times New Roman', size=16, bold=True)
        normal_font = Font(name='Times New Roman', size=12)
        bold_font = Font(name='Times New Roman', size=12, bold=True)
        small_font = Font(name='Times New Roman', size=11)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Lấy thông tin
        so = picking.sale_id or (picking.move_ids and picking.move_ids[0].sale_line_id and picking.move_ids[0].sale_line_id.order_id)
        dt_src = picking.date_done or picking.scheduled_date or (so and so.date_order) or picking.create_date
        current_date = datetime.now()
        
        # PO number extraction
        po_raw = (so and so.origin) or picking.origin or ''
        po_upper = (po_raw or '').upper()
        if 'PO' in po_upper:
            po_number = po_upper.split('PO')[-1].strip().split()[0]
        else:
            po_number = ''.join([c for c in (po_raw or '') if c.isdigit()])

        # Set column widths - TỐI ƯU HÓA CHO IN (không dùng fitToPage nên có thể rộng hơn)
        col_a_width = 5    # STT
        col_b_width = 12   # Số PR
        ws.column_dimensions['A'].width = col_a_width
        ws.column_dimensions['B'].width = col_b_width
        ws.column_dimensions['C'].width = 38   # Tên hàng
        ws.column_dimensions['D'].width = 8    # DVT
        ws.column_dimensions['E'].width = 7    # SL
        ws.column_dimensions['F'].width = 13   # Đơn giá
        ws.column_dimensions['G'].width = 8    # VAT%
        ws.column_dimensions['H'].width = 13   # Thuế VAT
        ws.column_dimensions['I'].width = 14   # Thành tiền
        ws.column_dimensions['J'].width = 12   # Ghi chú

        # Header - Logo và thông tin công ty
        row = 1

        # Tính toán kích thước khung A1:B4 cho logo
        # Excel column width công thức chính xác cho merged cells:
        # pixels = ((256 * width + 128/7)/256) * 7 + 5 (nếu width > 1)
        # Với Times New Roman và merged cells A1:B4
        def excel_column_width_to_pixels(width):
            """Chuyển đổi Excel column width sang pixels (chính xác)"""
            if width <= 1:
                return int(((256 * width + 18) / 256) * 7)
            else:
                return int(((256 * width + 18) / 256) * 7) + 5
        
        logo_cell_width_px = excel_column_width_to_pixels(col_a_width) + excel_column_width_to_pixels(col_b_width)
        
        # Excel row height: 1 point = 96/72 pixels (96 DPI)
        # 1 point = 1.333... pixels
        logo_row_height_points = 30
        logo_cell_height_px = 4 * logo_row_height_points * 96 / 72  # = 160 pixels
        
        # Điều chỉnh chiều cao các hàng header
        ws.row_dimensions[1].height = logo_row_height_points
        ws.row_dimensions[2].height = logo_row_height_points
        ws.row_dimensions[3].height = logo_row_height_points
        ws.row_dimensions[4].height = logo_row_height_points

        # Thêm logo nếu có (fit vào khung A1:B4)
        if picking.company_id.logo and XLImage:
            try:
                logo_data = base64.b64decode(picking.company_id.logo)
                logo_stream = BytesIO(logo_data)
                img = XLImage(logo_stream)
                
                if getattr(img, "width", None) and getattr(img, "height", None) and img.width > 0:
                    # Tính tỷ lệ để fit vào khung, giữ aspect ratio (logo vuông)
                    width_ratio = logo_cell_width_px / float(img.width)
                    height_ratio = logo_cell_height_px / float(img.height)
                    ratio = min(width_ratio, height_ratio)  # Giữ logo vừa trong khung
                    
                    # Scale logo với padding nhỏ (2%)
                    new_width = int(img.width * ratio * 0.98)
                    new_height = int(img.height * ratio * 0.98)
                    img.width = new_width
                    img.height = new_height
                    
                    # Tính toán offset để căn giữa logo trong khung A1:B4
                    # Khoảng trống bên trái = (chiều rộng khung - chiều rộng logo) / 2
                    offset_x = (logo_cell_width_px - new_width) / 2
                    offset_y = (logo_cell_height_px - new_height) / 2
                    
                    # Anchor với offset để căn giữa
                    # openpyxl anchor format: từ ô A1, dịch chuyển theo pixels
                    img.anchor = 'A1'
                    # Điều chỉnh vị trí bằng cách set anchor properties
                    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
                    # Tạo anchor marker với offset
                    img.anchor = f'A1'
                    
                    # Thêm logo vào worksheet
                    ws.add_image(img, 'A1')
                    
                    # Điều chỉnh anchor sau khi thêm để căn giữa
                    if hasattr(img, '_positioned'):
                        # Set offset từ góc trên-trái của cell A1
                        img.anchor._from.colOff = int(offset_x * 9525)  # Convert px to EMU (1px = 9525 EMU)
                        img.anchor._from.rowOff = int(offset_y * 9525)
                else:
                    # Fallback: logo vuông, căn giữa
                    size = int(min(logo_cell_width_px, logo_cell_height_px) * 0.98)
                    img.width = size
                    img.height = size
                    
                    offset_x = (logo_cell_width_px - size) / 2
                    offset_y = (logo_cell_height_px - size) / 2
                    
                    ws.add_image(img, 'A1')
                    if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                        img.anchor._from.colOff = int(offset_x * 9525)
                        img.anchor._from.rowOff = int(offset_y * 9525)

                ws.merge_cells('A1:B4')
            except Exception as e:
                pass

        # Thông tin công ty bên phải
        # Row 1: Tên công ty (font 14, bold)
        ws.merge_cells(f'C{row}:J{row}')
        ws[f'C{row}'] = 'CÔNG TY TNHH VI NA HOÀNG LONG VŨ'
        ws[f'C{row}'].font = header_font
        ws[f'C{row}'].alignment = left_align

        # Row 2: Địa chỉ
        row += 1
        ws.merge_cells(f'C{row}:J{row}')
        addr = picking.company_id.partner_id._display_address(without_company=True).replace('\n', ' ') or ''
        ws[f'C{row}'] = f'Địa chỉ: {addr}'
        ws[f'C{row}'].font = normal_font
        ws[f'C{row}'].alignment = left_align

        # Row 3: Điện thoại và Di động (merge để tránh bị kéo dãn bởi cột C)
        row += 1
        ws.merge_cells(f'C{row}:D{row}')
        ws[f'C{row}'] = f'Điện thoại: {picking.company_id.phone or ""}'
        ws[f'C{row}'].font = normal_font
        ws[f'C{row}'].alignment = left_align
        
        ws.merge_cells(f'E{row}:J{row}')
        ws[f'E{row}'] = f'Di động: {picking.company_id.mobile or picking.company_id.partner_id.mobile or ""}'
        ws[f'E{row}'].font = normal_font
        ws[f'E{row}'].alignment = left_align

        # Row 4: Email và Website (merge để tránh bị kéo dãn bởi cột C)
        row += 1
        ws.merge_cells(f'C{row}:D{row}')
        ws[f'C{row}'] = f'Email: {picking.company_id.email or ""}'
        ws[f'C{row}'].font = normal_font
        ws[f'C{row}'].alignment = left_align
        
        ws.merge_cells(f'E{row}:J{row}')
        ws[f'E{row}'] = f'Website: {picking.company_id.website or ""}'
        ws[f'E{row}'].font = normal_font
        ws[f'E{row}'].alignment = left_align

        # Tiêu đề
        row += 2
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = 'BIÊN BẢN GIAO NHẬN HÀNG HÓA'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].alignment = center_align
        ws.row_dimensions[row].height = 30

        row += 1
        ws.merge_cells(f'A{row}:J{row}')
        year_str = dt_src.strftime('%Y') if dt_src else current_date.strftime('%Y')
        ws[f'A{row}'] = f'(No.: {picking.name} ; Date {current_date.strftime("%d/%m/%Y")})'
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = center_align

        # Căn cứ đơn hàng
        row += 2
        ws.merge_cells(f'A{row}:J{row}')
        partner_name = so.partner_id.commercial_partner_id.name if so and so.partner_id else ''
        ws[f'A{row}'] = f'Căn cứ đơn đặt hàng ngày {current_date.strftime("%d/%m/%Y")} PO số: {po_number} của {partner_name}'
        ws[f'A{row}'].font = Font(name='Times New Roman', size=12, italic=True)
        ws[f'A{row}'].alignment = left_align

        row += 1
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = f'Hôm nay, ngày {current_date.strftime("%d/%m/%Y")} tại {partner_name}, Chúng tôi gồm:'
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = left_align

        # BÊN A
        row += 2
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = f'BÊN A (Bên nhận hàng): {picking.partner_id.commercial_partner_id.name if picking.partner_id else ""}'
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].alignment = left_align

        row += 1
        ws.merge_cells(f'A{row}:J{row}')
        p = picking.partner_id.commercial_partner_id or picking.partner_id
        p_addr = p._display_address(without_company=True).replace('\n', ' ') if p else ''
        ws[f'A{row}'] = f'    Địa chỉ: {p_addr}'
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = left_align

        row += 1
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = '    Điện thoại: ................................................................    Fax: ................................................................'
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = left_align

        row += 1
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = '    Đại diện Ông/bà: ................................................................    Chức vụ: ................................................................'
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = left_align

        # BÊN B
        row += 2
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = 'BÊN B (Bên giao hàng): CÔNG TY TNHH VI NA HOÀNG LONG VŨ'
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].alignment = left_align

        row += 1
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = f'    Địa chỉ: {addr}'
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = left_align

        row += 1
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = f'    Điện thoại: {picking.company_id.phone or ""}    ................................................................    Fax: ................................................................'
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = left_align

        row += 1
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = '    Đại diện Ông/bà: ................................................................    Chức vụ: ................................................................'
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = left_align

        # Thống nhất giao hàng
        row += 2
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = 'Hai bên cùng thống nhất số lượng giao hàng như sau:'
        ws[f'A{row}'].font = Font(name='Times New Roman', size=12, italic=True)
        ws[f'A{row}'].alignment = left_align

        # Bảng hàng hóa - Header
        row += 1
        header_row = row
        headers = ['STT', 'Số PR', 'Tên hàng', 'DVT', 'SL', 'Đơn giá', 'Vat(%)', 'Vat', 'Thành Tiền', 'Ghi chú']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')

        # Data rows - XỬ LÝ COMBO ĐÚNG: Parent có giá, Children chỉ hiển thị tên thụt vào
        row += 1
        idx = 1
        total_amount_untaxed = 0.0  # Tổng tiền hàng (chưa thuế)
        total_tax = 0.0  # Tổng thuế VAT
        
        for line_data in enriched_lines:
            line_type = line_data['type']
            is_child = line_data['is_combo_child']
            qty = line_data['qty'] or 0.0
            
            # Chỉ tính giá cho PARENT và STANDALONE (không tính cho CHILD)
            if line_type == 'parent' or line_type == 'standalone':
                unit_price = line_data.get('price_unit', 0.0) or 0.0
                tax_percent = line_data.get('tax_percent', 0.0)
                line_total = qty * unit_price
                line_tax_value = line_total * tax_percent / 100.0
                
                # Cộng dồn tổng
                total_amount_untaxed += line_total
                total_tax += line_tax_value
            else:
                # CHILD: không có giá
                unit_price = 0.0
                tax_percent = 0.0
                line_total = 0.0
                line_tax_value = 0.0
            
            # STT - Chỉ đánh số cho parent và standalone
            cell = ws.cell(row=row, column=1)
            if line_type != 'child':
                cell.value = idx
                idx += 1
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = normal_font

            # Số PR - Lấy từ sale.order.line.note (chỉ cho parent và standalone)
            cell = ws.cell(row=row, column=2)
            if line_type != 'child':
                sol = line_data.get('sol')
                cell.value = (sol and sol.note) or ''
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = normal_font

            # Tên hàng - CHILD thụt vào 4 spaces
            cell = ws.cell(row=row, column=3)
            product_name = line_data['product_name']
            if is_child:
                product_name = '    ' + product_name  # Indent child
            cell.value = product_name
            cell.alignment = left_align
            cell.border = thin_border
            cell.font = normal_font

            # DVT
            cell = ws.cell(row=row, column=4)
            cell.value = line_data['uom']
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = normal_font

            # SL
            cell = ws.cell(row=row, column=5)
            cell.value = round(qty, 2)
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = normal_font

            # Đơn giá - CHỈ hiển thị cho parent và standalone
            cell = ws.cell(row=row, column=6)
            if line_type != 'child':
                cell.value = round(unit_price, 0)
                cell.number_format = '#,##0'
            cell.alignment = right_align
            cell.border = thin_border
            cell.font = normal_font

            # VAT(%) - CHỈ hiển thị cho parent và standalone
            cell = ws.cell(row=row, column=7)
            if line_type != 'child':
                cell.value = f"{tax_percent}%"
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = normal_font

            # Thuế VAT - CHỈ hiển thị cho parent và standalone
            cell = ws.cell(row=row, column=8)
            if line_type != 'child':
                cell.value = round(line_tax_value, 0)
                cell.number_format = '#,##0'
            cell.alignment = right_align
            cell.border = thin_border
            cell.font = normal_font

            # Thành tiền - CHỈ hiển thị cho parent và standalone
            cell = ws.cell(row=row, column=9)
            if line_type != 'child':
                cell.value = round(line_total, 0)
                cell.number_format = '#,##0'
            cell.alignment = right_align
            cell.border = thin_border
            cell.font = normal_font

            # Ghi chú (trống)
            cell = ws.cell(row=row, column=10)
            cell.border = thin_border
            cell.font = normal_font

            row += 1

        # Nếu không có dòng nào
        if not enriched_lines:
            ws.merge_cells(f'A{row}:J{row}')
            cell = ws.cell(row=row, column=1)
            cell.value = 'Không có dòng hàng.'
            cell.font = Font(name='Times New Roman', size=12, italic=True)
            cell.alignment = center_align
            cell.border = thin_border
            row += 1

        # 3 hàng tổng - ĐIỀN GIÁ TRỊ
        grey_fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
        
        # Tính tổng tiền thanh toán
        total_amount_total = total_amount_untaxed + total_tax
        
        # Row 1: Tổng tiền hàng
        for col in range(1, 6):  # A-E
            for r in range(row, row + 3):
                cell = ws.cell(row=r, column=col)
                cell.fill = grey_fill
                cell.border = thin_border
        ws.merge_cells(f'A{row}:E{row+2}')

        ws.merge_cells(f'F{row}:H{row}')
        for col in range(6, 9):  # F-H
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
        cell = ws.cell(row=row, column=6)
        cell.value = 'Tổng tiền hàng (VNĐ)'
        cell.font = bold_font
        cell.alignment = left_align

        cell = ws.cell(row=row, column=9)
        cell.value = round(total_amount_untaxed, 0)
        cell.number_format = '#,##0'
        cell.alignment = right_align
        cell.font = bold_font
        cell.border = thin_border

        for r in range(row, row + 3):
            cell = ws.cell(row=r, column=10)
            cell.fill = grey_fill
            cell.border = thin_border
        ws.merge_cells(f'J{row}:J{row+2}')

        row += 1

        # Row 2: Tổng thuế VAT
        ws.merge_cells(f'F{row}:H{row}')
        for col in range(6, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
        cell = ws.cell(row=row, column=6)
        cell.value = 'Tổng thuế VAT (VNĐ)'
        cell.font = bold_font
        cell.alignment = left_align

        cell = ws.cell(row=row, column=9)
        cell.value = round(total_tax, 0)
        cell.number_format = '#,##0'
        cell.alignment = right_align
        cell.font = bold_font
        cell.border = thin_border

        row += 1

        # Row 3: Tổng tiền thanh toán
        ws.merge_cells(f'F{row}:H{row}')
        for col in range(6, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
        cell = ws.cell(row=row, column=6)
        cell.value = 'Tổng tiền thanh toán (VNĐ)'
        cell.font = bold_font
        cell.alignment = left_align

        cell = ws.cell(row=row, column=9)
        cell.value = round(total_amount_total, 0)
        cell.number_format = '#,##0'
        cell.alignment = right_align
        cell.font = bold_font
        cell.border = thin_border

        row += 1

        # Bằng chữ - CHUYỂN ĐỔI SỐ THÀNH CHỮ
        ws.merge_cells(f'A{row}:E{row}')
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = grey_fill
            cell.border = thin_border
        cell = ws.cell(row=row, column=1)
        cell.value = 'Bằng chữ:'
        cell.font = bold_font
        cell.alignment = left_align

        ws.merge_cells(f'F{row}:J{row}')
        for col in range(6, 11):
            cell = ws.cell(row=row, column=col)
            cell.fill = grey_fill
            cell.border = thin_border
        # Chuyển số thành chữ (Vietnamese) - Sử dụng hlv.amount
        amount_text = self.env['hlv.amount'].amount_to_text_vn(total_amount_total)
        cell = ws.cell(row=row, column=6)
        cell.value = amount_text
        cell.font = Font(name='Times New Roman', size=12, italic=True)
        cell.alignment = left_align

        # Xác nhận
        row += 2
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = 'Bên A xác nhận Bên B đã giao cho Bên A đúng chủng loại và đủ số lượng hàng như trên.'
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = left_align

        row += 1
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = 'Hai bên đồng ý, thống nhất ký tên. Biên bản được lập thành 05 bản, bên mua giữ 04 bản, bên bán giữ 01 bản và có giá trị pháp lý như nhau.'
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = left_align

        row += 2
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = 'Giao hàng tại kho: ................... Người nhận hàng: ................... SĐT người nhận hàng: ...................'
        ws[f'A{row}'].font = Font(name='Times New Roman', size=12, italic=True, bold=True)
        ws[f'A{row}'].alignment = left_align

        # Chữ ký
        row += 2
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = 'Đại diện bên nhận hàng'
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].alignment = center_align

        ws.merge_cells(f'E{row}:J{row}')
        ws[f'E{row}'] = 'Đại diện bên giao hàng'
        ws[f'E{row}'].font = bold_font
        ws[f'E{row}'].alignment = center_align

        # Thêm spacing cho chữ ký (5 rows trống)
        row += 5

        # ===== CẤU HÌNH IN PDF TỐI ƯU =====
        # Set page setup - KHÔNG dùng fitToPage để tránh nội dung bị thu nhỏ quá
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # Khổ dọc
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        
        # KHÔNG set fitToHeight/fitToWidth để giữ kích thước font gốc
        # ws.page_setup.fitToHeight = 1  
        # ws.page_setup.fitToWidth = 1
        
        # Set print area (chỉ nội dung chính, không bao gồm footer)
        ws.print_area = f'A1:J{row}'
        
        # Margins - Vừa phải để tối ưu không gian nhưng không làm mất nội dung
        ws.page_margins.left = 0.7
        ws.page_margins.right = 0.7
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 1.0  # Tăng bottom margin cho footer
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.5
        
        # Print options
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = False
        
        # KHÔNG dùng fitToPage - để giữ kích thước font gốc
        # ws.sheet_properties.pageSetUpPr.fitToPage = True
        
        # Đặt footer ở cuối trang (sẽ luôn nằm ở cuối mọi trang khi in)
        footer_text = 'PHÂN PHỐI CHÍNH HÃNG: SKF-NSK-KOYO-NTN-ASAHI-IKO-MITSUBOSHI-LS-HANYOUNG-BOSCH-MAKITA-MILWAUKEE-DEWALT'
        ws.oddFooter.center.text = footer_text
        ws.oddFooter.center.size = 10
        ws.oddFooter.center.font = "Times New Roman"
        
        # Footer cho trang chẵn (nếu in 2 mặt)
        ws.evenFooter.center.text = footer_text
        ws.evenFooter.center.size = 10
        ws.evenFooter.center.font = "Times New Roman"

        # Tạo file
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Tạo attachment
        filename = f'BBGN_{picking.name}_{current_date.strftime("%d%m%Y")}.xlsx'
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': 'stock.picking',
            'res_id': picking.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }