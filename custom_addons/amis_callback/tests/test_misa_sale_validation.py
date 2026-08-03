# -*- coding: utf-8 -*-
from openpyxl import Workbook

from odoo.tests.common import TransactionCase

from odoo.addons.misa_fetch_po_button.utils.misa_api_utils import (
    MisaApiUtils,
    _build_crm_combo_payload,
)

from ..models.misa_sale_validation_wizard import (
    _analyze_validation_sheet,
    _build_result_workbook,
    _find_validation_header,
)


class TestMisaSaleValidation(TransactionCase):

    def test_exact_crm_product_lookup_uses_exact_result(self):
        class FakeResponse:
            ok = True
            status_code = 200
            text = "ok"

            @staticmethod
            def json():
                return {"Success": True, "Data": [
                    {"ID": 1, "ProductCode": "COMP-A-OTHER"},
                    {
                        "ID": 46720,
                        "ProductCode": "COMP-A",
                        "ProductName": "Component A",
                        "UsageUnitID": 4,
                        "UsageUnitIDText": "Cái",
                    },
                ]}

        class FakeSession:
            payload = None

            def post(self, _url, headers, json, timeout):
                self.payload = json
                self.headers = headers
                self.timeout = timeout
                return FakeResponse()

        class FakeApi:
            session = FakeSession()

            def _get_retry_session(self):
                return self.session

        fake_api = FakeApi()
        result = MisaApiUtils._find_exact_crm_product_by_code(
            fake_api, "comp-a", headers={"Authorization": "Bearer test"},
        )

        self.assertEqual(result["misa_id"], 46720)
        self.assertEqual(result["unit_id"], 4)
        self.assertEqual(fake_api.session.payload["Filters"][0]["Operator"], 1)

    def test_build_crm_combo_payload(self):
        payload = _build_crm_combo_payload({
            "code": "CB-TEST",
            "name": "Combo test",
            "category_id": 164,
            "category_name": "Hàng hóa",
            "unit_id": 3,
            "unit_name": "Bộ",
            "tax_id": "5",
            "tax_name": "8%",
            "form_layout_id": 128,
        }, [{
            "misa_id": 46720,
            "code": "COMP-A",
            "name": "Component A",
            "unit_id": 4,
            "unit_name": "Cái",
            "quantity": 2,
        }])

        self.assertEqual(payload["FormLayoutID"], 128)
        self.assertEqual(payload["IsSetProduct"], "1")
        custom_table = payload["CustomTables"][0]
        self.assertEqual(custom_table["TableName"], "set_product")
        self.assertEqual(custom_table["Summary"]["AmountSummary"], 2)
        self.assertEqual(custom_table["Data"][0]["ProductID"], 46720)
        self.assertEqual(custom_table["Data"][0]["ProductIDText"], "COMP-A")

    @staticmethod
    def _validation_sheet():
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append([
            "Tình Trạng",
            "Chi tiết lỗi",
            "Hình thức bán hàng",
            "Số chứng từ (*)",
            "Mã hàng (*)",
            "Thuộc combo",
            "ĐVT",
        ])
        worksheet.append([None, None, "Bán hàng", "BH01", "CB-A", None, "Bộ"])
        worksheet.append([None, "Lỗi liên quan đến dòng <2>.", "Bán hàng", "BH01", "P1", "CB-A", "Cái"])
        worksheet.append([
            "Không hợp lệ",
            "Mã hàng <CB-MISSING> không có trong danh mục",
            "Bán hàng", "BH01", "CB-MISSING", None, "Bộ",
        ])
        worksheet.append([
            None, "Combo <CB-MISSING> chưa được khai báo trong danh mục.",
            "Bán hàng", "BH01", "P2", "CB-MISSING", "Cái",
        ])
        worksheet.append([
            "Không hợp lệ",
            "Trên 1 chứng từ bán hàng mỗi combo chỉ được nhập khẩu dữ liệu 1 lần",
            "Bán hàng", "BH01", "CB-A", None, "Bộ",
        ])
        worksheet.append([None, "Lỗi liên quan đến dòng <6>.", "Bán hàng", "BH01", "P1", "CB-A", "Cái"])
        worksheet.append([
            "Không hợp lệ",
            "Đơn vị tính <Bộ> không phải là đơn vị tính chính hoặc là đơn vị chuyển đổi của mặt hàng <11-921T>",
            "Bán hàng", "BH01", "11-921T", None, "Bộ",
        ])
        return workbook, worksheet

    def test_analyze_and_split_validation_result(self):
        source_wb, worksheet = self._validation_sheet()
        header_row, columns = _find_validation_header(worksheet)
        analysis = _analyze_validation_sheet(worksheet, header_row, columns)

        self.assertEqual(analysis["missing_codes"], ["CB-MISSING"])
        self.assertEqual(analysis["uom_fixes"], {8: "Bộ."})
        self.assertEqual(len(analysis["duplicate_occurrences"]), 1)
        self.assertEqual(analysis["duplicate_rows"], {6, 7})
        self.assertEqual(analysis["related_error_rows"], {3, 7})

        result_wb = _build_result_workbook(
            worksheet, header_row, columns, analysis,
        )
        self.assertEqual(result_wb.sheetnames, ["Phieu ban hang", "Combo bi trung"])
        self.assertEqual(result_wb["Phieu ban hang"].max_column, 5)
        self.assertEqual(result_wb["Phieu ban hang"].max_row, 6)
        self.assertEqual(result_wb["Phieu ban hang"].cell(6, 5).value, "Bộ.")
        self.assertEqual(result_wb["Combo bi trung"].max_row, 3)

        result_wb.close()
        source_wb.close()
