# -*- coding: utf-8 -*-
import base64
import datetime
import logging
from io import BytesIO

from odoo import api, fields, models
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None
    Workbook = None

try:
    import xlrd
except ImportError:
    xlrd = None

_logger = logging.getLogger(__name__)

# ── 52 cột mẫu MISA Phiếu Bán Hàng ─────────────────────────────────────────
MISA_COLUMNS = [
    'Hình thức bán hàng',
    'Phương thức thanh toán',
    'Kiêm phiếu xuất kho',
    'Lập kèm hóa đơn',
    'Đã lập hóa đơn',
    'Ngày hạch toán (*)',
    'Ngày chứng từ (*)',
    'Số chứng từ (*)',
    'Số phiếu xuất',
    'Mẫu số HĐ',
    'Ký hiệu HĐ',
    'Số hóa đơn',
    'Ngày hóa đơn',
    'Mã khách hàng',
    'Tên khách hàng',
    'Địa chỉ',
    'Mã số thuế',
    'Đơn vị giao đại lý',
    'Người nộp',
    'Nộp vào TK',
    'Tên ngân hàng',
    'Diễn giải/Lý do nộp',
    'Lý do xuất',
    'Mã hàng (*)',
    'Thuộc combo',
    'Tên hàng',
    'Là dòng ghi chú',
    'Hàng khuyến mại',
    'Chiết khấu thương mại',
    'TK Tiền/Chi phí/Nợ (*)',
    'TK Doanh thu/Có (*)',
    'ĐVT',
    'Số lượng',
    'Đơn giá',
    'Thành tiền',
    'Tỷ lệ CK (%)',
    'Tiền chiết khấu',
    'TK chiết khấu',
    'Giá tính thuế XK',
    '% thuế xuất khẩu',
    'Tiền thuế xuất khẩu',
    'TK thuế xuất khẩu',
    '% thuế GTGT',
    '% thuế suất KHAC',
    'Tiền thuế GTGT',
    'TK thuế GTGT',
    'HH không TH trên tờ khai thuế GTGT',
    'Mã kho',
    'TK giá vốn',
    'TK Kho',
    'Đơn giá vốn',
    'Tiền vốn',
    'Hàng hóa giữ hộ/bán hộ',
]


class ShopeeWalletImportWizard(models.TransientModel):
    _name = 'shopee.wallet.import.wizard'
    _description = 'Import Báo Cáo Ví Shopee → Xuất Phiếu Bán Hàng MISA'

    file_data = fields.Binary(
        string='File báo cáo ví Shopee',
        required=True,
        attachment=False,
        help='Upload file Excel xuất từ Shopee Seller Center (Báo cáo tài chính / Ví Shopee).',
    )
    file_name = fields.Char(string='Tên file')
    result_summary = fields.Text(string='Kết quả', readonly=True)

    # ── Đọc file ────────────────────────────────────────────────────────────

    def _read_shopee_file(self):
        """Đọc file XLS/XLSX, trả về list[list] các rows."""
        if not self.file_data:
            raise UserError('Vui lòng upload file báo cáo ví Shopee.')
        raw = base64.b64decode(self.file_data)
        fname = (self.file_name or '').lower().strip()
        rows = []

        if fname.endswith('.xlsx') or fname.endswith('.xlsm'):
            if not openpyxl:
                raise UserError('Thiếu thư viện openpyxl để đọc file XLSX.')
            wb = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            wb.close()
        else:
            # XLS (mặc định)
            if not xlrd:
                raise UserError('Thiếu thư viện xlrd để đọc file XLS.')
            book = xlrd.open_workbook(file_contents=raw)
            sheet = book.sheet_by_index(0)
            for r in range(sheet.nrows):
                rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
        return rows

    def _find_header_row_idx(self, rows):
        """Tìm index của row header (chứa 'Mã đơn hàng' ở cột D hoặc gần đó)."""
        for i, row in enumerate(rows):
            # Quét 5 cột đầu để tìm 'mã đơn hàng'
            for cell in list(row)[:8]:
                if cell and 'mã đơn hàng' in str(cell).lower():
                    return i
        return 17  # fallback: row 18 (0-based = 17)

    def _extract_shopee_codes(self, rows):
        """Trích xuất list mã đơn Shopee duy nhất từ cột D (index 3)."""
        header_idx = self._find_header_row_idx(rows)
        codes = []
        seen = set()
        for row in rows[header_idx + 1:]:
            if len(row) <= 3:
                continue
            code = str(row[3] or '').strip()
            if code and code.lower() not in ('', 'none', '-', 'nan') and code not in seen:
                codes.append(code)
                seen.add(code)
        return codes

    def _get_customer_map(self):
        """
        Trả về dict: shop_identifier → {'name': str, 'code': str}
        Lấy từ amis.callback.config.shopee_customer_map_ids (persistent).
        Fallback hardcode nếu config chưa có.
        """
        DEFAULTS = {
            '1357810112': {'name': 'KHÁCH HÀNG KHÔNG CUNG CẤP THÔNG TIN_SHOPEE DEWALT', 'code': ''},
            '796817584':  {'name': 'KHÁCH HÀNG KHÔNG CUNG CẤP THÔNG TIN_SHOPEE MILWAUKEE', 'code': ''},
            '326259406':  {'name': 'KHÁCH HÀNG KHÔNG CUNG CẤP THÔNG TIN_SHOPEE HLV', 'code': ''},
        }
        config = self.env['amis.callback.config'].sudo().search([], limit=1)
        if not config or not config.shopee_customer_map_ids:
            return DEFAULTS
        result = dict(DEFAULTS)  # bắt đầu từ default, config ghi đè
        for m in config.shopee_customer_map_ids:
            ident = (m.shop_identifier or '').strip()
            if ident:
                result[ident] = {
                    'name': m.customer_name or '',
                    'code': m.customer_code or '',
                }
        return result

    def _find_sale_orders(self, shopee_codes):
        """Tìm sale.order theo shopee_order_ref. Trả về dict code→SO (hoặc False)."""
        result = {}
        if not shopee_codes:
            return result
        sos = self.env['sale.order'].sudo().search([
            ('shopee_order_ref', 'in', shopee_codes),
        ])
        so_map = {}
        for so in sos:
            ref = getattr(so, 'shopee_order_ref', '') or ''
            if ref and ref not in so_map:
                so_map[ref] = so
        for code in shopee_codes:
            result[code] = so_map.get(code, False)
        return result

    # ── Helpers ──────────────────────────────────────────────────────────────

    def _get_kit_bom(self, product):
        """
        Trả về mrp.bom đầu tiên có type='phantom' (Kit) cho sản phẩm.
        Ưu tiên BOM khớp cả product_id, fallback về product_tmpl_id.
        Trả về False nếu không có.
        """
        MrpBom = self.env['mrp.bom'].sudo()
        bom = MrpBom.search([
            ('type', '=', 'phantom'),
            ('product_id', '=', product.id),
        ], limit=1)
        if not bom:
            bom = MrpBom.search([
                ('type', '=', 'phantom'),
                ('product_tmpl_id', '=', product.product_tmpl_id.id),
                ('product_id', '=', False),
            ], limit=1)
        return bom or False

    # ── Xây dựng rows MISA ───────────────────────────────────────────────────

    def _build_misa_rows(self, code_so_map):
        """Xây dựng danh sách rows cho MISA Phiếu Bán Hàng."""
        misa_rows = []

        # ── Hằng số cho cả batch xuất ────────────────────────────────────────
        today = datetime.date.today()
        today_str = today.strftime('%d/%m/%Y')
        date_tag = today.strftime('%d%m%Y')        # DDMMYYYY
        so_ref = 'BH%sSP' % date_tag              # VD: BH12052026SP  (12 ký tự)
        xk_ref = 'XK%sSP' % date_tag              # VD: XK12052026SP
        customer_map = self._get_customer_map()    # dict identifier → {name, code}

        for shopee_code, so in code_so_map.items():
            if not so:
                # Không tìm thấy SO: 1 dòng cảnh báo
                row = [''] * len(MISA_COLUMNS)
                row[7] = so_ref
                row[21] = 'KHÔNG TÌM THẤY SO | Shopee: %s' % shopee_code
                misa_rows.append(row)
                continue

            # ── Thông tin shop ─────────────────────────────────────────────
            shop_name = ''
            shop_identifier = ''
            if hasattr(so, 'shopee_shop_id') and so.shopee_shop_id:
                shop_name = so.shopee_shop_id.name or ''
                shop_identifier = str(getattr(so.shopee_shop_id, 'shop_identifier', '') or '').strip()

            # ── Tên/mã khách hàng từ map theo shop ────────────────────────
            cust_info = customer_map.get(shop_identifier, {})
            customer_name = cust_info.get('name', '') or shop_name or 'Khách hàng không cung cấp thông tin'
            customer_code = cust_info.get('code', '')

            # ── Thông tin đơn xuất (picking) ──────────────────────────────
            pickings = so.picking_ids.filtered(lambda p: p.state == 'done' and p.picking_type_code == 'outgoing')
            picking_names = ', '.join(p.name for p in pickings) if pickings else ''

            # ── meInvoice ─────────────────────────────────────────────────
            inv_no = getattr(so, 'misa_meinvoice_inv_no', '') or ''
            inv_date_val = getattr(so, 'misa_meinvoice_inv_date', False)
            inv_date = ''
            if inv_date_val:
                inv_date = inv_date_val.strftime('%d/%m/%Y') if hasattr(inv_date_val, 'strftime') else str(inv_date_val)[:10]
            da_lap_hd = 'Có' if getattr(so, 'misa_meinvoice_synced', False) else '' # khúc này xem lại hard cho chưa lập luôn

            # ── Diễn giải ─────────────────────────────────────────────────
            dien_giai = 'Đơn bán: %s | Shopee: %s | Shop: %s' % (so.name, shopee_code, shop_name)
            ly_do_xuat = ('Đơn xuất: %s' % picking_names) if picking_names else ''

            # ── Từng dòng sản phẩm ────────────────────────────────────────
            for line in so.order_line:
                if line.display_type:  # skip section/note
                    continue

                product = line.product_id
                uom_name = line.product_uom.name if line.product_uom else ''
                line_qty = line.product_uom_qty

                # Thuế GTGT (từ dòng SO cha)
                vat_rate = 0
                for tax in (line.tax_id or []):
                    if tax.amount_type == 'percent' and tax.amount > 0:
                        vat_rate = tax.amount
                        break

                discount_pct = line.discount or 0
                price = line.price_unit
                subtotal = line.price_total
                discount_amt = round(price * discount_pct / (100 + vat_rate), 2) if discount_pct else 0
                vat_amount = round(subtotal * vat_rate / 100, 2) if vat_rate else 0

                # ── Kiểm tra BOM Kit (combo) ──────────────────────────────
                kit_bom = self._get_kit_bom(product) if product else False

                if kit_bom:
                    # Xuất dòng đầu: sản phẩm cha với giá gốc
                    row = [''] * len(MISA_COLUMNS)
                    row[0] = 'Bán hàng hóa trong nước'
                    row[1] = 'Chưa thu tiền'
                    row[2] = 'Có'
                    row[3] = 'Không' # Lập kèm hóa đơn
                    row[4] = 'Chưa lập'
                    row[5] = today_str
                    row[6] = today_str
                    row[7] = so_ref
                    row[8] = xk_ref
                    row[9] = '1'
                    row[10] = '1C26TLV'
                    row[11] = inv_no
                    row[12] = inv_date
                    row[13] = customer_code
                    row[14] = customer_name
                    row[15] = ''
                    row[21] = dien_giai
                    row[22] = ly_do_xuat
                    row[23] = product.default_code or ''
                    row[24] = ''                           # Thuộc combo (SP cha — để trống)
                    row[25] = product.name or line.name or ''
                    row[29] = '131'
                    row[30] = '5111'
                    row[31] = uom_name
                    row[32] = line_qty
                    row[33] = price
                    row[34] = subtotal
                    row[35] = discount_pct if discount_pct else ''
                    row[36] = discount_amt if discount_amt else ''
                    row[42] = vat_rate if vat_rate else ''
                    row[44] = vat_amount if vat_amount else ''
                    row[45] = '3331' if vat_rate else ''
                    row[47] = 'HLV'
                    row[48] = '632'
                    row[49] = '1561'
                    misa_rows.append(row)

                    # Xuất các dòng con (BOM components) với giá 0
                    for bom_line in kit_bom.bom_line_ids:
                        comp = bom_line.product_id
                        if not comp:
                            continue
                        comp_qty = bom_line.product_qty * line_qty
                        comp_uom = bom_line.product_uom_id.name if bom_line.product_uom_id else ''
                        comp_row = [''] * len(MISA_COLUMNS)
                        comp_row[0] = 'Bán hàng hóa trong nước'
                        comp_row[1] = 'Chưa thu tiền'
                        comp_row[2] = 'Có'
                        comp_row[3] = 'Không'                         # Lập kèm hóa đơn
                        # comp_row[4] = da_lap_hd
                        comp_row[4] =  "Chưa lập"                # Đã lập hóa đơn
                        comp_row[5] = today_str
                        comp_row[6] = today_str
                        comp_row[7] = so_ref
                        comp_row[8] = xk_ref
                        comp_row[9] = '1'
                        comp_row[10] = '1C26TLV'
                        comp_row[11] = inv_no
                        comp_row[12] = inv_date
                        comp_row[13] = customer_code
                        comp_row[14] = customer_name
                        comp_row[15] = ''
                        comp_row[21] = dien_giai
                        comp_row[22] = ly_do_xuat
                        comp_row[23] = comp.default_code or ''
                        comp_row[24] = product.default_code or ''  # Thuộc combo (mã SP cha)
                        comp_row[25] = comp.name or ''
                        comp_row[29] = '131'
                        comp_row[30] = '5111'
                        comp_row[31] = comp_uom
                        comp_row[32] = comp_qty
                        comp_row[33] = 0
                        comp_row[34] = 0
                        comp_row[47] = 'HLV'
                        comp_row[48] = '632'
                        comp_row[49] = '1561'
                        misa_rows.append(comp_row)

                else:
                    # Sản phẩm thường — xuất bình thường
                    product_code = product.default_code or '' if product else ''
                    product_name = (product.name or line.name or '') if product else (line.name or '')

                    row = [''] * len(MISA_COLUMNS)
                    row[0] = 'Bán hàng hóa trong nước'  # Hình thức bán hàng
                    row[1] = 'Chưa thu tiền'                    # Phương thức thanh toán
                    row[2] = 'Có'                       # Kiêm phiếu xuất kho
                    row[3] = 'Không'                         # Lập kèm hóa đơn
                    # row[4] = da_lap_hd                  # Đã lập hóa đơn
                    row[4] =  "Chưa lập"                # Đã lập hóa đơn
                    row[5] = today_str                  # Ngày hạch toán (ngày hiện tại)
                    row[6] = today_str                  # Ngày chứng từ (ngày hiện tại)
                    row[7] = so_ref                     # Số chứng từ (VD: BH12052026SP)
                    row[8] = xk_ref                     # Số phiếu xuất (VD: XK12052026SP)
                    row[9] = '1'                        # Mẫu số HĐ
                    row[10] = '1C26TLV'                 # Ký hiệu HĐ
                    row[11] = inv_no                    # Số hóa đơn
                    row[12] = inv_date                  # Ngày hóa đơn
                    row[13] = customer_code             # Mã khách hàng (từ map)
                    row[14] = customer_name             # Tên khách hàng (từ map)
                    row[15] = ''                        # Địa chỉ (để trống)
                    row[17] = ''                    # Đơn vị giao đại lý
                    row[18] = ''                    # Người nộp
                    row[19] = ''                    # Nộp vào TK
                    row[20] = ''                    # Tên ngân hàng
                    row[21] = dien_giai             # Diễn giải/Lý do nộp
                    row[22] = ly_do_xuat            # Lý do xuất
                    row[23] = product_code          # Mã hàng
                    row[24] = ''                    # Thuộc combo (SP thường — để trống)
                    row[25] = product_name          # Tên hàng
                    row[26] = ''                    # Là dòng ghi chú
                    row[27] = ''                    # Hàng khuyến mại
                    row[28] = ''                    # Chiết khấu thương mại
                    row[29] = '131'                 # TK Tiền/Chi phí/Nợ
                    row[30] = '5111'                # TK Doanh thu/Có
                    row[31] = uom_name              # ĐVT
                    row[32] = line_qty              # Số lượng
                    row[33] = price                 # Đơn giá
                    row[34] = subtotal              # Thành tiền
                    row[35] = discount_pct if discount_pct else ''   # Tỷ lệ CK (%) — giữ thập phân
                    row[36] = discount_amt if discount_amt else ''   # Tiền chiết khấu
                    row[37] = ''                    # TK chiết khấu
                    row[38] = ''                    # Giá tính thuế XK
                    row[39] = ''                    # % thuế xuất khẩu
                    row[40] = ''                    # Tiền thuế xuất khẩu
                    row[41] = ''                    # TK thuế xuất khẩu
                    row[42] = vat_rate if vat_rate else ''     # % thuế GTGT
                    row[43] = ''                    # % thuế suất KHAC
                    row[44] = vat_amount if vat_amount else '' # Tiền thuế GTGT
                    row[45] = '3331' if vat_rate else ''       # TK thuế GTGT
                    row[46] = ''                    # HH không TH trên tờ khai thuế GTGT
                    row[47] = 'HLV'                 # Mã kho
                    row[48] = '632'                 # TK giá vốn
                    row[49] = '1561'                # TK Kho
                    row[50] = ''                    # Đơn giá vốn
                    row[51] = ''                    # Tiền vốn
                    row[52] = ''                    # Hàng hóa giữ hộ/bán hộ

                    misa_rows.append(row)

        return misa_rows

    # ── Tạo Excel ────────────────────────────────────────────────────────────

    def _generate_excel(self, misa_rows):
        if not Workbook:
            raise UserError('Thiếu thư viện openpyxl để tạo file Excel.')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Phieu ban hang'

        header_font = Font(name='Arial', size=10, bold=True)
        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border_side = Side(style='thin', color='000000')
        border = Border(
            left=border_side, right=border_side,
            top=border_side, bottom=border_side,
        )
        cell_align = Alignment(horizontal='left', vertical='center', wrap_text=False)
        num_align = Alignment(horizontal='right', vertical='center')

        # Column widths
        col_widths = {
            0: 22, 1: 22, 2: 15, 3: 15, 4: 15,
            5: 15, 6: 15, 7: 20, 8: 30, 9: 12,
            10: 12, 11: 12, 12: 14, 13: 15, 14: 30,
            15: 35, 16: 15, 21: 60, 22: 35, 23: 15,
            24: 18, 25: 35, 31: 8, 32: 10, 33: 12, 34: 14,
        }

        # Header
        for col_idx, col_name in enumerate(MISA_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
            w = col_widths.get(col_idx - 1, 14)
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        ws.row_dimensions[1].height = 35
        ws.freeze_panes = 'A2'

        # Data
        for row_idx, row_data in enumerate(misa_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value if value != '' else None
                cell.border = border
                if isinstance(value, (int, float)):
                    cell.alignment = num_align
                    cell.number_format = '#,##0' if col_idx in (33, 34, 36, 44, 51) else '#,##0.##'
                else:
                    cell.alignment = cell_align

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()

    # ── Main action ──────────────────────────────────────────────────────────

    def action_export(self):
        self.ensure_one()
        rows = self._read_shopee_file()
        shopee_codes = self._extract_shopee_codes(rows)

        if not shopee_codes:
            raise UserError('Không tìm thấy mã đơn hàng Shopee nào trong file.\n'
                            'Kiểm tra lại file: cột D từ row 19 phải chứa mã đơn Shopee.')

        code_so_map = self._find_sale_orders(shopee_codes)
        found = sum(1 for v in code_so_map.values() if v)
        not_found = [c for c, v in code_so_map.items() if not v]

        _logger.info(
            'Shopee wallet import: %d codes, %d found, %d not found',
            len(shopee_codes), found, len(not_found),
        )
        if not_found:
            _logger.warning('Không tìm thấy SO cho: %s', ', '.join(not_found[:20]))

        misa_rows = self._build_misa_rows(code_so_map)
        if not misa_rows:
            raise UserError('Không có dữ liệu để xuất.')

        excel_data = self._generate_excel(misa_rows)

        summary = 'Tìm thấy %d/%d đơn hàng.' % (found, len(shopee_codes))
        if not_found:
            summary += '\nKhông tìm thấy (%d): %s%s' % (
                len(not_found),
                ', '.join(not_found[:5]),
                '...' if len(not_found) > 5 else '',
            )
        self.write({'result_summary': summary})

        filename = 'PhieuBanHang_Shopee_%s.xlsx' % fields.Date.today().strftime('%Y%m%d')
        attachment = self.env['ir.attachment'].sudo().create({
            'name': filename,
            'type': 'binary',
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'datas': base64.b64encode(excel_data),
            'res_model': self._name,
            'res_id': self.id,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%d?download=true' % attachment.id,
            'target': 'new',
        }
