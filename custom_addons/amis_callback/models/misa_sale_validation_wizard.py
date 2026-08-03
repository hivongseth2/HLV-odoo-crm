# -*- coding: utf-8 -*-
import base64
import logging
import re
import unicodedata
from copy import copy
from io import BytesIO

from odoo import _, fields, models
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - handled with a user-facing error
    openpyxl = None
    Workbook = None


_logger = logging.getLogger(__name__)

MISSING_PRODUCT_RE = re.compile(
    r"M[aã]\s*h[aà]ng\s*<([^>]+)>\s*kh[oô]ng\s*c[oó]\s*trong\s*danh\s*m[uụ]c",
    re.IGNORECASE,
)
UOM_ERROR_RE = re.compile(
    r"[ĐD]ơn\s*v[iị]\s*t[ií]nh\s*<([^>]+)>.*?m[aặ]t\s*h[aà]ng\s*<([^>]+)>",
    re.IGNORECASE,
)
DUPLICATE_COMBO_TEXT = "tren 1 chung tu ban hang moi combo chi duoc nhap khau du lieu 1 lan"

REQUIRED_HEADERS = {
    "error": "chi tiet loi",
    "document": "so chung tu (*)",
    "product_code": "ma hang (*)",
    "combo_parent": "thuoc combo",
    "uom": "dvt",
}


def _plain_text(value):
    """Lower-case and remove Vietnamese accents for stable header matching."""
    value = str(value or "").strip().lower().replace("đ", "d")
    return "".join(
        char for char in unicodedata.normalize("NFD", value)
        if unicodedata.category(char) != "Mn"
    )


def _find_validation_header(ws):
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        normalized = {
            _plain_text(ws.cell(row=row_idx, column=col_idx).value): col_idx
            for col_idx in range(1, ws.max_column + 1)
        }
        if all(name in normalized for name in REQUIRED_HEADERS.values()):
            return row_idx, {
                key: normalized[name] for key, name in REQUIRED_HEADERS.items()
            }
    raise UserError(_(
        "Không nhận diện được file Kết quả kiểm tra MISA. "
        "File phải có các cột Chi tiết lỗi, Số chứng từ, Mã hàng, "
        "Thuộc combo và ĐVT."
    ))


def _cell_text(ws, row_idx, col_idx):
    return str(ws.cell(row=row_idx, column=col_idx).value or "").strip()


def _analyze_validation_sheet(ws, header_row, columns):
    """Return target repairs without changing the source workbook."""
    error_col = columns["error"]
    document_col = columns["document"]
    code_col = columns["product_code"]
    parent_col = columns["combo_parent"]

    combo_codes = {
        _cell_text(ws, row_idx, parent_col)
        for row_idx in range(header_row + 1, ws.max_row + 1)
        if _cell_text(ws, row_idx, parent_col)
    }

    # A parent may have no usable child row in a malformed file. The explicit
    # duplicate error still lets us recognize it as a combo parent.
    for row_idx in range(header_row + 1, ws.max_row + 1):
        error = _plain_text(_cell_text(ws, row_idx, error_col))
        code = _cell_text(ws, row_idx, code_col)
        if code and DUPLICATE_COMBO_TEXT in error:
            combo_codes.add(code)

    occurrences = []
    row_idx = header_row + 1
    while row_idx <= ws.max_row:
        code = _cell_text(ws, row_idx, code_col)
        parent = _cell_text(ws, row_idx, parent_col)
        if code and not parent and code in combo_codes:
            group_rows = [row_idx]
            child_idx = row_idx + 1
            while child_idx <= ws.max_row:
                if _cell_text(ws, child_idx, parent_col) != code:
                    break
                group_rows.append(child_idx)
                child_idx += 1
            occurrences.append({
                "key": (_cell_text(ws, row_idx, document_col), code),
                "code": code,
                "rows": group_rows,
            })
            row_idx = child_idx
            continue
        row_idx += 1

    seen = set()
    duplicate_occurrences = []
    duplicate_rows = set()
    for occurrence in occurrences:
        if occurrence["key"] in seen:
            duplicate_occurrences.append(occurrence)
            duplicate_rows.update(occurrence["rows"])
        else:
            seen.add(occurrence["key"])

    missing_codes = []
    missing_seen = set()
    uom_fixes = {}
    recognized_error_rows = set()
    related_error_rows = set()

    for row_idx in range(header_row + 1, ws.max_row + 1):
        error = _cell_text(ws, row_idx, error_col)
        normalized_error = _plain_text(error)
        if normalized_error.startswith("loi lien quan den dong"):
            related_error_rows.add(row_idx)
            continue

        missing_match = MISSING_PRODUCT_RE.search(error)
        if missing_match:
            code = missing_match.group(1).strip()
            if code and code not in missing_seen:
                missing_seen.add(code)
                missing_codes.append(code)
            recognized_error_rows.add(row_idx)

        uom_match = UOM_ERROR_RE.search(error)
        if uom_match:
            bad_uom = uom_match.group(1).strip()
            if _plain_text(bad_uom) == "bo":
                uom_fixes[row_idx] = "Bộ."
            recognized_error_rows.add(row_idx)

        if DUPLICATE_COMBO_TEXT in normalized_error:
            recognized_error_rows.add(row_idx)

    return {
        "duplicate_occurrences": duplicate_occurrences,
        "duplicate_rows": duplicate_rows,
        "missing_codes": missing_codes,
        "uom_fixes": uom_fixes,
        "recognized_error_rows": recognized_error_rows,
        "related_error_rows": related_error_rows,
    }


def _copy_cell(source_cell, target_cell, value=None):
    target_cell.value = source_cell.value if value is None else value
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.font:
        target_cell.font = copy(source_cell.font)
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.border:
        target_cell.border = copy(source_cell.border)
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.protection:
        target_cell.protection = copy(source_cell.protection)


def _copy_column_dimensions(source_ws, target_ws, source_columns):
    for target_idx, source_idx in enumerate(source_columns, 1):
        source_letter = get_column_letter(source_idx)
        target_letter = get_column_letter(target_idx)
        source_dim = source_ws.column_dimensions[source_letter]
        target_dim = target_ws.column_dimensions[target_letter]
        target_dim.width = source_dim.width
        target_dim.hidden = source_dim.hidden
        target_dim.bestFit = source_dim.bestFit


def _build_result_workbook(source_ws, header_row, columns, analysis):
    result_wb = Workbook()
    main_ws = result_wb.active
    main_ws.title = "Phieu ban hang"

    # MISA's validation result prepends Tình trạng and Chi tiết lỗi.
    # Remove those two metadata columns from the sheet that will be imported.
    error_col = columns["error"]
    status_col = error_col - 1
    import_columns = [
        col_idx for col_idx in range(1, source_ws.max_column + 1)
        if col_idx not in (status_col, error_col)
    ]
    _copy_column_dimensions(source_ws, main_ws, import_columns)

    source_rows = list(range(1, header_row + 1)) + [
        row_idx for row_idx in range(header_row + 1, source_ws.max_row + 1)
        if row_idx not in analysis["duplicate_rows"]
    ]

    for target_row, source_row in enumerate(source_rows, 1):
        if source_ws.row_dimensions[source_row].height:
            main_ws.row_dimensions[target_row].height = source_ws.row_dimensions[source_row].height
        for target_col, source_col in enumerate(import_columns, 1):
            value = source_ws.cell(source_row, source_col).value
            if source_col == columns["uom"] and source_row in analysis["uom_fixes"]:
                value = analysis["uom_fixes"][source_row]
            _copy_cell(
                source_ws.cell(source_row, source_col),
                main_ws.cell(target_row, target_col),
                value=value,
            )

    main_ws.freeze_panes = "A2" if header_row == 1 else None
    main_ws.auto_filter.ref = main_ws.dimensions

    if analysis["duplicate_rows"]:
        duplicate_ws = result_wb.create_sheet("Combo bi trung")
        all_columns = list(range(1, source_ws.max_column + 1))
        _copy_column_dimensions(source_ws, duplicate_ws, all_columns)
        duplicate_source_rows = list(range(1, header_row + 1)) + sorted(analysis["duplicate_rows"])
        for target_row, source_row in enumerate(duplicate_source_rows, 1):
            if source_ws.row_dimensions[source_row].height:
                duplicate_ws.row_dimensions[target_row].height = source_ws.row_dimensions[source_row].height
            for col_idx in all_columns:
                value = source_ws.cell(source_row, col_idx).value
                if col_idx == columns["uom"] and source_row in analysis["uom_fixes"]:
                    value = analysis["uom_fixes"][source_row]
                _copy_cell(
                    source_ws.cell(source_row, col_idx),
                    duplicate_ws.cell(target_row, col_idx),
                    value=value,
                )
        duplicate_ws.freeze_panes = "A2" if header_row == 1 else None
        duplicate_ws.auto_filter.ref = duplicate_ws.dimensions

    return result_wb


class MisaSaleValidationWizard(models.TransientModel):
    _name = "misa.sale.validation.wizard"
    _description = "Xử lý kết quả kiểm tra Phiếu bán hàng MISA"

    file_data = fields.Binary(
        string="File Kết quả kiểm tra MISA",
        required=True,
        attachment=False,
    )
    file_name = fields.Char(string="Tên file")
    sync_missing_combos = fields.Boolean(
        string="Tạo combo thiếu trên CRM",
        default=True,
        help="Tạo combo từ combo kit/BOM Odoo; CRM sẽ đồng bộ sang MISA.",
    )
    result_file = fields.Binary(string="File đã xử lý", readonly=True, attachment=False)
    result_file_name = fields.Char(string="Tên file kết quả", readonly=True)
    result_summary = fields.Text(string="Kết quả", readonly=True)

    def _load_workbook(self):
        if not openpyxl:
            raise UserError(_("Thiếu thư viện openpyxl để xử lý file XLSX."))
        if not self.file_data:
            raise UserError(_("Vui lòng upload file Kết quả kiểm tra MISA."))
        if not (self.file_name or "").lower().endswith((".xlsx", ".xlsm")):
            raise UserError(_("Chức năng này chỉ nhận file XLSX/XLSM."))
        try:
            return openpyxl.load_workbook(
                BytesIO(base64.b64decode(self.file_data)),
                data_only=False,
            )
        except Exception as exc:
            raise UserError(_("Không đọc được file Excel: %s") % exc) from exc

    def _find_source_sheet(self, workbook):
        errors = []
        for worksheet in workbook.worksheets:
            try:
                header_row, columns = _find_validation_header(worksheet)
                return worksheet, header_row, columns
            except UserError as exc:
                errors.append(str(exc))
        raise UserError(errors[-1] if errors else _("File Excel không có sheet dữ liệu."))

    def _find_odoo_combo(self, code):
        Product = self.env["product.product"].sudo().with_context(active_test=False)
        product = Product.search([("default_code", "=", code)], limit=1)
        if not product:
            product = Product.search([("default_code", "=ilike", code)], limit=1)
        return product

    def _get_combo_components(self, product):
        template = product.product_tmpl_id
        components = []

        if "combo_product_id" in template._fields and template.combo_product_id:
            for line in template.combo_product_id:
                component = line.product_id
                components.append({
                    "product_id": component.id,
                    "code": (component.default_code or "").strip(),
                    "name": component.name or "",
                    "quantity": line.product_quantity or 1.0,
                    "uom": line.uom_id.name or component.uom_id.name or "",
                })
        elif "mrp.bom" in self.env:
            Bom = self.env["mrp.bom"].sudo()
            bom = Bom.search([
                ("type", "=", "phantom"),
                ("product_id", "=", product.id),
            ], limit=1)
            if not bom:
                bom = Bom.search([
                    ("type", "=", "phantom"),
                    ("product_tmpl_id", "=", template.id),
                    ("product_id", "=", False),
                ], limit=1)
            for line in bom.bom_line_ids if bom else []:
                component = line.product_id
                components.append({
                    "product_id": component.id,
                    "code": (component.default_code or "").strip(),
                    "name": component.name or "",
                    "quantity": line.product_qty or 1.0,
                    "uom": line.product_uom_id.name or component.uom_id.name or "",
                })

        missing_component_codes = [item["name"] for item in components if not item["code"]]
        if missing_component_codes:
            raise UserError(_(
                "Combo %s có sản phẩm con chưa có Mã nội bộ: %s"
            ) % (product.default_code, ", ".join(missing_component_codes)))
        if not components:
            raise UserError(_("Không tìm thấy combo kit/BOM cho %s.") % product.default_code)
        return components

    def _create_combo_on_crm(self, product, components):
        """Call the project CRM combo API once its supported by misa.api.utils."""
        if "misa.api.utils" not in self.env:
            raise UserError(_("Model misa.api.utils chưa được cài đặt."))
        api = self.env["misa.api.utils"].sudo()
        create_method = getattr(api, "create_combo_product_misa", None)
        if not create_method:
            raise UserError(_(
                "Chưa có API tạo combo CRM trong misa.api.utils. "
                "Cần bổ sung payload request/response mẫu."
            ))
        return create_method(product.id, components=components)

    def _sync_missing_combo_codes(self, codes):
        result = {"created": [], "existing": [], "not_found": [], "failed": []}
        for code in codes:
            product = self._find_odoo_combo(code)
            if not product:
                result["not_found"].append(code)
                continue
            try:
                components = self._get_combo_components(product)
                with self.env.cr.savepoint():
                    crm_result = self._create_combo_on_crm(product, components)
                target = "existing" if isinstance(crm_result, dict) and not crm_result.get("created", True) else "created"
                result[target].append(code)
            except Exception as exc:
                _logger.exception("Không tạo được combo CRM %s", code)
                result["failed"].append((code, str(exc)))
        return result

    def action_process(self):
        self.ensure_one()
        source_wb = self._load_workbook()
        try:
            source_ws, header_row, columns = self._find_source_sheet(source_wb)
            analysis = _analyze_validation_sheet(source_ws, header_row, columns)

            sync_result = {"created": [], "existing": [], "not_found": [], "failed": []}
            if self.sync_missing_combos and analysis["missing_codes"]:
                sync_result = self._sync_missing_combo_codes(analysis["missing_codes"])

            result_wb = _build_result_workbook(
                source_ws, header_row, columns, analysis,
            )
            output = BytesIO()
            result_wb.save(output)
            result_wb.close()
        finally:
            source_wb.close()

        duplicate_rows = len(analysis["duplicate_rows"])
        duplicate_count = len(analysis["duplicate_occurrences"])
        has_combo_sync_errors = bool(
            (analysis["missing_codes"] and not self.sync_missing_combos)
            or sync_result["not_found"]
            or sync_result["failed"]
        )
        summary_lines = [
            _(
                "File đã xử lý nhưng chưa nên import; hãy xử lý các lỗi combo bên dưới."
                if has_combo_sync_errors
                else "File đã sẵn sàng để import lại MISA."
            ),
            _("- Đã sửa ĐVT: %d dòng") % len(analysis["uom_fixes"]),
            _("- Combo trùng đã chuyển sang sheet 'Combo bi trung': %d combo / %d dòng")
            % (duplicate_count, duplicate_rows),
            _("- Mã combo thiếu trong danh mục: %d") % len(analysis["missing_codes"]),
        ]
        if self.sync_missing_combos:
            summary_lines.append(_("- Đã tạo trên CRM: %d") % len(sync_result["created"]))
            if sync_result["created"]:
                summary_lines.append(_("- Hãy chờ CRM đồng bộ combo mới sang MISA trước khi import file."))
            if sync_result["existing"]:
                summary_lines.append(_("- Combo đã có trên CRM: %s") % ", ".join(sync_result["existing"]))
            if sync_result["not_found"]:
                summary_lines.append(_("- Không tìm thấy sản phẩm Odoo: %s") % ", ".join(sync_result["not_found"]))
            for code, error in sync_result["failed"]:
                summary_lines.append(_("- Chưa tạo được %s: %s") % (code, error))
        elif analysis["missing_codes"]:
            summary_lines.append(_("- Chưa tạo combo CRM do tùy chọn này đang tắt."))

        result_name = "PhieuBanHang_MISA_DaXuLy_%s.xlsx" % fields.Date.today().strftime("%Y%m%d")
        self.write({
            "result_file": base64.b64encode(output.getvalue()),
            "result_file_name": result_name,
            "result_summary": "\n".join(summary_lines),
        })

        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }
