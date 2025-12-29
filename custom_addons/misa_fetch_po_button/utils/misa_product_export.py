# -*- coding: utf-8 -*-
"""
Script xuất tất cả sản phẩm từ MISA CRM ra file Excel
Sử dụng làm template để import vào POS Odoo

Cách sử dụng trong Odoo Shell:
    from odoo.addons.misa_fetch_po_button.utils.misa_product_export import MisaProductExporter
    exporter = MisaProductExporter(env)
    exporter.export_all_products_to_excel('/path/to/output.xlsx')

Hoặc gọi qua API endpoint:
    POST /api/misa/product/export
"""
import requests
import logging
import json
import re
from datetime import datetime
from io import BytesIO

_logger = logging.getLogger(__name__)


class MisaProductExporter:
    """Class xuất sản phẩm từ MISA CRM ra Excel"""
    
    def __init__(self, env):
        self.env = env
        self.misa_config = env['misa.config']
        self.misa_utils = env['misa.api.utils']
    
    def _get_token(self):
        """Lấy token đăng nhập MISA"""
        return self.misa_utils._fetch_login_crm_token()
    
    def _get_headers(self, token):
        """Chuẩn bị headers cho API calls"""
        headers = self.misa_config.get_crm_header(token)
        headers.update({"LayoutCode": "product", "X-Misa-Language": "vi-VN"})
        return headers
    
    def fetch_all_products(self, page_size=1000, max_pages=1000, columns=None):
        """
        Lấy tất cả sản phẩm từ MISA CRM.
        
        Args:
            page_size: Số sản phẩm mỗi trang (tăng lên 1000 cho nhanh)
            max_pages: Số trang tối đa
            columns: Chuỗi các cột cần lấy (base64 encoded hoặc raw string). 
                     Nếu None sẽ lấy mặc định.
            
        Returns:
            list: Danh sách sản phẩm
        """
        import uuid
        
        token = self._get_token()
        if not token:
            raise Exception("Không lấy được Token MISA")
        
        headers = self._get_headers(token)
        url = "https://amisapp.misa.vn/crm/g2/api/business/Product/Grid"
        
        # Default columns payload
        default_columns = "SUQsUHJvZHVjdENvZGUsUHJvZHVjdE5hbWUsUHJvZHVjdENhdGVnb3J5SUQsUHJvZHVjdENhdGVnb3J5SURUZXh0LFVzYWdlVW5pdElELFVzYWdlVW5pdElEVGV4dCxVbml0UHJpY2UsVGF4SUQsVGF4SURUZXh0LElzU2V0UHJvZHVjdCxGb3JtTGF5b3V0SUQsRm9ybUxheW91dElEVGV4dCxPd25lcklELE93bmVySURUZXh0LElzU3lzdGVtLEF2YXRhcg=="
        
        all_products = []
        page = 1
        
        while page <= max_pages:
            _logger.info(f"📥 Đang lấy trang {page}...")
            
            # Payload theo format của search_product_by_name đang hoạt động
            payload = {
                "Columns": columns or default_columns,
                "Sorts": [],
                "Start": (page - 1) * page_size,
                "Page": page,
                "PageSize": page_size,
                "Filters": [],
                "Formula": "",
                "LayoutCode": "Product",
                "DefaultTotal": False,
                "IsMappingData": False,
                "MappingValueObject": {},
                "IsApproved": False,
                "CustomPagingData": {},
                "IsUsedELTS": True,
                "ListGmailPage": [],
                "ListFacebookPage": {},
                "IsListPaging": True,
                "IsGetCache": True,
                "IsCheckInactive": False,
                "IsConverted": False,
                "SessionID": str(uuid.uuid4()),
                "LayoutCodeCheckPermission": "Product",
                "AISearchKeyword": ""
            }
            
            try:
                session = self.misa_utils._get_retry_session()
                res = session.post(url, headers=headers, json=payload, timeout=60)
                
                if res.status_code != 200:
                    _logger.error(f"HTTP {res.status_code}: {res.text[:200]}")
                    break
                
                data = res.json()
                if not data.get("Success"):
                    _logger.error(f"MISA Error: {data}")
                    break
                
                products = data.get("Data", [])
                total = data.get("Total", 0)
                
                if not products:
                    _logger.info(f"Không còn sản phẩm, dừng lại")
                    break
                
                all_products.extend(products)
                _logger.info(f"✅ Đã lấy {len(products)} sản phẩm (tổng: {len(all_products)}/{total})")
                
                # Kiểm tra đã lấy hết chưa
                if len(all_products) >= total or len(products) < page_size:
                    break
                
                page += 1
                
            except Exception as e:
                _logger.exception(f"Lỗi lấy trang {page}: {e}")
                break
        
        _logger.info(f"🎉 Hoàn thành! Tổng số sản phẩm: {len(all_products)}")
        return all_products
    
    def fetch_all_categories(self):
        """Lấy tất cả danh mục sản phẩm từ MISA"""
        token = self._get_token()
        if not token:
            raise Exception("Không lấy được Token MISA")
        
        headers = self._get_headers(token)
        
        # Xóa content headers cho GET request
        get_headers = headers.copy()
        for k in ['content-length', 'Content-Length', 'content-type', 'Content-Type']:
            get_headers.pop(k, None)
        
        url = "https://amisapp.misa.vn/crm/g1/api/business/ProductCategory/tree/0/false"
        
        try:
            session = self.misa_utils._get_retry_session()
            res = session.get(url, headers=get_headers, timeout=30)
            
            if not res.ok:
                _logger.error(f"Lỗi lấy danh mục: {res.status_code}")
                return []
            
            data = res.json()
            if not data.get("Success"):
                return []
            
            raw_data = data.get("Data")
            nodes = json.loads(raw_data) if isinstance(raw_data, str) else raw_data
            
            # Flatten tree thành list
            categories = []
            def flatten(node_list, parent_name=""):
                for node in node_list:
                    cat_name = node.get("ProductCategoryName", "")
                    categories.append({
                        "id": node.get("ID"),
                        "name": cat_name,
                        "parent": parent_name,
                        "full_path": f"{parent_name} / {cat_name}" if parent_name else cat_name
                    })
                    children = node.get("Children")
                    if children:
                        flatten(children, cat_name)
            
            if isinstance(nodes, list):
                flatten(nodes)
            
            return categories
            
        except Exception as e:
            _logger.exception(f"Lỗi lấy danh mục: {e}")
            return []
    
    def export_all_products_to_excel(self, file_path=None):
        """
        Xuất tất cả sản phẩm ra file Excel.
        
        Args:
            file_path: Đường dẫn file output. Nếu None, trả về bytes.
            
        Returns:
            str hoặc bytes: Đường dẫn file hoặc nội dung bytes
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise Exception("Vui lòng cài thư viện: pip install openpyxl")
        
        # Lấy dữ liệu
        _logger.info("📥 Đang lấy sản phẩm từ MISA...")
        products = self.fetch_all_products()
        
        _logger.info("📥 Đang lấy danh mục từ MISA...")
        categories = self.fetch_all_categories()
        
        # Tạo workbook
        wb = openpyxl.Workbook()
        
        # === SHEET 1: SẢN PHẨM ===
        ws = wb.active
        ws.title = "Sản phẩm MISA"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers - Tương thích với POS import
        headers = [
            'ID MISA',
            'Mã sản phẩm',
            'Tên sản phẩm',
            'Giá bán',
            'Thuế %',
            'Đơn vị tính',
            'Danh mục MISA',
            'Loại sản phẩm',
            'Danh mục POS (điền thêm)',
            'Có trong POS',
            'Hoạt động',
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data rows
        for row_idx, p in enumerate(products, 2):
            # Parse thuế từ text (ví dụ: "10%" -> 10)
            tax_text = p.get("TaxIDText", "") or ""
            tax_num = 0
            tax_match = re.findall(r"(\d+(?:\.\d+)?)", tax_text)
            if tax_match:
                try:
                    tax_num = float(tax_match[0])
                except:
                    pass
            
            # Loại sản phẩm
            prop_text = p.get("ProductPropertiesIDText", "") or ""
            if "dịch vụ" in prop_text.lower():
                p_type = "service"
            else:
                p_type = "product"
            
            row_data = [
                p.get("ID") or p.get("ProductID"),
                p.get("ProductCode", ""),
                p.get("ProductName", ""),
                p.get("UnitPrice", 0) or 0,
                tax_num,
                p.get("UsageUnitIDText", "Cái"),
                p.get("ProductCategoryIDText", ""),
                p_type,
                "",  # Danh mục POS - user điền thêm
                "X" if p.get("Active", True) else "",
                "X" if p.get("Active", True) else "",
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        # Độ rộng cột
        col_widths = [15, 20, 45, 15, 10, 15, 25, 15, 25, 12, 12]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
        # === SHEET 2: DANH MỤC MISA ===
        ws_cat = wb.create_sheet("Danh mục MISA")
        cat_headers = ['ID', 'Tên danh mục', 'Danh mục cha', 'Đường dẫn đầy đủ']
        for col, header in enumerate(cat_headers, 1):
            cell = ws_cat.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row_idx, cat in enumerate(categories, 2):
            ws_cat.cell(row=row_idx, column=1, value=cat.get("id"))
            ws_cat.cell(row=row_idx, column=2, value=cat.get("name"))
            ws_cat.cell(row=row_idx, column=3, value=cat.get("parent"))
            ws_cat.cell(row=row_idx, column=4, value=cat.get("full_path"))
        
        ws_cat.column_dimensions['A'].width = 15
        ws_cat.column_dimensions['B'].width = 30
        ws_cat.column_dimensions['C'].width = 25
        ws_cat.column_dimensions['D'].width = 50
        
        # === SHEET 3: HƯỚNG DẪN ===
        ws_guide = wb.create_sheet("Hướng dẫn Import POS")
        guide = [
            ["HƯỚNG DẪN IMPORT SẢN PHẨM VÀO POS ODOO"],
            [""],
            ["BƯỚC 1: Điền danh mục POS"],
            ["- Cột 'Danh mục POS' cần điền tên danh mục POS đã tạo trong Odoo"],
            ["- Danh mục phải tồn tại trước khi import"],
            [""],
            ["BƯỚC 2: Đánh dấu sản phẩm cần import"],
            ["- Cột 'Có trong POS': Đánh X nếu muốn sản phẩm hiển thị trong POS"],
            ["- Cột 'Hoạt động': Đánh X nếu sản phẩm đang hoạt động"],
            [""],
            ["BƯỚC 3: Import vào Odoo"],
            ["- Vào Point of Sale > Configuration > Products"],
            ["- Click Action > Import"],
            ["- Chọn file Excel đã chỉnh sửa"],
            [""],
            ["LƯU Ý:"],
            ["- Mã sản phẩm phải trùng với Internal Reference trong Odoo"],
            ["- Đơn vị tính phải tồn tại trong Odoo"],
            ["- Kiểm tra lại thuế % phù hợp với cấu hình thuế Odoo"],
            [""],
            [f"File được tạo lúc: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"],
            [f"Tổng số sản phẩm: {len(products)}"],
            [f"Tổng số danh mục: {len(categories)}"],
        ]
        for row_idx, row in enumerate(guide, 1):
            ws_guide.cell(row=row_idx, column=1, value=row[0] if row else "")
        ws_guide.column_dimensions['A'].width = 80
        
        # Lưu file
        if file_path:
            wb.save(file_path)
            _logger.info(f"✅ Đã lưu file: {file_path}")
            return file_path
        else:
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
    
    def export_to_base64(self):
        """Xuất ra base64 string để trả về qua API"""
        import base64
        content = self.export_all_products_to_excel()
        return base64.b64encode(content).decode('utf-8')
