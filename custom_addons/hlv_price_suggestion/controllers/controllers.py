import io
import json
import logging

from odoo import http
from odoo.http import request, content_disposition

_logger = logging.getLogger(__name__)


class PriceChatExcelController(http.Controller):

    @http.route('/price_chat/export_excel/<int:session_id>',
                type='http', auth='user', methods=['GET'])
    def export_excel(self, session_id, **kwargs):
        """Xuất file Excel chứa dữ liệu đề xuất giá từ phiên chat."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return request.make_response(
                'Cần cài thư viện openpyxl. Chạy: pip install openpyxl',
                headers=[('Content-Type', 'text/plain')],
            )

        session = request.env['price.chat.session'].browse(session_id)
        if not session.exists():
            return request.not_found()

        data = session._generate_excel_data()

        wb = openpyxl.Workbook()

        # ── Style definitions ──
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
        subheader_fill = PatternFill(start_color='A3D5FF', end_color='A3D5FF', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # ══════════════════════════════════════
        # Sheet 1: Tổng hợp đề xuất giá
        # ══════════════════════════════════════
        ws = wb.active
        ws.title = 'Tổng hợp giá'

        headers = [
            'Mã SP', 'Tên sản phẩm', 'Giá bán hiện tại',
            'Giá nhập gần nhất', 'Giá nhập TB',
            'Tồn kho', 'Tồn sẵn sàng',
            'Bán 30 ngày', 'TB bán/ngày',
            'Số ngày tồn còn',
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        for row_idx, item in enumerate(data, 2):
            stock = item.get('ton_kho', {})
            sales = item.get('luot_ban_30_ngay', {})

            # Giá nhập gần nhất
            last_purchase = item['gia_nhap'][0]['gia'] if item.get('gia_nhap') else 0
            # Giá nhập trung bình
            purchase_prices = item.get('gia_nhap', [])
            if purchase_prices:
                total_cost = sum(p['gia'] * p['so_luong'] for p in purchase_prices)
                total_qty = sum(p['so_luong'] for p in purchase_prices)
                avg_purchase = total_cost / total_qty if total_qty else 0
            else:
                avg_purchase = 0

            days_left = item.get('so_ngay_ton_kho_con', '')
            if isinstance(days_left, str):
                days_left = 'N/A'

            row_data = [
                item.get('ma_sp', ''),
                item.get('san_pham', ''),
                item.get('gia_ban_hien_tai', 0),
                last_purchase,
                round(avg_purchase),
                stock.get('tong', 0),
                stock.get('san_sang', 0),
                sales.get('tong_da_ban', 0),
                sales.get('trung_binh_ngay', 0),
                days_left,
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx in (3, 4, 5):  # Format tiền
                    cell.number_format = '#,##0'

        # Auto width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # ══════════════════════════════════════
        # Sheet 2: Chi tiết giá nhập (PO)
        # ══════════════════════════════════════
        ws2 = wb.create_sheet('Chi tiết giá nhập')
        po_headers = ['Sản phẩm', 'Đơn mua hàng', 'Ngày', 'Nhà cung cấp', 'Giá nhập', 'Số lượng']
        for col_idx, header in enumerate(po_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        row_idx = 2
        for item in data:
            for po in item.get('gia_nhap', []):
                ws2.cell(row=row_idx, column=1, value=item['san_pham']).border = thin_border
                ws2.cell(row=row_idx, column=2, value=po['don_hang']).border = thin_border
                ws2.cell(row=row_idx, column=3, value=po['ngay']).border = thin_border
                ws2.cell(row=row_idx, column=4, value=po['nha_cung_cap']).border = thin_border
                c = ws2.cell(row=row_idx, column=5, value=po['gia'])
                c.number_format = '#,##0'
                c.border = thin_border
                ws2.cell(row=row_idx, column=6, value=po['so_luong']).border = thin_border
                row_idx += 1

        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # ══════════════════════════════════════
        # Sheet 3: Chi tiết giá bán theo công ty
        # ══════════════════════════════════════
        ws3 = wb.create_sheet('Giá bán theo công ty')
        so_headers = [
            'Sản phẩm', 'Công ty', 'Đơn hàng', 'Ngày',
            'Khách hàng', 'Giá bán', 'Số lượng', 'Chiết khấu %',
        ]
        for col_idx, header in enumerate(so_headers, 1):
            cell = ws3.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        row_idx = 2
        for item in data:
            for cty in item.get('gia_ban_theo_cty', []):
                for so in cty.get('don_hang', []):
                    ws3.cell(row=row_idx, column=1, value=item['san_pham']).border = thin_border
                    ws3.cell(row=row_idx, column=2, value=cty['cong_ty']).border = thin_border
                    ws3.cell(row=row_idx, column=3, value=so['don_hang']).border = thin_border
                    ws3.cell(row=row_idx, column=4, value=so['ngay']).border = thin_border
                    ws3.cell(row=row_idx, column=5, value=so['khach_hang']).border = thin_border
                    c = ws3.cell(row=row_idx, column=6, value=so['gia_ban'])
                    c.number_format = '#,##0'
                    c.border = thin_border
                    ws3.cell(row=row_idx, column=7, value=so['so_luong']).border = thin_border
                    ws3.cell(row=row_idx, column=8, value=so.get('chiet_khau', 0)).border = thin_border
                    row_idx += 1

        for col in ws3.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws3.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # ══════════════════════════════════════
        # Sheet 4: Lịch sử chat
        # ══════════════════════════════════════
        ws4 = wb.create_sheet('Lịch sử chat')
        chat_headers = ['Thời gian', 'Vai trò', 'Nội dung']
        for col_idx, header in enumerate(chat_headers, 1):
            cell = ws4.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for row_idx, msg in enumerate(session.message_ids.sorted('create_date'), 2):
            ws4.cell(
                row=row_idx, column=1,
                value=str(msg.create_date)[:19],
            ).border = thin_border
            role_label = 'Bạn' if msg.role == 'user' else 'AI'
            ws4.cell(row=row_idx, column=2, value=role_label).border = thin_border
            ws4.cell(row=row_idx, column=3, value=msg.content).border = thin_border

        ws4.column_dimensions['A'].width = 22
        ws4.column_dimensions['B'].width = 12
        ws4.column_dimensions['C'].width = 80

        # ── Write to stream ──
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'de_xuat_gia_{session.id}.xlsx'
        headers = [
            ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            ('Content-Disposition', content_disposition(filename)),
        ]
        return request.make_response(output.read(), headers=headers)
