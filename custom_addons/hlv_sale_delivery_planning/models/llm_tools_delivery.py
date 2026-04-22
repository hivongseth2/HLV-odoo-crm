# -*- coding: utf-8 -*-
"""
Delivery Planner — LLM Data Tools (read-only)
=============================================
Tập hợp các tool ``@llm_tool`` (data-only) cho AI Dispatcher trong floating
chat của Delivery Planner.

NGUYÊN TẮC:
- 100% read-only — KHÔNG có hành động ghi (assign shipper, đổi state…).
  AI chỉ gợi ý, thủ kho tự bấm.
- Tự nhận filter user đang xem Kanban qua ``dp_active_filter`` (cache theo
  user hiện tại) — bằng nhịp này AI không cần được ép data trong system
  prompt mà có thể tự query khi cần.
- Tránh trả response quá to: list_orders mặc định ``limit=30``, address /
  product list bị cắt ngắn.

Tất cả tool được gắn lên thread khi ``ensure_chat_thread`` chạy
(xem ``delivery_suggestion.py``).
"""
import base64
import json
import logging
import re
from collections import defaultdict
from datetime import timedelta

from odoo import api, fields, models

from odoo.addons.llm_tool.decorators import llm_tool

_logger = logging.getLogger(__name__)

# ir.config_parameter key chứa snapshot filter hiện tại của user (JSON).
_FILTER_PARAM_PREFIX = 'hlv_dp.chat.filters.uid_'

# Whitelist filter keys hợp lệ truyền cho ``get_dashboard_data``.
_PASSTHROUGH_FILTERS = {
    'search_query', 'filter_warehouse_id',
    'filter_delivery_status', 'filter_stock_status',
    'filter_packing_status',
    'filter_date_from', 'filter_date_to',
    'filter_done_date_from', 'filter_done_date_to',
    'filter_po_date_from', 'filter_po_date_to', 'filter_po_status',
    'filter_saler_code', 'filter_htgh',
    'filter_delivery_type', 'filter_tag_ids',
    'show_completed', 'filter_need_transfer',
    'filter_new_orders', 'filter_print_status',
    'filter_shipper_received',
}


def _money(v):
    try:
        return f"{int(v or 0):,}".replace(',', '.')
    except Exception:
        return str(v)


class HlvDeliveryPlannerTools(models.AbstractModel):
    """Bộ tool LLM (read-only) cho Delivery Planner.

    Đặt trên một AbstractModel riêng, không trộn với ``hlv.delivery.suggestion``
    để tách rõ "logic chat / prompt builder" vs "data tools mà AI gọi".
    """
    _name = 'hlv.delivery.planner.tools'
    _description = 'HLV Delivery Planner — LLM Data Tools (read-only)'

    # ──────────────────────────────────────────────────────────────────
    # Cache filter từ Kanban (snoop FE → server)
    # ──────────────────────────────────────────────────────────────────
    @api.model
    def _filter_param_key(self, uid=None):
        return f"{_FILTER_PARAM_PREFIX}{uid or self.env.uid}"

    @api.model
    def set_user_dashboard_context(self, filters):
        """Lưu snapshot filter Kanban hiện tại của user vào ir.config_parameter.
        Gọi từ FE mỗi khi ``_buildFetchKwargs`` chạy (debounce ở client).
        """
        clean = {}
        for k, v in (filters or {}).items():
            if k in _PASSTHROUGH_FILTERS and v not in (None, '', 'all', False):
                clean[k] = v
        ICP = self.env['ir.config_parameter'].sudo()
        ICP.set_param(self._filter_param_key(), json.dumps(clean))
        return True

    @api.model
    def _get_user_dashboard_context(self):
        ICP = self.env['ir.config_parameter'].sudo()
        raw = ICP.get_param(self._filter_param_key(), '')
        if not raw:
            return {}
        try:
            return json.loads(raw) or {}
        except Exception:
            return {}

    @api.model
    def _get_archived_so_ids(self):
        """Trả về set id các đơn user đã 'cất' (đóng gói chờ KH xác nhận).
        Các đơn này KHÔNG được đưa vào kế hoạch giao hôm nay.
        """
        Pref = self.env['hlv.delivery.planner.user.pref'].sudo()
        rec = Pref.search([('user_id', '=', self.env.uid)], limit=1)
        if not rec:
            return set()
        return set(rec.archived_so_ids.ids)

    # ──────────────────────────────────────────────────────────────────
    # Internal helpers
    # ──────────────────────────────────────────────────────────────────
    def _service(self):
        return self.env['hlv.delivery.planner.service']

    def _resolve_warehouse_id(self, warehouse_name):
        """Cho phép AI truyền tên kho thay vì id."""
        if not warehouse_name:
            return None
        if isinstance(warehouse_name, int) or (
            isinstance(warehouse_name, str) and warehouse_name.isdigit()
        ):
            return int(warehouse_name)
        wh = self.env['stock.warehouse'].search(
            [('name', '=ilike', warehouse_name)], limit=1,
        )
        return wh.id if wh else None

    def _resolve_tag_ids(self, route_or_tag):
        if not route_or_tag:
            return ''
        if isinstance(route_or_tag, (list, tuple)):
            tokens = list(route_or_tag)
        else:
            tokens = [t.strip() for t in str(route_or_tag).split(',') if t.strip()]
        ids = []
        for tok in tokens:
            if tok.isdigit():
                ids.append(int(tok))
                continue
            tag = self.env['crm.tag'].search([('name', '=ilike', tok)], limit=1)
            if tag:
                ids.append(tag.id)
        return ','.join(str(i) for i in ids) if ids else ''

    def _build_kwargs(self, packing_status=None, warehouse_name=None,
                      route_or_tag=None, htgh=None, search=None,
                      use_active_filter=True, limit=30, offset=0):
        kwargs = {
            'limit': int(limit),
            'offset': int(offset),
            'include_stats': False,
        }
        if use_active_filter:
            kwargs.update(self._get_user_dashboard_context())
        if packing_status:
            kwargs['filter_packing_status'] = packing_status
        if warehouse_name:
            wh_id = self._resolve_warehouse_id(warehouse_name)
            if wh_id:
                kwargs['filter_warehouse_id'] = wh_id
        if route_or_tag:
            tag_ids = self._resolve_tag_ids(route_or_tag)
            if tag_ids:
                kwargs['filter_tag_ids'] = tag_ids
        if htgh:
            kwargs['filter_htgh'] = htgh
        if search:
            kwargs['search_query'] = search
        # Default an toàn: nếu chưa có status nào, tập trung "đã đóng, chờ giao"
        kwargs.setdefault('filter_packing_status', 'packed_waiting_ship')
        return kwargs

    # ──────────────────────────────────────────────────────────────────
    # TOOL 1 — Filter snapshot
    # ──────────────────────────────────────────────────────────────────
    @llm_tool(name='dp_active_filter', read_only_hint=True)
    def tool_active_filter(self) -> dict:
        """Trả về snapshot filter Kanban Delivery Planner mà user đang xem.

        Dùng khi user nói "trong filter của tao" / "đơn tao đang lọc". Các
        tool khác mặc định đã ăn filter này, nhưng có thể gọi tool này để
        xác nhận hoặc liệt kê lại các filter cho user.
        """
        df = self._get_user_dashboard_context()
        out = {'has_filter': bool(df), 'filters': {}}
        for k, v in df.items():
            label = v
            if k == 'filter_warehouse_id':
                wh = self.env['stock.warehouse'].browse(int(v))
                label = wh.name if wh.exists() else v
                out['filters']['warehouse'] = label
            elif k == 'filter_tag_ids':
                try:
                    ids = [int(x) for x in str(v).split(',') if x.strip()]
                    names = self.env['crm.tag'].browse(ids).mapped('name')
                    out['filters']['tags'] = names
                except Exception:
                    out['filters']['tags'] = v
            else:
                out['filters'][k] = v
        return out

    # ──────────────────────────────────────────────────────────────────
    # TOOL 2 — Dashboard summary (KPI)
    # ──────────────────────────────────────────────────────────────────
    @llm_tool(name='dp_dashboard_summary', read_only_hint=True)
    def tool_dashboard_summary(self, packing_status: str = '',
                               warehouse_name: str = '',
                               use_active_filter: bool = True) -> dict:
        """Tóm tắt KPI: tổng số đơn, tổng giá trị, phân bổ theo kho / tuyến.

        Args:
            packing_status: 'packed_waiting_ship' (mặc định nếu trống),
                'all', 'pending'... — bỏ trống = ưu tiên status từ active
                filter, fallback 'packed_waiting_ship'.
            warehouse_name: lọc theo tên kho (không phân biệt hoa thường).
            use_active_filter: True = áp filter user đang lọc trên Kanban.
        """
        kwargs = self._build_kwargs(
            packing_status=packing_status or None,
            warehouse_name=warehouse_name or None,
            use_active_filter=use_active_filter,
            limit=200, offset=0,
        )
        try:
            data = self._service().get_dashboard_data(**kwargs)
        except Exception as e:
            _logger.exception("dp_dashboard_summary failed")
            return {'error': str(e), 'orders': []}
        orders = data.get('orders') or []
        archived = self._get_archived_so_ids()
        if archived:
            orders = [o for o in orders if o.get('id') not in archived]
        archived_skipped = len(data.get('orders') or []) - len(orders)
        by_wh = defaultdict(lambda: {'count': 0, 'value': 0.0})
        by_route = defaultdict(lambda: {'count': 0, 'value': 0.0})
        total_value = 0.0
        for o in orders:
            v = o.get('amount_total') or 0.0
            total_value += v
            wh = o.get('warehouse_id')
            wh_name = wh[1] if isinstance(wh, (list, tuple)) and len(wh) > 1 else 'Khác'
            by_wh[wh_name]['count'] += 1
            by_wh[wh_name]['value'] += v
            tag_pairs = o.get('tag_ids') or []
            if not tag_pairs:
                by_route['(chưa phân tuyến)']['count'] += 1
                by_route['(chưa phân tuyến)']['value'] += v
            for t in tag_pairs:
                if isinstance(t, (list, tuple)) and len(t) > 1:
                    by_route[t[1]]['count'] += 1
                    by_route[t[1]]['value'] += v
        return {
            'total_orders': max((data.get('total_count') or len(orders)) - archived_skipped, len(orders)),
            'sample_size': len(orders),
            'archived_excluded': archived_skipped,
            'total_value': total_value,
            'total_value_str': _money(total_value),
            'by_warehouse': [
                {'warehouse': k, 'count': v['count'],
                 'value': v['value'], 'value_str': _money(v['value'])}
                for k, v in sorted(by_wh.items(), key=lambda kv: -kv[1]['count'])
            ],
            'by_route': [
                {'route': k, 'count': v['count'],
                 'value': v['value'], 'value_str': _money(v['value'])}
                for k, v in sorted(by_route.items(), key=lambda kv: -kv[1]['count'])
            ],
            'filter_applied': self._get_user_dashboard_context() if use_active_filter else {},
        }

    # ──────────────────────────────────────────────────────────────────
    # TOOL 3 — List orders (paginated)
    # ──────────────────────────────────────────────────────────────────
    @llm_tool(name='dp_list_orders', read_only_hint=True)
    def tool_list_orders(self, packing_status: str = '',
                         warehouse_name: str = '',
                         route_or_tag: str = '',
                         htgh: str = '',
                         search: str = '',
                         use_active_filter: bool = True,
                         limit: int = 60,
                         offset: int = 0) -> dict:
        """Liệt kê các đơn bán theo filter.

        Trả về 1 dict: {orders: [...], total, has_more}. Mỗi order chứa
        fields gọn: id, name, partner, address, route, htgh, amount,
        commitment_date, warehouse, shipper, product_count.

        Args:
            packing_status: ví dụ 'packed_waiting_ship', 'pending', 'all'.
            warehouse_name: tên kho (=ilike).
            route_or_tag: tên tag hoặc danh sách tag, phân tách dấu phẩy.
            htgh: lọc theo hình thức giao hàng.
            search: tìm theo tên KH, mã đơn, số điện thoại.
            use_active_filter: True = nối thêm filter user đang xem Kanban.
            limit: tối đa 100. Mặc định 30.
            offset: pagination.
        """
        limit = max(1, min(int(limit or 60), 200))
        kwargs = self._build_kwargs(
            packing_status=packing_status or None,
            warehouse_name=warehouse_name or None,
            route_or_tag=route_or_tag or None,
            htgh=htgh or None,
            search=search or None,
            use_active_filter=use_active_filter,
            limit=limit, offset=int(offset or 0),
        )
        try:
            data = self._service().get_dashboard_data(**kwargs)
        except Exception as e:
            _logger.exception("dp_list_orders failed")
            return {'error': str(e), 'orders': []}
        raw = data.get('orders') or []
        archived = self._get_archived_so_ids()
        archived_in_page = 0
        if archived:
            filtered = []
            for o in raw:
                if o.get('id') in archived:
                    archived_in_page += 1
                    continue
                filtered.append(o)
            raw = filtered
        out = []
        for o in raw:
            wh = o.get('warehouse_id')
            wh_name = wh[1] if isinstance(wh, (list, tuple)) and len(wh) > 1 else ''
            tag_pairs = o.get('tag_ids') or []
            tags = [t[1] for t in tag_pairs if isinstance(t, (list, tuple)) and len(t) > 1]
            partner = o.get('partner_id')
            partner_name = partner[1] if isinstance(partner, (list, tuple)) and len(partner) > 1 else ''
            shipper = ''
            picking_names = []
            for p in (o.get('pickings') or []):
                picking_names.append(p.get('name') or '')
                su = p.get('shipper_user_id')
                if isinstance(su, (list, tuple)) and len(su) > 1 and not shipper:
                    shipper = su[1]
            address = (o.get('misa_shipping_address') or '').strip()
            if len(address) > 120:
                address = address[:117] + '...'
            out.append({
                'id': o.get('id'),
                'name': o.get('name'),
                'partner': partner_name,
                'address': address,
                'route': ' / '.join(tags),
                'htgh': o.get('x_studio_htgh') or '',
                'amount': o.get('amount_total') or 0.0,
                'amount_str': _money(o.get('amount_total') or 0),
                'commitment_date': o.get('commitment_date') or '',
                'warehouse': wh_name,
                'shipper': shipper,
                'pickings': [n for n in picking_names if n][:5],
            })
        total = max((data.get('total_count') or len(raw)) - len(archived), len(raw))
        return {
            'orders': out,
            'total': total,
            'returned': len(out),
            'archived_excluded_in_page': archived_in_page,
            'archived_excluded_total': len(archived),
            'offset': int(offset or 0),
            'has_more': (int(offset or 0) + len(out) + archived_in_page) < (data.get('total_count') or len(out)),
            'filter_applied': self._get_user_dashboard_context() if use_active_filter else {},
        }

    # ──────────────────────────────────────────────────────────────────
    # TOOL 4 — Order detail
    # ──────────────────────────────────────────────────────────────────
    @llm_tool(name='dp_order_detail', read_only_hint=True)
    def tool_order_detail(self, order_id_or_name: str) -> dict:
        """Chi tiết 1 đơn bán: products, address, pickings, shipper, PO.

        Args:
            order_id_or_name: id (số) hoặc mã đơn (vd 'S00123').
        """
        SO = self.env['sale.order']
        rec = SO.browse()
        if str(order_id_or_name).isdigit():
            rec = SO.browse(int(order_id_or_name)).exists()
        if not rec:
            rec = SO.search([('name', '=', str(order_id_or_name))], limit=1)
        if not rec:
            return {'error': f"Không tìm thấy đơn '{order_id_or_name}'"}

        partner = rec.partner_shipping_id or rec.partner_id
        lines = []
        for ml in rec.order_line:
            if ml.product_id.type == 'service':
                continue
            lines.append({
                'product': ml.product_id.display_name,
                'ordered': ml.product_uom_qty or 0,
                'delivered': ml.qty_delivered or 0,
                'pending': max((ml.product_uom_qty or 0) - (ml.qty_delivered or 0), 0),
                'uom': ml.product_uom.name if ml.product_uom else '',
            })
        pickings = []
        for p in rec.picking_ids:
            shipper = p.shipper_user_id.name if p.shipper_user_id else ''
            pickings.append({
                'name': p.name,
                'type': p.picking_type_id.name if p.picking_type_id else '',
                'state': p.state,
                'scheduled_date': fields.Datetime.to_string(p.scheduled_date) if p.scheduled_date else '',
                'date_done': fields.Datetime.to_string(p.date_done) if p.date_done else '',
                'shipper': shipper,
            })
        pos = []
        for po in rec.purchase_order_ids if hasattr(rec, 'purchase_order_ids') else []:
            pos.append({'name': po.name, 'state': po.state,
                        'date_planned': fields.Datetime.to_string(po.date_planned) if po.date_planned else ''})
        return {
            'id': rec.id,
            'name': rec.name,
            'state': rec.state,
            'partner': partner.display_name if partner else '',
            'partner_phone': (partner.phone or partner.mobile) if partner else '',
            'address': rec.misa_shipping_address or (
                ', '.join(p for p in [
                    partner.street, partner.street2, partner.city,
                    partner.state_id.name if partner and partner.state_id else None,
                ] if p) if partner else ''
            ),
            'warehouse': rec.warehouse_id.name if rec.warehouse_id else '',
            'route': ' / '.join(rec.tag_ids.mapped('name')),
            'htgh': rec.x_studio_htgh or '',
            'delivery_type': rec.x_studio_delivery_type or '',
            'commitment_date': fields.Datetime.to_string(rec.commitment_date) if rec.commitment_date else '',
            'amount_total': rec.amount_total,
            'amount_total_str': _money(rec.amount_total),
            'currency': rec.currency_id.name if rec.currency_id else 'VND',
            'note': rec.origin or '',
            'lines': lines,
            'pickings': pickings,
            'purchase_orders': pos,
        }

    # ──────────────────────────────────────────────────────────────────
    # TOOL 5 — List routes (active)
    # ──────────────────────────────────────────────────────────────────
    @llm_tool(name='dp_list_routes', read_only_hint=True)
    def tool_list_routes(self, packing_status: str = '',
                         warehouse_name: str = '',
                         use_active_filter: bool = True) -> dict:
        """Danh sách các tuyến (tag_ids) đang có đơn ở trạng thái lọc.

        Trả về list tuyến + số đơn + tổng giá trị, sort giảm dần theo count.
        """
        kwargs = self._build_kwargs(
            packing_status=packing_status or None,
            warehouse_name=warehouse_name or None,
            use_active_filter=use_active_filter,
            limit=200, offset=0,
        )
        try:
            data = self._service().get_dashboard_data(**kwargs)
        except Exception as e:
            _logger.exception("dp_list_routes failed")
            return {'error': str(e), 'routes': []}
        routes = defaultdict(lambda: {'count': 0, 'value': 0.0, 'orders': []})
        no_route_key = '(chưa phân tuyến)'
        for o in data.get('orders') or []:
            v = o.get('amount_total') or 0.0
            tag_pairs = o.get('tag_ids') or []
            keys = [t[1] for t in tag_pairs if isinstance(t, (list, tuple)) and len(t) > 1] or [no_route_key]
            for k in keys:
                routes[k]['count'] += 1
                routes[k]['value'] += v
                if len(routes[k]['orders']) < 8:
                    routes[k]['orders'].append(o.get('name'))
        return {
            'routes': [
                {'name': k, 'count': v['count'], 'value': v['value'],
                 'value_str': _money(v['value']), 'sample_orders': v['orders']}
                for k, v in sorted(routes.items(), key=lambda kv: -kv[1]['count'])
            ],
            'filter_applied': self._get_user_dashboard_context() if use_active_filter else {},
        }

    # ──────────────────────────────────────────────────────────────────
    # TOOL 6 — Shipper history
    # ──────────────────────────────────────────────────────────────────
    @llm_tool(name='dp_shipper_history', read_only_hint=True)
    def tool_shipper_history(self, days: int = 30,
                             shipper_name: str = '') -> dict:
        """Lịch sử shipper (read-only, để gợi ý — KHÔNG tự assign).

        Args:
            days: số ngày tra cứu (mặc định 30, max 180).
            shipper_name: nếu cung cấp, chỉ trả 1 shipper khớp tên.
        """
        days = max(1, min(int(days or 30), 180))
        date_from = fields.Datetime.now() - timedelta(days=days)
        domain = [
            ('picking_type_id.code', '=', 'outgoing'),
            ('state', '=', 'done'),
            ('date_done', '>=', date_from),
            ('shipper_user_id', '!=', False),
        ]
        if shipper_name:
            user = self.env['res.users'].search(
                [('name', 'ilike', shipper_name)], limit=1,
            )
            if not user:
                return {'shippers': [], 'note': f'Không tìm thấy shipper "{shipper_name}"'}
            domain.append(('shipper_user_id', '=', user.id))
        pickings = self.env['stock.picking'].search(domain, limit=3000)

        agg = {}
        for p in pickings:
            su = p.shipper_user_id
            entry = agg.setdefault(su.id, {
                'name': su.name, 'completed_orders': 0,
                '_total_hours': 0.0, '_count_with_duration': 0,
                'avg_delivery_hours': None,
                'on_time_count': 0, 'late_count': 0,
                'routes': defaultdict(int),
            })
            entry['completed_orders'] += 1
            start = None
            for fld in ('shipper_received_date', 'shipper_received_at',
                        'date_pack', 'scheduled_date'):
                if fld in p._fields:
                    val = getattr(p, fld, None)
                    if val:
                        start = val
                        break
            if start and p.date_done:
                try:
                    h = (p.date_done - start).total_seconds() / 3600.0
                    if 0 < h < 240:
                        entry['_total_hours'] += h
                        entry['_count_with_duration'] += 1
                except Exception:
                    pass
            try:
                if p.scheduled_date and p.date_done:
                    if p.date_done > p.scheduled_date:
                        entry['late_count'] += 1
                    else:
                        entry['on_time_count'] += 1
            except Exception:
                pass
            try:
                so = p.sale_id
                if so:
                    for t in so.tag_ids:
                        entry['routes'][t.name] += 1
            except Exception:
                pass

        out = []
        for v in agg.values():
            if v['_count_with_duration']:
                v['avg_delivery_hours'] = round(v['_total_hours'] / v['_count_with_duration'], 2)
            on_total = v['on_time_count'] + v['late_count']
            out.append({
                'name': v['name'],
                'completed_orders': v['completed_orders'],
                'avg_delivery_hours': v['avg_delivery_hours'],
                'on_time_rate': round(100 * v['on_time_count'] / on_total) if on_total else None,
                'top_routes': dict(sorted(v['routes'].items(),
                                          key=lambda kv: -kv[1])[:8]),
            })
        out.sort(key=lambda x: -x['completed_orders'])
        return {'days': days, 'shippers': out}

    # ──────────────────────────────────────────────────────────────────
    # TOOL 7 — Warehouse info (địa chỉ kho xuất phát)
    # ──────────────────────────────────────────────────────────────────
    @llm_tool(name='dp_warehouse_info', read_only_hint=True)
    def tool_warehouse_info(self, warehouse_name: str = '') -> dict:
        """Trả về thông tin kho (tên + địa chỉ đầy đủ).

        Dùng để biết điểm xuất phát của shipper khi lập tuyến.

        Args:
            warehouse_name: tên kho. Để trống = lấy theo filter user
                đang xem trên Kanban (filter_warehouse_id).
        """
        wh_id = None
        if warehouse_name:
            wh_id = self._resolve_warehouse_id(warehouse_name)
        if not wh_id:
            ctx = self._get_user_dashboard_context()
            wh_id = ctx.get('filter_warehouse_id')
        if not wh_id:
            return {'has_warehouse': False,
                    'message': 'User không filter kho — không xác định được điểm xuất phát.'}
        wh = self.env['stock.warehouse'].browse(int(wh_id)).exists()
        if not wh:
            return {'has_warehouse': False, 'message': f'Không tồn tại warehouse_id {wh_id}'}
        partner = wh.partner_id
        full_addr = ''
        street = (partner.street or '') if partner else ''
        if partner:
            parts = [partner.street, partner.street2,
                     partner.city,
                     partner.state_id.name if partner.state_id else '',
                     partner.country_id.name if partner.country_id else '']
            full_addr = ', '.join([p for p in parts if p])
        return {
            'has_warehouse': True,
            'id': wh.id,
            'name': wh.name,
            'code': wh.code,
            'street': street,
            'address_full': full_addr or street,
            'phone': partner.phone if partner else '',
        }

    # ──────────────────────────────────────────────────────────────────
    # TOOL 8 — Fleet (đội xe)
    # ──────────────────────────────────────────────────────────────────
    @llm_tool(name='dp_fleet', read_only_hint=True)
    def tool_fleet(self, warehouse_name: str = '') -> dict:
        """Liệt kê đội xe khả dụng để phân chuyến.

        Args:
            warehouse_name: lọc xe theo kho gốc. Để trống = lấy theo
                filter user (nếu có), nếu không có thì lấy tất cả xe.

        Trả về list xe với: type, capacity_kg, max_orders_per_trip,
        preferred_for, driver. AI dùng để chia chuyến cho phù hợp loại xe.
        """
        wh_id = None
        if warehouse_name:
            wh_id = self._resolve_warehouse_id(warehouse_name)
        if not wh_id:
            ctx = self._get_user_dashboard_context()
            wh_id = ctx.get('filter_warehouse_id')
        Vehicle = self.env['hlv.delivery.planner.vehicle']
        fleet = Vehicle.get_active_fleet(warehouse_id=wh_id)
        return {
            'count': len(fleet),
            'warehouse_id': wh_id,
            'vehicles': fleet,
            'planning_hint': (
                'Chia chuyến theo loại xe: xe máy → đơn nhẹ <30kg gần kho; '
                'sedan → đơn vừa nội thành; van/truck → đơn lớn / KCN xa, '
                'gom nhiều đơn cùng tuyến.'
            ),
        }

    # ──────────────────────────────────────────────────────────────────
    # TOOL 9 — Locality breakdown (phân tích địa chỉ thô — không cần geocode)
    # ──────────────────────────────────────────────────────────────────
    _LOCALITY_PATTERNS = [
        # Industrial parks first (most specific)
        (r'\bKCN\s+([A-ZÀ-Ỹ][A-Za-zÀ-ỹ\s\-\d]{1,30}?)(?=,|\s*-|\s*\.|$|\s+(?:huyện|xã|phường|quận|tỉnh|tp|thành phố))',
         'KCN'),
        (r'\b(?:Khu Công Nghiệp)\s+([A-ZÀ-Ỹ][A-Za-zÀ-ỹ\s\-\d]{1,30}?)(?=,|\s*-|\s*\.|$)', 'KCN'),
        # Districts
        (r'\b(?:Huyện|H\.)\s+([A-ZÀ-Ỹ][A-Za-zÀ-ỹ\s\-]{1,25}?)(?=,|\s*-|\s*\.|$)', 'Huyện'),
        (r'\b(?:Quận|Q\.)\s*([0-9]{1,2}|[A-ZÀ-Ỹ][A-Za-zÀ-ỹ\s\-]{1,25}?)(?=,|\s*-|\s*\.|$)', 'Quận'),
        (r'\b(?:Thị xã|TX\.)\s+([A-ZÀ-Ỹ][A-Za-zÀ-ỹ\s\-]{1,25}?)(?=,|\s*-|\s*\.|$)', 'TX'),
        # Wards / communes
        (r'\b(?:Phường|P\.)\s*([0-9]{1,2}|[A-ZÀ-Ỹ][A-Za-zÀ-ỹ\s\-]{1,25}?)(?=,|\s*-|\s*\.|$)', 'Phường'),
        (r'\b(?:Xã|X\.)\s+([A-ZÀ-Ỹ][A-Za-zÀ-ỹ\s\-]{1,25}?)(?=,|\s*-|\s*\.|$)', 'Xã'),
        # Province / city as last resort
        (r'\b(?:TP|Tp|Thành phố|T\.P\.)\s*([A-ZÀ-Ỹ][A-Za-zÀ-ỹ\s\-]{1,25}?)(?=,|\s*-|\s*\.|$)', 'TP'),
        (r'\b(?:Tỉnh|T\.)\s+([A-ZÀ-Ỹ][A-Za-zÀ-ỹ\s\-]{1,25}?)(?=,|\s*-|\s*\.|$)', 'Tỉnh'),
    ]

    def _extract_locality(self, address):
        """Trích locality token từ string địa chỉ tiếng Việt.

        Trả về (kind, name) — vd ('KCN', 'Nhơn Trạch 3'), ('Huyện', 'Nhơn Trạch'),
        ('Quận', '7'), hoặc (None, None) nếu không match.
        """
        if not address:
            return (None, None)
        s = ' ' + str(address).strip() + ' '
        for pat, kind in self._LOCALITY_PATTERNS:
            m = re.search(pat, s, flags=re.IGNORECASE)
            if m:
                name = m.group(1).strip(' ,.-').title()
                if name and len(name) <= 40:
                    return (kind, name)
        return (None, None)

    @llm_tool(name='dp_locality_breakdown', read_only_hint=True)
    def tool_locality_breakdown(self, use_active_filter: bool = True,
                                limit: int = 200) -> dict:
        """Gom các đơn theo locality (KCN / Huyện / Quận / Xã / Phường) trích
        từ ``misa_shipping_address``. Dùng để lập tuyến tối ưu **mà không cần
        geocode** — gom đơn cùng locality đi 1 chuyến.

        Args:
            use_active_filter: True = áp filter user đang xem.
            limit: số đơn tối đa quét (default 200).

        Trả về:
            {groups: [{key, kind, name, count, value, value_str,
                       sample_orders: [{name, partner, address}]}],
             unparsed_count, total}
        """
        kwargs = self._build_kwargs(
            use_active_filter=use_active_filter,
            limit=int(limit), offset=0,
        )
        try:
            data = self._service().get_dashboard_data(**kwargs)
        except Exception as e:
            _logger.exception("dp_locality_breakdown failed")
            return {'error': str(e), 'groups': []}
        archived = self._get_archived_so_ids()
        orders = [o for o in (data.get('orders') or []) if o.get('id') not in archived]

        groups = defaultdict(lambda: {'count': 0, 'value': 0.0, 'samples': []})
        unparsed = 0
        for o in orders:
            addr = o.get('misa_shipping_address') or ''
            kind, name = self._extract_locality(addr)
            if not name:
                unparsed += 1
                key = '(không xác định)'
                kind = None
                name = '(không xác định)'
            else:
                key = f"{kind}: {name}"
            g = groups[key]
            g['count'] += 1
            g['value'] += o.get('amount_total') or 0.0
            if len(g['samples']) < 3:
                partner = o.get('partner_id')
                pname = partner[1] if isinstance(partner, (list, tuple)) and len(partner) > 1 else ''
                g['samples'].append({
                    'name': o.get('name'),
                    'partner': pname,
                    'address': (addr or '')[:80],
                })
            g['_kind'] = kind
            g['_name'] = name

        out = []
        for key, g in sorted(groups.items(), key=lambda kv: -kv[1]['count']):
            out.append({
                'key': key,
                'kind': g['_kind'],
                'name': g['_name'],
                'count': g['count'],
                'value': g['value'],
                'value_str': _money(g['value']),
                'sample_orders': g['samples'],
            })
        return {
            'total': len(orders),
            'unparsed_count': unparsed,
            'groups': out,
            'hint': (
                'Gom các group cùng KCN / Huyện / Quận → 1 chuyến. Group nhỏ '
                '(1-2 đơn) cùng khu vực gần nhau có thể merge. Chuyến lớn '
                '(KCN xa, nhiều đơn) → ưu tiên van/truck. Group ít đơn nội '
                'thành → xe máy / sedan.'
            ),
        }

    # ──────────────────────────────────────────────────────────────────
    # TOOL 10 — Rich Excel export (Claude điều khiển style)
    # ──────────────────────────────────────────────────────────────────
    @llm_tool(name='dp_export_excel')
    def tool_export_excel(self, filename: str,
                          headers: list,
                          rows: list,
                          sheet_name: str = 'Sheet1',
                          merges: list = None,
                          row_styles: list = None,
                          cell_styles: list = None,
                          column_widths: list = None,
                          header_fill: str = '4472C4',
                          header_font_color: str = 'FFFFFF',
                          freeze_header: bool = True) -> str:
        """Xuất Excel với toàn quyền styling (merge, fill color, font color).

        Dùng tool này thay cho ``file_export`` mặc định khi cần bảng đẹp:
        merge cell theo nhóm tuyến, tô màu hàng cảnh báo, set độ rộng cột…

        Args:
            filename: tên file (auto thêm .xlsx).
            headers: list tên cột.
            rows: list các hàng, mỗi hàng = list cell value (str/int/float/None).
            sheet_name: tên sheet.
            merges: list dải merge dạng "A2:A5" hoặc dict {"range":"A2:A5"}.
                    Dùng để gộp ô cùng nhóm tuyến / cùng chuyến.
            row_styles: list dict điều khiển style theo HÀNG (1-based, header
                là row 1). Ví dụ:
                  [{"row": 5, "fill": "FFEB9C", "font_color": "9C5700",
                    "bold": false}]
                Dùng để tô màu hàng cảnh báo, đơn quá hạn…
            cell_styles: list dict điều khiển style theo Ô (cell-level, đè
                row_styles). Ví dụ:
                  [{"row": 5, "col": 6, "fill": "FFC7CE",
                    "font_color": "9C0006", "bold": true,
                    "number_format": "#,##0"}]
                ``col`` là 1-based.
            column_widths: list số (đơn vị Excel character width). Index khớp
                với headers. Để None = auto theo nội dung. Ví dụ
                ``[12, 8, 28, 24, 14, 12, 14, 18, 20]``.
            header_fill: màu nền header (hex 6 ký tự, mặc định 4472C4).
            header_font_color: màu chữ header (mặc định FFFFFF).
            freeze_header: True = khoá hàng header khi cuộn.

        Trả về 1 dòng tóm tắt (đã đính kèm file vào message).
        """
        if not headers:
            return 'Error: headers cannot be empty'
        if not rows:
            return 'Error: rows cannot be empty'

        try:
            import openpyxl
            from openpyxl.styles import (Alignment, Border, Font, PatternFill,
                                          Side)
            from openpyxl.utils import get_column_letter
        except ImportError:
            return 'Error: openpyxl chưa được cài đặt trên server.'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (sheet_name or 'Sheet1')[:31]

        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header
        h_font = Font(bold=True, color=(header_font_color or 'FFFFFF').lstrip('#'),
                      size=11)
        h_fill = PatternFill(
            start_color=(header_fill or '4472C4').lstrip('#'),
            end_color=(header_fill or '4472C4').lstrip('#'),
            fill_type='solid',
        )
        h_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = h_font
            c.fill = h_fill
            c.alignment = h_align
            c.border = border

        # Body
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                if ci > len(headers):
                    break
                cell_val = self._coerce_cell_value(val)
                c = ws.cell(row=ri, column=ci, value=cell_val)
                c.border = border
                if isinstance(cell_val, (int, float)) and not isinstance(cell_val, bool):
                    c.alignment = Alignment(horizontal='right', vertical='center')
                    c.number_format = '#,##0' if isinstance(cell_val, int) else '#,##0.00'
                else:
                    c.alignment = Alignment(vertical='center', wrap_text=True)

        # Row styles
        for rs in (row_styles or []):
            try:
                r = int(rs.get('row'))
                fill = rs.get('fill')
                font_color = rs.get('font_color')
                bold = bool(rs.get('bold'))
                italic = bool(rs.get('italic'))
                for ci in range(1, len(headers) + 1):
                    c = ws.cell(row=r, column=ci)
                    if fill:
                        hex_fill = fill.lstrip('#')
                        c.fill = PatternFill(start_color=hex_fill,
                                              end_color=hex_fill,
                                              fill_type='solid')
                    if font_color or bold or italic:
                        existing = c.font
                        c.font = Font(
                            name=existing.name,
                            size=existing.size,
                            bold=bold or existing.bold,
                            italic=italic or existing.italic,
                            color=(font_color or '').lstrip('#') or existing.color,
                        )
            except Exception:
                _logger.debug('row_style ignored: %s', rs, exc_info=True)

        # Cell styles (override row)
        for cs in (cell_styles or []):
            try:
                r = int(cs.get('row'))
                ci = int(cs.get('col'))
                c = ws.cell(row=r, column=ci)
                fill = cs.get('fill')
                if fill:
                    hex_fill = fill.lstrip('#')
                    c.fill = PatternFill(start_color=hex_fill,
                                          end_color=hex_fill,
                                          fill_type='solid')
                font_color = cs.get('font_color')
                bold = cs.get('bold')
                italic = cs.get('italic')
                if any(v is not None for v in (font_color, bold, italic)):
                    existing = c.font
                    c.font = Font(
                        name=existing.name, size=existing.size,
                        bold=bool(bold) if bold is not None else existing.bold,
                        italic=bool(italic) if italic is not None else existing.italic,
                        color=(font_color or '').lstrip('#') or existing.color,
                    )
                num_fmt = cs.get('number_format')
                if num_fmt:
                    c.number_format = num_fmt
                align_h = cs.get('align')
                if align_h:
                    c.alignment = Alignment(horizontal=align_h, vertical='center',
                                             wrap_text=True)
            except Exception:
                _logger.debug('cell_style ignored: %s', cs, exc_info=True)

        # Merges
        for m in (merges or []):
            rng = m if isinstance(m, str) else (m or {}).get('range')
            if not rng:
                continue
            try:
                ws.merge_cells(rng)
                # Center content in merged cell
                first = ws[rng.split(':')[0]]
                first.alignment = Alignment(horizontal='center',
                                             vertical='center',
                                             wrap_text=True)
            except Exception:
                _logger.debug('merge ignored: %s', rng, exc_info=True)

        # Column widths
        if column_widths:
            for ci, w in enumerate(column_widths, 1):
                if not w:
                    continue
                try:
                    ws.column_dimensions[get_column_letter(ci)].width = float(w)
                except Exception:
                    pass
        else:
            # Auto-fit (cap 50)
            for col in ws.columns:
                max_len = 0
                # col may contain MergedCell — use first non-merged cell
                col_letter = None
                for cell in col:
                    if hasattr(cell, 'column_letter'):
                        col_letter = cell.column_letter
                        break
                if not col_letter:
                    continue
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

        if freeze_header:
            ws.freeze_panes = 'A2'

        # Output
        import io
        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': base64.b64encode(content).decode(),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'type': 'binary',
        })
        # Đính vào message giống file_export làm
        msg = self.env.context.get('message')
        if msg:
            try:
                msg.write({'attachment_ids': [(4, attachment.id)]})
                if msg.model and msg.res_id:
                    attachment.write({
                        'res_model': msg.model,
                        'res_id': msg.res_id,
                    })
            except Exception:
                _logger.warning('Could not attach xlsx to message', exc_info=True)

        return (
            f"Đã xuất file '{filename}' ({len(rows)} hàng, "
            f"{len(headers)} cột, {len(merges or [])} merge, "
            f"{len(row_styles or [])} row style, "
            f"{len(cell_styles or [])} cell style)."
        )

    @staticmethod
    def _coerce_cell_value(value):
        if value is None or isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, str):
            s = value.strip()
            # Don't auto-convert codes that look numeric (DH001, 0123…)
            if not s or s.startswith('0') and len(s) > 1:
                return value
            try:
                if '.' not in s and ',' not in s:
                    return int(s)
            except (ValueError, TypeError):
                pass
            try:
                return float(s.replace(',', ''))
            except (ValueError, TypeError):
                pass
        return value
