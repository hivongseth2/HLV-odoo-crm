# -*- coding: utf-8 -*-
from odoo import api, fields, models, _
from odoo.exceptions import UserError
import base64
import datetime
import json
from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

def _to_date_str(val):
    if not val:
        return ""
    if isinstance(val, str):
        try:
            d = fields.Datetime.from_string(val)
            if d:
                return d.strftime("%A, %B %d, %Y")
        except Exception:
            try:
                d2 = fields.Date.from_string(val)
                if d2:
                    return d2.strftime("%A, %B %d, %Y")
            except Exception:
                return val
        return val
    if isinstance(val, datetime.datetime):
        return val.strftime("%A, %B %d, %Y")
    if isinstance(val, datetime.date):
        return val.strftime("%A, %B %d, %Y")
    return str(val)

class PickingExportSalesReportWizard(models.TransientModel):
    _name = "picking.export.sales.report.wizard"
    _description = "Báo cáo bán hàng (Excel)"

    date_from = fields.Date(string="Từ ngày", required=True)
    date_to = fields.Date(string="Đến ngày", required=True)

    warehouse_ids = fields.Many2many(
        "stock.warehouse",
        string="Kho xuất",
        help="Để trống = Tất cả kho. Chọn 1 hoặc nhiều kho để lọc cụ thể.",
    )

    state_filter = fields.Selection(
        [
            ("all", "Tất cả"),
            ("assigned", "Đã kiểm tra tồn (assigned)"),
            ("done", "Đã hoàn thành (done)"),
            ("confirmed", "Đã xác nhận (confirmed)"),
            ("waiting", "Chờ khác (waiting)"),
        ],
        string="Trạng thái",
        default="all",
    )
    
    def _harsh_warehouse_code(self, code):
        if code == "KBC":
            return "BENCAM"
        if code == "TSN":
            return "HCM"
        if code == "KHD":
            return "HIENDUC"
        if code == "TSNSR":
            return "HCM_SHOWROOM"
        return code

    def _get_warehouse_name_vietnamese(self, code):
        mapping = {
            "KBC": "BẾN CAM",
            "BENCAM": "BẾN CAM",
            "TSN": "HCM",
            "HCM": "HCM",
            "KHD": "HIỀN ĐỨC",
            "HIENDUC": "HIỀN ĐỨC",
            "TSNSR": "HCM",
            "HCM_SHOWROOM": "HCM",
            "DNA": "ĐÀ NẴNG",
            "DANANG": "ĐÀ NẴNG",
        }
        return mapping.get(code, code)

    def _find_sale_order(self, move, picking):
        if getattr(move, 'sale_line_id', False) and move.sale_line_id.order_id:
            return move.sale_line_id.order_id
        grp = getattr(move, 'group_id', False)
        if grp and getattr(grp, 'sale_id', False):
            return grp.sale_id
        if getattr(picking, 'sale_id', False):
            return picking.sale_id
        return False
        
    def _get_columns_definition(self):
        return [
            {'key': 'hinh_thuc_ban_hang', 'name': 'Hình thức bán hàng', 'width': 25},
            {'key': 'phuong_thuc_thanh_toan', 'name': 'Phương thức thanh toán', 'width': 25},
            {'key': 'hinh_thuc_giao_hang', 'name': 'Hình thức giao hàng', 'width': 25},
            {'key': 'hinh_thuc_thanh_toan_so', 'name': 'Hình thức thanh toán (SO)', 'width': 25},
            {'key': 'ben_tra_phi_van_chuyen', 'name': 'Bên trả phí vận chuyển', 'width': 25},
            {'key': 'kiem_phieu_xuat_kho', 'name': 'Kiêm phiếu xuất kho', 'width': 20},
            {'key': 'lap_kem_hoa_don', 'name': 'Lập kèm hóa đơn', 'width': 18},
            {'key': 'da_lap_hoa_don', 'name': 'Đã lập hóa đơn', 'width': 18},
            {'key': 'so_chung_tu', 'name': 'Số chứng từ (*)', 'width': 20},
            {'key': 'so_phieu_xuat', 'name': 'Số phiếu xuất', 'width': 20},
            {'key': 'mau_so_hd', 'name': 'Mẫu số HĐ', 'width': 15},
            {'key': 'ky_hieu_hd', 'name': 'Ký hiệu HĐ', 'width': 15},
            {'key': 'ngay_hoa_don', 'name': 'Ngày hóa đơn', 'width': 25},
            {'key': 'ma_khach_hang', 'name': 'Mã khách hàng', 'width': 15},
            {'key': 'ten_khach_hang', 'name': 'Tên khách hàng', 'width': 40},
            {'key': 'dia_chi', 'name': 'Địa chỉ', 'width': 50},
            {'key': 'ma_so_thue', 'name': 'Mã số thuế', 'width': 15},
            {'key': 'nguoi_nop', 'name': 'Người nộp', 'width': 25},
            {'key': 'dien_giai', 'name': 'Diễn giải/Lý do nộp', 'width': 40},
            {'key': 'ly_do_xuat', 'name': 'Lý do xuất', 'width': 40},
            {'key': 'ma_nhan_vien', 'name': 'Mã nhân viên bán hàng', 'width': 20},
            {'key': 'so_ct_phieu_xuat', 'name': 'Số chứng từ kèm theo (Phiếu xuất)', 'width': 25},
            {'key': 'ma_hang', 'name': 'Mã hàng (*)', 'width': 18},
            {'key': 'ten_hang', 'name': 'Tên hàng', 'width': 40},
            {'key': 'la_dong_ghi_chu', 'name': 'Là dòng ghi chú', 'width': 18},
            {'key': 'hang_khuyen_mai', 'name': 'Hàng khuyến mại', 'width': 18},
            {'key': 'tk_tien_no', 'name': 'TK Tiền/Chi phí/Nợ (*)', 'width': 22},
            {'key': 'tk_doanh_thu_co', 'name': 'TK Doanh thu/Có (*)', 'width': 20},
            {'key': 'dvt', 'name': 'ĐVT', 'width': 12},
            {'key': 'so_luong', 'name': 'Số lượng', 'width': 12},
            {'key': 'don_gia', 'name': 'Đơn giá', 'width': 15},
            {'key': 'thanh_tien', 'name': 'Thành tiền', 'width': 15},
            {'key': 'bao_gom_thue', 'name': 'Bao gồm thuế', 'width': 15},
            {'key': 'ty_le_ck', 'name': 'Tỷ lệ CK (%)', 'width': 12},
            {'key': 'tien_chiet_khau', 'name': 'Tiền chiết khấu', 'width': 15},
            {'key': 'ty_le_thue_gtgt', 'name': '% thuế GTGT', 'width': 12},
            {'key': 'tien_thue_gtgt', 'name': 'Tiền thuế GTGT', 'width': 15},
            {'key': 'tk_thue_gtgt', 'name': 'TK thuế GTGT', 'width': 15},
            {'key': 'hh_khong_th_tren_to_khai', 'name': 'HH không TH trên tờ khai thuế GTGT', 'width': 35},
            {'key': 'ma_don_vi', 'name': 'Mã đơn vị', 'width': 15},
            {'key': 'so_don_dat_hang', 'name': 'Số đơn đặt hàng', 'width': 20},
            {'key': 'so_hop_dong_ban', 'name': 'Số hợp đồng bán', 'width': 20},
            {'key': 'cp_khong_hop_ly', 'name': 'CP không hợp lý', 'width': 18},
            {'key': 'ma_kho', 'name': 'Mã kho', 'width': 15},
            {'key': 'tk_gia_von', 'name': 'TK giá vốn', 'width': 15},
            {'key': 'tk_kho', 'name': 'TK Kho', 'width': 12},
            {'key': 'don_gia_von', 'name': 'Đơn giá vốn', 'width': 15},
            {'key': 'tien_von', 'name': 'Tiền vốn', 'width': 15},
            {'key': 'misa_sync', 'name': 'Misa Sync', 'width': 15},
            {'key': 'so_phieu_tra_lai', 'name': 'Số phiếu trả lại', 'width': 25},
            {'key': 'ngay_tra_lai', 'name': 'Ngày trả lại', 'width': 25},
            {'key': 'sl_tra_lai', 'name': 'SL trả lại', 'width': 12},
            {'key': 'sl_thuc_ban', 'name': 'SL thực bán', 'width': 12},
        ]

    def _domain(self):
        self.ensure_one()
        if self.date_from > self.date_to:
            raise UserError(_("Khoảng ngày không hợp lệ."))

        domain = [
            ("picking_type_code", "=", "outgoing"),
            ("date_done", ">=", fields.Date.to_date(self.date_from)),
            ("date_done", "<=", fields.Date.to_date(self.date_to)),
            ("state", "=", "done"),
        ]

        if self.warehouse_ids:
            domain.append(("picking_type_id.warehouse_id", "in", self.warehouse_ids.ids))

        return domain

    def _partner_code(self, partner):
        if not partner:
            return ""
        # Ưu tiên ref từ commercial_partner > parent > partner
        if partner.commercial_partner_id and partner.commercial_partner_id.ref:
            return partner.commercial_partner_id.ref
        if partner.parent_id and partner.parent_id.ref:
            return partner.parent_id.ref
        return partner.ref or (partner.company_registry if hasattr(partner, "company_registry") else None) or partner.vat or str(partner.id) or ""

    def _get_warehouse_code(self, picking):
        pt = picking.picking_type_id
        if pt and pt.warehouse_id:
            code = pt.warehouse_id.code or pt.warehouse_id.name or ""
            return self._harsh_warehouse_code(code)
        return ""

    def _get_move_line_rows(self, picking):
        rows = []
        so = self._find_sale_order(picking.move_ids_without_package[0] if picking.move_ids_without_package else None, picking)
        
        scheduled_date_str = _to_date_str(picking.scheduled_date)
        picking_name = picking.name or ""
        partner = picking.partner_id
        ref_partner = so.partner_id if so else partner
        partner_code = self._partner_code(ref_partner)
        partner_name = (partner and partner.name) or ""
        partner_address = ""
        import unicodedata
        def normalize_addr(s):
            s = s.strip().lower()
            s = unicodedata.normalize('NFD', s)
            s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
            return s
            
        if partner:
            street = partner.street or ""
            city = partner.city or ""
            state = partner.state_id.name if partner.state_id else ""
            address_parts = []
            normalized = set()
            for part in [street, city, state]:
                norm = normalize_addr(part) if part else ""
                if part and norm not in normalized:
                    address_parts.append(part)
                    normalized.add(norm)
            partner_address = ", ".join(address_parts)
            
        partner_vat = (partner and partner.vat) or ""
        
        sale_name = so.name if so else (picking.origin or "")
        sale_user_code = ''
        if so:
            misa_code = getattr(so, 'x_studio_misa_saler_code', None)
            if misa_code:
                sale_user_code = misa_code
            elif so.user_id:
                sale_user_code = so.user_id.login or so.user_id.name or ''
        
        dien_giai = ""
        if so and so.origin:
            dien_giai = f"Xuất kho bán hàng cho {partner_name}"
        elif picking.note:
            dien_giai = picking.note
        else:
            dien_giai = f"Bán hàng {partner_name}"
        
        ly_do_xuat = dien_giai
        warehouse_code = self._get_warehouse_code(picking)

        pos_order = getattr(picking, 'pos_order_id', False)
        
        if pos_order:
            for pos_line in pos_order.lines:
                prod = pos_line.product_id
                if not prod:
                    continue
                
                move = None
                ml = None
                if picking.move_line_ids:
                    for move_line in picking.move_line_ids:
                        if move_line.product_id == prod:
                            ml = move_line
                            move = move_line.move_id
                            break
                else:
                    for mv in picking.move_ids_without_package:
                        if mv.product_id == prod:
                            move = mv
                            break
                
                row = self._build_row_data(
                    picking, so, prod, ml, move,
                    scheduled_date_str, picking_name, partner_code, partner_name,
                    partner_address, partner_vat, sale_name, sale_user_code,
                    dien_giai, ly_do_xuat, warehouse_code,
                    pos_line=pos_line
                )
                rows.append(row)
        else:
            aggregated_sols = {}
            processed_moves_without_sol = []
            
            for mv in picking.move_ids_without_package:
                sol = getattr(mv, 'sale_line_id', False)
                if sol:
                    if sol.id not in aggregated_sols:
                        aggregated_sols[sol.id] = {
                            'sol': sol,
                            'move': mv,
                            'prod': sol.product_id,
                        }
                else:
                    processed_moves_without_sol.append(mv)
                    
            for sol_id, data in aggregated_sols.items():
                sol = data['sol']
                mv = data['move']
                prod = data['prod']
                
                row = self._build_row_data(
                    picking, so, prod, None, mv,
                    scheduled_date_str, picking_name, partner_code, partner_name,
                    partner_address, partner_vat, sale_name, sale_user_code,
                    dien_giai, ly_do_xuat, warehouse_code,
                    sale_line=sol
                )
                rows.append(row)
                
            for mv in processed_moves_without_sol:
                prod = mv.product_id
                if not prod:
                    continue
                row = self._build_row_data(
                    picking, so, prod, None, mv,
                    scheduled_date_str, picking_name, partner_code, partner_name,
                    partner_address, partner_vat, sale_name, sale_user_code,
                    dien_giai, ly_do_xuat, warehouse_code
                )
                rows.append(row)

        return rows

    def _compute_return_fields(self, picking, prod, move, qty):
        """Tính toán các trường trả lại cho một dòng sản phẩm."""
        # Ưu tiên tìm qua move.returned_move_ids trước
        return_info = self._get_return_info_for_move(move)
        # Nếu không tìm được qua move, tìm qua picking.return_ids
        if not return_info['sl_tra_lai'] and picking:
            return_info = self._get_return_info_for_product(picking, prod)
        return_info['sl_thuc_ban'] = qty - return_info['sl_tra_lai']
        return return_info

    def _get_return_info_for_move(self, move):
        """Tìm thông tin trả lại cho một stock.move dựa trên returned_move_ids."""
        if not move:
            return {'so_phieu_tra_lai': '', 'ngay_tra_lai': '', 'sl_tra_lai': 0.0}
        returned_moves = move.returned_move_ids.filtered(lambda m: m.state == 'done')
        if not returned_moves:
            return {'so_phieu_tra_lai': '', 'ngay_tra_lai': '', 'sl_tra_lai': 0.0}
        total_returned_qty = sum(m.quantity for m in returned_moves)
        picking_names = list(set(m.picking_id.name for m in returned_moves if m.picking_id))
        dates = [m.picking_id.date_done for m in returned_moves if m.picking_id and m.picking_id.date_done]
        latest_date = max(dates) if dates else False
        return {
            'so_phieu_tra_lai': ', '.join(picking_names),
            'ngay_tra_lai': _to_date_str(latest_date) if latest_date else '',
            'sl_tra_lai': total_returned_qty,
        }

    def _get_return_info_for_product(self, picking, prod):
        """Tìm thông tin trả lại cho sản phẩm từ return pickings của phiếu xuất."""
        return_pickings = picking.return_ids.filtered(lambda p: p.state == 'done')
        if not return_pickings:
            return {'so_phieu_tra_lai': '', 'ngay_tra_lai': '', 'sl_tra_lai': 0.0}
        total_returned_qty = 0.0
        picking_names = []
        dates = []
        for rp in return_pickings:
            for rm in rp.move_ids_without_package:
                if rm.product_id == prod and rm.state == 'done':
                    total_returned_qty += rm.quantity
            if rp.name not in picking_names:
                picking_names.append(rp.name)
            if rp.date_done:
                dates.append(rp.date_done)
        if total_returned_qty == 0.0:
            return {'so_phieu_tra_lai': '', 'ngay_tra_lai': '', 'sl_tra_lai': 0.0}
        latest_date = max(dates) if dates else False
        return {
            'so_phieu_tra_lai': ', '.join(picking_names),
            'ngay_tra_lai': _to_date_str(latest_date) if latest_date else '',
            'sl_tra_lai': total_returned_qty,
        }

    def _build_row_data(self, picking, so, prod, ml, move,
                        scheduled_date_str, picking_name, partner_code, partner_name,
                        partner_address, partner_vat, sale_name, sale_user_code,
                        dien_giai, ly_do_xuat, warehouse_code, pos_line=None, sale_line=None, forced_qty=None):
        
        product_code = prod.default_code or (prod.barcode if hasattr(prod, 'barcode') else "") or ""
        product_name = prod.display_name or prod.name or ""
        
        if not pos_line:
            pos_order = getattr(picking, 'pos_order_id', False)
            if pos_order and prod:
                pos_lines = pos_order.lines.filtered(lambda l: l.product_id == prod)
                if pos_lines:
                    pos_line = pos_lines[0]
        
        if sale_line:
            sol = sale_line
        else:
            sol = getattr(move, 'sale_line_id', False) if move else False
        
        if pos_line:
            uom = prod.uom_id
            qty = pos_line.qty or 0.0
            thanh_tien = pos_line.price_subtotal or 0.0
            ty_le_ck = pos_line.discount or 0.0
            if qty != 0 and ty_le_ck != 100.0:
                 don_gia = thanh_tien / (qty * (1 - ty_le_ck / 100.0))
            else:
                 don_gia = pos_line.price_unit
            tien_chiet_khau = abs(don_gia * qty * ty_le_ck / 100)
            ty_le_thue_gtgt = 0.0
            if pos_line.tax_ids_after_fiscal_position:
                for tax in pos_line.tax_ids_after_fiscal_position:
                    ty_le_thue_gtgt = tax.amount or 0.0
                    break
            tien_thue_gtgt = abs(thanh_tien * ty_le_thue_gtgt / 100)
            price_total = pos_line.price_subtotal_incl or 0.0
            
        elif sol:
            uom = sol.product_uom or prod.uom_id
            qty = sol.qty_delivered or sol.product_uom_qty or 0.0
            don_gia = sol.price_unit or 0.0
            ty_le_ck = sol.discount or 0.0
            tien_chiet_khau = (don_gia * qty * ty_le_ck) / 100
            
            thanh_tien = don_gia * qty * (1 - ty_le_ck / 100.0)
            
            ty_le_thue_gtgt = 0.0
            if sol.tax_id:
                for tax in sol.tax_id:
                    ty_le_thue_gtgt = tax.amount or 0.0
                    break
            tien_thue_gtgt = (thanh_tien * ty_le_thue_gtgt) / 100
            price_total = thanh_tien + tien_thue_gtgt
            
        else:
            if forced_qty is not None:
                uom = ml.product_uom_id if ml else (move.product_uom if move else prod.uom_id)
                qty = forced_qty
            elif ml:
                uom = ml.product_uom_id or prod.uom_id
                qty = ml.qty_done or 0.0
            elif move:
                uom = move.product_uom or prod.uom_id
                qty = move.qty_done if hasattr(move, 'qty_done') else (move.product_uom_qty or 0.0)
            else:
                uom = prod.uom_id
                qty = 1.0
            
            don_gia = prod.list_price or 0.0
            thanh_tien = don_gia * qty
            tien_chiet_khau = 0.0
            ty_le_ck = 0.0
            ty_le_thue_gtgt = 0.0
            tien_thue_gtgt = 0.0
            price_total = thanh_tien
        
        uom_name = (uom and uom.name) or ""
        don_gia_von = prod.standard_price or 0.0
        tien_von = don_gia_von * qty

        return {
            'hinh_thuc_ban_hang': 'Bán hàng hóa trong nước',
            'phuong_thuc_thanh_toan': 'Chưa thu tiền',
            'hinh_thuc_giao_hang': getattr(so, 'x_studio_htgh', '') or '' if so else '',
            'hinh_thuc_thanh_toan_so': getattr(so, 'x_studio_httt', '') or '' if so else '',
            'ben_tra_phi_van_chuyen': getattr(so, 'x_studio_misa_delivery', '') or '' if so else '',
            'kiem_phieu_xuat_kho': 'Có',
            'lap_kem_hoa_don': 'Có',
            'da_lap_hoa_don': 'Đã lập',
            'so_chung_tu': picking_name,
            'so_phieu_xuat': sale_name,
            'mau_so_hd': '01GTKT0/001',
            'ky_hieu_hd': '1C25TLV',
            'ngay_hoa_don': scheduled_date_str,
            'ma_khach_hang': partner_code,
            'ten_khach_hang': partner_name,
            'dia_chi': partner_address,
            'ma_so_thue': partner_vat,
            'nguoi_nop': partner_name,
            'dien_giai': dien_giai,
            'ly_do_xuat': ly_do_xuat,
            'ma_nhan_vien': sale_user_code,
            'so_ct_phieu_xuat': sale_name,
            'ma_hang': product_code,
            'ten_hang': product_name,
            'la_dong_ghi_chu': 'không',
            'hang_khuyen_mai': 'Không',
            'tk_tien_no': '131',
            'tk_doanh_thu_co': '5111',
            'dvt': uom_name,
            'so_luong': qty,
            'don_gia': don_gia,
            'thanh_tien': thanh_tien,
            'bao_gom_thue': f"{price_total:,.2f}",
            'ty_le_ck': ty_le_ck,
            'tien_chiet_khau': tien_chiet_khau,
            'ty_le_thue_gtgt': ty_le_thue_gtgt,
            'tien_thue_gtgt': tien_thue_gtgt,
            'tk_thue_gtgt': '33311',
            'hh_khong_th_tren_to_khai': 'Không',
            'ma_don_vi': 'PKD',
            'so_don_dat_hang': sale_name,
            'so_hop_dong_ban': sale_name,
            'cp_khong_hop_ly': 'Không',
            'ma_kho': warehouse_code,
            'tk_gia_von': '632',
            'tk_kho': '156',
            'don_gia_von': don_gia_von,
            'tien_von': tien_von,
            'misa_sync': getattr(picking, 'x_studio_misa_sav', False),
            **self._compute_return_fields(picking, prod, move, qty),
        }

    def _create_excel_workbook(self, data_rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "Báo cáo bán hàng"

        columns = self._get_columns_definition()

        header_font = Font(name='Arial', size=10, bold=True)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border_side = Side(style='thin', color='000000')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        number_alignment = Alignment(horizontal='right', vertical='center')

        HEADER_ROW = 1
        DATA_START = 2

        for col_idx, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=HEADER_ROW, column=col_idx)
            cell.value = col_def['name']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def.get('width', 15)

        for row_idx, row_data in enumerate(data_rows, start=DATA_START):
            for col_idx, col_def in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(col_def['key'], "")

                if value is None:
                    value = ""

                cell.value = value
                cell.border = border

                if isinstance(value, (int, float)) and value != "":
                    cell.alignment = number_alignment
                    if col_def['key'] in ['don_gia', 'thanh_tien', 'tien_chiet_khau', 
                                         'tien_thue_gtgt', 'don_gia_von', 'tien_von']:
                        cell.number_format = '#,##0'
                    elif col_def['key'] in ['ty_le_ck', 'ty_le_thue_gtgt', 'ty_le_thue_xk']:
                        cell.number_format = '0.00'
                    elif col_def['key'] == 'so_luong':
                        cell.number_format = '#,##0.00'
                else:
                    cell.alignment = cell_alignment

        ws.row_dimensions[HEADER_ROW].height = 30

        return wb

    def action_export(self):
        self.ensure_one()
        if Workbook is None:
            raise UserError(_("Thiếu thư viện openpyxl. Vui lòng cài đặt 'openpyxl' cho Python."))

        pickings = self.env["stock.picking"].sudo().search(self._domain(), order="scheduled_date asc, id asc")
        if not pickings:
            raise UserError(_("Không tìm thấy phiếu xuất kho nào trong khoảng ngày đã chọn."))

        all_rows = []
        for picking in pickings:
            rows = self._get_move_line_rows(picking)
            all_rows.extend(rows)

        if not all_rows:
            raise UserError(_("Không có dữ liệu chi tiết để xuất."))

        wb = self._create_excel_workbook(all_rows)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        file_data = base64.b64encode(out.getvalue())
        
        attachment = self.env['ir.attachment'].create({
            'name': f"Bao_cao_ban_hang_{fields.Date.to_string(self.date_from)}_{fields.Date.to_string(self.date_to)}.xlsx",
            'type': 'binary',
            'datas': file_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'new',
        }
