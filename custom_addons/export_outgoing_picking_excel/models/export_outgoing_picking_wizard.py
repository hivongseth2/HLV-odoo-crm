# -*- coding: utf-8 -*-
from odoo import api, fields, models, _
from odoo.exceptions import UserError
import base64
import datetime
import json
from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


def _to_date_str(val, hour=None):
    if not val:
        return ""
    
    fmt = "%A, %B %d, %Y"
    if hour is not None:
        fmt += f" {hour:02d}:00:00"

    if isinstance(val, str):
        try:
            d = fields.Datetime.from_string(val)
            if d:
                return d.strftime(fmt)
        except Exception:
            try:
                d2 = fields.Date.from_string(val)
                if d2:
                    return d2.strftime(fmt)
            except Exception:
                return val
        return val
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime(fmt)
    return str(val)


class PickingExportWizard(models.TransientModel):
    def _harsh_warehouse_code(self, code):
        if code == "KBC":
            return "BENCAM"
        if code == "TSN":
            return "HCM"
        if code == "KHD":
            return "HIENDUC"
        if code == "TSNSR":
            return "HCM_SHOWROOM"
        return code
    
    def _get_warehouse_name_vietnamese(self, code):
        """Map warehouse code to Vietnamese name for POS CRM export"""
        mapping = {
            "KBC": "BẾN CAM",
            "BENCAM": "BẾN CAM",
            "TSN": "HCM",
            "HCM": "HCM",
            "KHD": "HIỀN ĐỨC",
            "HIENDUC": "HIỀN ĐỨC",
            "TSNSR": "HCM",
            "HCM_SHOWROOM": "HCM",
            "DNA": "ĐÀ NẴNG",
            "DANANG": "ĐÀ NẴNG",
        }
        return mapping.get(code, code)
    _name = "picking.export.wizard"
    _description = "Xuất Excel lệnh xuất kho theo template kế toán"

    date_from = fields.Date(string="Từ ngày", required=True)
    date_to = fields.Date(string="Đến ngày", required=True)

    warehouse_ids = fields.Many2many(
        "stock.warehouse",
        string="Kho xuất",
        help="Để trống = Tất cả kho. Chọn 1 hoặc nhiều kho để lọc cụ thể.",
    )

    state_filter = fields.Selection(
        [
            ("all", "Tất cả"),
            ("assigned", "Đã kiểm tra tồn (assigned)"),
            ("done", "Đã hoàn thành (done)"),
            ("confirmed", "Đã xác nhận (confirmed)"),
            ("waiting", "Chờ khác (waiting)"),
        ],
        string="Trạng thái",
        default="all",
    )

    def _find_sale_order(self, move, picking):
        # 1) Từ sale_line_id trực tiếp
        if getattr(move, 'sale_line_id', False) and move.sale_line_id.order_id:
            return move.sale_line_id.order_id
        # 2) Từ procurement group
        grp = getattr(move, 'group_id', False)
        if grp and getattr(grp, 'sale_id', False):
            return grp.sale_id
        # 3) Từ picking
        if getattr(picking, 'sale_id', False):
            return picking.sale_id
        return False

    def _find_source_pos_picking(self, picking):
        """
        Truy ngược từ phiếu OUT/PACK về phiếu PICK gốc có POS order.
        Dùng cho 3-step delivery: OUT không có pos_order_id / x_studio_pos_group,
        nhưng phiếu PICK gốc thì có.
        Trả về picking gốc hoặc False.
        """
        visited = set()
        pickings_to_check = picking
        
        while pickings_to_check:
            for p in pickings_to_check:
                if p.id in visited:
                    continue
                visited.add(p.id)
                
                # Kiểm tra phiếu này có POS order không
                pos_order = getattr(p, 'pos_order_id', False)
                if pos_order:
                    return p
            
            # Truy ngược qua move_orig_ids
            orig_moves = pickings_to_check.mapped('move_ids.move_orig_ids')
            upstream_pickings = orig_moves.mapped('picking_id').filtered(
                lambda p: p.id not in visited
            )
            
            if not upstream_pickings:
                break
            pickings_to_check = upstream_pickings
        
        return False

    def _get_columns_definition(self):
        """Định nghĩa các cột theo template kế toán"""
        return [
            {'key': 'hinh_thuc_ban_hang', 'name': 'Hình thức bán hàng', 'width': 25},
            {'key': 'phuong_thuc_thanh_toan', 'name': 'Phương thức thanh toán', 'width': 25},
            {'key': 'hinh_thuc_giao_hang', 'name': 'Hình thức giao hàng', 'width': 25},
            {'key': 'hinh_thuc_thanh_toan_so', 'name': 'Hình thức thanh toán (SO)', 'width': 25},
            {'key': 'ben_tra_phi_van_chuyen', 'name': 'Bên trả phí vận chuyển', 'width': 25},
            {'key': 'kiem_phieu_xuat_kho', 'name': 'Kiêm phiếu xuất kho', 'width': 20},
            {'key': 'lap_kem_hoa_don', 'name': 'Lập kèm hóa đơn', 'width': 18},
            {'key': 'da_lap_hoa_don', 'name': 'Đã lập hóa đơn', 'width': 18},
            {'key': 'ngay_hach_toan', 'name': 'Ngày hạch toán (*)', 'width': 25},
            {'key': 'ngay_chung_tu', 'name': 'Ngày chứng từ (*)', 'width': 25},
            {'key': 'so_chung_tu', 'name': 'Số chứng từ (*)', 'width': 20},
            {'key': 'so_phieu_xuat', 'name': 'Số phiếu xuất', 'width': 20},
            {'key': 'mau_so_hd', 'name': 'Mẫu số HĐ', 'width': 15},
            {'key': 'ky_hieu_hd', 'name': 'Ký hiệu HĐ', 'width': 15},
            {'key': 'so_hoa_don', 'name': 'Số hóa đơn', 'width': 15},
            {'key': 'ngay_hoa_don', 'name': 'Ngày hóa đơn', 'width': 25},
            {'key': 'ma_khach_hang', 'name': 'Mã khách hàng', 'width': 15},
            {'key': 'ten_khach_hang', 'name': 'Tên khách hàng', 'width': 40},
            {'key': 'dia_chi', 'name': 'Địa chỉ', 'width': 50},
            {'key': 'ma_so_thue', 'name': 'Mã số thuế', 'width': 15},
            {'key': 'don_vi_giao_dai_ly', 'name': 'Đơn vị giao đại lý', 'width': 30},
            {'key': 'nguoi_nop', 'name': 'Người nộp', 'width': 25},
            {'key': 'nop_vao_tk', 'name': 'Nộp vào TK', 'width': 15},
            {'key': 'ten_ngan_hang', 'name': 'Tên ngân hàng', 'width': 30},
            {'key': 'dien_giai', 'name': 'Diễn giải/Lý do nộp', 'width': 40},
            {'key': 'ly_do_xuat', 'name': 'Lý do xuất', 'width': 40},
            {'key': 'ma_nhan_vien', 'name': 'Mã nhân viên bán hàng', 'width': 20},
            {'key': 'so_ct_phieu_thu', 'name': 'Số chứng từ kèm theo (Phiếu thu)', 'width': 25},
            {'key': 'so_ct_phieu_xuat', 'name': 'Số chứng từ kèm theo (Phiếu xuất)', 'width': 25},
            {'key': 'han_thanh_toan', 'name': 'Hạn thanh toán', 'width': 20},
            {'key': 'ma_hang', 'name': 'Mã hàng (*)', 'width': 18},
            {'key': 'thuoc_combo', 'name': 'Thuộc combo', 'width': 15},
            {'key': 'ten_hang', 'name': 'Tên hàng', 'width': 40},
            {'key': 'la_dong_ghi_chu', 'name': 'Là dòng ghi chú', 'width': 18},
            {'key': 'hang_khuyen_mai', 'name': 'Hàng khuyến mại', 'width': 18},
            {'key': 'chiet_khau_thuong_mai', 'name': 'Chiết khấu thương mại', 'width': 25},
            {'key': 'tk_tien_no', 'name': 'TK Tiền/Chi phí/Nợ (*)', 'width': 22},
            {'key': 'tk_doanh_thu_co', 'name': 'TK Doanh thu/Có (*)', 'width': 20},
            {'key': 'dvt', 'name': 'ĐVT', 'width': 12},
            {'key': 'so_luong', 'name': 'Số lượng', 'width': 12},
            {'key': 'don_gia', 'name': 'Đơn giá', 'width': 15},
            {'key': 'thanh_tien', 'name': 'Thành tiền', 'width': 15},
            {'key': 'bao_gom_thue', 'name': 'Bao gồm thuế', 'width': 15},
            {'key': 'ty_le_ck', 'name': 'Tỷ lệ CK (%)', 'width': 12},
            {'key': 'tien_chiet_khau', 'name': 'Tiền chiết khấu', 'width': 15},
            {'key': 'tk_chiet_khau', 'name': 'TK chiết khấu', 'width': 15},
            {'key': 'gia_tinh_thue_xk', 'name': 'Giá tính thuế XK', 'width': 15},
            {'key': 'ty_le_thue_xk', 'name': '% thuế xuất khẩu', 'width': 15},
            {'key': 'tien_thue_xk', 'name': 'Tiền thuế xuất khẩu', 'width': 18},
            {'key': 'tk_thue_xk', 'name': 'TK thuế xuất khẩu', 'width': 18},
            {'key': 'ty_le_thue_gtgt', 'name': '% thuế GTGT', 'width': 12},
            {'key': 'ty_le_thue_khac', 'name': '% thuế suất KHAC', 'width': 18},
            {'key': 'tien_thue_gtgt', 'name': 'Tiền thuế GTGT', 'width': 15},
            {'key': 'tk_thue_gtgt', 'name': 'TK thuế GTGT', 'width': 15},
            {'key': 'bien_kiem_soat', 'name': 'Biển kiểm soát ', 'width': 18},
            {'key': 'hh_khong_th_tren_to_khai', 'name': 'HH không TH trên tờ khai thuế GTGT', 'width': 35},
            {'key': 'ma_khoan_muc_cp', 'name': 'Mã khoản mục chi phí', 'width': 22},
            {'key': 'ma_don_vi', 'name': 'Mã đơn vị', 'width': 15},
            {'key': 'ma_doi_tuong_thcp', 'name': 'Mã đối tượng THCP', 'width': 20},
            {'key': 'ma_cong_trinh', 'name': 'Mã công trình', 'width': 18},
            {'key': 'so_don_dat_hang', 'name': 'Số đơn đặt hàng', 'width': 20},
            {'key': 'so_hop_dong_ban', 'name': 'Số hợp đồng bán', 'width': 20},
            {'key': 'ma_thong_ke', 'name': 'Mã thống kê', 'width': 15},
            {'key': 'so_khe_uoc_cho_vay', 'name': 'Số khế ước cho vay', 'width': 20},
            {'key': 'cp_khong_hop_ly', 'name': 'CP không hợp lý', 'width': 18},
            {'key': 'ma_kho', 'name': 'Mã kho', 'width': 15},
            {'key': 'tk_gia_von', 'name': 'TK giá vốn', 'width': 15},
            {'key': 'tk_kho', 'name': 'TK Kho', 'width': 12},
            {'key': 'don_gia_von', 'name': 'Đơn giá vốn', 'width': 15},
            {'key': 'tien_von', 'name': 'Tiền vốn', 'width': 15},
            {'key': 'hang_hoa_giu_ho', 'name': 'Hàng hóa giữ hộ/bán hộ', 'width': 25},
            {'key': 'vi_tri', 'name': 'vị trí', 'width': 25},
            {'key': 'misa_sync', 'name': 'Misa Sync', 'width': 15},
        ]

    def _domain(self):
        self.ensure_one()
        if self.date_from > self.date_to:
            raise UserError(_("Khoảng ngày không hợp lệ."))

        domain = [
            ("picking_type_code", "=", "outgoing"),
            ("date_done", ">=", fields.Date.to_date(self.date_from)),
            ("date_done", "<=", fields.Date.to_date(self.date_to)),
            ("state", "=", "done"),
        ]

        if self.warehouse_ids:
            domain.append(("picking_type_id.warehouse_id", "in", self.warehouse_ids.ids))

        # Không cần lọc state_filter nữa vì đã cố định là 'done'

        return domain

    def _partner_code(self, partner):
        if not partner:
            return ""
        # Ưu tiên ref từ commercial_partner > parent > partner
        if partner.commercial_partner_id and partner.commercial_partner_id.ref:
            return partner.commercial_partner_id.ref
        if partner.parent_id and partner.parent_id.ref:
            return partner.parent_id.ref
        return partner.ref or (partner.company_registry if hasattr(partner, "company_registry") else None) or partner.vat or str(partner.id) or ""

    def _get_warehouse_code(self, picking):
        """Lấy mã kho"""
        pt = picking.picking_type_id
        if pt and pt.warehouse_id:
            code = pt.warehouse_id.code or pt.warehouse_id.name or ""
            return self._harsh_warehouse_code(code)
        return ""

    # ====== HELPERS: xác định parent combo cho 1 sale.order.line ======
    def _thuoc_combo_code_for_move(self, move):
        """
        Trả về mã combo cha (default_code) cho move nếu là dòng con của combo.
        Sử dụng logic BoM Kit (phantom) hoặc Service Combo (header).
        """
        if not move:
            return ''
            
        sol = getattr(move, 'sale_line_id', False)
        if not sol:
            return ''
        
        sol_product = sol.product_id
        if not sol_product:
            return ''

        # Nếu sản phẩm trên move trùng với sản phẩm trên SOL -> chính nó (không phải component)
        if sol_product.id == move.product_id.id:
            return ''
        
        # 1. Check Service Combo (Header là service, move là component)
        if sol_product.type == 'service':
            return sol_product.default_code or ''

        # 2. Check BoM Kit
        is_kit = self.env['mrp.bom'].search_count([
            ('product_tmpl_id', '=', sol_product.product_tmpl_id.id),
            ('type', '=', 'phantom'),
            ('active', '=', True)
        ])
        
        if is_kit:
            return sol_product.default_code or ''
            
        return ''

    def _get_move_line_rows(self, picking):
        rows = []
        so = self._find_sale_order(picking.move_ids_without_package[0] if picking.move_ids_without_package else None, picking)
        
        # Thông tin chung từ Sale Order hoặc Picking
        scheduled_date_str = _to_date_str(picking.scheduled_date)
        picking_name = picking.name or ""
        partner = picking.partner_id
        ref_partner = so.partner_id if so else partner
        partner_code = self._partner_code(ref_partner)
        partner_name = (partner and partner.name) or ""
        partner_address = ""
        import unicodedata
        def normalize_addr(s):
            # Chuẩn hóa: về chữ thường, loại bỏ dấu, loại bỏ khoảng trắng thừa
            s = s.strip().lower()
            s = unicodedata.normalize('NFD', s)
            s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
            return s
        if partner:
            street = partner.street or ""
            city = partner.city or ""
            state = partner.state_id.name if partner.state_id else ""
            address_parts = []
            normalized = set()
            for part in [street, city, state]:
                norm = normalize_addr(part) if part else ""
                if part and norm not in normalized:
                    address_parts.append(part)
                    normalized.add(norm)
            partner_address = ", ".join(address_parts)
        partner_vat = (partner and partner.vat) or ""
        
        # Lấy thông tin từ Sale Order
        sale_name = so.name if so else (picking.origin or "")
        # Lấy mã nhân viên sale từ trường x_studio_misa_saler_code của sale.order, nếu không có thì lấy từ user Odoo
        sale_user_code = ''
        if so:
            misa_code = getattr(so, 'x_studio_misa_saler_code', None)
            if misa_code:
                sale_user_code = misa_code
            elif so.user_id:
                sale_user_code = so.user_id.login or so.user_id.name or ''
        
        # Diễn giải
        dien_giai = ""
        if so and so.origin:
            dien_giai = f"Xuất kho bán hàng cho {partner_name}"
        elif picking.note:
            dien_giai = picking.note
        else:
            dien_giai = f"Bán hàng {partner_name}"
        
        ly_do_xuat = dien_giai
        
        warehouse_code = self._get_warehouse_code(picking)

        # KIỂM TRA: Nếu là đơn POS, loop qua pos.order.lines thay vì stock moves
        pos_order = getattr(picking, 'pos_order_id', False)
        
        if pos_order:
            # Loop qua từng POS order line (để đảm bảo xuất đủ số dòng)
            for pos_line in pos_order.lines:
                prod = pos_line.product_id
                if not prod:
                    continue
                
                # Kiểm tra xem sản phẩm có phải là combo (có BoM kit) không
                bom = self.env['mrp.bom']._bom_find(prod, bom_type='phantom')
                
                # Tìm stock move tương ứng (nếu cần location info)
                move = None
                ml = None
                if picking.move_line_ids:
                    for move_line in picking.move_line_ids:
                        if move_line.product_id == prod:
                            ml = move_line
                            move = move_line.move_id
                            break
                else:
                    for mv in picking.move_ids_without_package:
                        if mv.product_id == prod:
                            move = mv
                            break
                
                # Nếu là combo, xuất dòng cha trước
                if bom:
                    parent_row = self._build_row_data(
                        picking, so, prod, ml, move,
                        scheduled_date_str, picking_name, partner_code, partner_name,
                        partner_address, partner_vat, sale_name, sale_user_code,
                        dien_giai, ly_do_xuat, warehouse_code,
                        pos_line=pos_line  # Truyền pos_line vào để ưu tiên
                    )
                    rows.append(parent_row)
                    
                    # Xuất các dòng con của combo (từ stock moves)
                    # Tìm tất cả stock moves có cùng sale_line_id hoặc cùng combo parent
                    if picking.move_line_ids:
                        for move_line in picking.move_line_ids:
                            child_move = move_line.move_id
                            # Kiểm tra xem move này có phải là component của combo không
                            if child_move and getattr(child_move, 'sale_line_id', False):
                                sol = child_move.sale_line_id
                                # Nếu sale line product khác với move product → là component
                                if sol.product_id == prod and child_move.product_id != prod:
                                    child_row = self._build_row_data(
                                        picking, so, child_move.product_id, move_line, child_move,
                                        scheduled_date_str, picking_name, partner_code, partner_name,
                                        partner_address, partner_vat, sale_name, sale_user_code,
                                        dien_giai, ly_do_xuat, warehouse_code
                                    )
                                    rows.append(child_row)
                else:
                    # Không phải combo, xuất bình thường
                    row = self._build_row_data(
                        picking, so, prod, ml, move,
                        scheduled_date_str, picking_name, partner_code, partner_name,
                        partner_address, partner_vat, sale_name, sale_user_code,
                        dien_giai, ly_do_xuat, warehouse_code,
                        pos_line=pos_line
                    )
                    rows.append(row)
        else:
            # Logic cho non-POS orders: Xuất combo headers và stock moves
            # AGGREGATION LOGIC: Gộp các move lines cùng sản phẩm/giá/sol lại thành 1 dòng
            if picking.move_line_ids:
                # Key: (sale_line_id, product_id, uom_id)
                # Value: { 'qty': float, 'ml': record, 'move': record }
                aggregated_lines = {}
                ordered_keys = []
                
                for ml in picking.move_line_ids:
                    move = ml.move_id
                    prod = ml.product_id
                    if not prod:
                        continue

                    # Determine Group Key
                    sol = getattr(move, 'sale_line_id', False)
                    sol_id = sol.id if sol else 0
                    key = (sol_id, prod.id, ml.product_uom_id.id)
                    
                    if key not in aggregated_lines:
                        aggregated_lines[key] = {
                            'qty': 0.0,
                            'ml': ml,
                            'move': move,
                            'prod': prod
                        }
                        ordered_keys.append(key)
                    
                    # Sum quantity
                    aggregated_lines[key]['qty'] += (ml.qty_done or 0.0)
                
                # Render Aggregated Rows
                for key in ordered_keys:
                    data = aggregated_lines[key]
                    ml_agg = data['ml']
                    move_agg = data['move']
                    prod_agg = data['prod']
                    qty_agg = data['qty']
                    
                    if qty_agg == 0:
                        continue

                    row = self._build_row_data(
                        picking, so, prod_agg, ml_agg, move_agg,
                        scheduled_date_str, picking_name, partner_code, partner_name,
                        partner_address, partner_vat, sale_name, sale_user_code,
                        dien_giai, ly_do_xuat, warehouse_code,
                        forced_qty=qty_agg
                    )
                    rows.append(row)

            else:
                for mv in picking.move_ids_without_package:
                    prod = mv.product_id
                    if not prod:
                        continue
                    
                    sol = getattr(mv, 'sale_line_id', False)
                    
                    # Nếu có sale order line và chưa xử lý
                    if sol and sol.id not in processed_sale_lines:
                        processed_sale_lines.add(sol.id)
                        
                        # Kiểm tra BoM combo
                        if sol.product_id != prod:
                            combo_prod = sol.product_id
                            
                            # Xuất dòng combo cha (BoM)
                            parent_row = self._build_row_data(
                                picking, so, combo_prod, None, mv,
                                scheduled_date_str, picking_name, partner_code, partner_name,
                                partner_address, partner_vat, sale_name, sale_user_code,
                                dien_giai, ly_do_xuat, warehouse_code
                            )
                            rows.append(parent_row)
                    
                    # Xuất dòng hiện tại
                    row = self._build_row_data(
                        picking, so, prod, None, mv,
                        scheduled_date_str, picking_name, partner_code, partner_name,
                        partner_address, partner_vat, sale_name, sale_user_code,
                        dien_giai, ly_do_xuat, warehouse_code
                    )
                    rows.append(row)

        return rows

    # build row của crm
    def _build_row_data(self, picking, so, prod, ml, move,
                        scheduled_date_str, picking_name, partner_code, partner_name,
                        partner_address, partner_vat, sale_name, sale_user_code,
                        dien_giai, ly_do_xuat, warehouse_code, pos_line=None, sale_line=None, forced_qty=None):
        """Xây dựng dữ liệu cho 1 dòng"""
        
        product_code = prod.default_code or (prod.barcode if hasattr(prod, 'barcode') else "") or ""
        product_name = prod.display_name or prod.name or ""
        
        # Ưu tiên 1: Sử dụng pos_line được truyền vào (nếu có)
        # Nếu không có pos_line được truyền vào, tìm trong picking
        if not pos_line:
            pos_order = getattr(picking, 'pos_order_id', False)
            if pos_order and prod:
                # Tìm pos.order.line tương ứng với sản phẩm này
                pos_lines = pos_order.lines.filtered(lambda l: l.product_id == prod)
                if pos_lines:
                    pos_line = pos_lines[0]  # Lấy dòng đầu tiên nếu có nhiều
        
        # Ưu tiên 2: Sale Order Line (truyền vào hoặc tìm từ move)
        if sale_line:
            sol = sale_line
        else:
            sol = getattr(move, 'sale_line_id', False) if move else False
        
        # Logic lấy dữ liệu theo thứ tự ưu tiên
        if pos_line:
            # Lấy từ POS Order Line (số lượng đã có dấu âm cho đơn hoàn tiền)
            uom = prod.uom_id  # POS không có product_uom field
            qty = pos_line.qty or 0.0  # Số lượng từ POS (đã âm nếu là return)
            # Thành tiền chưa thuế
            thanh_tien = pos_line.price_subtotal or 0.0
            
            # Tính lại đơn giá chưa thuế để tương thích với chiết khấu
            ty_le_ck = pos_line.discount or 0.0
            if qty != 0 and ty_le_ck != 100.0:
                 don_gia = thanh_tien / (qty * (1 - ty_le_ck / 100.0))
            else:
                 don_gia = pos_line.price_unit# Fallback if qty is 0 or discount is 100%
            
            # Tính tiền chiết khấu
            tien_chiet_khau = abs(don_gia * qty * ty_le_ck / 100)
            
            # Thuế GTGT từ POS
            ty_le_thue_gtgt = 0.0
            if pos_line.tax_ids_after_fiscal_position:
                for tax in pos_line.tax_ids_after_fiscal_position:
                    ty_le_thue_gtgt = tax.amount or 0.0
                    break
            
            # Tiền thuế
            tien_thue_gtgt = abs(thanh_tien * ty_le_thue_gtgt / 100)
            
            # Price total (bao gồm thuế) - POS đã tính sẵn
            price_total = pos_line.price_subtotal_incl or 0.0
            
        elif sol:
            # Lấy từ Sale Order Line
            uom = sol.product_uom or prod.uom_id
            qty = sol.product_uom_qty or 0.0
            don_gia = sol.price_unit or 0.0
            ty_le_ck = sol.discount or 0.0
            
            # Tính tiền chiết khấu
            tien_chiet_khau = (don_gia * qty * ty_le_ck) / 100
            
            # Thành tiền = price_subtotal từ Sale Order Line (đã tính sẵn chiết khấu)
            thanh_tien = sol.price_subtotal or 0.0
            
            # Thuế GTGT
            ty_le_thue_gtgt = 0.0
            if sol.tax_id:
                for tax in sol.tax_id:
                    ty_le_thue_gtgt = tax.amount or 0.0
                    break
            
            # Tiền thuế = (Thành tiền sau chiết khấu) * % thuế
            tien_thue_gtgt = (thanh_tien * ty_le_thue_gtgt) / 100
            
            # Price total (bao gồm thuế)
            price_total = sol.price_total or (thanh_tien + tien_thue_gtgt)
            
        else:
            # Fallback: lấy từ picking/move line
            if forced_qty is not None:
                uom = ml.product_uom_id if ml else (move.product_uom if move else prod.uom_id)
                qty = forced_qty
            elif ml:
                uom = ml.product_uom_id or prod.uom_id
                qty = ml.qty_done or 0.0
            elif move:
                uom = move.product_uom or prod.uom_id
                qty = move.qty_done if hasattr(move, 'qty_done') else (move.product_uom_qty or 0.0)
            else:
                # Trường hợp không có move/ml (VD: service product không thông qua sale_line)
                uom = prod.uom_id
                qty = 1.0 # Default value
            
            # Fallback: lấy giá từ product
            don_gia = prod.list_price or 0.0
            thanh_tien = don_gia * qty
            tien_chiet_khau = 0.0
            ty_le_ck = 0.0
            ty_le_thue_gtgt = 0.0
            tien_thue_gtgt = 0.0
            price_total = thanh_tien
        
        # Location name (vẫn lấy từ move/move_line vì không có trong SOL)
        if ml:
            location_name = (ml.location_id and ml.location_id.complete_name) or ""
        elif move:
            location_name = (move.location_id and move.location_id.complete_name) or ""
        else:
            location_name = ""
        
        uom_name = (uom and uom.name) or ""
        
        # Đơn giá vốn và tiền vốn
        don_gia_von = prod.standard_price or 0.0
        tien_von = don_gia_von * qty

        return {
            # Hardcoded fields
            'hinh_thuc_ban_hang': 'Bán hàng hóa trong nước',
            'phuong_thuc_thanh_toan': 'Chưa thu tiền',
            # 3 cột mới từ sale.order (đặt ngay sau phương thức thanh toán)
            'hinh_thuc_giao_hang': getattr(so, 'x_studio_htgh', '') if so else '',
            'hinh_thuc_thanh_toan_so': getattr(so, 'x_studio_httt', '') if so else '',
            'ben_tra_phi_van_chuyen': getattr(so, 'x_studio_misa_delivery', '') if so else '',
            'kiem_phieu_xuat_kho': 'Có',
            'lap_kem_hoa_don': 'Có',
            'da_lap_hoa_don': 'Đã lập',
            
            # Date fields
            'ngay_hach_toan': _to_date_str(datetime.date.today()),
            'ngay_chung_tu': _to_date_str(datetime.date.today()),
            'so_chung_tu': picking_name,
            'so_phieu_xuat': sale_name,
            
            # Invoice fields - hardcoded examples
            'mau_so_hd': '01GTKT0/001',
            'ky_hieu_hd': '1C25TLV',
            'so_hoa_don': '',  # Để trống hoặc lấy từ invoice nếu có
            'ngay_hoa_don': scheduled_date_str,
            
            # Partner info
            'ma_khach_hang': partner_code,
            'ten_khach_hang': partner_name,
            'dia_chi': partner_address,
            'ma_so_thue': partner_vat,
            
            # Other info
            'don_vi_giao_dai_ly': '',
            'nguoi_nop': partner_name,
            'nop_vao_tk': '',
            'ten_ngan_hang': '',
            'dien_giai': dien_giai,
            'ly_do_xuat': ly_do_xuat,
            'ma_nhan_vien': sale_user_code,
            'so_ct_phieu_thu': '',
            'so_ct_phieu_xuat': sale_name,
            'han_thanh_toan': '',
            
            # Product info
            'ma_hang': product_code,
            'thuoc_combo': self._thuoc_combo_code_for_move(move),
            'ten_hang': product_name,
            'la_dong_ghi_chu': 'không',
            'hang_khuyen_mai': 'Không',
            'chiet_khau_thuong_mai': '',
            
            # Account codes - hardcoded examples
            'tk_tien_no': '131',
            'tk_doanh_thu_co': '5111',
            
            # Quantity and price
            'dvt': uom_name,
            'so_luong': qty,
            'don_gia': don_gia,
            'thanh_tien': thanh_tien,
           'bao_gom_thue': f"{price_total:,.2f}",
            'ty_le_ck': ty_le_ck,
            'tien_chiet_khau': tien_chiet_khau,
            'tk_chiet_khau': '',
            
            # Tax info
            'gia_tinh_thue_xk': '',
            'ty_le_thue_xk': '',
            'tien_thue_xk': '',
            'tk_thue_xk': '',
            'ty_le_thue_gtgt': ty_le_thue_gtgt,
            'ty_le_thue_khac': '',
            'tien_thue_gtgt': tien_thue_gtgt,
            'tk_thue_gtgt': '33311',
            
            # Other fields
            'bien_kiem_soat': '',
            'hh_khong_th_tren_to_khai': 'Không',
            'ma_khoan_muc_cp': '',
            'ma_don_vi': 'PKD',
            'ma_doi_tuong_thcp': '',
            'ma_cong_trinh': '',
            'so_don_dat_hang': sale_name,
            'so_hop_dong_ban': sale_name,
            'ma_thong_ke': '',
            'so_khe_uoc_cho_vay': '',
            'cp_khong_hop_ly': 'Không',
            
            # Warehouse and cost
            'ma_kho': warehouse_code,
            'tk_gia_von': '632',
            'tk_kho': '156',
            'don_gia_von': don_gia_von,
            'tien_von': tien_von,
            'hang_hoa_giu_ho': '',
            'vi_tri': '',
            'misa_sync': getattr(picking, 'x_studio_misa_sav', False),
        }

    def _create_excel_workbook(self, data_rows):
        """Tạo workbook Excel với header"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Xuất bán hàng hóa"

        columns = self._get_columns_definition()

        # Styles
        header_font = Font(name='Arial', size=10, bold=True)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border_side = Side(style='thin', color='000000')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        number_alignment = Alignment(horizontal='right', vertical='center')

        # Header row
        HEADER_ROW = 1
        DATA_START = 2

        for col_idx, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=HEADER_ROW, column=col_idx)
            cell.value = col_def['name']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def.get('width', 15)

        # Data rows
        for row_idx, row_data in enumerate(data_rows, start=DATA_START):
            for col_idx, col_def in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(col_def['key'], "")

                if value is None:
                    value = ""

                cell.value = value
                cell.border = border

                # Number formatting
                if isinstance(value, (int, float)) and value != "":
                    cell.alignment = number_alignment
                    if col_def['key'] in ['don_gia', 'thanh_tien', 'tien_chiet_khau', 
                                         'tien_thue_gtgt', 'don_gia_von', 'tien_von']:
                        cell.number_format = '#,##0'
                    elif col_def['key'] in ['ty_le_ck', 'ty_le_thue_gtgt', 'ty_le_thue_xk']:
                        cell.number_format = '0.00'
                    elif col_def['key'] == 'so_luong':
                        cell.number_format = '#,##0.00'
                else:
                    cell.alignment = cell_alignment

        ws.row_dimensions[HEADER_ROW].height = 30

        return wb

    def action_export(self):
        self.ensure_one()
        if Workbook is None:
            raise UserError(_("Thiếu thư viện openpyxl. Vui lòng cài đặt 'openpyxl' cho Python."))

        pickings = self.env["stock.picking"].sudo().search(self._domain(), order="scheduled_date asc, id asc")
        if not pickings:
            raise UserError(_("Không tìm thấy phiếu xuất kho nào trong khoảng ngày đã chọn."))

        # Tạo dữ liệu
        all_rows = []
        for picking in pickings:
            rows = self._get_move_line_rows(picking)
            all_rows.extend(rows)

        if not all_rows:
            raise UserError(_("Không có dữ liệu chi tiết để xuất."))

        # Tạo Excel workbook
        wb = self._create_excel_workbook(all_rows)

        # Xuất file
        out = BytesIO()
        wb.save(out)
        out.seek(0)

        filename = f"Xuat_ban_hang_hoa_{self.date_from}_{self.date_to}.xlsx"
        attachment = self.env["ir.attachment"].sudo().create({
            "name": filename,
            "type": "binary",
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "datas": base64.b64encode(out.getvalue()),
            "res_model": "picking.export.wizard",
            "res_id": self.id,
        })

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }

    def _get_json_data(self, picking):
        """
        Lấy dữ liệu JSON: mã đơn hàng và dict sản phẩm với số lượng
        Trả về dict: {'ma_don_hang': 'SO001', 'san_pham': {'PROD001': 10.0, 'PROD002': 5.0}}
        """
        so = self._find_sale_order(picking.move_ids_without_package[0] if picking.move_ids_without_package else None, picking)
        
        # Mã đơn hàng: ưu tiên sale order name, sau đó là picking origin
        sale_name = so.name if so else (picking.origin or picking.name or "")
        
        # Dict để lưu sản phẩm và số lượng
        san_pham_dict = {}
        
        # Xử lý từng move line hoặc move
        if picking.move_line_ids:
            for ml in picking.move_line_ids:
                prod = ml.product_id
                if not prod:
                    continue
                
                product_code = prod.default_code or (prod.barcode if hasattr(prod, 'barcode') else "") or ""
                qty = ml.qty_done or 0.0
                
                if product_code and qty > 0:
                    # Cộng dồn số lượng nếu sản phẩm đã có
                    if product_code in san_pham_dict:
                        san_pham_dict[product_code] += qty
                    else:
                        san_pham_dict[product_code] = qty
        else:
            for mv in picking.move_ids_without_package:
                prod = mv.product_id
                if not prod:
                    continue
                
                product_code = prod.default_code or (prod.barcode if hasattr(prod, 'barcode') else "") or ""
                qty = mv.qty_done if hasattr(mv, 'qty_done') else (mv.product_uom_qty or 0.0)
                
                if product_code and qty > 0:
                    # Cộng dồn số lượng nếu sản phẩm đã có
                    if product_code in san_pham_dict:
                        san_pham_dict[product_code] += qty
                    else:
                        san_pham_dict[product_code] = qty
        
        # Chỉ trả về nếu có sản phẩm
        if san_pham_dict:
            return {
                'ma_don_hang': sale_name,
                'san_pham': san_pham_dict
            }
        return None

    def action_export_json(self):
        """
        Xuất dữ liệu ra file JSON với format nhóm theo mã đơn hàng:
        - mã đơn hàng
        - san_pham: dict chứa tất cả sản phẩm và số lượng trong đơn hàng đó
        """
        self.ensure_one()
        
        pickings = self.env["stock.picking"].sudo().search(self._domain(), order="scheduled_date asc, id asc")
        if not pickings:
            raise UserError(_("Không tìm thấy phiếu xuất kho nào trong khoảng ngày đã chọn."))
        
        # Dict để nhóm theo mã đơn hàng
        orders_dict = {}
        
        # Thu thập dữ liệu từ tất cả pickings
        for picking in pickings:
            order_data = self._get_json_data(picking)
            if not order_data:
                continue
            
            ma_don_hang = order_data['ma_don_hang']
            san_pham = order_data['san_pham']
            
            # Nếu đơn hàng đã tồn tại, cộng dồn sản phẩm
            if ma_don_hang in orders_dict:
                for product_code, qty in san_pham.items():
                    if product_code in orders_dict[ma_don_hang]['san_pham']:
                        orders_dict[ma_don_hang]['san_pham'][product_code] += qty
                    else:
                        orders_dict[ma_don_hang]['san_pham'][product_code] = qty
            else:
                orders_dict[ma_don_hang] = {
                    'ma_don_hang': ma_don_hang,
                    'san_pham': san_pham.copy()
                }
        
        # Chuyển thành list
        all_rows = list(orders_dict.values())
        
        if not all_rows:
            raise UserError(_("Không có dữ liệu chi tiết để xuất."))
        
        # Chuyển đổi sang JSON
        json_data = json.dumps(all_rows, ensure_ascii=False, indent=2)
        json_bytes = json_data.encode('utf-8')
        
        # Tạo attachment
        filename = f"Xuat_ban_hang_hoa_{self.date_from}_{self.date_to}.json"
        attachment = self.env["ir.attachment"].sudo().create({
            "name": filename,
            "type": "binary",
            "mimetype": "application/json",
            "datas": base64.b64encode(json_bytes),
            "res_model": "picking.export.wizard",
            "res_id": self.id,
        })
        
        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }

    def _get_pos_columns_definition(self):
        """Định nghĩa cột cho mẫu POS (52 columns match template)"""
        return [
            {'key': 'hinh_thuc_ban_hang', 'name': 'Hình thức bán hàng', 'width': 25},
            {'key': 'phuong_thuc_thanh_toan', 'name': 'Phương thức thanh toán', 'width': 25},
            {'key': 'kiem_phieu_xuat_kho', 'name': 'Kiêm phiếu xuất kho', 'width': 20},
            {'key': 'lap_kem_hoa_don', 'name': 'Lập kèm hóa đơn', 'width': 18},
            {'key': 'da_lap_hoa_don', 'name': 'Đã lập hóa đơn', 'width': 18},
            {'key': 'ngay_hach_toan', 'name': 'Ngày hạch toán (*)', 'width': 25},
            {'key': 'ngay_chung_tu', 'name': 'Ngày chứng từ (*)', 'width': 25},
            {'key': 'so_chung_tu', 'name': 'Số chứng từ (*)', 'width': 20},
            {'key': 'so_phieu_xuat', 'name': 'Số phiếu xuất', 'width': 20},
            {'key': 'mau_so_hd', 'name': 'Mẫu số HĐ', 'width': 15},
            {'key': 'ky_hieu_hd', 'name': 'Ký hiệu HĐ', 'width': 15},
            {'key': 'so_hoa_don', 'name': 'Số hóa đơn', 'width': 15},
            {'key': 'ngay_hoa_don', 'name': 'Ngày hóa đơn', 'width': 20},
            {'key': 'ma_khach_hang', 'name': 'Mã khách hàng', 'width': 15},
            {'key': 'ten_khach_hang', 'name': 'Tên khách hàng', 'width': 40},
            {'key': 'dia_chi', 'name': 'Địa chỉ', 'width': 50},
            {'key': 'ma_so_thue', 'name': 'Mã số thuế', 'width': 15},
            {'key': 'don_vi_giao_dai_ly', 'name': 'Đơn vị giao đại lý', 'width': 30},
            {'key': 'nguoi_nop', 'name': 'Người nộp', 'width': 25},
            {'key': 'nop_vao_tk', 'name': 'Nộp vào TK', 'width': 15},
            {'key': 'ten_ngan_hang', 'name': 'Tên ngân hàng', 'width': 30},
            {'key': 'dien_giai', 'name': 'Diễn giải/Lý do nộp', 'width': 40},
            {'key': 'ly_do_xuat', 'name': 'Lý do xuất', 'width': 40},
            {'key': 'nhan_vien_ban_hang', 'name': 'Mã nhân viên bán hàng', 'width': 20},
            {'key': 'kem_theo', 'name': 'Kèm theo', 'width': 20},
            {'key': 'han_thanh_toan', 'name': 'Hạn thanh toán', 'width': 20},
            {'key': 'ma_hang', 'name': 'Mã hàng (*)', 'width': 18},
            {'key': 'thuoc_combo', 'name': 'Thuộc combo', 'width': 15},
            {'key': 'ten_hang', 'name': 'Tên hàng', 'width': 40},
            {'key': 'la_dong_ghi_chu', 'name': 'Là dòng ghi chú', 'width': 18},
            {'key': 'hang_khuyen_mai', 'name': 'Hàng khuyến mại', 'width': 18},
            {'key': 'chiet_khau_thuong_mai', 'name': 'Chiết khấu thương mại', 'width': 25},
            {'key': 'tk_tien_no', 'name': 'TK Tiền/Chi phí/Nợ (*)', 'width': 20},
            {'key': 'tk_doanh_thu_co', 'name': 'TK Doanh thu/Có (*)', 'width': 20},
            {'key': 'dvt', 'name': 'ĐVT', 'width': 12},
            {'key': 'so_luong', 'name': 'Số lượng', 'width': 12},
            {'key': 'don_gia', 'name': 'Đơn giá', 'width': 15},
            {'key': 'thanh_tien', 'name': 'Thành tiền', 'width': 15},
            {'key': 'ty_le_ck', 'name': 'Tỷ lệ CK (%)', 'width': 12},
            {'key': 'tien_chiet_khau', 'name': 'Tiền chiết khấu', 'width': 15},
            {'key': 'tk_chiet_khau', 'name': 'TK chiết khấu', 'width': 15},
            {'key': 'gia_tinh_thue_xk', 'name': 'Giá tính thuế XK', 'width': 15},
            {'key': 'ty_le_thue_xk', 'name': '% thuế xuất khẩu', 'width': 15},
            {'key': 'tien_thue_xk', 'name': 'Tiền thuế xuất khẩu', 'width': 15},
            {'key': 'tk_thue_xk', 'name': 'TK thuế xuất khẩu', 'width': 15},
            {'key': 'ty_le_thue_gtgt', 'name': '% thuế GTGT', 'width': 12},
            {'key': 'ty_le_thue_khac', 'name': '% thuế suất KHAC', 'width': 15},
            {'key': 'tien_thue_gtgt', 'name': 'Tiền thuế GTGT', 'width': 15},
            {'key': 'tk_thue_gtgt', 'name': 'TK thuế GTGT', 'width': 15},
            {'key': 'hh_khong_th_tren_to_khai', 'name': 'HH không TH trên tờ khai thuế GTGT', 'width': 25},
            {'key': 'ma_kho', 'name': 'Mã kho', 'width': 15},
            {'key': 'tk_gia_von', 'name': 'TK giá vốn', 'width': 15},
            {'key': 'tk_kho', 'name': 'TK Kho', 'width': 15},
            {'key': 'don_gia_von', 'name': 'Đơn giá vốn', 'width': 15},
            {'key': 'tien_von', 'name': 'Tiền vốn', 'width': 15},
            {'key': 'hang_hoa_giu_ho', 'name': 'Hàng hóa giữ hộ/bán hộ', 'width': 20},
            {'key': 'la_hoa_don_tu_may_tinh_tien', 'name': 'Là hóa đơn từ máy tính tiền', 'width': 25},
        ]


    #  build row của misa
    def _get_pos_row_data(self, picking):
        """Xây dựng rows cho mẫu POS"""
        rows = []
        so = self._find_sale_order(picking.move_ids_without_package[0] if picking.move_ids_without_package else None, picking)
        
        # --- Common Info ---
        date_done = picking.date_done or picking.scheduled_date or fields.Datetime.now()
        date_str = _to_date_str(date_done)
        date_hach_toan_str = _to_date_str(date_done, hour=18)
        
        # --- Xử lý hậu tố 2 ca ---
        import pytz
        user_tz = pytz.timezone(self.env.user.tz or 'Asia/Ho_Chi_Minh')
        dt = picking.date_done or picking.scheduled_date
        shift_suffix = ""
        if dt:
            dt_utc = pytz.utc.localize(dt) if not getattr(dt, 'tzinfo', None) else dt
            dt_vn = dt_utc.astimezone(user_tz)
            if dt_vn.hour < 16 or (dt_vn.hour == 16 and dt_vn.minute < 30):
                shift_suffix = "-1"
            else:
                shift_suffix = "-2"

        # Số chứng từ: x_studio_pos_group (VD: POS/050126)
        # Nếu chưa có thì fallback về picking name
        base_so_chung_tu = picking.x_studio_pos_group or picking.name
        so_chung_tu = f"{base_so_chung_tu}{shift_suffix}" if base_so_chung_tu else ""
        
        # Số phiếu xuất: thêm 'PXK' trước số chứng từ
        # VD: PXKPOS/050126
        # Lưu ý: user nói "có thêm PXK phía trước"
        so_phieu_xuat = ""
        if picking.x_studio_pos_group:
            # Xử lý format nếu cần, ở đây ghép chuỗi đơn giản
            if "POS/" in picking.x_studio_pos_group:
                 so_phieu_xuat = "PXK" + picking.x_studio_pos_group.replace("/", "") + shift_suffix
            else:
                 so_phieu_xuat = "PXK" + picking.x_studio_pos_group + shift_suffix
        else:
            so_phieu_xuat = "PXK" + picking.name + shift_suffix

        partner = picking.partner_id
        partner_code = self._partner_code(partner)
        partner_name = (partner and partner.name) or ""
        partner_address = ""
        # Get address parts same as before...
        # ... (Simplified for brevity, assuming existing logic from _get_move_line_rows works or copying it)
        # Copy logic for address:
        import unicodedata
        def normalize_addr(s):
            s = s.strip().lower()
            s = unicodedata.normalize('NFD', s)
            s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
            return s
        if partner:
            parts = [partner.street, partner.city, partner.state_id.name]
            valid_parts = [p for p in parts if p]
            partner_address = ", ".join(valid_parts)
            
        partner_vat = (partner and partner.vat) or ""

        # Sale info
        sale_name = so.name if so else (picking.origin or "")
        sale_user_code = ''
        if so and so.user_id:
            sale_user_code = so.user_id.login or so.user_id.name or ''
        
        dien_giai = picking.note or f"Xuất bán hàng {partner_name}"
        warehouse_code = self._get_warehouse_code(picking)

        # KIỂM TRA: Nếu là đơn POS, loop qua pos.order.lines thay vì stock moves
        pos_order = getattr(picking, 'pos_order_id', False)
        
        if pos_order:
            # Loop qua từng POS order line (để đảm bảo xuất đủ số dòng)
            for pos_line in pos_order.lines:
                prod = pos_line.product_id
                if not prod:
                    continue
                
                # Tìm stock move tương ứng (nếu cần)
                move = None
                if picking.move_line_ids:
                    for move_line in picking.move_line_ids:
                        if move_line.product_id == prod:
                            move = move_line.move_id
                            break
                else:
                    for mv in picking.move_ids_without_package:
                        if mv.product_id == prod:
                            move = mv
                            break
                
                # Lấy dữ liệu trực tiếp từ POS line
                qty = pos_line.qty or 0.0
                uom = prod.uom_id
                
                # Thành tiền chưa thuế
                price_subtotal = pos_line.price_subtotal or 0.0
                
                # Tính lại đơn giá chưa thuế để tương thích với chiết khấu
                discount = pos_line.discount or 0.0
                if qty != 0 and discount != 100.0:
                     price_unit = price_subtotal / (qty * (1 - discount / 100.0))
                else:
                     price_unit = pos_line.price_unit# Fallback if qty is 0 or discount is 100%
                
                tax_amount = 0.0
                if pos_line.tax_ids_after_fiscal_position:
                    tax_amount = pos_line.tax_ids_after_fiscal_position[0].amount
                
                # Computed fields
                tien_ck = abs(price_unit * qty * discount / 100)
                thanh_tien = price_subtotal
                tien_thue = abs(thanh_tien * tax_amount / 100) if tax_amount else 0
                
                # Mapping logic based on warehouse and payment method
                ma_khach_hang = partner_code
                phuong_thuc_excel = 'Chưa thu tiền'
                
                # Get payment method from picking field (try both possible field names)
                raw_payment_method = getattr(picking, 'x_studio_pos_payment_method', '') or getattr(picking, 'x_studio_payment_method', '') or ''
                is_multiple = (',' in str(raw_payment_method)) or ("kết hợp" in str(raw_payment_method).lower())
                payment_method_lower = str(raw_payment_method).lower()

                # Mapping for KBC (BENCAM)
                if warehouse_code in ["KBC", "BENCAM"]:
                    if "tiền mặt" in payment_method_lower and not is_multiple:
                        ma_khach_hang = "KH27182013179"
                        phuong_thuc_excel = "Thu tiền ngay - Tiền mặt"
                        partner_name = "KHÁCH CH BẾN CAM TT TIỀN MẶT"
                    elif "chuyển khoản" in payment_method_lower and not is_multiple:
                        ma_khach_hang = "KH27182013178"
                        phuong_thuc_excel = "Thu tiền ngay - Chuyển khoản"
                        partner_name = "KHÁCH CH BẾN CAM TT CHUYỂN KHOẢN"
                    elif is_multiple:
                        ma_khach_hang = "KHACHLE-BC"
                        phuong_thuc_excel = "Chưa thu tiền"
                        partner_name = "KHÁCH LẺ CỬA HÀNG BẾN CAM"
                
                # Mapping for TSN (HCM)
                elif warehouse_code in ["TSN", "HCM","TSNSR"]:
                    if "tiền mặt" in payment_method_lower and not is_multiple:
                        ma_khach_hang = "KH27182013176"
                        phuong_thuc_excel = "Thu tiền ngay - Tiền mặt"
                        partner_name = "KHÁCH GHÉ VP HCM TT TIỀN MẶT"
                    elif "chuyển khoản" in payment_method_lower and not is_multiple:
                        ma_khach_hang = "KH27182013177"
                        phuong_thuc_excel = "Thu tiền ngay - Chuyển khoản"
                        partner_name = "KHÁCH GHÉ VP HCM TT CHUYỂN KHOẢN"
                    elif is_multiple:
                        ma_khach_hang = "KHACHLE-HCM"
                        phuong_thuc_excel = "Chưa thu tiền"
                        partner_name = "KHÁCH LẺ VP HCM"

                # Build row dict
                row = {
                    'hinh_thuc_ban_hang': 'Bán hàng hóa trong nước',
                    'phuong_thuc_thanh_toan': phuong_thuc_excel,
                    'kiem_phieu_xuat_kho': 'Có',
                    'lap_kem_hoa_don': 'Có',
                    'da_lap_hoa_don': 'Đã lập',
                    'ngay_hach_toan': date_hach_toan_str,
                    'ngay_chung_tu': date_str,
                    'so_chung_tu': so_chung_tu,
                    'so_phieu_xuat': so_phieu_xuat,
                    'mau_so_hd': '1',
                    'ky_hieu_hd': '1C26TLV',
                    'so_hoa_don': '',
                    'ngay_hoa_don': date_str,
                    'ma_khach_hang': ma_khach_hang,
                    'ten_khach_hang': partner_name,
                    'dia_chi': partner_address,
                    'ma_so_thue': '',
                    'don_vi_giao_dai_ly': '',
                    'nguoi_nop': '',
                    'nop_vao_tk': '',
                    'ten_ngan_hang': '',
                    'dien_giai': dien_giai,
                    'ly_do_xuat': dien_giai,
                    'nhan_vien_ban_hang': sale_user_code,
                    'kem_theo': '',
                    'han_thanh_toan': '',
                    
                    'ma_hang': prod.default_code or '',
                    'thuoc_combo': self._thuoc_combo_code_for_move(move) if move else '',
                    'ten_hang': prod.name,
                    'la_dong_ghi_chu': 'không',
                    'hang_khuyen_mai': 'Không',
                    'chiet_khau_thuong_mai': '',
                    
                    'tk_tien_no': '131',
                    'tk_doanh_thu_co': '5111',
                    'dvt': uom.name if uom else '',
                    'so_luong': qty,
                    'don_gia': price_unit,
                    'thanh_tien': thanh_tien,
                    'ty_le_ck': discount,
                    'tien_chiet_khau': tien_ck,
                    'tk_chiet_khau': '',
                    
                    'gia_tinh_thue_xk': '',
                    'ty_le_thue_xk': '',
                    'tien_thue_xk': '',
                    'tk_thue_xk': '',
                    
                    'ty_le_thue_gtgt': tax_amount,
                    'ty_le_thue_khac': '',
                    'tien_thue_gtgt': tien_thue,
                    'tk_thue_gtgt': '33311',
                    'hh_khong_th_tren_to_khai': 'Không',
                    
                    'ma_kho': warehouse_code,
                    'tk_gia_von': '632',
                    'tk_kho': '1561',
                    'don_gia_von': prod.standard_price,
                    'tien_von': prod.standard_price * abs(qty),  # Use abs for cost calculation
                    'hang_hoa_giu_ho': '',
                    'la_hoa_don_tu_may_tinh_tien': 'Có',
                }
                rows.append(row)
                
        else:
            # Loop moves (non-POS logic)
            moves = picking.move_line_ids if picking.move_line_ids else picking.move_ids_without_package
            
            for line in moves:
                # Determine move first
                if line._name == 'stock.move.line':
                    prod = line.product_id
                    move = line.move_id
                else:
                    prod = line.product_id
                    move = line # Itself
                
                if not prod: continue

                # Ưu tiên: Sale Order Line
                sol = move.sale_line_id if move and hasattr(move, 'sale_line_id') else False
                
                # Logic lấy dữ liệu
                if sol:
                    # Lấy từ Sale Order Line
                    qty = sol.product_uom_qty or 0.0
                    uom = sol.product_uom or prod.uom_id
                    price_unit = sol.price_unit or 0.0
                    discount = sol.discount or 0.0
                    price_subtotal = sol.price_subtotal or 0.0
                    
                    tax_amount = 0.0
                    if sol.tax_id:
                        tax_amount = sol.tax_id[0].amount
                        
                else:
                    # Fallback: lấy từ picking/move line
                    if line._name == 'stock.move.line':
                        qty = line.qty_done
                        uom = line.product_uom_id
                    else:
                        qty = line.quantity_done if hasattr(line, 'quantity_done') else line.product_uom_qty
                        uom = line.product_uom
                    
                    # Fallback to product list price if no SOL
                    price_unit = prod.list_price
                    discount = 0.0
                    price_subtotal = 0.0
                    tax_amount = 0.0

                # Computed fields
                tien_ck = abs(price_unit * qty * discount / 100)
                # Thành tiền: Ưu tiên price_subtotal từ SOL, nếu không có thì tính
                if sol and price_subtotal:
                    thanh_tien = price_subtotal
                else:
                    thanh_tien = (price_unit * qty) - tien_ck
            
            # Tiền thuế = (Thành tiền sau chiết khấu) * % thuế
            tien_thue = (thanh_tien * tax_amount / 100) if tax_amount else 0
            
            # Mapping logic based on warehouse and payment method
            ma_khach_hang = partner_code
            phuong_thuc_excel = 'Chưa thu tiền'
            
            # Get payment method from picking field (try both possible field names)
            raw_payment_method = getattr(picking, 'x_studio_pos_payment_method', '') or getattr(picking, 'x_studio_payment_method', '') or ''
            is_multiple = (',' in str(raw_payment_method)) or ("kết hợp" in str(raw_payment_method).lower())
            payment_method_lower = str(raw_payment_method).lower()

            # Mapping for KBC (BENCAM)
            if warehouse_code in ["KBC", "BENCAM"]:
                if "tiền mặt" in payment_method_lower and not is_multiple:
                    ma_khach_hang = "KH27182013179"
                    phuong_thuc_excel = "Thu tiền ngay - Tiền mặt"
                    partner_name = "KHÁCH CH BẾN CAM TT TIỀN MẶT"
                elif "chuyển khoản" in payment_method_lower and not is_multiple:
                    ma_khach_hang = "KH27182013178"
                    phuong_thuc_excel = "Thu tiền ngay - Chuyển khoản"
                    partner_name = "KHÁCH CH BẾN CAM TT CHUYỂN KHOẢN"
                elif is_multiple:
                    ma_khach_hang = "KHACHLE-BC"
                    phuong_thuc_excel = "Chưa thu tiền"
                    partner_name = "KHÁCH LẺ CỬA HÀNG BẾN CAM"
            
            # Mapping for TSN (HCM)
            elif warehouse_code in ["TSN", "HCM"]:
                if "tiền mặt" in payment_method_lower and not is_multiple:
                    ma_khach_hang = "KH27182013176"
                    phuong_thuc_excel = "Thu tiền ngay - Tiền mặt"
                    partner_name = "KHÁCH GHÉ VP HCM TT TIỀN MẶT"
                elif "chuyển khoản" in payment_method_lower and not is_multiple:
                    ma_khach_hang = "KH27182013177"
                    phuong_thuc_excel = "Thu tiền ngay - Chuyển khoản"
                    partner_name = "KHÁCH GHÉ VP HCM TT CHUYỂN KHOẢN"
                elif is_multiple:
                    ma_khach_hang = "KHACHLE-HCM"
                    phuong_thuc_excel = "Chưa thu tiền"
                    partner_name = "KHÁCH LẺ VP HCM"

            # Start building dict based on NEW columns
            row = {
                'hinh_thuc_ban_hang': 'Bán hàng hóa trong nước',
                'phuong_thuc_thanh_toan': phuong_thuc_excel,
                'kiem_phieu_xuat_kho': 'Có',
                'lap_kem_hoa_don': 'Không',
                'da_lap_hoa_don': 'Chưa lập',
                'ngay_hach_toan': date_hach_toan_str,
                'ngay_chung_tu': date_str,
                'so_chung_tu': so_chung_tu,
                'so_phieu_xuat': so_phieu_xuat,
                'mau_so_hd': '',
                'ky_hieu_hd': '',
                'so_hoa_don': '', # Customize if needed
                'ngay_hoa_don': '',
                'ma_khach_hang': ma_khach_hang,
                'ten_khach_hang': partner_name,
                'dia_chi': partner_address,
                'ma_so_thue': '',
                'don_vi_giao_dai_ly': '',
                'nguoi_nop': '',
                'nop_vao_tk': '',
                'ten_ngan_hang': '',
                'dien_giai': dien_giai,
                'ly_do_xuat': dien_giai,
                'nhan_vien_ban_hang': sale_user_code,
                'kem_theo': '',
                'han_thanh_toan': '',
                
                'ma_hang': prod.default_code or '',
                'thuoc_combo': self._thuoc_combo_code_for_move(move) if move else '',
                'ten_hang': prod.name,
                'la_dong_ghi_chu': 'không',
                'hang_khuyen_mai': 'Không',
                'chiet_khau_thuong_mai': '',
                
                'tk_tien_no': '131',
                'tk_doanh_thu_co': '5111',
                'dvt': uom.name if uom else '',
                'so_luong': qty,
                'don_gia': price_unit,
                'thanh_tien': thanh_tien,
                'ty_le_ck': '',
                'tien_chiet_khau': tien_ck,
                'tk_chiet_khau': '',
                
                'gia_tinh_thue_xk': '',
                'ty_le_thue_xk': '',
                'tien_thue_xk': '',
                'tk_thue_xk': '',
                
                'ty_le_thue_gtgt': tax_amount,
                'ty_le_thue_khac': '',
                'tien_thue_gtgt': tien_thue,
                'tk_thue_gtgt': '33311',
                'hh_khong_th_tren_to_khai': 'Không',
                
                'ma_kho': 'HLV',
                'tk_gia_von': '632',
                'tk_kho': '1561', # From template sample
                'don_gia_von': prod.standard_price,
                'tien_von': prod.standard_price * qty,
                'hang_hoa_giu_ho': '',
                'la_hoa_don_tu_may_tinh_tien': 'Có',
            }
            rows.append(row)
        return rows

    def _create_pos_excel_workbook(self, pickings):
        """Tạo workbook Excel mẫu POS với header và hướng dẫn"""
        wb = Workbook()
        ws = wb.active
        ws.title = "DS Hóa Đơn - POS"

        columns = self._get_pos_columns_definition()

        # Styles
        title_font = Font(name='Arial', size=16, bold=True)
        instruction_font = Font(name='Arial', size=10, italic=True, color='FF0000')
        header_font = Font(name='Arial', size=10, bold=True)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border_side = Side(style='thin', color='000000')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        number_alignment = Alignment(horizontal='right', vertical='center')

        # --- HEADER ROW (Row 1) ---
        HEADER_ROW = 1
        DATA_START = 2

        for col_idx, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=HEADER_ROW, column=col_idx)
            cell.value = col_def['name']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def.get('width', 15)

        # --- DATA ROWS ---
        current_row = DATA_START
        for picking in pickings:
            row_data_list = self._get_pos_row_data(picking)
            for row_data in row_data_list:
                for col_idx, col_def in enumerate(columns, start=1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    value = row_data.get(col_def['key'], "")
                    
                    if value is None: value = ""
                    
                    cell.value = value
                    cell.border = border
                    
                    # Number fmt
                    if isinstance(value, (int, float)) and value != "":
                        cell.alignment = number_alignment
                        if 'ty_le' in col_def['key'] or 'so_luong' in col_def['key']:
                             cell.number_format = '#,##0.00'
                        elif 'tien' in col_def['key'] or 'gia' in col_def['key']:
                             cell.number_format = '#,##0'
                    else:
                        cell.alignment = cell_alignment
                
                current_row += 1

        ws.row_dimensions[HEADER_ROW].height = 30
        return wb

    def action_export_pos_template(self):
        self.ensure_one()
        if Workbook is None:
            raise UserError(_("Thiếu thư viện openpyxl. Vui lòng cài đặt 'openpyxl' cho Python."))

        domain = self._domain()
        # Filter: Either has Group OR has POS Session (ungrouped POS orders)
        domain.append('|')
        domain.append(('x_studio_pos_group', '!=', False))
        domain.append(('pos_session_id', '!=', False))
        pickings = self.env["stock.picking"].sudo().search(domain, order="scheduled_date asc, id asc")
        if not pickings:
            raise UserError(_("Không tìm thấy phiếu xuất kho nào trong khoảng ngày đã chọn."))

        import pytz
        import zipfile
        
        user_tz = pytz.timezone(self.env.user.tz or 'Asia/Ho_Chi_Minh')
        
        groups = {}
        for p in pickings:
            dt = p.date_done or p.scheduled_date
            if not dt:
                key = "KhongCoNgay_Ca1"
            else:
                dt_utc = pytz.utc.localize(dt) if not getattr(dt, 'tzinfo', None) else dt
                dt_vn = dt_utc.astimezone(user_tz)
                d_str = dt_vn.strftime('%d-%m-%Y')
                
                # Check 0h-16h30 vs 16h30-24h
                if dt_vn.hour < 16 or (dt_vn.hour == 16 and dt_vn.minute < 30):
                    shift = 'Ca1_0h-16h30'
                else:
                    shift = 'Ca2_16h30-24h'
                
                key = f"{d_str}_{shift}"
            
            if key not in groups:
                groups[key] = self.env["stock.picking"].sudo().browse()
            groups[key] |= p

        # If only 1 group, just export 1 excel
        if len(groups) == 1:
            key, group_pickings = list(groups.items())[0]
            wb = self._create_pos_excel_workbook(group_pickings)
            out = BytesIO()
            wb.save(out)
            out.seek(0)
            filename = f"Xuat_POS_{key}.xlsx"
            mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            datas = base64.b64encode(out.getvalue())
        else:
            # Create a ZIP file containing multiple excel files
            out_zip = BytesIO()
            with zipfile.ZipFile(out_zip, 'w', zipfile.ZIP_DEFLATED) as zf:
                for key, group_pickings in groups.items():
                    wb = self._create_pos_excel_workbook(group_pickings)
                    out = BytesIO()
                    wb.save(out)
                    out.seek(0)
                    zf.writestr(f"Xuat_POS_{key}.xlsx", out.getvalue())
            
            out_zip.seek(0)
            filename = f"Xuat_POS_{self.date_from}_den_{self.date_to}.zip"
            mimetype = "application/zip"
            datas = base64.b64encode(out_zip.getvalue())

        attachment = self.env["ir.attachment"].sudo().create({
            "name": filename,
            "type": "binary",
            "mimetype": mimetype,
            "datas": datas,
            "res_model": "picking.export.wizard",
            "res_id": self.id,
        })

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }

    # ====== POS CRM EXPORT METHODS ======
    
    def _get_pos_crm_columns_sheet1(self):
        """Định nghĩa 49 cột cho Sheet 1: Nhập khẩu Đơn hàng"""
        return [
            {'key': 'su_dung_ngoai_te', 'name': 'Sử dụng ngoại tệ', 'width': 18},
            {'key': 'loai_tien', 'name': 'Loại tiền', 'width': 12},
            {'key': 'ty_gia', 'name': 'Tỷ giá', 'width': 12},
            {'key': 'so_don_hang', 'name': 'Số đơn hàng (*)', 'width': 20},
            {'key': 'ngay_dat_hang', 'name': 'Ngày đặt hàng (*)', 'width': 20},
            {'key': 'khach_hang', 'name': 'Khách hàng', 'width': 30},
            {'key': 'lien_he', 'name': 'Liên hệ', 'width': 25},
            {'key': 'don_hang_cha', 'name': 'Đơn hàng cha', 'width': 20},
            {'key': 'co_hoi', 'name': 'Cơ hội', 'width': 20},
            {'key': 'gia_tri_don_hang', 'name': 'Giá trị đơn hàng (*)', 'width': 18},
            {'key': 'bao_gia', 'name': 'Báo giá', 'width': 20},
            {'key': 'khach_tt_truoc', 'name': 'Khách TT trước', 'width': 18},
            {'key': 'loai_don_hang', 'name': 'Loại đơn hàng', 'width': 20},
            {'key': 'so_ngay_duoc_no', 'name': 'Số ngày được nợ', 'width': 18},
            {'key': 'han_giao_hang', 'name': 'Hạn giao hàng (*)', 'width': 20},
            {'key': 'han_thanh_toan', 'name': 'Hạn thanh toán (*)', 'width': 20},
            {'key': 'dien_giai', 'name': 'Diễn giải', 'width': 40},
            {'key': 'tinh_trang_kh', 'name': 'Tình trạng KH (*)', 'width': 20},
            {'key': 'tinh_trang', 'name': 'Tình trạng (*)', 'width': 20},
            {'key': 'ngay_ghi_so', 'name': 'Ngày ghi sổ', 'width': 20},
            {'key': 'thuc_thu', 'name': 'Thực thu', 'width': 15},
            {'key': 'tinh_trang_giao_hang', 'name': 'Tình trạng giao hàng', 'width': 22},
            {'key': 'du_kien_chi', 'name': 'Dự kiến chi', 'width': 15},
            {'key': 'tinh_trang_thanh_toan', 'name': 'Tình trạng thanh toán', 'width': 22},
            {'key': 'han_san_xuat', 'name': 'Hạn sản xuất', 'width': 18},
            {'key': 'da_xuat_hoa_don', 'name': 'Đã xuất hóa đơn', 'width': 18},
            {'key': 'khach_hang_hoa_don', 'name': 'Khách hàng (Hóa đơn)', 'width': 30},
            {'key': 'nguoi_mua_hang', 'name': 'Người mua hàng', 'width': 30},
            {'key': 'quoc_gia_hoa_don', 'name': 'Quốc gia (Hóa đơn)', 'width': 20},
            {'key': 'tinh_thanh_pho_hoa_don', 'name': 'Tỉnh/Thành phố (Hóa đơn)', 'width': 25},
            {'key': 'quan_huyen_hoa_don', 'name': 'Quận/Huyện (Hóa đơn)', 'width': 25},
            {'key': 'phuong_xa_hoa_don', 'name': 'Phường/Xã (Hóa đơn)', 'width': 25},
            {'key': 'so_nha_duong_pho_hoa_don', 'name': 'Số nhà, Đường phố (Hóa đơn)', 'width': 35},
            {'key': 'ma_vung_hoa_don', 'name': 'Mã vùng (Hóa đơn)', 'width': 18},
            {'key': 'dia_chi_hoa_don', 'name': 'Địa chỉ (Hóa đơn)', 'width': 50},
            {'key': 'nguoi_nhan_hang', 'name': 'Người nhận hàng', 'width': 30},
            {'key': 'dien_thoai', 'name': 'Điện thoại', 'width': 18},
            {'key': 'dia_chi_giao_hang', 'name': 'Địa chỉ (Giao hàng)', 'width': 50},
            {'key': 'nhan_vien_kho', 'name': 'Nhân viên kho', 'width': 25},
            {'key': 'hinh_thuc_giao_hang', 'name': 'Hình thức giao hàng', 'width': 22},
            {'key': 'ngay_giao_du_kien', 'name': 'Ngày giao dự kiến', 'width': 20},
            {'key': 'ben_tra_phi_van_chuyen', 'name': 'Bên trả phí vận chuyển', 'width': 25},
            {'key': 'hinh_thuc_thanh_toan', 'name': 'Hình thức thanh toán', 'width': 22},
            {'key': 'mo_ta', 'name': 'Mô tả', 'width': 40},
            {'key': 'nguoi_thuc_hien', 'name': 'Người thực hiện', 'width': 25},
            {'key': 'dung_chung', 'name': 'Dùng chung', 'width': 15},
            {'key': 'ngung_theo_doi', 'name': 'Ngừng theo dõi', 'width': 18},
            {'key': 'doi_tac_ctv_gioi_thieu', 'name': 'Đối tác/CTV giới thiệu', 'width': 30},
            {'key': 'dong_bo_don_gia_sau_ck', 'name': 'Đồng bộ đơn giá sau CK', 'width': 25},
        ]

    def _get_pos_crm_columns_sheet2(self):
        """Định nghĩa 18 cột cho Sheet 2: nhập khẩu hàng hóa"""
        return [
            {'key': 'ma_hang_hoa', 'name': 'Mã hàng hóa', 'width': 18},
            {'key': 'dien_giai', 'name': 'Diễn giải', 'width': 40},
            {'key': 'mo_ta', 'name': 'Mô tả', 'width': 40},
            {'key': 'kho', 'name': 'Kho (*)', 'width': 15},
            {'key': 'kho_odoo', 'name': 'Kho Odoo (*)', 'width': 20},
            {'key': 'tinh_trang_hang', 'name': 'Tình trạng hàng (*)', 'width': 20},
            {'key': 'don_vi_tinh', 'name': 'Đơn vị tính', 'width': 12},
            {'key': 'so_luong', 'name': 'Số lượng', 'width': 12},
            {'key': 'don_gia_sau_thue', 'name': 'Đơn giá sau thuế', 'width': 15},
            {'key': 'don_gia', 'name': 'Đơn giá', 'width': 15},
            {'key': 'thanh_tien', 'name': 'Thành tiền', 'width': 15},
            {'key': 'ty_le_chiet_khau', 'name': 'Tỷ lệ chiết khấu', 'width': 18},
            {'key': 'tien_chiet_khau', 'name': 'Tiền chiết khấu', 'width': 15},
            {'key': 'thue_suat', 'name': 'Thuế suất', 'width': 12},
            {'key': 'tien_thue', 'name': 'Tiền thuế', 'width': 15},
            {'key': 'tong_tien', 'name': 'Tổng tiền', 'width': 15},
            {'key': 'don_gia_mua_bat_buoc', 'name': 'Đơn giá mua bắt buộc', 'width': 22},
            {'key': 'don_hang', 'name': 'Đơn hàng (*)', 'width': 20},
        ]

    def _get_pos_crm_order_data(self, picking):
        """
        Mapping dữ liệu từ picking/POS order sang row Sheet 1 (49 cột)
        Trả về dict với key tương ứng columns sheet 1
        
        Hỗ trợ 3-step delivery: nếu phiếu OUT không có pos_order_id / x_studio_pos_group,
        truy ngược về phiếu PICK gốc để lấy dữ liệu POS.
        """
        pos_order = getattr(picking, 'pos_order_id', False)
        
        # Fallback: truy ngược về phiếu PICK gốc nếu OUT không có POS data
        source_picking = picking
        if not pos_order and not picking.x_studio_pos_group:
            source_pos_picking = self._find_source_pos_picking(picking)
            if source_pos_picking:
                source_picking = source_pos_picking
                pos_order = getattr(source_picking, 'pos_order_id', False)

        partner = picking.partner_id
        
        # Dates
        date_done = picking.date_done or picking.scheduled_date or fields.Datetime.now()
        date_str = _to_date_str(date_done)
        scheduled_date_str = _to_date_str(picking.scheduled_date) if picking.scheduled_date else date_str
        
        # Order reference - use x_studio_pos_group (from source picking if needed)
        so_don_hang = (picking.x_studio_pos_group 
                       or source_picking.x_studio_pos_group 
                       or (pos_order and pos_order.name) 
                       or picking.name or '')
        
        # Partner info
        partner_name = (partner and partner.name) or ''
        partner_phone = ''
        if partner:
            partner_phone = partner.phone or partner.mobile or ''
        
        # Address
        import unicodedata
        def normalize_addr(s):
            s = s.strip().lower()
            s = unicodedata.normalize('NFD', s)
            s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
            return s
        
        dia_chi_giao_hang = ''
        if partner:
            parts = []
            for p in [partner.street, partner.city, partner.state_id.name if partner.state_id else '']:
                if p:
                    parts.append(p)
            dia_chi_giao_hang = ', '.join(parts)
        
        # Order amount
        gia_tri_don_hang = 0.0
        thuc_thu = 0.0
        if pos_order:
            gia_tri_don_hang = pos_order.amount_total or 0.0
            thuc_thu = pos_order.amount_paid or 0.0
        
        # Payment method (try current picking first, then source picking)
        payment_method = (getattr(picking, 'x_studio_pos_payment_method', '') 
                          or getattr(picking, 'x_studio_payment_method', '') 
                          or getattr(source_picking, 'x_studio_pos_payment_method', '')
                          or getattr(source_picking, 'x_studio_payment_method', '') 
                          or '')
        
        # Update partner_name based on warehouse and payment method
        warehouse_code = self._get_warehouse_code(picking)
        is_multiple = (',' in str(payment_method)) or ("kết hợp" in str(payment_method).lower())
        payment_method_lower = str(payment_method).lower()

        # Mapping for KBC (BENCAM)
        if warehouse_code in ["KBC", "BENCAM"]:
            if "tiền mặt" in payment_method_lower and not is_multiple:
                partner_name = "KH27182013179"
            elif "chuyển khoản" in payment_method_lower and not is_multiple:
                partner_name = "KH27182013178"
            elif is_multiple:
                partner_name = "KHACHLE-BC"
        
        # Mapping for TSN (HCM)
        elif warehouse_code in ["TSN", "HCM"]:
            if "tiền mặt" in payment_method_lower and not is_multiple:
                partner_name = "KH27182013176"
            elif "chuyển khoản" in payment_method_lower and not is_multiple:
                partner_name = "KH27182013177"
            elif is_multiple:
                partner_name = "KHACHLE-HCM"
        
        # Status
        tinh_trang_giao_hang = 'Đã giao hàng' if picking.state == 'done' else 'Chưa giao hàng'
        tinh_trang_thanh_toan = 'Đã thanh toán' if (pos_order and pos_order.state == 'paid') else 'Chưa thanh toán'
        
        return {
            'su_dung_ngoai_te': 'Không',
            'loai_tien': 'VND',
            'ty_gia': 1,
            'so_don_hang': so_don_hang,
            'ngay_dat_hang': date_str,
            'khach_hang': partner_name,
            'lien_he': '',
            'don_hang_cha': '',
            'co_hoi': '',
            'gia_tri_don_hang': gia_tri_don_hang,
            'bao_gia': '',
            'khach_tt_truoc': '',
            'loai_don_hang': '',
            'so_ngay_duoc_no': 0,
            'han_giao_hang': date_str,
            'han_thanh_toan': date_str,
            'dien_giai': f'Bán hàng POS {partner_name}',
            'tinh_trang_kh': 'KH mới',
            'tinh_trang': 'Đang thực hiện',
            'ngay_ghi_so': date_str,
            'thuc_thu': thuc_thu,
            'tinh_trang_giao_hang': tinh_trang_giao_hang,
            'du_kien_chi': '',
            'tinh_trang_thanh_toan': tinh_trang_thanh_toan,
            'han_san_xuat': '',
            'da_xuat_hoa_don': '',
            'khach_hang_hoa_don': partner_name,
            'nguoi_mua_hang': '',
            'quoc_gia_hoa_don': '',
            'tinh_thanh_pho_hoa_don': '',
            'quan_huyen_hoa_don': '',
            'phuong_xa_hoa_don': '',
            'so_nha_duong_pho_hoa_don': '',
            'ma_vung_hoa_don': '',
            'dia_chi_hoa_don': dia_chi_giao_hang,
            'nguoi_nhan_hang': partner_name,
            'dien_thoai': partner_phone,
            'dia_chi_giao_hang': dia_chi_giao_hang,
            'nhan_vien_kho': '',
            'hinh_thuc_giao_hang': '',
            'ngay_giao_du_kien': scheduled_date_str,
            'ben_tra_phi_van_chuyen': '',
            'hinh_thuc_thanh_toan': payment_method,
            'mo_ta': '',
            'nguoi_thuc_hien': '',
            'dung_chung': '',
            'ngung_theo_doi': '',
            'doi_tac_ctv_gioi_thieu': '',
            'dong_bo_don_gia_sau_ck': '',
        }

    def _get_pos_crm_line_data(self, picking, pos_line, source_picking=None):
        """
        Mapping dữ liệu từ pos.order.line sang row Sheet 2 (18 cột)
        pos_line: pos.order.line record
        source_picking: phiếu PICK gốc (nếu picking là OUT trong 3-step delivery)
        Trả về dict với key tương ứng columns sheet 2
        """
        src = source_picking or picking
        pos_order = getattr(src, 'pos_order_id', False)
        prod = pos_line.product_id
        
        # Order reference (FK to Sheet 1) - use x_studio_pos_group
        so_don_hang = (picking.x_studio_pos_group 
                       or src.x_studio_pos_group 
                       or (pos_order and pos_order.name) 
                       or picking.name or '')
        
        # Warehouse
        raw_warehouse_code = self._get_warehouse_code(picking)
        warehouse_code_vietnamese = self._get_warehouse_name_vietnamese(raw_warehouse_code)
        
        # Product info
        ma_hang_hoa = prod.default_code or ''
        ten_hang = prod.display_name or prod.name or ''
        don_vi_tinh = (prod.uom_id and prod.uom_id.name) or ''
        
        # Quantity and prices
        so_luong = pos_line.qty or 0.0
        don_gia_sau_thue = pos_line.price_unit or 0.0
        discount = pos_line.discount or 0.0
        
        # Tax
        thue_suat = 0.0
        if pos_line.tax_ids_after_fiscal_position:
            thue_suat = pos_line.tax_ids_after_fiscal_position[0].amount or 0.0
        
        # Calculations
        # Đơn giá (trước thuế) = Đơn giá sau thuế / (1 + thuế suất%)
        don_gia = don_gia_sau_thue / (1 + thue_suat / 100) if thue_suat else don_gia_sau_thue
        
        # Thành tiền (chưa thuế, chưa chiết khấu)
        thanh_tien_truoc_ck = don_gia * so_luong
        
        # Tiền chiết khấu
        tien_chiet_khau = thanh_tien_truoc_ck * discount / 100
        
        # Thành tiền (sau chiết khấu, chưa thuế)
        thanh_tien = thanh_tien_truoc_ck - tien_chiet_khau
        
        # Tiền thuế
        tien_thue = thanh_tien * thue_suat / 100
        
        # Tổng tiền (sau thuế)
        tong_tien = thanh_tien + tien_thue
        
        return {
            'ma_hang_hoa': ma_hang_hoa,
            'dien_giai': ten_hang,
            'mo_ta': '',
            'kho': 'HLV',
            'kho_odoo': warehouse_code_vietnamese,
            'tinh_trang_hang': 'Bình thường',
            'don_vi_tinh': don_vi_tinh,
            'so_luong': so_luong,
            'don_gia_sau_thue': don_gia_sau_thue,
            'don_gia': don_gia,
            'thanh_tien': thanh_tien,
            'ty_le_chiet_khau': discount,
            'tien_chiet_khau': tien_chiet_khau,
            'thue_suat': thue_suat,
            'tien_thue': tien_thue,
            'tong_tien': tong_tien,
            'don_gia_mua_bat_buoc': '',
            'don_hang': so_don_hang,
        }

    def _create_pos_crm_excel_workbook(self, pickings):
        """
        Tạo workbook Excel với 2 sheets theo template SaleOrder_Template_Short.xlsx
        Sheet 1: Nhập khẩu Đơn hàng (49 cột - 1 row per order)
        Sheet 2: nhập khẩu hàng hóa (18 cột - multiple rows per order)
        """
        wb = Workbook()
        
        # Styles
        header_font = Font(name='Arial', size=10, bold=True)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border_side = Side(style='thin', color='000000')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        number_alignment = Alignment(horizontal='right', vertical='center')
        
        # ===== SHEET 1 =====
        ws1 = wb.active
        ws1.title = "Nhập khẩu Đơn hàng"
        columns_s1 = self._get_pos_crm_columns_sheet1()
        
        # Header row
        for col_idx, col_def in enumerate(columns_s1, start=1):
            cell = ws1.cell(row=1, column=col_idx)
            cell.value = col_def['name']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            ws1.column_dimensions[get_column_letter(col_idx)].width = col_def.get('width', 15)
        
        # Data rows - Grouped by so_don_hang (Sheet 1)
        # Gộp tất cả các dòng có cùng số đơn hàng lại thành 1 dòng duy nhất
        grouped_orders = {}
        order_keys = [] # To keep order
        
        for picking in pickings:
            row_data = self._get_pos_crm_order_data(picking)
            so_don_hang = row_data.get('so_don_hang')
            
            if not so_don_hang:
                continue
                
            if so_don_hang not in grouped_orders:
                grouped_orders[so_don_hang] = row_data
                order_keys.append(so_don_hang)
            else:
                # Sum values
                grouped_orders[so_don_hang]['gia_tri_don_hang'] += row_data.get('gia_tri_don_hang', 0.0)
                grouped_orders[so_don_hang]['thuc_thu'] += row_data.get('thuc_thu', 0.0)
                
        current_row = 2
        for so_don_hang in order_keys:
            row_data = grouped_orders[so_don_hang]
            for col_idx, col_def in enumerate(columns_s1, start=1):
                cell = ws1.cell(row=current_row, column=col_idx)
                value = row_data.get(col_def['key'], '')
                if value is None:
                    value = ''
                cell.value = value
                cell.border = border
                if isinstance(value, (int, float)) and value != '':
                    cell.alignment = number_alignment
                    if 'gia' in col_def['key'] or 'tien' in col_def['key'] or 'thu' in col_def['key']:
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = cell_alignment
            current_row += 1
        
        ws1.row_dimensions[1].height = 30
        
        # ===== SHEET 2 =====
        ws2 = wb.create_sheet(title="nhập khẩu hàng hóa")
        columns_s2 = self._get_pos_crm_columns_sheet2()
        
        # Header row
        for col_idx, col_def in enumerate(columns_s2, start=1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.value = col_def['name']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            ws2.column_dimensions[get_column_letter(col_idx)].width = col_def.get('width', 15)
        
        # Data rows - multiple rows per picking (1 per pos_line)
        current_row = 2
        for picking in pickings:
            pos_order = getattr(picking, 'pos_order_id', False)
            source_picking = None
            
            # Fallback: truy ngược từ OUT → PICK để lấy POS order
            if not pos_order:
                source_pos_picking = self._find_source_pos_picking(picking)
                if source_pos_picking:
                    source_picking = source_pos_picking
                    pos_order = getattr(source_picking, 'pos_order_id', False)
            
            if not pos_order:
                continue
            
            for pos_line in pos_order.lines:
                if not pos_line.product_id:
                    continue
                
                row_data = self._get_pos_crm_line_data(picking, pos_line, source_picking=source_picking)
                for col_idx, col_def in enumerate(columns_s2, start=1):
                    cell = ws2.cell(row=current_row, column=col_idx)
                    value = row_data.get(col_def['key'], '')
                    if value is None:
                        value = ''
                    cell.value = value
                    cell.border = border
                    if isinstance(value, (int, float)) and value != '':
                        cell.alignment = number_alignment
                        if 'so_luong' in col_def['key'] or 'ty_le' in col_def['key'] or 'thue_suat' in col_def['key']:
                            cell.number_format = '#,##0.00'
                        elif 'gia' in col_def['key'] or 'tien' in col_def['key']:
                            cell.number_format = '#,##0'
                    else:
                        cell.alignment = cell_alignment
                current_row += 1
        
        ws2.row_dimensions[1].height = 30
        
        return wb

    def action_export_pos_crm(self):
        """Xuất file Excel POS CRM với 2 sheets theo template"""
        self.ensure_one()
        if Workbook is None:
            raise UserError(_("Thiếu thư viện openpyxl. Vui lòng cài đặt 'openpyxl' cho Python."))
        
        # Phase 1: Tìm pickings có POS fields trực tiếp
        domain = self._domain()
        domain.append('|')
        domain.append(('x_studio_pos_group', '!=', False))
        domain.append(('pos_session_id', '!=', False))
        
        direct_pickings = self.env["stock.picking"].sudo().search(domain, order="scheduled_date asc, id asc")
        
        # Phase 2: Tìm thêm pickings outgoing chưa có POS fields nhưng trace lại được POS order
        # (3-step delivery: OUT picking không có x_studio_pos_group nhưng PICK gốc có)
        base_domain = self._domain()  # outgoing, done, date range
        all_outgoing = self.env["stock.picking"].sudo().search(base_domain, order="scheduled_date asc, id asc")
        
        # Lọc: pickings chưa có trong direct_pickings mà trace ngược lại có POS order
        traced_pickings = self.env["stock.picking"]
        already_found_ids = set(direct_pickings.ids)
        
        for picking in all_outgoing:
            if picking.id in already_found_ids:
                continue
            # Dùng _find_source_pos_picking để kiểm tra
            source = self._find_source_pos_picking(picking)
            if source:
                traced_pickings |= picking
        
        # Gộp lại
        pickings = direct_pickings | traced_pickings
        pickings = pickings.sorted(key=lambda p: (p.scheduled_date or p.date_done, p.id))
        
        if not pickings:
            raise UserError(_("Không tìm thấy phiếu xuất POS nào trong khoảng ngày đã chọn."))
        
        # Create workbook
        wb = self._create_pos_crm_excel_workbook(pickings)
        
        # Save to BytesIO
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        filename = f"POS_CRM_{self.date_from}_{self.date_to}.xlsx"
        attachment = self.env["ir.attachment"].sudo().create({
            "name": filename,
            "type": "binary",
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "datas": base64.b64encode(out.getvalue()),
            "res_model": "picking.export.wizard",
            "res_id": self.id,
        })
        
        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }
