# -*- coding: utf-8 -*-
from odoo import api, fields, models, _
from odoo.exceptions import UserError
import base64
import datetime
from io import BytesIO
from collections import defaultdict
import logging

_logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
except ImportError:
    Workbook = None


class InventoryReportWizard(models.TransientModel):
    _name = "inventory.report.wizard"
    _description = "Xuất báo cáo tồn kho Excel"

    report_date = fields.Date(
        string="Ngày báo cáo", 
        required=True, 
        default=fields.Date.context_today,
        help="Báo cáo sẽ tính tồn đầu ngày (0h) và tồn cuối tại thời điểm xuất file"
    )
    
    warehouse_ids = fields.Many2many(
        "stock.warehouse",
        string="Kho",
        help="Để trống = Tất cả kho. Chọn 1 hoặc nhiều kho để lọc cụ thể.",
    )

    def _get_start_of_day(self, date_val):
        """Lấy thời điểm bắt đầu ngày (0h)"""
        return datetime.datetime.combine(date_val, datetime.time.min)
    
    def _get_end_of_day(self, date_val):
        """Lấy thời điểm cuối ngày (23h59:59)"""
        return datetime.datetime.combine(date_val, datetime.time.max)
    
    def _get_current_datetime(self):
        """Lấy thời điểm hiện tại"""
        return fields.Datetime.now()
    
    def _get_report_end_datetime(self):
        """
        Lấy thời điểm kết thúc cho báo cáo:
        - Nếu ngày báo cáo = hôm nay: lấy thời điểm hiện tại
        - Nếu ngày báo cáo < hôm nay: lấy 23h59:59 của ngày đó
        """
        today = fields.Date.context_today(self)
        if self.report_date >= today:
            return self._get_current_datetime()
        else:
            return self._get_end_of_day(self.report_date)

    def _get_warehouse_locations(self):
        """
        Lấy danh sách location của các kho được chọn, bao gồm TẤT CẢ child locations
        NHƯNG PHẢI loại trừ transit locations để tính xuất/nhập đúng
        
        ⚠️ QUAN TRỌNG: Nếu transit location nằm trong location_ids:
        - Move KHD → Transit: cả source và dest đều trong location_ids → KHÔNG tính xuất
        - Move Transit → KBC: cả source và dest đều trong location_ids → KHÔNG tính nhập
        → KẾT QUẢ: Mất hết dữ liệu inter-warehouse transfer!
        """
        if self.warehouse_ids:
            warehouses = self.warehouse_ids
        else:
            warehouses = self.env['stock.warehouse'].search([])
        
        # Lấy tất cả location thuộc kho, bao gồm cả sub-locations
        location_ids = []
        excluded_count = 0
        
        for wh in warehouses:
            # Lấy view_location_id để tìm tất cả location con
            if wh.view_location_id:
                # Tìm TẤT CẢ location con có usage = internal HOẶC transit
                # (Vì một số hệ thống đặt transit usage = 'internal')
                child_locs = self.env['stock.location'].search([
                    ('id', 'child_of', wh.view_location_id.id),
                    '|',
                    ('usage', '=', 'internal'),
                    ('usage', '=', 'transit')
                ])
                
                # 🔧 LOẠI BỎ TRANSIT: Kiểm tra nhiều điều kiện
                # Bao gồm: usage='transit' HOẶC tên chứa 'transit', 'inter-warehouse', etc.
                filtered_locs = child_locs.filtered(
                    lambda loc: (
                        loc.usage != 'transit'  # ✓ Loại bỏ usage = transit
                        and 'transit' not in loc.complete_name.lower() 
                        and 'inter-warehouse' not in loc.complete_name.lower()
                        and 'inter warehouse' not in loc.complete_name.lower()
                        and not loc.name.lower().startswith('inter')
                        # Thêm điều kiện: location có parent là warehouse transit
                        and (not loc.location_id or 'transit' not in loc.location_id.complete_name.lower())
                    )
                )
                
                excluded_locs = child_locs - filtered_locs
                excluded_count += len(excluded_locs)
                
                location_ids.extend(filtered_locs.ids)
        
        return location_ids

    def _get_product_qty_at_datetime(self, product_id, location_ids, target_datetime):
        """
        Tính số lượng tồn kho của product tại thời điểm target_datetime
        Sử dụng stock.quant và stock.move để tính toán
        """
        # Lấy tồn kho hiện tại
        quants = self.env['stock.quant'].search([
            ('product_id', '=', product_id),
            ('location_id', 'in', location_ids),
        ])
        
        current_qty = sum(quants.mapped('quantity'))
        
        # Tìm các stock.move đã hoàn thành SAU target_datetime
        # để trừ ngược lại và tính tồn tại target_datetime
        moves_after = self.env['stock.move'].search([
            ('product_id', '=', product_id),
            ('state', '=', 'done'),
            ('date', '>', target_datetime),
            '|',
            ('location_id', 'in', location_ids),
            ('location_dest_id', 'in', location_ids),
        ])
        
        adjustment = 0
        for move in moves_after:
            # Nếu move vào kho (location_dest_id in location_ids) -> trừ đi
            if move.location_dest_id.id in location_ids and move.location_id.id not in location_ids:
                adjustment -= move.product_uom_qty
            # Nếu move ra khỏi kho (location_id in location_ids) -> cộng lại
            elif move.location_id.id in location_ids and move.location_dest_id.id not in location_ids:
                adjustment += move.product_uom_qty
            # Nếu move nội bộ (cả 2 đều trong location_ids) -> không ảnh hưởng
        
        return current_qty + adjustment

    def _is_virtual_location(self, location_id):
        """
        Kiểm tra xem location có phải là virtual location hay không
        Virtual location là location được sử dụng cho inventory adjustment
        Bao gồm: loss, inventory, scrap, production, etc. (usage != 'internal' và != 'transit')
        """
        location = self.env['stock.location'].browse(location_id)
        # Virtual locations có usage là: loss, inventory, scrap, production, etc.
        # Không phải internal hoặc transit
        return location.usage not in ['internal', 'transit']

    def _get_product_adjustment_details(self, product_id, location_ids, start_datetime, end_datetime):
        """
        Lấy danh sách điều chỉnh tồn kho (Inventory Adjustment) của sản phẩm
        
        🔧 LOGIC: 
        - Tìm các stock.move có source_location là virtual location (loss, inventory, scrap, etc.)
        - destination_location trong location_ids
        - Hoặc ngược lại: source trong location_ids, dest là virtual location (cho trường hợp xuất hàng hỏng)
        
        Trả về: {
            'incoming': [{'qty': ..., 'from_location': ..., 'move': ..., 'picking': ...}, ...],
            'outgoing': [{'qty': ..., 'to_location': ..., 'move': ..., 'picking': ...}, ...],
            'total_qty': số lượng net điều chỉnh
        }
        """
        moves = self.env['stock.move'].search([
            ('product_id', '=', product_id),
            ('state', '=', 'done'),
            ('date', '>=', start_datetime),
            ('date', '<=', end_datetime),
            '|',
            ('location_id', 'in', location_ids),
            ('location_dest_id', 'in', location_ids),
        ])
        
        adjustments = {
            'incoming': [],  # Điều chỉnh tăng (từ virtual location vào kho)
            'outgoing': [],  # Điều chỉnh giảm (từ kho ra virtual location)
            'total_qty': 0
        }
        
        for move in moves:
            # Bỏ qua moves không có picking_id hoặc moves nội bộ
            if not move.move_line_ids:
                continue
            
            for line in move.move_line_ids:
                location_id = line.location_id.id
                location_dest_id = line.location_dest_id.id
                qty = line.qty_done
                
                # Bỏ qua: location_id = location_dest_id
                if location_id == location_dest_id:
                    continue
                
                # 🔧 ĐIỀU CHỈNH TĂNG: Virtual location → Kho thực
                # Ví dụ: Inventory / Stock Adjustment → Internal Location
                if (self._is_virtual_location(location_id) and 
                    location_dest_id in location_ids):
                    
                    from_location = self.env['stock.location'].browse(location_id)
                    adjustments['incoming'].append({
                        'qty': qty,
                        'from_location': from_location.name,
                        'from_location_usage': from_location.usage,
                        'move': move,
                        'picking': move.picking_id,
                        'date': move.date
                    })
                    adjustments['total_qty'] += qty
                
                # 🔧 ĐIỀU CHỈNH GIẢM: Kho thực → Virtual location
                # Ví dụ: Internal Location → Loss / Scrap
                elif (location_id in location_ids and 
                      self._is_virtual_location(location_dest_id)):
                    
                    to_location = self.env['stock.location'].browse(location_dest_id)
                    adjustments['outgoing'].append({
                        'qty': qty,
                        'to_location': to_location.name,
                        'to_location_usage': to_location.usage,
                        'move': move,
                        'picking': move.picking_id,
                        'date': move.date
                    })
                    adjustments['total_qty'] -= qty
        
        return adjustments

    def _get_outgoing_qty_between(self, product_id, location_ids, start_datetime, end_datetime, log_locations=False):
        """
        Tính tổng số lượng xuất kho từ start_datetime đến end_datetime
        
        🔧 LOGIC MỚI: Dựa trên stock.move.line (CHÍNH XÁC hơn stock.move)
        - Xuất = move line từ internal location (trong location_ids) → location khác (ngoài location_ids)
        - Loại bỏ: move line có location_id = location_dest_id (same location)
        - GIỮ LẠI: move line có location_id khác location_dest_id dù cả 2 đều trong location_ids
        """
        # Log location_ids để debug (chỉ log 1 lần khi được yêu cầu)
        # Location IDs logged at debug level only if needed
        
        # Tìm các stock.move có move_line_ids
        moves = self.env['stock.move'].search([
            ('product_id', '=', product_id),
            ('state', '=', 'done'),
            ('date', '>=', start_datetime),
            ('date', '<=', end_datetime),
        ])
        
        total_qty = 0
        outgoing_moves_info = []
        internal_moves_info = []
        
        for move in moves:
            if not move.move_line_ids:
                continue
            
            # Duyệt qua TỪNG move.line để lấy location chính xác
            for line in move.move_line_ids:
                location_id = line.location_id.id
                location_dest_id = line.location_dest_id.id
                qty = line.qty_done
                
                # Loại bỏ: location_id = location_dest_id (move trong cùng 1 location)
                if location_id == location_dest_id:
                    continue
                
                # ✅ Xuất: từ location_ids → ngoài location_ids
                if location_id in location_ids and location_dest_id not in location_ids:
                    total_qty += qty
                    outgoing_moves_info.append({
                        'move': move,
                        'line': line,
                        'qty': qty
                    })
                # ⚠️ Internal: move giữa các location trong location_ids (bỏ qua)
                elif location_id in location_ids and location_dest_id in location_ids:
                    internal_moves_info.append({
                        'move': move,
                        'line': line,
                        'qty': qty
                    })
        
        # Debug logging - ENHANCED
        if outgoing_moves_info or internal_moves_info:
            _logger.warning(
                f"📤 Product {product_id}: {len(outgoing_moves_info)} OUTGOING, "
                f"{len(internal_moves_info)} INTERNAL (ignored)"
            )
        
        return total_qty

    def _get_incoming_qty_between(self, product_id, location_ids, start_datetime, end_datetime):
        """
        Tính tổng số lượng nhập kho từ start_datetime đến end_datetime
        
        🔧 LOGIC MỚI: Dựa trên stock.move.line (CHÍNH XÁC hơn stock.move)
        - Nhập = move line từ location khác (ngoài location_ids) → internal location (trong location_ids)
        - Loại bỏ: move line có location_id = location_dest_id (same location)
        - GIỮ LẠI: move line có location_id khác location_dest_id dù cả 2 đều trong location_ids
        """
        # Tìm các stock.move có move_line_ids
        moves = self.env['stock.move'].search([
            ('product_id', '=', product_id),
            ('state', '=', 'done'),
            ('date', '>=', start_datetime),
            ('date', '<=', end_datetime),
        ])
        
        total_qty = 0
        incoming_moves_info = []
        internal_moves_info = []
        
        for move in moves:
            if not move.move_line_ids:
                continue
            
            # Duyệt qua TỪNG move.line để lấy location chính xác
            for line in move.move_line_ids:
                location_id = line.location_id.id
                location_dest_id = line.location_dest_id.id
                qty = line.qty_done
                
                # Loại bỏ: location_id = location_dest_id (move trong cùng 1 location)
                if location_id == location_dest_id:
                    continue
                
                # ✅ Nhập: từ ngoài location_ids → vào location_ids
                if location_id not in location_ids and location_dest_id in location_ids:
                    total_qty += qty
                    incoming_moves_info.append({
                        'move': move,
                        'line': line,
                        'qty': qty
                    })
                # ⚠️ Internal: move giữa các location trong location_ids (bỏ qua)
                elif location_id in location_ids and location_dest_id in location_ids:
                    internal_moves_info.append({
                        'move': move,
                        'line': line,
                        'qty': qty
                    })
        
        # Debug logging - ENHANCED
        if incoming_moves_info or internal_moves_info:
            _logger.warning(
                f"📥 Product {product_id}: {len(incoming_moves_info)} INCOMING, "
                f"{len(internal_moves_info)} INTERNAL (ignored)"
            )
        
        return total_qty

    def _get_all_products_with_movement(self, location_ids, start_datetime):
        """
        Lấy tất cả sản phẩm có tồn kho hoặc có phát sinh từ start_datetime
        """
        # Sản phẩm có tồn kho hiện tại
        quants = self.env['stock.quant'].search([
            ('location_id', 'in', location_ids),
            ('quantity', '!=', 0),
        ])
        product_ids = set(quants.mapped('product_id').ids)
        
        # Sản phẩm có phát sinh từ start_datetime
        moves = self.env['stock.move'].search([
            ('state', '=', 'done'),
            ('date', '>=', start_datetime),
            '|',
            ('location_id', 'in', location_ids),
            ('location_dest_id', 'in', location_ids),
        ])
        product_ids.update(moves.mapped('product_id').ids)
        
        return list(product_ids)

    def _is_return_picking(self, picking):
        """
        Kiểm tra xem picking có phải là return order không
        Return picking: picking_type_id.code = 'crm_return' hoặc 'incoming_crm' (tùy cấu hình)
        hoặc picking có tên chứa 'Return', 'RMA', 'Credit Note', etc.
        """
        if not picking:
            return False
        
        # Kiểm tra picking type code
        if picking.picking_type_id:
            code = picking.picking_type_id.code
            if code in ['crm_return', 'incoming_crm', 'incoming_return', 'return']:
                return True
        
        # Kiểm tra tên picking
        picking_name = picking.name.upper()
        return_keywords = ['RETURN', 'RMA', 'CREDIT', 'RETOUR']
        return any(keyword in picking_name for keyword in return_keywords)

    def _get_product_outgoing_picking_names(self, product_id, location_ids, start_datetime, end_datetime):
        """
        Lấy danh sách tên (mã) các picking xuất kho của sản phẩm trong khoảng thời gian
        Bao gồm:
        - Picking xuất thường (xuất bán, xuất kho khác, etc.)
        - Điều chỉnh tồn kho (Inventory Adjustment) nếu có giảm tồn
        
        ⚠️ LOẠI BỎ: Return Order (trả hàng) - vì chúng là incoming, không phải outgoing
        
        Trả về: string danh sách mã đơn cách nhau bởi dấu phẩy
        
        🔧 LOGIC MỚI: Dựa trên stock.move.line
        - Lấy TẤT CẢ picking có move line xuất ra khỏi location_ids
        - NHƯNG loại bỏ return picking (là incoming, không phải outgoing)
        - Nếu xuất sang inter-warehouse transit, tìm thêm picking nhận ở kho đích
        - Thêm thông tin adjustment (tồn kho bị điều chỉnh giảm sang virtual location)
        """
        # Tìm TẤT CẢ các stock.move
        moves = self.env['stock.move'].search([
            ('product_id', '=', product_id),
            ('state', '=', 'done'),
            ('date', '>=', start_datetime),
            ('date', '<=', end_datetime),
        ], order='date asc')
        
        picking_names = []
        seen_picking_ids = set()
        return_picking_ids = set()  # Để loại bỏ return orders
        
        for move in moves:
            if not move.picking_id or not move.move_line_ids:
                continue
            
            picking = move.picking_id
            
            # 🔧 LOẠI BỎ: Return Order từ danh sách xuất
            if self._is_return_picking(picking):
                return_picking_ids.add(picking.id)
                continue
            
            # Kiểm tra xem move có line nào là outgoing không
            has_outgoing = False
            is_inter_warehouse = False
            
            for line in move.move_line_ids:
                location_id = line.location_id.id
                location_dest_id = line.location_dest_id.id
                
                # Bỏ qua same location
                if location_id == location_dest_id:
                    continue
                
                # Xuất: từ location_ids → ngoài location_ids
                if location_id in location_ids and location_dest_id not in location_ids:
                    has_outgoing = True
                    
                    # Check inter-warehouse transfer
                    dest_usage = line.location_dest_id.usage
                    if (dest_usage == 'transit' or 
                        'transit' in line.location_dest_id.complete_name.lower() or
                        'inter-warehouse' in line.location_dest_id.complete_name.lower()):
                        is_inter_warehouse = True
                    
                    break
            
            if has_outgoing:
                # Thêm picking xuất
                if picking.id not in seen_picking_ids:
                    picking_names.append(picking.name)
                    seen_picking_ids.add(picking.id)
                    
                    # Nếu là inter-warehouse, tìm picking nhận ở kho đích
                    if is_inter_warehouse or move.move_dest_ids:
                        dest_picking = self._find_destination_picking_from_move(move)
                        # 🔧 Kiểm tra return picking trước khi thêm
                        if (dest_picking and dest_picking.id not in seen_picking_ids and 
                            not self._is_return_picking(dest_picking)):
                            picking_names.append(dest_picking.name)
                            seen_picking_ids.add(dest_picking.id)
                        elif dest_picking and self._is_return_picking(dest_picking):
                            return_picking_ids.add(dest_picking.id)
        
        # 🆕 Thêm thông tin Inventory Adjustment (điều chỉnh giảm)
        adjustments = self._get_product_adjustment_details(product_id, location_ids, start_datetime, end_datetime)
        
        if adjustments['outgoing']:
            _logger.warning(
                f"Product {product_id}: {len(adjustments['outgoing'])} adjustments (outgoing)"
            )
            
            for adj in adjustments['outgoing']:
                # Tạo mô tả điều chỉnh
                move = adj['move']
                picking = adj['picking']
                
                # Lấy picking reference nếu có, hoặc tạo string mô tả
                if picking and picking.id not in seen_picking_ids:
                    picking_names.append(picking.name)
                    seen_picking_ids.add(picking.id)
                elif not picking:
                    # Nếu không có picking, tạo mô tả trực tiếp từ move
                    adjustment_desc = f"Điều chỉnh sang {adj['to_location']} - {adj['qty']} cái"
                    picking_names.append(adjustment_desc)
        
        return ', '.join(picking_names) if picking_names else ''

    def _find_destination_picking_from_move(self, outgoing_move):
        """
        Tìm picking nhận ở kho đích cho move xuất qua transit
        Sử dụng move_dest_ids để tìm chính xác
        """
        # Phương pháp 1: Qua move_dest_ids (linked moves) - ĐÁNG TIN CẬY NHẤT
        if outgoing_move.move_dest_ids:
            for dest_move in outgoing_move.move_dest_ids:
                if dest_move.picking_id and dest_move.state == 'done':
                    return dest_move.picking_id
        
        # Phương pháp 2: Tìm qua location_dest_id của outgoing_move
        # Nếu outgoing_move đi đến transit, tìm move tiếp theo từ transit
        transit_location_id = outgoing_move.location_dest_id.id
        
        # 🔧 Nới rộng điều kiện tìm kiếm để bắt được các move nhận từ transit
        dest_moves = self.env['stock.move'].search([
            ('product_id', '=', outgoing_move.product_id.id),
            ('state', '=', 'done'),
            ('location_id', '=', transit_location_id),
            ('date', '>=', outgoing_move.date),  # Phải sau hoặc cùng lúc với move xuất
            ('date', '<=', outgoing_move.date + datetime.timedelta(days=2)),  # Trong vòng 2 ngày
        ], order='date asc', limit=1)
        
        if dest_moves and dest_moves.picking_id:
            return dest_moves.picking_id
        
        _logger.warning(
            f"✗ Could not find destination picking for outgoing move {outgoing_move.picking_id.name} "
            f"to transit location {outgoing_move.location_dest_id.complete_name}"
        )
        return None

    def _get_product_incoming_picking_names(self, product_id, location_ids, start_datetime, end_datetime):
        """
        Lấy danh sách tên (mã) các picking nhập kho của sản phẩm trong khoảng thời gian
        Bao gồm: 
        - Picking nhập thường (từ nhà cung cấp, kho khác, etc.)
        - Return Order (trả hàng) - nếu có hàng nhập về từ khách hàng
        - Điều chỉnh tồn kho (Inventory Adjustment) nếu có tăng tồn
        
        Trả về: string danh sách mã đơn cách nhau bởi dấu phẩy, ví dụ: "WH/IN/00123, WH/IN/00124, RMA/00001"
        
        🔧 LOGIC MỚI: Dựa trên stock.move.line
        - Lấy TẤT CẢ picking có move line nhập vào location_ids
        - BỎ QUA: Non-return moves xuất ra khỏi kho (đó là outgoing)
        - GIỮ LẠI: Return Order (là incoming, hàng về từ khách hàng)
        - Thêm thông tin adjustment (tồn kho được điều chỉnh từ virtual location)
        """
        # Tìm TẤT CẢ các stock.move
        moves = self.env['stock.move'].search([
            ('product_id', '=', product_id),
            ('state', '=', 'done'),
            ('date', '>=', start_datetime),
            ('date', '<=', end_datetime),
        ], order='date asc')
        
        # Lấy danh sách picking names
        picking_names = []
        seen_picking_ids = set()
        outgoing_non_return_ids = set()  # Để loại bỏ non-return outgoing
        
        for move in moves:
            if not move.picking_id or not move.move_line_ids:
                continue
            
            picking = move.picking_id
            is_return = self._is_return_picking(picking)
            
            # Kiểm tra xem move có line nào là incoming không
            has_incoming = False
            has_outgoing = False
            
            for line in move.move_line_ids:
                location_id = line.location_id.id
                location_dest_id = line.location_dest_id.id
                
                # Bỏ qua same location
                if location_id == location_dest_id:
                    continue
                
                # Nhập: từ ngoài location_ids → vào location_ids
                if location_id not in location_ids and location_dest_id in location_ids:
                    has_incoming = True
                    break
                
                # Xuất: từ location_ids → ngoài location_ids (chỉ cho non-return)
                if not is_return and location_id in location_ids and location_dest_id not in location_ids:
                    has_outgoing = True
            
            # 🔧 LOGIC: Chỉ thêm nếu có incoming HOẶC là return picking
            if has_incoming or is_return:
                # Nếu là non-return outgoing, bỏ qua
                if has_outgoing and not is_return:
                    outgoing_non_return_ids.add(picking.id)
                    continue
                
                if picking.id not in seen_picking_ids:
                    picking_names.append(picking.name)
                    seen_picking_ids.add(picking.id)
        
        # Add inventory adjustments if any
        adjustments = self._get_product_adjustment_details(product_id, location_ids, start_datetime, end_datetime)
        
        if adjustments['incoming']:
            
            for adj in adjustments['incoming']:
                # Tạo mô tả điều chỉnh
                # Format: "Điều chỉnh từ {from_location} - {qty} {uom}"
                move = adj['move']
                picking = adj['picking']
                
                # Lấy picking reference nếu có, hoặc tạo string mô tả
                if picking and picking.id not in seen_picking_ids:
                    picking_names.append(picking.name)
                    seen_picking_ids.add(picking.id)
                elif not picking:
                    # Nếu không có picking, tạo mô tả trực tiếp từ move
                    adjustment_desc = f"Điều chỉnh tồn kho - {adj['qty']} cái"
                    picking_names.append(adjustment_desc)
        
        return ', '.join(picking_names) if picking_names else ''

    def _create_excel_workbook(self, data_rows):
        """Tạo workbook Excel với hyperlink đến picking"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Báo cáo tồn kho"

        # Định nghĩa cột
        columns = [
            {'key': 'stt', 'name': 'STT', 'width': 8},
            {'key': 'product_code', 'name': 'Mã hàng', 'width': 20},
            {'key': 'product_name', 'name': 'Tên hàng', 'width': 40},
            {'key': 'uom', 'name': 'ĐVT', 'width': 12},
            {'key': 'qty_start', 'name': 'Tồn đầu ngày (0h)', 'width': 20},
            {'key': 'qty_in', 'name': 'Số lượng nhập', 'width': 18},
            {'key': 'qty_out', 'name': 'Số lượng xuất', 'width': 18},
            {'key': 'qty_current', 'name': 'Tồn hiện tại', 'width': 18},
            {'key': 'incoming_picking_names', 'name': 'Chi tiết đơn nhập', 'width': 50},
            {'key': 'outgoing_picking_names', 'name': 'Chi tiết đơn xuất', 'width': 50},
        ]

        # Styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border_side = Side(style='thin', color='000000')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        number_alignment = Alignment(horizontal='right', vertical='center')

        # Header row
        HEADER_ROW = 1
        DATA_START = 2

        for col_idx, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=HEADER_ROW, column=col_idx)
            cell.value = col_def['name']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def.get('width', 15)

        # Data rows
        for row_idx, row_data in enumerate(data_rows, start=DATA_START):
            for col_idx, col_def in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Lấy giá trị
                value = row_data.get(col_def['key'], "")
                if value is None:
                    value = ""
                cell.value = value
                
                cell.border = border

                # Number formatting
                if col_def['key'] in ['qty_start', 'qty_in', 'qty_out', 'qty_current']:
                    cell.alignment = number_alignment
                    cell.number_format = '#,##0.00'
                elif col_def['key'] == 'stt':
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = cell_alignment

        # Wrap text for specific columns
        for row_idx in range(DATA_START, len(data_rows) + DATA_START):
            # Wrap text for 'product_name' column
            product_name_cell = ws.cell(row=row_idx, column=3)  # Column 3 is 'product_name'
            product_name_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # Wrap text for 'incoming_picking_names' column
            incoming_picking_names_cell = ws.cell(row=row_idx, column=9)  # Column 9 is 'incoming_picking_names'
            incoming_picking_names_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # Wrap text for 'outgoing_picking_names' column
            outgoing_picking_names_cell = ws.cell(row=row_idx, column=10)  # Column 10 is 'outgoing_picking_names'
            outgoing_picking_names_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        ws.row_dimensions[HEADER_ROW].height = 35

        return wb

    def action_export(self):
        self.ensure_one()
        if Workbook is None:
            raise UserError(_("Thiếu thư viện openpyxl. Vui lòng cài đặt 'pip install openpyxl'."))

        # Lấy thông tin thời gian
        start_of_day = self._get_start_of_day(self.report_date)
        end_of_period = self._get_report_end_datetime()
        
        # Lấy danh sách location
        location_ids = self._get_warehouse_locations()
        if not location_ids:
            raise UserError(_("Không tìm thấy location kho nào."))

        # Lấy danh sách sản phẩm
        product_ids = self._get_all_products_with_movement(location_ids, start_of_day)
        if not product_ids:
            raise UserError(_("Không có sản phẩm nào có tồn kho hoặc phát sinh."))

        products = self.env['product.product'].browse(product_ids)
        
        # Tạo dữ liệu báo cáo
        data_rows = []
        stt = 1
        is_first_product = True
        
        for product in products.sorted(key=lambda p: p.default_code or p.name):
            # Tính tồn đầu ngày
            qty_start = self._get_product_qty_at_datetime(product.id, location_ids, start_of_day)
            
            # Tính số lượng nhập từ đầu ngày đến end_of_period
            qty_in = self._get_incoming_qty_between(product.id, location_ids, start_of_day, end_of_period)
            
            # Tính số lượng xuất từ đầu ngày đến end_of_period
            qty_out = self._get_outgoing_qty_between(product.id, location_ids, start_of_day, end_of_period, log_locations=is_first_product)
            is_first_product = False
            
            # Tính tồn tại end_of_period
            qty_current = self._get_product_qty_at_datetime(product.id, location_ids, end_of_period)
            
            # Bỏ qua sản phẩm không có tồn và không có xuất nhập
            if qty_start == 0 and qty_in == 0 and qty_out == 0 and qty_current == 0:
                continue
            
            # Lấy danh sách mã đơn nhập kho (chỉ picking nhập vào kho hiện tại)
            incoming_picking_names = self._get_product_incoming_picking_names(
                product.id, location_ids, start_of_day, end_of_period
            )
            
            # Lấy danh sách mã đơn xuất kho (bao gồm cả xuất sang transit/kho khác)
            outgoing_picking_names = self._get_product_outgoing_picking_names(
                product.id, location_ids, start_of_day, end_of_period
            )
            
            row = {
                'stt': stt,
                'product_code': product.default_code or '',
                'product_name': product.name or '',
                'uom': product.uom_id.name if product.uom_id else '',
                'qty_start': qty_start,
                'qty_in': qty_in,
                'qty_out': qty_out,
                'qty_current': qty_current,
                'incoming_picking_names': incoming_picking_names,
                'outgoing_picking_names': outgoing_picking_names,
            }
            data_rows.append(row)
            stt += 1

        if not data_rows:
            raise UserError(_("Không có dữ liệu để xuất báo cáo."))

        # Tạo Excel workbook
        wb = self._create_excel_workbook(data_rows)

        # Xuất file
        out = BytesIO()
        wb.save(out)
        out.seek(0)

        warehouse_names = ", ".join(self.warehouse_ids.mapped('name')) if self.warehouse_ids else "TatCaKho"
        filename = f"BaoCao_TonKho_{self.report_date}_{warehouse_names}.xlsx"
        
        attachment = self.env["ir.attachment"].sudo().create({
            "name": filename,
            "type": "binary",
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "datas": base64.b64encode(out.getvalue()),
            "res_model": "inventory.report.wizard",
            "res_id": self.id,
        })

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }

