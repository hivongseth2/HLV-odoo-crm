# -*- coding: utf-8 -*-
from odoo import fields, models, _
from odoo.exceptions import UserError
import base64
from datetime import date, datetime as dt
from io import BytesIO
import unicodedata

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


def _to_date_str(val):
    """Trả về 'Monday, January 01, 2024' từ date/datetime/chuỗi ngày; rỗng nếu không hợp lệ."""
    if not val:
        return ""
    if isinstance(val, (date, dt)):
        d = val.date() if isinstance(val, dt) else val
    else:
        try:
            d = fields.Date.to_date(val)
        except Exception:
            return str(val)
    return d.strftime("%A, %B %d, %Y")


class PurchaseExportWizard(models.TransientModel):
    _name = "purchase.export.wizard"
    _description = "Xuất Excel lệnh mua hàng theo template kế toán"

    _WH_MAP = {
        "KBC": "BENCAM",
        "TSN": "HCM",
        "KHD": "HIENDUC",
        "TSNSR": "HCM_SHOWROOM",
    }

    date_from = fields.Date(string="Từ ngày", required=True)
    date_to   = fields.Date(string="Đến ngày", required=True)

    def _harsh_warehouse_code(self, code):
        return self._WH_MAP.get(code, code)

    def _get_columns_definition(self):
        """Định nghĩa cột cố định theo template mua hàng."""
        return [
            {'key': 'hinh_thuc_mua_hang', 'name': 'Hình thức mua hàng', 'width': 25},
            {'key': 'phuong_thuc_thanh_toan', 'name': 'Phương thức thanh toán', 'width': 25},
            {'key': 'nhan_kem_hoa_don', 'name': 'Nhận kèm hóa đơn', 'width': 20},
            {'key': 'ngay_hach_toan', 'name': 'Ngày hạch toán (*)', 'width': 25},
            {'key': 'ngay_chung_tu', 'name': 'Ngày chứng từ (*)', 'width': 25},
            {'key': 'so_phieu_nhap', 'name': 'Số phiếu nhập (*)', 'width': 20},
            {'key': 'so_ct_ghi_no', 'name': 'Số chứng từ ghi nợ/Số chứng từ thanh toán', 'width': 35},
            {'key': 'mau_so_hd', 'name': 'Mẫu số HĐ', 'width': 15},
            {'key': 'ky_hieu_hd', 'name': 'Ký hiệu HĐ', 'width': 15},
            {'key': 'so_hoa_don', 'name': 'Số hóa đơn', 'width': 15},
            {'key': 'ngay_hoa_don', 'name': 'Ngày hóa đơn', 'width': 25},
            {'key': 'so_tk_chi', 'name': 'Số tài khoản chi', 'width': 18},
            {'key': 'ten_ngan_hang_chi', 'name': 'Tên ngân hàng chi', 'width': 30},
            {'key': 'ma_nha_cung_cap', 'name': 'Mã nhà cung cấp', 'width': 18},
            {'key': 'ten_nha_cung_cap', 'name': 'Tên nhà cung cấp', 'width': 40},
            {'key': 'dia_chi', 'name': 'Địa chỉ', 'width': 50},
            {'key': 'ma_so_thue', 'name': 'Mã số thuế', 'width': 15},
            {'key': 'nguoi_giao_hang', 'name': 'Người giao hàng', 'width': 25},
            {'key': 'dien_giai', 'name': 'Diễn giải', 'width': 40},
            {'key': 'so_tk_nhan', 'name': 'Số tài khoản nhận', 'width': 18},
            {'key': 'ten_ngan_hang_nhan', 'name': 'Tên ngân hàng nhận', 'width': 30},
            {'key': 'ly_do_chi', 'name': 'Lý do chi/nội dung thanh toán', 'width': 40},
            {'key': 'ma_nhan_vien', 'name': 'Mã nhân viên mua hàng', 'width': 20},
            {'key': 'so_luong_ct_kem_theo', 'name': 'Số lượng chứng từ kèm theo', 'width': 25},
            {'key': 'han_thanh_toan', 'name': 'Hạn thanh toán', 'width': 20},
            {'key': 'ma_hang', 'name': 'Mã hàng (*)', 'width': 18},
            {'key': 'ten_hang', 'name': 'Tên hàng', 'width': 40},
            {'key': 'la_dong_ghi_chu', 'name': 'Là dòng ghi chú', 'width': 18},
            {'key': 'ma_kho', 'name': 'Mã kho', 'width': 15},
            {'key': 'hang_hoa_giu_ho', 'name': 'Hàng hóa giữ hộ/bán hộ', 'width': 25},
            {'key': 'tk_kho', 'name': 'TK kho/TK chi phí (*)', 'width': 22},
            {'key': 'tk_cong_no', 'name': 'TK công nợ/TK tiền (*)', 'width': 22},
            {'key': 'dvt', 'name': 'ĐVT', 'width': 12},
            {'key': 'so_luong', 'name': 'Số lượng', 'width': 12},
            {'key': 'don_gia', 'name': 'Đơn giá', 'width': 15},
            {'key': 'thanh_tien', 'name': 'Thành tiền', 'width': 15},
            {'key': 'ty_le_ck', 'name': 'Tỷ lệ CK (%)', 'width': 12},
            {'key': 'tien_chiet_khau', 'name': 'Tiền chiết khấu', 'width': 15},
            {'key': 'ty_le_thue_gtgt', 'name': '% thuế GTGT', 'width': 12},
            {'key': 'ty_le_thue_khac', 'name': '% thuế suất KHAC', 'width': 18},
            {'key': 'tien_thue_gtgt', 'name': 'Tiền thuế GTGT', 'width': 15},
            {'key': 'tk_thue_gtgt', 'name': 'TK thuế GTGT', 'width': 15},
            {'key': 'phi_hang_ve_kho', 'name': 'Phí hàng về kho/Chi phí mua hàng', 'width': 30},
            {'key': 'nhom_hhdv_mua_vao', 'name': 'Nhóm HHDV mua vào', 'width': 20},
            {'key': 'so_lenh_san_xuat', 'name': 'Số Lệnh sản xuất', 'width': 20},
            {'key': 'ma_khoan_muc_cp', 'name': 'Mã khoản mục chi phí', 'width': 22},
            {'key': 'ma_don_vi', 'name': 'Mã đơn vị', 'width': 15},
            {'key': 'ma_doi_tuong_thcp', 'name': 'Mã đối tượng THCP', 'width': 20},
            {'key': 'ma_cong_trinh', 'name': 'Mã công trình', 'width': 18},
            {'key': 'so_don_dat_hang', 'name': 'Số đơn đặt hàng', 'width': 20},
            {'key': 'so_don_mua_hang', 'name': 'Số đơn mua hàng', 'width': 20},
            {'key': 'misa_purchase_order_org_ref_detail_id', 'name': 'MISA org_ref_detail_id', 'width': 30},
            {'key': 'so_hop_dong_mua', 'name': 'Số hợp đồng mua', 'width': 20},
            {'key': 'so_hop_dong_ban', 'name': 'Số hợp đồng bán', 'width': 20},
            {'key': 'ma_thong_ke', 'name': 'Mã thống kê', 'width': 15},
            {'key': 'so_khe_uoc_di_vay', 'name': 'Số khế ước đi vay', 'width': 20},
            {'key': 'so_khe_uoc_cho_vay', 'name': 'Số khế ước cho vay', 'width': 20},
            {'key': 'cp_khong_hop_ly', 'name': 'CP không hợp lý', 'width': 18},
            {'key': 'misa_sync', 'name': 'Đã lập chứng từ', 'width': 15},
        ]

    def _domain(self):
        self.ensure_one()
        if self.date_from > self.date_to:
            raise UserError(_("Khoảng ngày không hợp lệ."))
        return [
            ("picking_type_code", "=", "incoming"),
            ("state", "=", "done"),
            ("date_done", ">=", fields.Date.to_date(self.date_from)),
            ("date_done", "<=", fields.Date.to_date(self.date_to)),
            ("purchase_id", "!=", False),
        ]

    def _partner_code(self, partner):
        if not partner:
            return ""
        return partner.ref or getattr(partner, "barcode", None) or partner.vat or str(partner.id) or ""

    def _get_warehouse_code(self, picking):
        """Lấy mã kho từ picking_type.warehouse_id."""
        if not picking or not picking.picking_type_id or not picking.picking_type_id.warehouse_id:
            return ""
        code = picking.picking_type_id.warehouse_id.code or picking.picking_type_id.warehouse_id.name or ""
        return self._harsh_warehouse_code(code)

    @staticmethod
    def _normalize_addr_token(s):
        s = (s or "").strip().lower()
        s = unicodedata.normalize("NFD", s)
        return "".join(c for c in s if unicodedata.category(c) != "Mn")

    def _compose_partner_address(self, partner):
        if not partner:
            return ""
        parts = [partner.street or "", partner.city or "", partner.state_id.name if partner.state_id else ""]
        seen, out = set(), []
        for p in parts:
            token = self._normalize_addr_token(p)
            if p and token not in seen:
                out.append(p)
                seen.add(token)
        return ", ".join(out)

    def _get_returned_qty_for_move(self, move):
        """Tính tổng số lượng đã trả lại cho 1 stock.move nhập kho (chỉ lấy phiếu trả hàng ở trạng thái done)."""
        returned_moves = move.returned_move_ids.filtered(lambda m: m.state == 'done')
        if not returned_moves:
            returned_moves = self.env['stock.move'].sudo().search([
                ('origin_returned_move_id', '=', move.id),
                ('state', '=', 'done')
            ])
        total_returned = 0.0
        for rm in returned_moves:
            qty = rm.quantity_done if hasattr(rm, 'quantity_done') else getattr(rm, 'quantity', rm.product_uom_qty)
            if rm.product_uom and move.product_uom and rm.product_uom != move.product_uom:
                qty = rm.product_uom._compute_quantity(qty, move.product_uom)
            total_returned += qty
        return total_returned

    def _get_picking_line_rows(self, picking):
        receipt_date_str = _to_date_str(picking.date_done)
        picking_name = picking.name or ""
        purchase_name = picking.origin or "" # Số đơn đặt hàng
        
        partner = picking.partner_id
        partner_code = self._partner_code(partner)
        partner_name = partner.name or ""
        partner_vat = partner.vat or ""
        partner_addr = self._compose_partner_address(partner)
        ma_kho = self._get_warehouse_code(picking)

        rows = []
        
        # Pre-compute net quantity per stock.move in this picking
        move_net_qty_map = {}
        for move in picking.move_ids:
            if move.state != 'done':
                continue
            orig_qty = move.quantity_done if hasattr(move, 'quantity_done') else getattr(move, 'quantity', move.product_uom_qty)
            ret_qty = self._get_returned_qty_for_move(move)
            move_net_qty_map[move.id] = max(0.0, orig_qty - ret_qty)

        # Helper to process a move/move_line
        def process_line(move, qty, uom):
            if not move.product_id:
                return
            
            # Purchase Order Line information
            pol = move.purchase_line_id
            
            # Adjust qty based on returned qty
            orig_move_qty = move.quantity_done if hasattr(move, 'quantity_done') else getattr(move, 'quantity', move.product_uom_qty)
            net_move_qty = move_net_qty_map.get(move.id, orig_move_qty)
            
            if orig_move_qty > 0:
                net_qty = (qty / orig_move_qty) * net_move_qty
            else:
                net_qty = qty

            rows.append(self._build_row_data(
                picking, pol, move.product_id, net_qty, uom,
                receipt_date_str, picking_name, purchase_name, partner_code, partner_name,
                partner_addr, partner_vat, ma_kho, move=move
            ))

        # Iterate over move lines if available, otherwise moves without package
        if picking.move_line_ids:
            for ml in picking.move_line_ids:
                process_line(ml.move_id, ml.qty_done, ml.product_uom_id)
        else:
            for move in picking.move_ids_without_package:
                process_line(move, move.quantity_done if hasattr(move, 'quantity_done') else getattr(move, 'quantity', move.product_uom_qty), move.product_uom)
                
        return rows

    def _build_row_data(self, picking, pol, prod, qty, uom,
                        receipt_date_str, picking_name, purchase_name, partner_code, partner_name,
                        partner_address, partner_vat, ma_kho, move=None):
        product_code = prod.default_code or getattr(prod, 'barcode', '') or ""
        product_name = prod.display_name or prod.name or ""
        uom_name     = uom.name if uom else ""
        
        don_gia      = 0.0
        thanh_tien   = 0.0
        ty_le_ck     = 0.0
        ty_le_thue_gtgt = 0.0
        tien_thue_gtgt = 0.0
        
        purchase = False
        
        if pol:
            purchase = pol.order_id
            don_gia = pol.price_unit or 0.0
            thanh_tien = don_gia * qty
            ty_le_ck = getattr(pol, "discount", 0.0) or 0.0
            
            # Calculate taxes based on the ratio from the PO line
            ty_le_thue_gtgt = next((t.amount or 0.0 for t in pol.taxes_id), 0.0)
            tien_thue_gtgt  = thanh_tien * ty_le_thue_gtgt / 100.0

        misa_org_ref_id = ""
        if move:
            misa_org_ref_id = getattr(move, 'misa_purchase_order_org_ref_detail_id', '') or ''
        if not misa_org_ref_id and pol:
            misa_org_ref_id = getattr(pol, 'misa_purchase_order_org_ref_detail_id', '') or ''

        return {
            # Fixed fields
            'hinh_thuc_mua_hang': 'Mua hàng hóa trong nước',
            'phuong_thuc_thanh_toan': 'Chưa thanh toán',
            'nhan_kem_hoa_don': 'Nhận kèm hóa đơn',

            # Dates (ngày hiện tại)
            'ngay_hach_toan': _to_date_str(date.today()),
            'ngay_chung_tu': _to_date_str(date.today()),
            'so_phieu_nhap': picking_name, # Changed: Picking Name
            'so_ct_ghi_no': '',

            # Invoice-ish fields
            'mau_so_hd': '01GTKT0/001',
            'ky_hieu_hd': 'AB/20E',
            'so_hoa_don': purchase.origin or "" if purchase else "", # Or invoice number if linked? Keep as origin/PO name for now or empty
            'ngay_hoa_don': receipt_date_str, # Changed: Receipt Date

            # Bank chi
            'so_tk_chi': '04080082835',
            'ten_ngan_hang_chi': 'Ngân hàng quốc tế Việt Nam',

            # Partner
            'ma_nha_cung_cap': partner_code,
            'ten_nha_cung_cap': partner_name,
            'dia_chi': partner_address,
            'ma_so_thue': partner_vat,

            # Misc fixed
            'nguoi_giao_hang': 'Vũ Thị Bích Thủy',
            'dien_giai': purchase.origin or "" if purchase else "",
            'so_tk_nhan': '0486523679',
            'ten_ngan_hang_nhan': 'Ngân hàng Nông nghiệp và Phát triển Nông thôn Việt Nam',
            'ly_do_chi': '',
            'ma_nhan_vien': 'DINHTRANTHIKIMQUYEN',
            'so_luong_ct_kem_theo': '',
            'han_thanh_toan': '',

            # Product
            'ma_hang': product_code,
            'ten_hang': product_name,
            'la_dong_ghi_chu': '',
            'ma_kho': ma_kho,
            'hang_hoa_giu_ho': '',

            # Accounts (hardcoded)
            'tk_kho': '156',
            'tk_cong_no': '331',

            # Qty/price
            'dvt': uom_name,
            'so_luong': qty,
            'don_gia': don_gia,
            'thanh_tien': thanh_tien,
            'ty_le_ck': ty_le_ck,
            'tien_chiet_khau': 0.0,

            # Tax
            'ty_le_thue_gtgt': ty_le_thue_gtgt,
            'ty_le_thue_khac': '',
            'tien_thue_gtgt': tien_thue_gtgt,
            'tk_thue_gtgt': '1331',

            # Others
            'phi_hang_ve_kho': '',
            'nhom_hhdv_mua_vao': '',
            'so_lenh_san_xuat': '',
            'ma_khoan_muc_cp': '',
            'ma_don_vi': '',
            'ma_doi_tuong_thcp': '',
            'ma_cong_trinh': '',
            'so_don_dat_hang': '',
            'so_don_mua_hang': purchase_name, # Changed: PO Name
            'misa_purchase_order_org_ref_detail_id': misa_org_ref_id,
            'so_hop_dong_mua': '',
            'so_hop_dong_ban': '',
            'ma_thong_ke': '',
            'so_khe_uoc_di_vay': '',
            'so_khe_uoc_cho_vay': '',
            'cp_khong_hop_ly': 'Không',
            'misa_sync': getattr(picking, 'x_studio_misa_sav', False) if picking else False,
        }

    def _create_excel_workbook(self, data_rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "Mua hàng hóa"

        columns = self._get_columns_definition()

        # Styles
        header_font = Font(name='Arial', size=10, bold=True)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border_side = Side(style='thin', color='000000')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        number_alignment = Alignment(horizontal='right', vertical='center')

        HEADER_ROW, DATA_START = 1, 2

        # Header
        for col_idx, col_def in enumerate(columns, 1):
            cell = ws.cell(row=HEADER_ROW, column=col_idx)
            cell.value = col_def['name']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def.get('width', 15)

        # Data
        for row_idx, row_data in enumerate(data_rows, DATA_START):
            for col_idx, col_def in enumerate(columns, 1):
                v = row_data.get(col_def['key'], "")
                v = "" if v is None else v
                cell = ws.cell(row=row_idx, column=col_idx, value=v)
                cell.border = border
                if isinstance(v, (int, float)):
                    cell.alignment = number_alignment
                    if col_def['key'] in {'don_gia', 'thanh_tien', 'tien_chiet_khau', 'tien_thue_gtgt'}:
                        cell.number_format = '#,##0'
                    elif col_def['key'] in {'ty_le_ck', 'ty_le_thue_gtgt', 'ty_le_thue_khac'}:
                        cell.number_format = '0.00'
                    elif col_def['key'] == 'so_luong':
                        cell.number_format = '#,##0.00'
                else:
                    cell.alignment = cell_alignment

        ws.row_dimensions[HEADER_ROW].height = 30
        return wb

    def action_export(self):
        self.ensure_one()
        if Workbook is None:
            raise UserError(_("Thiếu thư viện openpyxl. Vui lòng cài đặt 'openpyxl' cho Python."))

        pickings = self.env["stock.picking"].sudo().search(self._domain(), order="date_done asc, id asc")
        if not pickings:
            raise UserError(_("Không tìm thấy phiếu nhập kho nào trong khoảng ngày đã chọn."))

        # Tạo dữ liệu
        all_rows = []
        for picking in pickings:
            all_rows.extend(self._get_picking_line_rows(picking))

        if not all_rows:
            raise UserError(_("Không có dữ liệu chi tiết để xuất."))

        # Tạo Excel workbook
        wb = self._create_excel_workbook(all_rows)

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        filename = f"Phieu_mua_hang_trong_nuoc_{self.date_from}_{self.date_to}.xlsx"
        attachment = self.env["ir.attachment"].sudo().create({
            "name": filename,
            "type": "binary",
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "datas": base64.b64encode(out.getvalue()),
            "res_model": "purchase.export.wizard",
            "res_id": self.id,
        })

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }
