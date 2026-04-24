# -*- coding: utf-8 -*-
from odoo import api, fields, models, _
from odoo.exceptions import UserError
import base64
import datetime
import json
import re
from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

import logging
_logger = logging.getLogger(__name__)

import logging
_logger = logging.getLogger(__name__)



def _to_date_str(val, hour=None):
    if not val:
        return ""
    
    fmt = "%A, %B %d, %Y"
    if hour is not None:
        fmt += f" {hour:02d}:00:00"

    if isinstance(val, str):
        try:
            d = fields.Datetime.from_string(val)
            if d:
                return d.strftime(fmt)
        except Exception:
            try:
                d2 = fields.Date.from_string(val)
                if d2:
                    return d2.strftime(fmt)
            except Exception:
                return val
        return val
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime(fmt)
    return str(val)


class StockExportWizard(models.TransientModel):
    _name = "stock.export.wizard"
    _description = "Xuất Excel Kho (Nội bộ & Xuất bán)"

    date_from = fields.Date(string="Từ ngày", required=True)
    date_to = fields.Date(string="Đến ngày", required=True)
    
    warehouse_ids = fields.Many2many(
        "stock.warehouse", string="Kho",
        help="Để trống để lấy tất cả kho trong công ty hiện tại."
    )
    

    state_filter = fields.Selection([
        ('done', 'Hoàn thành'),
        ('assigned', 'Sẵn sàng'),
        ('all', 'Tất cả')
    ], string="Trạng thái", default='done')

    exclude_shopee = fields.Boolean(
        string="Không xuất đơn Shopee",
        default=True,
        help="Chọn để loại bỏ các đơn hàng có đối tác chứa từ khóa 'shopee' khỏi file xuất."
    )


    def _get_warehouse_code(self, picking):
        """Lấy mã kho"""
        if picking.picking_type_id and picking.picking_type_id.warehouse_id:
            return picking.picking_type_id.warehouse_id.code
        # fallback source location
        loc = picking.location_id
        if loc and loc.warehouse_id:
            return loc.warehouse_id.code
        return ""

    def _partner_code(self, partner):
        if not partner:
            return ""
        
        # Ưu tiên 1: Lấy Company Registry (của công ty mẹ hoặc chính nó)
        ref = partner.commercial_partner_id.company_registry or partner.company_registry
        
        # Ưu tiên 2: Nếu không có Registry, mới tìm đến Partner Ref
        if not ref:
            ref = partner.commercial_partner_id.ref or partner.ref
            
        return ref or ""

    def _find_sale_order(self, move, picking):
        # 1) Từ sale_line_id trực tiếp
        if getattr(move, 'sale_line_id', False) and move.sale_line_id.order_id:
            return move.sale_line_id.order_id
        # 2) Từ procurement group
        grp = getattr(move, 'group_id', False)
        if grp and getattr(grp, 'sale_id', False):
            return grp.sale_id
        # 3) Từ picking
        if getattr(picking, 'sale_id', False):
            return picking.sale_id
        return False

    def _get_misa_formatted_name(self, picking, so):
        """
        Logic đổi tên phiếu xuất kho cho Excel MISA.
        Định dạng: [Mã SO (8 số đầu)]-[Số thứ tự phiếu]
        """
        if not picking:
            return ""
            
        # Lấy tên SO/văn bản gốc
        so_name = so.name if so else (picking.origin or "")
        # Nếu không có nguồn gốc SO thì giữ nguyên tên picking
        if not so_name:
            return picking.name or ""
            
        # 1. Trích xuất đoạn mã SO (lấy đến chữ số thứ 8)
        so_match = re.search(r'^(.*?\d{8})', so_name)
        if not so_match:
            # Fallback nếu không khớp regex (VD: SO ngắn hoặc POS)
            so_segment = so_name[:10]
        else:
            so_segment = so_match.group(1)
            
        # 2. Lấy số thứ tự từ tên hiện tại (ví dụ: TSN/OUT/07823 -> 07823)
        current_name = picking.name or ""
        
        # Nếu tên đã đúng định dạng (có dấu gạch ngang nối với SO segment và kết thúc bằng số), trả về luôn
        if '-' in current_name and current_name.startswith(so_segment):
            suffix = current_name.split('-')[-1]
            if suffix.isdigit():
                return current_name

        # Lấy phần số thứ tự từ tên gốc Odoo (thường là phần cuối sau dấu /)
        name_parts = current_name.split('/')
        seq_num = name_parts[-1] if name_parts else ''
        
        if not seq_num or not any(char.isdigit() for char in seq_num):
            # Fallback dùng ID nếu tên không chứa số
            seq_num = str(picking.id).zfill(5)
        else:
            # Nếu seq_num có chuỗi phi số ở trước (VD: OUT/08839), lấy phần số cuối
            digit_match = re.search(r'(\d+)$', seq_num)
            if digit_match:
                seq_num = digit_match.group(1)
        
        return f"{so_segment}-{seq_num}"

    def _domain(self):
        self.ensure_one()
        if self.date_from > self.date_to:
            raise UserError(_("Khoảng ngày không hợp lệ."))

        domain = [
            ("date_done", ">=", fields.Date.to_date(self.date_from)),
            ("date_done", "<=", fields.Date.to_date(self.date_to)),
        ]
        
        # Picking Type Filter by Sequence Code (OUT only)
        # Only export "Xuất bán hàng"
        domain.append(("picking_type_id.sequence_code", "=", "OUT"))

        # State Filter
        if self.state_filter and self.state_filter != 'all':
            domain.append(("state", "=", self.state_filter))
        else:
            domain.append(("state", "in", ["done", "assigned"]))

        if self.warehouse_ids:
            domain.append(("picking_type_id.warehouse_id", "in", self.warehouse_ids.ids))

        return domain

    # ====== STOCK EXPORT TEMPLATE ======
    
    def _get_stock_export_columns(self):
        """Định nghĩa cột cho mẫu Xuất Kho (22 columns)"""
        return [
            {'key': 'loai_xuat_kho', 'name': 'Loại xuất kho', 'width': 20},
            {'key': 'ngay_hach_toan', 'name': 'Ngày hạch toán (*)', 'width': 18},
            {'key': 'ngay_chung_tu', 'name': 'Ngày chứng từ (*)', 'width': 18},
            {'key': 'so_chung_tu', 'name': 'Số chứng từ (*)', 'width': 20},
            {'key': 'don_hang_goc', 'name': 'Đơn hàng gốc', 'width': 20},
            {'key': 'ma_doi_tuong', 'name': 'Mã đối tượng', 'width': 15},
            {'key': 'ten_doi_tuong', 'name': 'Tên đối tượng', 'width': 30},
            {'key': 'khach_hang', 'name': 'Khách hàng', 'width': 30},
            {'key': 'dia_chi', 'name': 'Địa chỉ/Bộ phận', 'width': 40},
            {'key': 'ly_do_xuat', 'name': 'Lý do xuất', 'width': 30},
            {'key': 'ma_hang', 'name': 'Mã hàng (*)', 'width': 18},
            {'key': 'ten_hang', 'name': 'Tên hàng', 'width': 35},
            {'key': 'la_dong_ghi_chu', 'name': 'Là dòng ghi chú', 'width': 15},
            {'key': 'hang_khuyen_mai', 'name': 'Hàng khuyến mại', 'width': 15},
            {'key': 'ma_kho', 'name': 'Mã kho', 'width': 15},
            {'key': 'tk_no', 'name': 'TK Nợ (*)', 'width': 12},
            {'key': 'tk_co', 'name': 'TK Có (*)', 'width': 12},
            {'key': 'dvt', 'name': 'ĐVT', 'width': 10},
            {'key': 'so_luong', 'name': 'Số lượng', 'width': 12},
            {'key': 'don_gia', 'name': 'Đơn giá', 'width': 15},
            {'key': 'thue_suat', 'name': 'Thuế suất (%)', 'width': 12},
            {'key': 'thanh_tien', 'name': 'Thành tiền', 'width': 15},
            {'key': 'thanh_tien_sau_thue', 'name': 'Thành tiền sau thuế', 'width': 18},
            {'key': 'so_lenh_sx', 'name': 'Số lệnh sản xuất', 'width': 15},
            {'key': 'ma_khoan_muc_cp', 'name': 'Mã khoản mục chi phí', 'width': 18},
            {'key': 'ma_doi_tuong_thcp', 'name': 'Mã đối tượng THCP', 'width': 18},
            {'key': 'ma_don_shopee', 'name': 'Mã đơn Shopee', 'width': 20},

        ]

    def _get_stock_export_row_data(self, picking):
        """Xây dựng rows cho mẫu Xuất Kho"""
        rows = []
        
        # --- Common Info ---
        date_done = picking.date_done or picking.scheduled_date or fields.Datetime.now()
        date_str = _to_date_str(date_done)
        date_hach_toan_str = _to_date_str(date_done, hour=18)
        
        partner = picking.partner_id
        partner_code = self._partner_code(partner)
        partner_name = (partner and partner.name) or ""
        
        # Address
        partner_address = ""
        if partner:
            parts = []
            for p in [partner.street, partner.city, partner.state_id.name if partner.state_id else '']:
                if p: parts.append(p)
            partner_address = ", ".join(parts)
            
        # warehouse_code = self._get_warehouse_code(picking) # requested fixed HLV
        # ly_do_xuat = picking.note or picking.name
        ly_do_xuat = "Xuất kho bán hàng cho " + partner_name
        loai_xuat = 'Xuất kho bán hàng'
        
        # Determine moves
        moves = picking.move_line_ids if picking.move_line_ids else picking.move_ids_without_package
        
        # Determine move & product (to find SO)
        first_move = moves[0] if moves else None
        if hasattr(first_move, '_name') and first_move._name == 'stock.move.line':
            first_move_id = first_move.move_id
        else:
            first_move_id = first_move 
        
        so = self._find_sale_order(first_move_id, picking)
        don_hang_goc = so.name if so else (picking.origin or "")

        # --- EARLY SHOPEE EXCLUSION CHECK ---
        if self.exclude_shopee:
            is_shopee = False
            p_name = str(partner_name).lower() if partner_name else ''
            p_code = str(partner_code).lower() if partner_code else ''
            
            if 'shopee' in p_name or 'shopee' in p_code:
                is_shopee = True
            elif getattr(picking, 'shopee_order_ref', False):
                is_shopee = True
            elif so:
                if getattr(so, 'shopee_shop_id', False):
                    is_shopee = True
                elif so.partner_id and 'shopee' in str(so.partner_id.name).lower():
                    is_shopee = True
            
            if is_shopee:
                return [] # Skip this picking entirely
        
        # Customer Name (Khách hàng) - Priority: SO Partner -> Picking Commercial Partner -> Picking Partner
        khach_hang = ""
        if so and so.partner_id:
            khach_hang = so.partner_id.name
            # Ưu tiên company_registry, nếu không có mới lấy ref
            # Thứ tự ưu tiên mới: Parent (để lấy DOTHANH) -> Commercial Partner -> Partner
            p_ref = False
            
            # DEBUG LOG
            _logger.info(f"DEBUG EXPORT: SO {so.name} - Partner {so.partner_id.name} (ID: {so.partner_id.id})")
            _logger.info(f"--- Commercial Partner: {so.partner_id.commercial_partner_id.name} (Reg: {so.partner_id.commercial_partner_id.company_registry}, Ref: {so.partner_id.commercial_partner_id.ref})")
            _logger.info(f"--- Parent: {so.partner_id.parent_id.name if so.partner_id.parent_id else 'None'} (Reg: {so.partner_id.parent_id.company_registry if so.partner_id.parent_id else 'None'}, Ref: {so.partner_id.parent_id.ref if so.partner_id.parent_id else 'None'})")
            _logger.info(f"--- Self Reg: {so.partner_id.company_registry} - Self Ref: {so.partner_id.ref}")

            # Ưu tiên 1: Check Parent Company directly (Sẽ bắt được chữ DOTHANH ở đây và dừng lại)
            # Ưu tiên 1: Check Parent Company directly
            if so.partner_id.parent_id and (so.partner_id.parent_id.ref or so.partner_id.parent_id.company_registry):
                p_ref = so.partner_id.parent_id.ref or so.partner_id.parent_id.company_registry
                
            # Ưu tiên 2: Check Commercial Partner
            elif so.partner_id.commercial_partner_id and (so.partner_id.commercial_partner_id.ref or so.partner_id.commercial_partner_id.company_registry):
                p_ref = so.partner_id.commercial_partner_id.ref or so.partner_id.commercial_partner_id.company_registry
                
            # Ưu tiên 3: Check Partner itself
            elif so.partner_id.ref or so.partner_id.company_registry:
                p_ref = so.partner_id.ref or so.partner_id.company_registry
            
            if p_ref:
                partner_code = p_ref
            else:
                _logger.info("--- NO COMPANY REGISTRY OR REF FOUND!")
                
            # --- SHOPEE OVERRIDE LOGIC ---
            if hasattr(so, 'shopee_shop_id') and so.shopee_shop_id:
                shop = so.shopee_shop_id
                # Check Account Name contains 2014645
                account = getattr(shop, 'account_id', False)
                if account and '2014645' in getattr(account, 'name', ''):
                    shop_id = getattr(shop, 'shop_identifier', 0)
                    target_pid = False
                    
                    if shop_id == 796817584:
                        target_pid = 9715 # MILWAUKEE
                    elif shop_id == 1357810112:
                        target_pid = 9720 # DEWALT
                    elif shop_id == 326259406:
                        target_pid = 9701 # HLV
                    
                    if target_pid:
                        target_partner = self.env['res.partner'].browse(target_pid)
                        if target_partner.exists():
                            khach_hang = target_partner.name
                            ly_do_xuat = "Xuất kho bán hàng cho " + target_partner.name
                            
                            # Recalculate Partner Code for this new Customer
                            s_ref = False
                            if target_partner.commercial_partner_id and target_partner.commercial_partner_id.ref:
                                s_ref = target_partner.commercial_partner_id.ref
                            elif target_partner.parent_id and target_partner.parent_id.ref:
                                s_ref = target_partner.parent_id.ref
                            elif target_partner.ref:
                                s_ref = target_partner.ref
                            
                            if s_ref:
                                partner_code = s_ref
                            
                            _logger.info(f"SHOPEE OVERRIDE: Shop {shop_id} -> Partner {target_partner.name} (Code: {partner_code})")

        elif partner:
             khach_hang = partner.commercial_partner_id.name or partner.name

        for line in moves:
            # Determine move & product
            if line._name == 'stock.move.line':
                prod = line.product_id
                move = line.move_id
                qty = line.qty_done
                uom = line.product_uom_id
            else:
                prod = line.product_id
                move = line
                qty = line.quantity_done if hasattr(line, 'quantity_done') else line.product_uom_qty
                uom = line.product_uom
            
            if not prod: continue

            # Values
            # standard_price = prod.standard_price or 0.0
            # cost_value = standard_price * qty
            
            # --- Price Logic from Sale Order ---
            price_unit = 0.0
            tax_rate = 0.0
            price_subtotal = 0.0
            price_total = 0.0
            
            sol = getattr(move, 'sale_line_id', False)
            if sol:
                # Use Sale Order Price
                price_unit = sol.price_unit
                
                # Get Tax Rate
                if sol.tax_id:
                     # Taking the first tax as representative (common for single VAT rate)
                     tax = sol.tax_id[0]
                     tax_rate = tax.amount
                
                # Calculate Subtotal (Thành tiền) based on Exported Qty
                # Formula: Unit Price * Qty * (1 - Discount/100)
                discount_factor = 1.0 - (sol.discount or 0.0) / 100.0
                price_subtotal = price_unit * qty * discount_factor
                
                # Calculate Total After Tax
                # Taxes are calculated on price_subtotal
                price_total = price_subtotal * (1.0 + tax_rate / 100.0)
            else:
                # Fallback if no SO line (e.g. manual move)
                # Keep 0 or use Product List Price? 
                # Request says "không được lấy ở sản phẩm". 
                # Yet if no SO, we have no choice but 0 or product price.
                # Let's default to standard logic if absolutely necessary, but prompt implies SO focus.
                # We will check if we can fallback to picking Valuation if needed, 
                # but "price_unit" usually implies Sales Price.
                # Let's try to get from move price_unit if it exists and is relevant?
                # For outgoing stock moves, price_unit might be cost. 
                # So best to leave 0 if no SO to avoid "wrong product price".
                # But to be safe for display, maybe list_price as last resort?
                # "Unit price taken from sale order (not from product)".
                # Implies strictly SO. So 0.0 if no SO.
                price_unit = 0.0
                price_subtotal = 0.0
                price_total = 0.0
            
            row = {
                'loai_xuat_kho': loai_xuat,
                'ngay_hach_toan': date_hach_toan_str,
                'ngay_chung_tu': date_str,
                'so_chung_tu': self._get_misa_formatted_name(picking, so),
                'don_hang_goc': don_hang_goc,
                'ma_doi_tuong': partner_code,
                'ten_doi_tuong': partner_name,
                'khach_hang': khach_hang,
                'dia_chi': partner_address,
                'dia_chi': partner_address,
                'ly_do_xuat': ly_do_xuat,
                'ma_hang': prod.default_code or '',
                'ten_hang': prod.name,
                'la_dong_ghi_chu': 'Không',
                'hang_khuyen_mai': 'Không',
                'ma_kho': 'HLV',
                'tk_no': '632',
                'tk_co': '1561',
                'dvt': uom.name if uom else '',
                'so_luong': qty,
                'don_gia': price_unit,
                'thue_suat': tax_rate,
                'thanh_tien': price_subtotal,
                'thanh_tien_sau_thue': price_total,
                'so_lenh_sx': '',
                'ma_khoan_muc_cp': '',
                'ma_doi_tuong_thcp': '',
                'ma_don_shopee': picking.shopee_order_ref or '',

            }
            rows.append(row)
            
        return rows

    def _create_stock_export_workbook(self, pickings):
        """Tạo workbook Excel mẫu Xuất Kho"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Phiếu Xuất Kho"

        columns = self._get_stock_export_columns()

        # Styles
        header_font = Font(name='Arial', size=10, bold=True)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border_side = Side(style='thin', color='000000')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        number_alignment = Alignment(horizontal='right', vertical='center')

        # Header
        for col_idx, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_def['name']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def.get('width', 15)

        # Data
        current_row = 2
        for picking in pickings:
             rows_data = self._get_stock_export_row_data(picking)
             for row_data in rows_data:
                 for col_idx, col_def in enumerate(columns, start=1):
                     cell = ws.cell(row=current_row, column=col_idx)
                     value = row_data.get(col_def['key'], '')
                     if value is None: value = ''
                     
                     cell.value = value
                     cell.border = border
                     
                     if isinstance(value, (int, float)) and value != '':
                         cell.alignment = number_alignment
                         if 'so_luong' in col_def['key']:
                             cell.number_format = '#,##0.00'
                         elif 'tien' in col_def['key'] or 'gia' in col_def['key']:
                             cell.number_format = '#,##0'
                         elif col_def['key'] == 'thue_suat':
                             cell.number_format = '0.00'
                     else:
                         cell.alignment = cell_alignment
                 current_row += 1
        
        ws.row_dimensions[1].height = 30
        return wb

    def action_export_stock_template(self):
        """Action xuất Excel mẫu Xuất Kho"""
        self.ensure_one()
        if Workbook is None:
             raise UserError(_("Thiếu thư viện openpyxl."))

        domain = self._domain()
        
        pickings = self.env["stock.picking"].sudo().search(domain, order="scheduled_date asc, id asc")
        if not pickings:
            raise UserError(_("Không tìm thấy phiếu xuất kho trong khoảng thời gian này."))

        wb = self._create_stock_export_workbook(pickings)
        
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        filename = f"Xuat_Kho_{self.date_from}_{self.date_to}.xlsx"
        attachment = self.env["ir.attachment"].sudo().create({
            "name": filename,
            "type": "binary",
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "datas": base64.b64encode(out.getvalue()),
            "res_model": "stock.export.wizard",
            "res_id": self.id,
        })
        
        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }
