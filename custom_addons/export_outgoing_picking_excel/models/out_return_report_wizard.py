# -*- coding: utf-8 -*-
from odoo import api, fields, models, _
from odoo.exceptions import UserError
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


class OutReturnReportLine(models.Model):
    """Dòng báo cáo OUT + Trả hàng - Lưu trữ lâu dài"""
    _name = "out.return.report.line"
    _description = "Dòng báo cáo xuất kho và trả hàng"
    _order = "out_picking_date desc, id"
    _rec_name = "out_picking_name"

    # Phiếu OUT
    out_picking_id = fields.Many2one("stock.picking", string="Phiếu xuất kho", index=True)
    out_picking_name = fields.Char(string="Mã phiếu xuất", index=True)
    out_picking_origin = fields.Char(string="Mã báo giá")
    out_picking_date = fields.Datetime(string="Ngày xuất", index=True)
    partner_id = fields.Many2one("res.partner", string="Khách hàng")
    partner_name = fields.Char(string="Tên khách hàng")
    out_qty_total = fields.Float(string="Tổng SL xuất")
    
    # Phiếu trả hàng
    return_picking_id = fields.Many2one("stock.picking", string="Phiếu trả hàng")
    return_picking_name = fields.Char(string="Mã phiếu trả")
    return_picking_date = fields.Datetime(string="Ngày trả")
    has_return = fields.Boolean(string="Có trả hàng", default=True)
    return_qty_total = fields.Float(string="Tổng SL trả")
    
    # Chi tiết sản phẩm
    product_id = fields.Many2one("product.product", string="Sản phẩm")
    product_code = fields.Char(string="Mã SP")
    product_name = fields.Char(string="Tên SP")
    product_uom = fields.Char(string="ĐVT")
    out_qty = fields.Float(string="SL xuất")
    return_qty = fields.Float(string="SL trả")
    
    note = fields.Char(string="Ghi chú")

    @api.model
    def action_open_export_wizard(self):
        """Mở wizard xuất Excel"""
        return {
            'type': 'ir.actions.act_window',
            'name': 'Xuất Excel báo cáo',
            'res_model': 'out.return.report.wizard',
            'view_mode': 'form',
            'target': 'new',
        }

    @api.model
    def _cron_update_report_data(self):
        """Cron job: Cập nhật dữ liệu báo cáo hàng ngày"""
        # Lấy ngày 3 tháng trước để quét
        date_from = fields.Date.today() - relativedelta(months=3)
        self._generate_report_data(date_from, fields.Date.today())

    @api.model
    def _generate_report_data(self, date_from, date_to, warehouse_ids=None):
        """Tạo/cập nhật dữ liệu báo cáo cho khoảng thời gian"""
        domain = [
            ("date_done", ">=", date_from),
            ("date_done", "<=", date_to),
            ("picking_type_id.code", "=", "outgoing"),
            ("state", "=", "done"),
        ]
        
        if warehouse_ids:
            domain.append(("picking_type_id.warehouse_id", "in", warehouse_ids))
        
        out_pickings = self.env["stock.picking"].sudo().search(domain, order="date_done desc")
        
        for out_picking in out_pickings:
            # Tính max_date = ngày OUT + 1 tháng
            max_date = out_picking.date_done + relativedelta(months=1)
            
            # Tìm phiếu trả hàng
            return_pickings = self._find_return_pickings(out_picking, max_date)
            
            if not return_pickings:
                continue  # Chỉ lưu các phiếu có return
            
            # Tính tổng SL xuất
            out_qty_total = sum(out_picking.move_ids.mapped(
                lambda m: m.quantity_done if hasattr(m, 'quantity_done') else m.product_uom_qty
            ))
            
            for return_picking in return_pickings:
                return_qty_total = sum(return_picking.move_ids.mapped(
                    lambda m: m.quantity_done if hasattr(m, 'quantity_done') else m.product_uom_qty
                ))
                
                for move in return_picking.move_ids:
                    product = move.product_id
                    return_qty = move.quantity_done if hasattr(move, 'quantity_done') else move.product_uom_qty
                    out_qty = self._get_product_qty_in_picking(out_picking, product.id)
                    
                    # Kiểm tra xem đã có record chưa
                    existing = self.search([
                        ('out_picking_id', '=', out_picking.id),
                        ('return_picking_id', '=', return_picking.id),
                        ('product_id', '=', product.id),
                    ], limit=1)
                    
                    vals = {
                        'out_picking_id': out_picking.id,
                        'out_picking_name': out_picking.name,
                        'out_picking_origin': out_picking.origin or '',
                        'out_picking_date': out_picking.date_done,
                        'partner_id': out_picking.partner_id.id if out_picking.partner_id else False,
                        'partner_name': out_picking.partner_id.name if out_picking.partner_id else '',
                        'out_qty_total': out_qty_total,
                        'return_picking_id': return_picking.id,
                        'return_picking_name': return_picking.name,
                        'return_picking_date': return_picking.date_done,
                        'has_return': True,
                        'return_qty_total': return_qty_total,
                        'product_id': product.id,
                        'product_code': product.default_code or '',
                        'product_name': product.name,
                        'product_uom': move.product_uom.name if move.product_uom else '',
                        'out_qty': out_qty,
                        'return_qty': return_qty,
                    }
                    
                    if existing:
                        existing.write(vals)
                    else:
                        self.create(vals)

    @api.model
    def _find_return_pickings(self, out_picking, max_date):
        """Tìm các phiếu trả hàng liên kết với phiếu OUT"""
        return_pickings = self.env["stock.picking"]
        
        # Phương pháp 1: Qua move_dest_ids
        for move in out_picking.move_ids:
            for dest_move in move.move_dest_ids:
                if dest_move.picking_id and dest_move.picking_id.state == 'done':
                    picking = dest_move.picking_id
                    if picking.picking_type_id.code == 'incoming':
                        if picking.date_done and picking.date_done <= max_date:
                            return_pickings |= picking
        
        # Phương pháp 2: Tìm theo origin
        origin_returns = self.env["stock.picking"].search([
            ("origin", "ilike", out_picking.name),
            ("picking_type_id.code", "=", "incoming"),
            ("state", "=", "done"),
            ("date_done", "<=", max_date),
        ])
        return_pickings |= origin_returns
        
        # Phương pháp 3: Tìm theo partner và sản phẩm
        if out_picking.partner_id:
            product_ids = out_picking.move_ids.mapped('product_id').ids
            if product_ids:
                related_returns = self.env["stock.picking"].search([
                    ("partner_id", "=", out_picking.partner_id.id),
                    ("picking_type_id.code", "=", "incoming"),
                    ("state", "=", "done"),
                    ("date_done", ">", out_picking.date_done),
                    ("date_done", "<=", max_date),
                    ("move_ids.product_id", "in", product_ids),
                ])
                return_pickings |= related_returns
        
        return return_pickings

    @api.model
    def _get_product_qty_in_picking(self, picking, product_id):
        """Lấy số lượng của 1 sản phẩm trong picking"""
        qty = 0.0
        for move in picking.move_ids:
            if move.product_id.id == product_id:
                qty += move.quantity_done if hasattr(move, 'quantity_done') else move.product_uom_qty
        return qty


class OutReturnReportWizard(models.TransientModel):
    """Wizard xuất Excel báo cáo OUT và trả hàng"""
    _name = "out.return.report.wizard"
    _description = "Xuất Excel báo cáo xuất kho và trả hàng"

    date_from = fields.Date(string="Từ ngày", required=True, default=lambda self: fields.Date.today() - relativedelta(months=1))
    date_to = fields.Date(string="Đến ngày", required=True, default=fields.Date.today)
    
    warehouse_ids = fields.Many2many(
        "stock.warehouse", string="Kho",
        help="Để trống để lấy tất cả kho"
    )

    def action_export_excel(self):
        """Xuất báo cáo ra file Excel"""
        self.ensure_one()
        
        if Workbook is None:
            raise UserError(_("Thiếu thư viện openpyxl."))
        
        if self.date_from > self.date_to:
            raise UserError(_("Khoảng ngày không hợp lệ."))
        
        # Lấy dữ liệu từ out.return.report.line
        domain = [
            ("out_picking_date", ">=", self.date_from),
            ("out_picking_date", "<=", self.date_to),
        ]
        if self.warehouse_ids:
            domain.append(("out_picking_id.picking_type_id.warehouse_id", "in", self.warehouse_ids.ids))
        
        lines = self.env["out.return.report.line"].search(domain, order="out_picking_date desc")
        
        if not lines:
            raise UserError(_("Không tìm thấy dữ liệu trong khoảng thời gian này."))
        
        wb = Workbook()
        
        # Styles
        header_font = Font(name='Arial', size=10, bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        date_alignment = Alignment(horizontal='center', vertical='center')
        border_side = Side(style='thin', color='000000')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # ===== SHEET 1: CHI TIẾT PHIẾU XUẤT (OUT) =====
        ws1 = wb.active
        ws1.title = "Chi tiết phiếu xuất"
        
        columns1 = [
            {'name': 'STT', 'width': 5},
            {'name': 'Mã phiếu xuất', 'width': 15},
            {'name': 'Mã báo giá', 'width': 15},
            {'name': 'Ngày xuất', 'width': 15},
            {'name': 'Khách hàng', 'width': 25},
            {'name': 'Mã SP', 'width': 15},
            {'name': 'Tên SP', 'width': 40},
            {'name': 'ĐVT', 'width': 8},
            {'name': 'SL xuất', 'width': 10},
            {'name': 'Có trả lại', 'width': 10},
        ]
        
        # Header Sheet 1
        for col_idx, col_def in enumerate(columns1, start=1):
            cell = ws1.cell(row=1, column=col_idx)
            cell.value = col_def['name']
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
            ws1.column_dimensions[get_column_letter(col_idx)].width = col_def['width']
        
        # Data Sheet 1
        out_picking_ids = lines.mapped('out_picking_id')
        row_idx = 2
        stt = 1
        
        for out_picking in out_picking_ids:
            for move in out_picking.move_ids:
                has_return = any(l.out_picking_id.id == out_picking.id and l.has_return for l in lines)
                
                ws1.cell(row=row_idx, column=1).value = stt
                ws1.cell(row=row_idx, column=2).value = out_picking.name
                ws1.cell(row=row_idx, column=3).value = out_picking.origin or ''
                ws1.cell(row=row_idx, column=4).value = out_picking.date_done.strftime('%d/%m/%Y') if out_picking.date_done else ''
                ws1.cell(row=row_idx, column=4).alignment = date_alignment
                ws1.cell(row=row_idx, column=5).value = out_picking.partner_id.name or ''
                ws1.cell(row=row_idx, column=6).value = move.product_id.default_code or ''
                ws1.cell(row=row_idx, column=7).value = move.product_id.name
                ws1.cell(row=row_idx, column=8).value = move.product_uom.name
                ws1.cell(row=row_idx, column=9).value = move.quantity_done if hasattr(move, 'quantity_done') else move.product_uom_qty
                ws1.cell(row=row_idx, column=10).value = 'Có' if has_return else 'Không'
                
                for col in range(1, 11):
                    ws1.cell(row=row_idx, column=col).border = border
                
                row_idx += 1
                stt += 1

        # ===== SHEET 2: CHI TIẾT PHIẾU TRẢ (RETURN) =====
        ws2 = wb.create_sheet("Chi tiết phiếu trả")
        
        columns2 = [
            {'name': 'STT', 'width': 5},
            {'name': 'Mã phiếu xuất gốc', 'width': 15},
            {'name': 'Mã phiếu trả', 'width': 15},
            {'name': 'Ngày trả', 'width': 15},
            {'name': 'Khách hàng', 'width': 25},
            {'name': 'Mã SP', 'width': 15},
            {'name': 'Tên SP', 'width': 40},
            {'name': 'ĐVT', 'width': 8},
            {'name': 'SL trả', 'width': 10},
        ]
        
        # Header Sheet 2
        for col_idx, col_def in enumerate(columns2, start=1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.value = col_def['name']
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
            ws2.column_dimensions[get_column_letter(col_idx)].width = col_def['width']
        
        # Data Sheet 2
        row_idx = 2
        stt = 1
        sorted_lines = lines.sorted(key=lambda r: r.return_picking_date or fields.Datetime.now(), reverse=True)
        
        for line in sorted_lines:
            if not line.has_return:
                continue
                
            ws2.cell(row=row_idx, column=1).value = stt
            ws2.cell(row=row_idx, column=2).value = line.out_picking_name
            ws2.cell(row=row_idx, column=3).value = line.return_picking_name
            ws2.cell(row=row_idx, column=4).value = line.return_picking_date.strftime('%d/%m/%Y') if line.return_picking_date else ''
            ws2.cell(row=row_idx, column=4).alignment = date_alignment
            ws2.cell(row=row_idx, column=5).value = line.partner_name
            ws2.cell(row=row_idx, column=6).value = line.product_code
            ws2.cell(row=row_idx, column=7).value = line.product_name
            ws2.cell(row=row_idx, column=8).value = line.product_uom
            ws2.cell(row=row_idx, column=9).value = line.return_qty
            
            for col in range(1, 10):
                ws2.cell(row=row_idx, column=col).border = border
            
            row_idx += 1
            stt += 1
        
        # Xuất file
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        filename = f"BaoCao_XuatKho_TraHang_{self.date_from}_{self.date_to}.xlsx"
        attachment = self.env["ir.attachment"].sudo().create({
            "name": filename,
            "type": "binary",
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "datas": base64.b64encode(out.getvalue()),
            "res_model": "out.return.report.wizard",
            "res_id": self.id,
        })
        
        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }
