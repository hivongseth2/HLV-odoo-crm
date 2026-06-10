# -*- coding: utf-8 -*-
"""
split_sale_orders_by_customer_code_from_excel.py
================================================
Move sale orders to the correct root customer by an Excel mapping.

Use case:
  - Column B: sale order number.
  - Column I: MISA customer code.
  - Orders were incorrectly merged into DONGJIN-BS.
  - Rows whose customer code is DONGJINTEXTILE must be moved back to the
    DONGJINTEXTILE root partner.

Run in odoo shell:
    $env:SALE_ORDER_SPLIT_XLSX="D:\\path\\orders.xlsx"
    odoo-bin shell -c <odoo.conf> --no-http < bin/split_sale_orders_by_customer_code_from_excel.py

Review output first with DRY_RUN=True. Set DRY_RUN=False to write.
"""

import os
import re


DRY_RUN = True

EXCEL_PATH = os.environ.get("SALE_ORDER_SPLIT_XLSX", "").strip()
SHEET_NAME = os.environ.get("SALE_ORDER_SPLIT_SHEET", "").strip() or None

HEADER_ROW = 1
ORDER_COL = 2          # B
CUSTOMER_CODE_COL = 9  # I

# Safety guards for the current Dongjin recovery.
SOURCE_CODE = "DONGJIN-BS"
TARGET_CODES = {"DONGJINTEXTILE"}

UPDATE_SALE_ORDER = True
UPDATE_PICKINGS = True
UPDATE_DRAFT_INVOICES = True
UPDATE_POSTED_INVOICES = False

SEP = "=" * 100
SEP2 = "-" * 100


def section(title):
    print("\n%s\n  %s\n%s" % (SEP, title, SEP))


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def norm_code(value):
    return clean(value).upper().replace(" ", "")


def partner_codes(partner):
    return set(filter(None, [
        norm_code(partner.ref),
        norm_code(partner.company_registry),
    ]))


def has_code(partner, code):
    code = norm_code(code)
    if not partner:
        return False
    candidates = partner
    if partner.commercial_partner_id:
        candidates |= partner.commercial_partner_id
    return any(code in partner_codes(p) for p in candidates)


def or_domain(clauses):
    if not clauses:
        return []
    if len(clauses) == 1:
        return clauses
    return ["|"] * (len(clauses) - 1) + clauses


def split_order_values(value):
    value = clean(value)
    if not value:
        return []
    parts = re.split(r"[\n,;]+", value)
    return [part.strip() for part in parts if part.strip()]


def load_excel_rows():
    if not EXCEL_PATH:
        raise Exception("Missing SALE_ORDER_SPLIT_XLSX environment variable.")
    if not os.path.exists(EXCEL_PATH):
        raise Exception("Excel file not found: %s" % EXCEL_PATH)

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise Exception("Missing python package openpyxl in this Odoo environment.")

    workbook = load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    sheet = workbook[SHEET_NAME] if SHEET_NAME else workbook.active

    rows = []
    for row_index, row in enumerate(sheet.iter_rows(), start=1):
        if row_index <= HEADER_ROW:
            continue
        order_cell = row[ORDER_COL - 1].value if len(row) >= ORDER_COL else None
        code_cell = row[CUSTOMER_CODE_COL - 1].value if len(row) >= CUSTOMER_CODE_COL else None
        code = clean(code_cell)
        for order_no in split_order_values(order_cell):
            if order_no or code:
                rows.append({
                    "row": row_index,
                    "order_no": order_no,
                    "customer_code": code,
                })
    return rows


def find_root_partner_by_code(code):
    code = clean(code)
    if not code:
        return env["res.partner"]

    Partner = env["res.partner"].sudo().with_context(active_test=False)
    partners = Partner.search([
        ("parent_id", "=", False),
        "|",
        ("ref", "=", code),
        ("company_registry", "=", code),
    ], order="active desc, is_company desc, id asc")
    partners = partners.filtered(lambda p: norm_code(code) in partner_codes(p))

    if not partners:
        partners = Partner.search([
            ("parent_id", "=", False),
            "|",
            ("ref", "!=", False),
            ("company_registry", "!=", False),
        ], order="active desc, is_company desc, id asc").filtered(
            lambda p: norm_code(code) in partner_codes(p)
        )

    active_partners = partners.filtered(lambda p: p.active)
    return (active_partners or partners)[:1]


def find_sale_order(order_no):
    SaleOrder = env["sale.order"].sudo()
    clauses = [("name", "=", order_no)]
    for field_name in ("client_order_ref", "origin", "misa_order_no", "x_studio_misa_order_no"):
        if field_name in SaleOrder._fields:
            clauses.append((field_name, "=", order_no))
    return SaleOrder.search(or_domain(clauses), order="id asc")


def matching_child_for_shipping(source_shipping, target_partner):
    if not source_shipping or source_shipping == target_partner:
        return target_partner
    if source_shipping.parent_id != source_shipping.commercial_partner_id:
        # This is not a normal child address. Keep the target root.
        return target_partner

    Partner = env["res.partner"].sudo().with_context(active_test=False)
    domain = [
        ("parent_id", "=", target_partner.id),
        ("type", "=", source_shipping.type or "delivery"),
        ("name", "=", source_shipping.name or target_partner.name),
    ]
    if source_shipping.street:
        domain.append(("street", "=", source_shipping.street))
    existing = Partner.search(domain, order="active desc, id asc", limit=1)
    if existing:
        return existing

    vals = {
        "parent_id": target_partner.id,
        "type": source_shipping.type or "delivery",
        "name": source_shipping.name or target_partner.name,
        "street": source_shipping.street,
        "street2": source_shipping.street2,
        "city": source_shipping.city,
        "zip": source_shipping.zip,
        "state_id": source_shipping.state_id.id if source_shipping.state_id else False,
        "country_id": source_shipping.country_id.id if source_shipping.country_id else False,
        "phone": source_shipping.phone,
        "mobile": source_shipping.mobile,
        "email": source_shipping.email,
    }
    if DRY_RUN:
        print("       DRY create shipping child under target: %s" % vals)
        return target_partner
    return Partner.create(vals)


def describe_partner(partner):
    return "id=%s name=%s ref=%r company_registry=%r vat=%r" % (
        partner.id,
        partner.display_name,
        partner.ref,
        partner.company_registry,
        partner.vat,
    )


section("SPLIT SALE ORDERS BY EXCEL CUSTOMER CODE")
print("  DRY_RUN=%s" % DRY_RUN)
print("  EXCEL_PATH=%s" % (EXCEL_PATH or "(missing)"))
print("  SOURCE_CODE=%s" % SOURCE_CODE)
print("  TARGET_CODES=%s" % sorted(TARGET_CODES))

target_partner_by_code = {}
for code in TARGET_CODES:
    partner = find_root_partner_by_code(code)
    if not partner:
        raise Exception("Target partner not found for code %s" % code)
    target_partner_by_code[norm_code(code)] = partner
    print("  Target %s -> %s" % (code, describe_partner(partner)))

source_partner = find_root_partner_by_code(SOURCE_CODE)
if not source_partner:
    raise Exception("Source partner not found for code %s" % SOURCE_CODE)
print("  Source %s -> %s" % (SOURCE_CODE, describe_partner(source_partner)))

rows = load_excel_rows()
selected_rows = [
    row for row in rows
    if norm_code(row["customer_code"]) in {norm_code(c) for c in TARGET_CODES}
]
order_numbers = []
seen_orders = set()
for row in selected_rows:
    key = row["order_no"]
    if key not in seen_orders:
        seen_orders.add(key)
        order_numbers.append(key)

section("INPUT")
print("  Excel rows read       : %s" % len(rows))
print("  Rows matching targets : %s" % len(selected_rows))
print("  Unique sale orders    : %s" % len(order_numbers))

stats = {
    "moved": 0,
    "already_target": 0,
    "missing_order": 0,
    "multiple_order": 0,
    "wrong_source": 0,
    "invoice_skipped": 0,
    "errors": 0,
}

section("PROCESS")
for order_no in order_numbers:
    row = next(r for r in selected_rows if r["order_no"] == order_no)
    target_code = norm_code(row["customer_code"])
    target = target_partner_by_code[target_code]

    print("\n  %s" % SEP2)
    print("  Row=%s SO=%s target_code=%s" % (row["row"], order_no, row["customer_code"]))
    orders = find_sale_order(order_no)
    if not orders:
        stats["missing_order"] += 1
        print("  [SKIP] Sale order not found")
        continue
    if len(orders) > 1:
        stats["multiple_order"] += 1
        print("  [SKIP] Multiple sale orders found: %s" % orders.ids)
        continue

    order = orders[0]
    current_partner = order.partner_id.commercial_partner_id or order.partner_id
    print("  Current partner: %s" % describe_partner(current_partner))
    print("  Target partner : %s" % describe_partner(target))

    if current_partner.id == target.id:
        stats["already_target"] += 1
        print("  OK already on target")
        continue

    if SOURCE_CODE and not has_code(current_partner, SOURCE_CODE):
        stats["wrong_source"] += 1
        print("  [SKIP] Current partner is not source code %s" % SOURCE_CODE)
        continue

    try:
        shipping_target = target
        if "partner_shipping_id" in order._fields and order.partner_shipping_id:
            shipping_target = matching_child_for_shipping(order.partner_shipping_id, target)

        vals = {}
        if UPDATE_SALE_ORDER:
            vals["partner_id"] = target.id
            if "partner_invoice_id" in order._fields:
                vals["partner_invoice_id"] = target.id
            if "partner_shipping_id" in order._fields:
                vals["partner_shipping_id"] = shipping_target.id

        print("  Sale order vals: %s" % vals)
        if vals and not DRY_RUN:
            order.write(vals)
            order.message_post(body=(
                "Moved sale order partner from %s to %s by Excel customer code %s."
            ) % (current_partner.display_name, target.display_name, row["customer_code"]))

        if UPDATE_PICKINGS and "picking_ids" in order._fields:
            pickings = order.picking_ids.filtered(lambda p: p.state != "cancel")
            print("  Pickings to update: %s" % pickings.mapped("name"))
            if pickings and not DRY_RUN:
                pickings.write({"partner_id": target.id})

        if "invoice_ids" in order._fields:
            invoices = order.invoice_ids.filtered(lambda m: m.state != "cancel")
            draft_invoices = invoices.filtered(lambda m: m.state != "posted")
            posted_invoices = invoices - draft_invoices
            if posted_invoices and not UPDATE_POSTED_INVOICES:
                stats["invoice_skipped"] += len(posted_invoices)
                print("  [WARN] Posted invoices not updated: %s" % posted_invoices.mapped("name"))
            invoice_to_update = draft_invoices if UPDATE_DRAFT_INVOICES else env["account.move"]
            if UPDATE_POSTED_INVOICES:
                invoice_to_update |= posted_invoices
            if invoice_to_update:
                print("  Invoices to update: %s" % invoice_to_update.mapped("name"))
                if not DRY_RUN:
                    invoice_to_update.write({"partner_id": target.id})

        stats["moved"] += 1
        print("  %s" % ("DRY moved" if DRY_RUN else "MOVED"))
    except Exception as exc:
        stats["errors"] += 1
        print("  [ERROR] %s" % exc)

if not DRY_RUN:
    env.cr.commit()

section("SUMMARY")
for key in sorted(stats):
    print("  %-16s: %s" % (key, stats[key]))
print("  Commit          : %s" % ("NO (DRY_RUN)" if DRY_RUN else "YES"))
print("\n  Done.")
